from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()
ws = wb.active
ws.title = "생산처방"

# Row 1-2 merges
ws.merge_cells('A1:F2')
ws['A1'] = '생산지시서'
ws['A1'].font = Font(size=14, bold=True)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws.merge_cells('G1:H2')
ws['G1'] = '결재방'
ws['G1'].alignment = Alignment(horizontal='center', vertical='center')

# Row 3
ws.merge_cells('A3:D3')
ws['A3'] = '세럼'
ws['A3'].alignment = Alignment(horizontal='center')
ws['E3'] = '생산코드'
ws['F3'] = 'BA112147'
ws['G3'] = '제조자'

# Row 4
ws['A4'] = '지시일'
ws['B4'] = '2025-11-02 14:18'
ws['C4'] = '적용일'
ws['D4'] = '2025-11-01'
ws['E4'] = '생산량(kg)'
ws['F4'] = '1.0 kg'
ws['G4'] = '수득량'

# Row 5 empty

# Row 6 headers
ws['A6'] = 'Ph'
ws['B6'] = '구분'
ws['C6'] = '코드'
ws['D6'] = '원료명'
ws['E6'] = '함량(%)'
ws['F6'] = '생산량(kg)'
ws['G6'] = '제조공정'
ws['H6'] = '공정검사'

# A phase rows 7-10
ws['A7'] = 'A'
ws.merge_cells('A7:A10')
ws['A7'].alignment = Alignment(vertical='center', horizontal='center')
ws['B7'] = '1'
ws['C7'] = '3036'
ws['D7'] = 'Zemea Select Propanediol'
ws['E7'] = '3.4'
ws['F7'] = '0.034'
ws.merge_cells('G7:G10')
ws['G7'] = '공정1. 수상 용해 (A상)\nA상을 유화조에 투입하여 완전히 용해.\n* H/M : 1500rpm\n* P/M : 20rpm\n시간 :\n온도 :             ℃\nH/M:              rpm\nP/M:              rpm'
ws['G7'].alignment = Alignment(wrap_text=True, vertical='top')
ws.merge_cells('H7:H10')
ws['H7'].alignment = Alignment(vertical='center')

# Row 8
ws['B8'] = '2'
ws['C8'] = '3001'
ws['D8'] = 'Glycerin'
ws['E8'] = '1.5'
ws['F8'] = '0.015'

# Row 9
ws['B9'] = '3'
ws['C9'] = '7094'
ws['D9'] = 'Peptosome Neo'
ws['E9'] = '60.08'
ws['F9'] = '0.6008'

# Row 10
ws['B10'] = '4'
ws['C10'] = '7095'
ws['D10'] = 'Glucan 3M'
ws['E10'] = '35'
ws['F10'] = '0.35'

# B phase rows 11-12
ws['A11'] = 'B'
ws.merge_cells('A11:A12')
ws['A11'].alignment = Alignment(vertical='center', horizontal='center')
ws['B11'] = '5'
ws['C11'] = '1052'
ws['D11'] = 'TREHA'
ws['E11'] = '0.01'
ws['F11'] = '0.0001'
ws.merge_cells('G11:G12')
ws['G11'] = '공정2. 점증제 분산 (B상)\nB상을 별도의 용기에서 디스퍼믹서 10분간 팽윤시킨 후 유화조에 서서히 투입 하여 75℃ 가온하면서 교반.\n* 75℃, 5분\n* H/M : 2000rpm\n* P/M : 20rpm\n공정3. 45℃ 냉각\n시간 :\n온도 :             ℃\nH/M:              rpm'
ws['G11'].alignment = Alignment(wrap_text=True, vertical='top')
ws.merge_cells('H11:H12')
ws['H11'].alignment = Alignment(vertical='center')

# Row 12
ws['B12'] = '6'
ws['C12'] = '7096'
ws['D12'] = 'Hydrolyzed Sodium Hyaluronate'
ws['E12'] = '0.01'
ws['F12'] = '0.0001'

# Row 13
ws['A13'] = '합계 (Total)'
ws.merge_cells('A13:D13')
ws['A13'].alignment = Alignment(horizontal='center')
ws['E13'] = '100'

# Auto-adjust columns (병합된 셀 처리)
for col_idx in range(1, ws.max_column + 1):
    column = get_column_letter(col_idx)
    max_length = 0
    
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        # 병합된 셀이 아닌 경우만 처리
        if cell.value and not isinstance(cell, type(ws['A1']).__bases__[0]):
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
    
    # 최소 10, 최대 50으로 제한
    adjusted_width = min(max(max_length + 2, 10), 50)
    ws.column_dimensions[column].width = adjusted_width

# 행 높이 조정
ws.row_dimensions[1].height = 30
ws.row_dimensions[7].height = 120
ws.row_dimensions[11].height = 120

# Save
filename = f'생산처방_세럼_A_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
wb.save(filename)
print(f"생성 완료: {filename}")