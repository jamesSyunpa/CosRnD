import customtkinter as ctk
from openpyxl import load_workbook
import os
from tkinter import messagebox
import tkinter.filedialog as fd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

class FinishedProductReportFrame(ctk.CTkFrame):
    """완제품 시험성적서 UI 클래스 (엑셀 내용 자동 반영)"""
    
    def __init__(self, master, excel_path="_완제품 시험성적서.xlsx"):
        super().__init__(master)
        self.excel_path = excel_path
        
        # 기본 프레임 구조
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        
        # --- [1] 상단 기본 정보 ---
        info_frame = ctk.CTkFrame(self, fg_color="transparent")
        info_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 5))
        info_frame.grid_columnconfigure((1, 3), weight=1) # Entry 컬럼들이 확장되도록 설정
        
        self.entries = {}
        
        # 요청하신 레이아웃에 맞춰 라벨과 입력 필드 재구성
        info_items = [
            ("제 품 명", 0, 0, 3),
            ("반제품 제조일자", 1, 0, 1), ("반제품제조번호(LOT)", 1, 2, 1),
            ("포장 일자", 2, 0, 1), ("완제품제조번호(LOT)", 2, 2, 1),
            ("사용 기한", 3, 0, 1), ("단위 용량 (ml)", 3, 2, 1),
            ("샘플링 방법", 4, 0, 1), ("시험 일자", 4, 2, 1),
        ]

        for text, r, c, colspan in info_items:
            # 라벨 생성
            label = ctk.CTkLabel(info_frame, text=text, font=ctk.CTkFont(weight="bold"))
            label.grid(row=r, column=c, padx=10, pady=5, sticky="w")
            
            # 입력 필드(Entry) 생성
            entry = ctk.CTkEntry(info_frame)
            entry.grid(row=r, column=c + 1, columnspan=colspan, padx=10, pady=5, sticky="ew")
            self.entries[text] = entry

        
        # --- [2] 내용목록(시험항목 테이블) ---
        self.table_scroll_frame = ctk.CTkScrollableFrame(self, label_text="내용 목록")
        self.table_scroll_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
        self.table_scroll_frame.grid_columnconfigure(1, weight=2)
        self.table_scroll_frame.grid_columnconfigure(2, weight=3)
        self.table_scroll_frame.grid_columnconfigure(3, weight=3)
        self.table_scroll_frame.grid_columnconfigure(4, weight=1)
        self.table_scroll_frame.grid_columnconfigure(5, weight=0)
        
        headers = ["구분", "시험항목", "시험기준", "시험결과", "비고", "삭제"]
        for i, h in enumerate(headers):
            ctk.CTkLabel(self.table_scroll_frame, text=h, font=ctk.CTkFont(weight="bold")).grid(row=0, column=i, sticky="ew", padx=2, pady=2)
        
        self.test_item_rows = []
        self._create_initial_rows()
        
        # --- [3] 판정/시험자 ---
        conclusion_frame = ctk.CTkFrame(self, fg_color="transparent")
        conclusion_frame.grid(row=2, column=0, sticky="ew", padx=10, pady=5)
        conclusion_frame.grid_columnconfigure((1, 3, 5), weight=1)
        
        # '일자' 항목 추가 및 순서 변경: 시험자, 검토자, 종합판정
        for idx, (label, key) in enumerate([("시험자", "시험자"), ("검토자", "검토자"), ("종합판정", "종합판정")]):
            ctk.CTkLabel(conclusion_frame, text=label, font=ctk.CTkFont(weight="bold")).grid(row=0, column=idx*2, padx=10, pady=5, sticky="w")
            self.entries[key] = ctk.CTkEntry(conclusion_frame)
            self.entries[key].grid(row=0, column=idx*2+1, padx=10, pady=5, sticky="ew")
        
        # --- [4] 버튼 ---
        button_frame = ctk.CTkFrame(self, fg_color="transparent")
        button_frame.grid(row=3, column=0, sticky="ew", padx=10, pady=10)
        
        ctk.CTkButton(button_frame, text="행 추가", command=self._add_row).pack(side="left", padx=5)
        right = ctk.CTkFrame(button_frame, fg_color="transparent")
        right.pack(side="right")
        ctk.CTkButton(right, text="엑셀 보고서 생성", command=self.generate_report, fg_color="#3B8ED0").pack(side="left", padx=5)
        ctk.CTkButton(right, text="초기화", command=self.clear_form, fg_color="gray50").pack(side="left", padx=5)

    def _create_initial_rows(self):
        """초기 시험 항목 행들을 생성합니다."""
        initial_items = [
            {"id": "1", "item": "성 상", "spec": "유백색 크림상", "result": "", "note": "관능검사 / 표준품과 비교"},
            {"id": "2", "item": "향 취", "spec": "표준품과 일치", "result": "", "note": "관능검사 / 표준품과 비교"},
            {"id": "3", "item": "이물질", "spec": "미발견", "result": "", "note": "관능검사 / 표준품과 비교"},
            {"id": "4", "item": "내용량", "spec": "표시량의 97% 이상", "result": "", "note": "화장품 기준 및 시험방법 중 2.시험방법 1)내용량 가)용량으로 표시된 제품에 따른다"},
            {"id": "5", "item": "사용감", "spec": "표준품과 일치", "result": "", "note": "관능검사 / 표준품과 비교"},
            {"id": "6", "item": "pH(25℃)", "spec": "5.80 ± 1.00", "result": "", "note": "pH Meter 사용"},
            {"id": "7", "item": "점도(25℃)", "spec": "19,000 ± 4,000", "result": "", "note": "Helipath, 50rpm, Spindle-E, 1min"},
            {"id": "8", "item": "비중(25℃)", "spec": "1.010 ± 0.050", "result": "", "note": "비중병 사용"},
            {"id": "9", "item": "외관 및 용기상태", "spec": "표준품과 비교", "result": "", "note": "관능검사 / 표준품과 비교"},
            {"id": "10", "item": "인쇄상태", "spec": "육안 식별", "result": "", "note": "관능검사 / 표준품과 비교"},
            {"id": "11", "item": "법적 표시사항", "spec": "법적 규정에 적합", "result": "", "note": "화장품법 및 화장품법 시행규칙 참고"},
            {"id": "12", "item": "아데노신 함량 시험", "spec": "표시량의 90% 이상 (0.04%)", "result": "", "note": "기능성화장품 기준 및 시험방법 참고"},
            {"id": "13", "item": "아데노신 확인 시험", "spec": "검체와 표준액의 RT가 같다", "result": "", "note": "기능성화장품 기준 및 시험방법 참고"},
            {"id": "14", "item": "미생물(일반세균)", "spec": "100 cfu/ml 이하", "result": "", "note": "화장품의 미생물 한도 기준을 참고. 3M™ Petrifilm Plate법 사용"},
            {"id": "", "item": "미생물(진균(효모/곰팡이))", "spec": "10 cfu/ml 이하", "result": "", "note": ""},
            {"id": "", "item": "미생물(대장균)", "spec": "불검출", "result": "", "note": ""},
            {"id": "특이사항", "item": "", "spec": "", "result": "", "note": ""},
        ]
        for item_data in initial_items:
            self._add_row(item_data, is_initial=True)
    
    # --- 행 추가 ---
    def _add_row(self, data=None, is_initial=False):
        if data is None:
            data = {"id": "", "item": "", "spec": "", "result": "", "note": ""}
        row_index = len(self.test_item_rows) + 1
        row_widgets = {"id_label": ctk.CTkLabel(self.table_scroll_frame, text=data.get("id", ""))}
        
        for key in ["item", "spec", "result", "note"]:
            e = ctk.CTkEntry(self.table_scroll_frame, corner_radius=0, border_width=0)
            e.insert(0, data.get(key, ""))
            row_widgets[f"{key}_entry"] = e
        
        rm_btn = ctk.CTkButton(self.table_scroll_frame, text="삭제", width=40, fg_color="gray50", hover_color="gray35",
                               command=lambda r=row_widgets: self._remove_row(r))
        row_widgets["remove_button"] = rm_btn
        
        self.test_item_rows.append(row_widgets)
        self._grid_row_widgets(row_widgets, row_index)
        if not is_initial:
            self._redraw_table()
    
    # --- 행 배치 ---
    def _grid_row_widgets(self, row_widgets, row_index):
        row_widgets["id_label"].grid(row=row_index, column=0, sticky="ew", padx=2, pady=2)
        row_widgets["item_entry"].grid(row=row_index, column=1, sticky="ew", padx=2, pady=2)
        row_widgets["spec_entry"].grid(row=row_index, column=2, sticky="ew", padx=2, pady=2)
        row_widgets["result_entry"].grid(row=row_index, column=3, sticky="ew", padx=2, pady=2)
        row_widgets["note_entry"].grid(row=row_index, column=4, sticky="ew", padx=2, pady=2)
        row_widgets["remove_button"].grid(row=row_index, column=5, padx=5, pady=2)
    
    # --- 행 삭제 및 재정렬 ---
    def _remove_row(self, row):
        for w in row.values():
            w.destroy()
        self.test_item_rows.remove(row)
        self._redraw_table()
    
    def _redraw_table(self):
        for i, row in enumerate(self.test_item_rows, start=1):
            row["id_label"].configure(text=str(i))
            self._grid_row_widgets(row, i)

    def clear_form(self):
        """폼의 모든 입력 필드를 초기화하고 테이블을 기본값으로 되돌립니다."""
        # 상단 정보 초기화
        for entry in self.entries.values():
            entry.delete(0, "end")
        
        # 테이블의 모든 동적 행 제거
        for row_widgets in self.test_item_rows[:]:
            for widget in row_widgets.values():
                widget.destroy()
        self.test_item_rows.clear()
            
        # 초기 행 다시 생성
        self._create_initial_rows()
        
        messagebox.showinfo("알림", "양식이 초기화되었습니다.", parent=self)

    def generate_report(self):
        """입력된 데이터를 기반으로 완제품 시험성적서 엑셀 파일을 생성합니다."""
        try:
            # --- 데이터 수집 ---
            info_data = {key: entry.get() for key, entry in self.entries.items()}
            
            if not all(info_data.get(key) for key in ["제 품 명", "완제품제조번호(LOT)", "종합판정"]):
                messagebox.showwarning("입력 오류", "제품명, 완제품 LOT, 종합판정은 필수 입력 항목입니다.", parent=self)
                return

            test_items = []
            for row_widgets in self.test_item_rows:
                item_id = row_widgets["id_label"].cget("text")
                item = row_widgets["item_entry"].get()
                spec = row_widgets["spec_entry"].get()
                result = row_widgets["result_entry"].get()
                note = row_widgets["note_entry"].get()
                test_items.append({"id": item_id, "item": item, "spec": spec, "result": result, "note": note})

            # --- Excel 워크북 및 스타일 정의 ---
            wb = Workbook()
            ws = wb.active
            ws.title = "완제품 시험성적서"

            # 스타일
            title_font = Font(name='맑은 고딕', size=18, bold=True)
            header_font = Font(name='맑은 고딕', size=11, bold=True, color="FFFFFF")
            label_font = Font(name='맑은 고딕', size=10, bold=True)
            cell_font = Font(name='맑은 고딕', size=10)
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_align = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            label_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

            def apply_style(cell, font=None, border=None, fill=None, alignment=None):
                if font: cell.font = font
                if border: cell.border = border
                if fill: cell.fill = fill
                if alignment: cell.alignment = alignment

            # --- 문서 제목 ---
            ws.merge_cells('A1:E2')
            ws['A1'] = "완제품 시험성적서"
            apply_style(ws['A1'], font=title_font, alignment=center_align)

            # --- 기본 정보 ---
            info_layout = [
                ("제 품 명", info_data.get("제 품 명")), ("반제품 제조일자", info_data.get("반제품 제조일자")), ("반제품제조번호(LOT)", info_data.get("반제품제조번호(LOT)")),
                ("포장 일자", info_data.get("포장 일자")), ("완제품제조번호(LOT)", info_data.get("완제품제조번호(LOT)")), ("사용 기한", info_data.get("사용 기한")),
                ("단위 용량 (ml)", info_data.get("단위 용량 (ml)")), ("샘플링 방법", info_data.get("샘플링 방법")), ("시험 일자", info_data.get("시험 일자"))
            ]
            row = 4
            # 제품명 행
            apply_style(ws.cell(row, 1, info_layout[0][0]), font=label_font, fill=label_fill, alignment=center_align, border=thin_border)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
            apply_style(ws.cell(row, 2, info_layout[0][1]), font=cell_font, alignment=left_align, border=thin_border)
            for c in range(3, 6): apply_style(ws.cell(row, c), border=thin_border) # 병합된 셀 테두리 적용
            row += 1
            # 나머지 정보 행
            for i in range(1, len(info_layout), 2):
                apply_style(ws.cell(row, 1, info_layout[i][0]), font=label_font, fill=label_fill, alignment=center_align, border=thin_border)
                apply_style(ws.cell(row, 2, info_layout[i][1]), font=cell_font, alignment=left_align, border=thin_border)
                if i + 1 < len(info_layout):
                    apply_style(ws.cell(row, 3, info_layout[i+1][0]), font=label_font, fill=label_fill, alignment=center_align, border=thin_border)
                    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
                    apply_style(ws.cell(row, 4, info_layout[i+1][1]), font=cell_font, alignment=left_align, border=thin_border)
                    apply_style(ws.cell(row, 5), border=thin_border) # 병합된 셀 테두리 적용
                row += 1

            # --- 시험 항목 테이블 ---
            ws.append([])
            table_start_row = ws.max_row + 1
            headers = ["구분", "시험항목", "시험기준", "시험결과", "비고"]
            ws.append(headers)
            for col_idx, header_text in enumerate(headers, 1):
                apply_style(ws.cell(table_start_row, col_idx), font=header_font, fill=header_fill, alignment=center_align, border=thin_border)

            for item in test_items:
                ws.append([item["id"], item["item"], item["spec"], item["result"], item["note"]])
                for col in range(1, 6):
                    apply_style(ws.cell(ws.max_row, col), font=cell_font, border=thin_border, alignment=center_align)

            # --- 종합판정 및 시험자 ---
            ws.append([])
            conclusion_row = ws.max_row + 1
            apply_style(ws.cell(conclusion_row, 1, "시험자"), font=label_font, fill=label_fill, alignment=center_align, border=thin_border)
            apply_style(ws.cell(conclusion_row, 2, info_data.get("시험자")), font=cell_font, alignment=center_align, border=thin_border)
            apply_style(ws.cell(conclusion_row, 3, "검토자"), font=label_font, fill=label_fill, alignment=center_align, border=thin_border)
            apply_style(ws.cell(conclusion_row, 4, info_data.get("검토자")), font=cell_font, alignment=center_align, border=thin_border)
            ws.cell(conclusion_row, 5).border = thin_border # 빈칸 테두리
            
            conclusion_row += 1
            apply_style(ws.cell(conclusion_row, 1, "종합판정"), font=label_font, fill=label_fill, alignment=center_align, border=thin_border)
            ws.merge_cells(start_row=conclusion_row, start_column=2, end_row=conclusion_row, end_column=5)
            apply_style(ws.cell(conclusion_row, 2, info_data.get("종합판정")), font=Font(name='맑은 고딕', size=10, bold=True), alignment=center_align, border=thin_border)
            for c in range(3, 6): apply_style(ws.cell(conclusion_row, c), border=thin_border) # 병합된 셀 테두리 적용

            # --- 열 너비 조정 ---
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 35

            # --- 파일 저장 ---
            default_filename = f"{info_data['제 품 명']}_{info_data['완제품제조번호(LOT)']}_완제품시험성적서.xlsx"
            file_path = fd.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel 통합 문서", "*.xlsx")],
                initialfile=default_filename,
                title="보고서 다른 이름으로 저장"
            )

            if file_path:
                wb.save(file_path)
                messagebox.showinfo("성공", f"보고서가 성공적으로 저장되었습니다:\n{file_path}", parent=self)

        except Exception as e:
            messagebox.showerror("오류", f"보고서 생성 중 오류가 발생했습니다:\n{e}", parent=self)