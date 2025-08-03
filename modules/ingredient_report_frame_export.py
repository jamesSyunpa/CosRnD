import customtkinter as ctk
from openpyxl import Workbook
import tkinter.filedialog as fd
from tkinter import messagebox
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

class IngredientReportFrame_Export(ctk.CTkFrame):
    """화장품 원료목록 보고 내보내기용 UI (엑셀 저장 전용)"""
    
    def __init__(self, master):
        super().__init__(master)
        
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        
        # === [1] 상단 정보 입력 ===
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.grid(row=0, column=0, sticky="ew", padx=10, pady=(10,5))
        header.grid_columnconfigure((1, 3, 5), weight=1) # 입력 필드가 확장되도록 설정
        
        # --- 데이터 ---
        self.product_types = {
            "가. 3세 이하의 영유아용 제품류": {"가1": "영유아용 샴푸, 린스", "가2": "영유아용 로션, 크림", "가3": "영유아용 오일", "가4": "영유아용 인체 세정용 제품", "가5": "영유아용 목욕용 제품"},
            "나. 목욕용 제품류": {"나1": "목욕용 오일·정제·캡슐", "나2": "목욕용 소금류", "나3": "버블 배스", "나4": "그 밖의 목욕용 제품류"},
            "다. 인체세정용 제품류": {"다1": "폼 클렌저", "다2": "바디 클렌저", "다3": "액체 비누", "다3-1": "화장 비누(고체 형태의 세안용 비누)", "다4": "외음부 세정제", "다5": "물휴지", "다6": "그 밖의 인체 세정용 제품류"},
            "라. 눈화장용 제품류": {"라1": "아이브로 제품", "라2": "아이 라이너", "라3": "아이 섀도", "라4": "마스카라", "라5": "아이 메이크업 리무버", "라6": "그 밖의 눈화장용 제품류", "라7": "속눈썹용 퍼머넌트 웨이브"},
            "마. 방향용 제품류": {"마1": "향수", "마4": "콜롱", "마5": "그 밖의 방향용 제품류"},
            "바. 두발 염색용 제품류": {"바1": "헤어 틴트", "바2": "헤어 컬라 스프레이", "바3": "그 밖의 염모용 제품류", "바4": "염모제", "바5": "탈염·탈색용 제품"},
            "사. 색조화장용 제품류": {"사1": "볼연지", "사2": "페이스 파우더", "사3": "리퀴드, 크림, 케이크 파운데이션", "사4": "메이크업 베이스", "사5": "메이크업 픽서티브", "사6": "립스틱, 립라이너", "사7": "립글로스, 립밤", "사8": "바디페인팅, 페이스페인팅, 분장용 제품", "사9": "그 밖의 메이크업 제품류"},
            "아. 두발용 제품류": {"아1": "헤어 컨디셔너, 트리트먼트, 헤어 팩, 린스", "아2": "헤어 토닉, 헤어 에센스", "아3": "헤어 그루밍에이드", "아4": "헤어 크림, 로션", "아5": "헤어 오일", "아6": "포마드", "아7": "헤어 스프레이·무스·왁스·젤", "아8": "샴푸", "아9": "헤어 퍼머넌트 웨이브", "아10": "헤어 스트레이트너", "아11": "그 밖의 두발용 제품류", "아12": "흑채"},
            "자. 손발톱용 제품류": {"자1": "베이스코트, 언더코트", "자2": "네일폴리시, 네일에나멜", "자3": "탑코트", "자4": "네일 크림·로션·에센스, 오일", "자5": "네일폴리시·네일에나멜 리무버", "자6": "그 밖의 손발톱용 제품류"},
            "차. 면도용 제품류": {"차1": "애프터셰이브 로션", "차3": "프리셰이브 로션", "차4": "세이빙 크림", "차5": "세이빙 폼", "차6": "그 밖의 면도용 제품류"},
            "카. 기초화장용 제품류": {"카1": "수렴·유연·영양화장수", "카2": "마사지 크림", "카3": "에센스, 오일", "카4": "파우더", "카5": "바디 제품", "카6": "팩, 마스크", "카7": "눈 주위 제품", "카8": "로션, 크림", "카9": "손·발의 피부연화 제품", "카10": "클렌징워터·클렌징오일·클렌징로션·클렌징크림 등 메이크업 리무버", "카11": "그 밖의 기초화장용 제품류"},
            "타. 체취방지용 제품류": {"타1": "데오도런트", "타2": "그 밖의 체취 방지용 제품류"},
            "파. 체모 제거용 제품류": {"파1": "제모제", "파2": "그 밖의 체모 제거용 제품류", "파3": "제모왁스"}
        }
        self.functional_types = {
            "": "기능성 아님", "F1": "미백", "F2": "주름개선", "F3": "자외선차단", "F4": "미백+주름개선",
            "F5": "미백+자외선차단", "F6": "주름개선+자외선차단", "F7": "미백+주름개선+자외선차단",
            "F8": "염모, 탈염탈색", "F9": "제모 (물리적 제거 제외)", "F10": "탈모증상 완화",
            "F11": "여드름성 피부 완화 (인체세정용 제품류)", "F12": "피부 장벽 기능 회복하여 가려움 완화",
            "F13": "튼살로 인한 붉은 선 완화", "F14": "기타 복합유형"
        }
        
        # --- 1행: 보고유형, 제품명 ---
        ctk.CTkLabel(header, text="보고유형", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.report_type = ctk.CTkComboBox(header, values=["신규보고", "변경보고", "삭제보고"], state="readonly")
        self.report_type.set("신규보고")
        self.report_type.grid(row=0, column=1, padx=5, pady=5)
        
        ctk.CTkLabel(header, text="제품명", font=ctk.CTkFont(weight="bold")).grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.product_name = ctk.CTkEntry(header)
        self.product_name.grid(row=0, column=3, columnspan=3, padx=5, pady=5, sticky="ew")
        
        # --- 2행: 유형 카테고리, 유형표시, 기능성 유형 ---
        ctk.CTkLabel(header, text="유형 카테고리", font=ctk.CTkFont(weight="bold")).grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.category_combo = ctk.CTkComboBox(header, values=list(self.product_types.keys()), command=self.update_type_combo, state="readonly")
        self.category_combo.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        ctk.CTkLabel(header, text="유형표시", font=ctk.CTkFont(weight="bold")).grid(row=1, column=2, padx=5, pady=5, sticky="w")
        self.type_combo = ctk.CTkComboBox(header, values=[], state="readonly")
        self.type_combo.grid(row=1, column=3, padx=5, pady=5, sticky="ew")

        ctk.CTkLabel(header, text="기능성 유형", font=ctk.CTkFont(weight="bold")).grid(row=1, column=4, padx=5, pady=5, sticky="w")
        functional_values = [f"{k} - {v}" if k else v for k, v in self.functional_types.items()]
        self.functional_combo = ctk.CTkComboBox(header, values=functional_values, state="readonly")
        self.functional_combo.set("기능성 아님")
        self.functional_combo.grid(row=1, column=5, padx=5, pady=5, sticky="ew")

        # --- 3행: 기능성품목코드, 제조업자상호 ---
        ctk.CTkLabel(header, text="기능성품목코드", font=ctk.CTkFont(weight="bold")).grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.functional_code_entry = ctk.CTkEntry(header)
        self.functional_code_entry.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

        ctk.CTkLabel(header, text="제조업자상호", font=ctk.CTkFont(weight="bold")).grid(row=2, column=2, padx=5, pady=5, sticky="w")
        self.manufacturer_entry = ctk.CTkEntry(header)
        self.manufacturer_entry.grid(row=2, column=3, padx=5, pady=5, sticky="ew")

        # --- 4행: 용도, 맞춤형 내용물 ---
        ctk.CTkLabel(header, text="용도(수출전용 여부)", font=ctk.CTkFont(weight="bold")).grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.usage_combo = ctk.CTkComboBox(header, values=["", "E"], state="readonly")
        self.usage_combo.grid(row=3, column=1, padx=5, pady=5, sticky="ew")

        ctk.CTkLabel(header, text="맞춤형 내용물(C1/C2)", font=ctk.CTkFont(weight="bold")).grid(row=3, column=2, padx=5, pady=5, sticky="w")
        self.custom_content_combo = ctk.CTkComboBox(header, values=["", "C1", "C2"], state="readonly")
        self.custom_content_combo.grid(row=3, column=3, padx=5, pady=5, sticky="ew")

        # --- 5행: 원료성분명 붙여넣기 ---
        ctk.CTkLabel(header, text="원료성분명\n(붙여넣기)", font=ctk.CTkFont(weight="bold")).grid(row=4, column=0, padx=5, pady=5, sticky="w")
        self.ingredient_paste_textbox = ctk.CTkTextbox(header, height=100)
        self.ingredient_paste_textbox.grid(row=4, column=1, columnspan=5, padx=5, pady=5, sticky="ew")

        
        # === [2] 테이블 ===
        self.table = ctk.CTkScrollableFrame(self, label_text="원료목록")
        self.table.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
        for i in range(9):
            self.table.grid_columnconfigure(i, weight=1)
        
        headers = [
            "일련번호", "제품명", "유형표시", "기능성화장품\n유형",
            "기능성화장품\n품목코드", "제조업자\n상호", "원료성분명",
            "용도", "맞춤형 내용물\n(혼합용‘C1',소분용'C2')"
        ]
        for i, h in enumerate(headers):
            ctk.CTkLabel(self.table, text=h, font=ctk.CTkFont(weight="bold")).grid(row=0, column=i, padx=3, pady=3, sticky="ew")
        
        self.rows = []
        self._add_row() # UI 생성 시 바로 첫 행 추가
        
        # === [3] 하단 버튼 ===
        bottom = ctk.CTkFrame(self, fg_color="transparent")
        bottom.grid(row=2, column=0, sticky="ew", padx=10, pady=10)
        
        ctk.CTkButton(bottom, text="행 추가", command=self._add_row).pack(side="left", padx=5)
        ctk.CTkButton(bottom, text="행 삭제", fg_color="gray50", command=self._remove_row).pack(side="left", padx=5)
        ctk.CTkButton(bottom, text="엑셀로 내보내기", fg_color="#3B8ED0", command=self._export_excel).pack(side="right", padx=5)
    
    def _add_row(self, ingredient_name=""):
        """테이블에 행을 추가합니다. 원료성분명을 인자로 받을 수 있습니다."""
        r = len(self.rows) + 1
        widgets = {}
        for i in range(9):
            e = ctk.CTkEntry(self.table, corner_radius=0, border_width=0)
            if i == 0:
                e.insert(0, str(r))
                e.configure(state="readonly")
            elif i == 6: # 원료성분명
                e.insert(0, ingredient_name)
            e.grid(row=r, column=i, padx=(1,0), pady=(1,0), sticky="ew")
            widgets[f"col{i}"] = e
        self.rows.append(widgets)
    
    def _remove_row(self):
        if not self.rows: return
        last = self.rows.pop()
        for w in last.values():
            w.destroy()
    
    def _export_excel(self):
        """엑셀로 내보내기 (수정된 로직)"""
        if not self.product_name.get():
            messagebox.showwarning("입력 오류", "제품명을 입력해주세요.", parent=self)
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "원료목록보고"
        
        # --- 스타일 정의 ---
        header_font = Font(name='맑은 고딕', size=11, bold=True, color="FFFFFF")
        cell_font = Font(name='맑은 고딕', size=10)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        headers = ["일련번호", "제품명", "유형표시", "기능성화장품유형", "기능성화장품품목코드", "제조업자상호", "원료성분명", "용도", "맞춤형 내용물(혼합용‘C1',소분용'C2')"]
        ws.append(headers)
        for col_idx, header_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border
        
        # --- 상단 입력 데이터 수집 ---
        report_type = self.report_type.get()
        product_name = self.product_name.get()
        
        type_selected = self.type_combo.get()
        type_code = type_selected.split(' - ')[0] if ' - ' in type_selected else ""
        
        functional_selected = self.functional_combo.get()
        functional_type_code = functional_selected.split(' - ')[0] if ' - ' in functional_selected else ""

        functional_code = self.functional_code_entry.get()
        manufacturer = self.manufacturer_entry.get()
        
        usage = self.usage_combo.get()
        custom_content = self.custom_content_combo.get()

        # --- 원료성분명 붙여넣기 처리 (쉼표 또는 줄바꿈으로 분리) ---
        pasted_text = self.ingredient_paste_textbox.get("1.0", "end-1c").strip()
        if ', ' in pasted_text:
            pasted_ingredients = pasted_text.split(', ')
        elif ',' in pasted_text:
            pasted_ingredients = pasted_text.split(',')
        else:
            pasted_ingredients = pasted_text.split('\n')
        pasted_ingredients = [line.strip() for line in pasted_ingredients if line.strip()]

        if not pasted_ingredients:
            messagebox.showwarning("입력 오류", "원료성분명을 입력해주세요.", parent=self)
            return

        # --- 테이블 재생성 및 데이터 채우기 ---
        for row_widget in self.rows:
            for widget in row_widget.values():
                widget.destroy()
        self.rows.clear()

        for name in pasted_ingredients:
            self._add_row(ingredient_name=name)

        for row_widgets in self.rows:
            # UI 업데이트
            row_widgets["col1"].insert(0, product_name)
            row_widgets["col2"].insert(0, type_code)
            row_widgets["col3"].insert(0, functional_type_code)
            row_widgets["col4"].insert(0, functional_code)
            row_widgets["col5"].insert(0, manufacturer)
            row_widgets["col7"].insert(0, usage)
            row_widgets["col8"].insert(0, custom_content)
            
            # 엑셀 데이터 준비 및 스타일 적용
            excel_row = [row_widgets[f"col{i}"].get() for i in range(9)]
            current_row_num = ws.max_row + 1
            for col_idx, value in enumerate(excel_row, 1):
                cell = ws.cell(row=current_row_num, column=col_idx, value=value)
                cell.font = cell_font; cell.border = thin_border; cell.alignment = left_align

        # --- 열 너비 자동 조절 ---
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter

            # 헤더의 길이를 먼저 계산
            header_cell = ws[f"{column_letter}1"]
            if header_cell.value:
                # 헤더는 여러 줄일 수 있으므로 각 줄의 최대 길이를 계산
                header_lines = str(header_cell.value).split('\n')
                for line in header_lines:
                    if len(line) * 1.2 > max_length: # 한글 폰트 너비를 고려하여 가중치 부여
                        max_length = len(line) * 1.2

            for cell in col:
                if cell.value:
                    # 한글/영문 길이를 다르게 계산
                    # 한글은 2, 영문/숫자는 1로 계산하여 길이를 추정
                    length = sum(2 if '\uac00' <= char <= '\ud7a3' else 1 for char in str(cell.value))
                    if length > max_length:
                        max_length = length
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        default_name = f"원료목록보고_{self.product_name.get()}.xlsx"
        file_path = fd.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], initialfile=default_name, title="엑셀로 저장")
        if file_path:
            wb.save(file_path)
            messagebox.showinfo("성공", f"엑셀 파일이 성공적으로 저장되었습니다:\n{file_path}", parent=self)
    
    def update_type_combo(self, category):
        """카테고리 선택 시 세부 유형 콤보박스 업데이트"""
        if category in self.product_types:
            types = self.product_types[category]
            type_values = [f"{k} - {v}" for k, v in types.items()]
            self.type_combo.configure(values=type_values)
            self.type_combo.set(type_values[0] if type_values else "")