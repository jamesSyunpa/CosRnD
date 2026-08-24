# modules/excel_handler.py
import openpyxl
from openpyxl import Workbook
from tkinter import filedialog, messagebox
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import configparser
from datetime import datetime
import os
from PIL import Image, ImageDraw, ImageFont
import time
import shutil
import tempfile

# 프로젝트 루트 경로
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# --- 경로 설정을 읽기 위한 설정 ---
def _get_excel_config_path(app_dir_name: str = 'CosRQD') -> str:
    appdata_dir = os.path.join(os.getenv('APPDATA', os.path.expanduser('~')), app_dir_name)
    target_config = os.path.join(appdata_dir, 'config.ini')
    if os.path.exists(target_config):
        return target_config
    root_config = os.path.join(PROJECT_ROOT, 'config.ini')
    if os.path.exists(root_config):
        return root_config
    return target_config

CONFIG_FILE_PATH = _get_excel_config_path()

def _get_display_length(text):
    """Calculate display length of a string, considering East Asian characters."""
    if not text:
        return 0
    text = str(text)
    length = 0
    # East Asian characters (Hangul, Japanese, Chinese) are wider.
    for char in text:
        if '\uac00' <= char <= '\uD7A3' or \
           '\u3040' <= char <= '\u309F' or \
           '\u30a0' <= char <= '\u30FF' or \
           '\u4E00' <= char <= '\u9FFF':
            length += 2
        else:
            length += 1
    return length

def _load_workbook_robust(file_path, data_only=True):
    """
    엑셀 파일을 읽어옵니다.
    PermissionError(파일 열림 등) 발생 시, 임시 폴더로 복사하여 읽기를 시도합니다.
    """
    try:
        return openpyxl.load_workbook(file_path, data_only=data_only)
    except PermissionError:
        try:
            # 임시 파일 생성 (고유한 이름 보장)
            temp_dir = tempfile.gettempdir()
            temp_path = os.path.join(temp_dir, f"temp_import_{datetime.now().strftime('%Y%m%d%H%M%S')}_{os.path.basename(file_path)}")
            
            # 파일 복사
            shutil.copy2(file_path, temp_path)
            
            # 복사본 읽기
            wb = openpyxl.load_workbook(temp_path, data_only=data_only)
            
            # (중요) 읽기 후 바로 삭제하지 않고, OS가 정리하도록 두거나 나중에 cleanup할 수 있지만,
            # 여기서는 wb 객체가 파일을 잡고 있을 수 있으므로 반환만 함.
            # openpyxl은 load 시점에 메모리로 올리는 것이 기본이므로(read_only=False),
            # 여기서 닫거나 삭제 시도를 하면 에러가 날 수 있음. 
            # 단순히 복사본을 읽어서 반환.
            return wb
        except Exception as e:
            # 복사조차 실패하면 원본 에러를 던짐 (혹은 복사 에러 로그)
            print(f"[ERROR] Temp copy failed: {e}")
            raise # 원본 PermissionError가 아니라 복사 실패 에러가 나감, 하지만 상위에서 잡음
    except Exception:
        raise

def get_excel_path():
    """config.ini에서 엑셀 기본 경로를 읽어옵니다."""
    config = configparser.ConfigParser()
    cfg_path = _get_excel_config_path()
    if os.path.exists(cfg_path):
        config.read(cfg_path, encoding='utf-8')
    path = config.get('Paths', 'excel_dir', fallback='').strip()
    if not path or not os.path.isdir(path):
        default_path = os.path.join(os.path.expanduser('~'), 'Documents', 'CosRQD', 'ExcelData')
        os.makedirs(default_path, exist_ok=True)
        return default_path
    return path

def save_excel_path(path):
    """선택된 경로를 config.ini에 저장합니다."""
    cfg_path = _get_excel_config_path()
    config = configparser.ConfigParser()
    if os.path.exists(cfg_path):
        config.read(cfg_path, encoding='utf-8')
    if not config.has_section('Paths'):
        config.add_section('Paths')
    config.set('Paths', 'excel_dir', path)
    os.makedirs(os.path.dirname(cfg_path), exist_ok=True)
    with open(cfg_path, 'w', encoding='utf-8') as configfile:
        config.write(configfile)

def get_timestamped_filename(default_filename):
    """파일 이름에 타임스탬프를 추가합니다 (확장자 앞)."""
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    name, ext = os.path.splitext(default_filename)
    return f"{name}_{now}{ext}"

def export_template(headers, default_filename="template.xlsx"):
    """단일 시트를 가진 엑셀 템플릿을 내보냅니다."""
    initial_dir = get_excel_path()
    timestamped_filename = get_timestamped_filename(default_filename)
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialdir=initial_dir,
        initialfile=timestamped_filename,
        title="엑셀 폼 저장"
    )
    if not file_path:
        return

    save_excel_path(os.path.dirname(file_path))

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(name='맑은 고딕', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')

    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "양식"
        sheet.row_dimensions[1].height = 28

        for col_idx, header_text in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
            # 너비 자동 조절
            col_letter = get_column_letter(col_idx)
            header_len = _get_display_length(header_text)
            sheet.column_dimensions[col_letter].width = max(header_len + 5, 14)

        workbook.save(file_path)
        messagebox.showinfo("성공", f"엑셀 폼이 '{file_path}'에 성공적으로 저장되었습니다.")
        try:
            os.startfile(os.path.abspath(file_path))
        except Exception:
            pass
    except Exception as e:
        messagebox.showerror("오류", f"파일 저장 중 오류가 발생했습니다: {e}")

def export_multisheet_template(sheets_with_headers, default_filename="template.xlsx"):
    """여러 시트를 가진 엑셀 템플릿을 내보냅니다."""
    initial_dir = get_excel_path()
    timestamped_filename = get_timestamped_filename(default_filename)
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialdir=initial_dir,
        initialfile=timestamped_filename,
        title="엑셀 폼 저장"
    )
    if not file_path:
        return

    save_excel_path(os.path.dirname(file_path))

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(name='맑은 고딕', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')

    try:
        workbook = Workbook()
        workbook.remove(workbook.active) # 기본 시트 제거

        for sheet_name, headers in sheets_with_headers.items():
            sheet = workbook.create_sheet(title=sheet_name)
            sheet.row_dimensions[1].height = 28
            
            for col_idx, header_text in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_idx, value=header_text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
                
                col_letter = get_column_letter(col_idx)
                header_len = _get_display_length(header_text)
                sheet.column_dimensions[col_letter].width = max(header_len + 5, 14)

        workbook.save(file_path)
        messagebox.showinfo("성공", f"다중 시트 엑셀 폼이 '{file_path}'에 성공적으로 저장되었습니다.")
        try:
            os.startfile(os.path.abspath(file_path))
        except Exception:
            pass
    except Exception as e:
        messagebox.showerror("내보내기 오류", f"파일 저장 중 오류가 발생했습니다: {e}")

def export_production_formulation_revised_to_excel(
    production_data,
    default_filename: str = "production_revised.xlsx",
    file_path: str | None = None,
    open_print_preview: bool = False,
):
    """수정본 템플릿(생산처방ui.py의 병합/모양 준용)으로 단일 시트 내보내기.
    - 상단 레이아웃을 생산처방ui.py처럼 구성:
        • 제목: A1:F2 병합, 중앙 정렬("생산지시서")
        • 결재란 자리: G1:H2 병합(텍스트 라벨만), 기존 색상/테두리는 유지
        • 기본정보: 3행/4행에 제품명과 핵심 정보 배치 (병합 위치는 ui.py 참고)
    - 본문 표: A..H = [Ph, 구분, 코드, 원료명, 함량(%), 생산량(kg), 제조공정, 공정검사]
        • 같은 Ph 구간은 A열을 세로 병합
        • 제조공정/공정검사도 Ph 구간 단위로 각각 세로 병합
        • 각 구간의 텍스트는 해당 구간 아이템들 중 가장 긴 내용으로 채움
    - 색상/테두리/폰트 등 스타일은 기존(원래) 정의를 그대로 사용
    - 인쇄 설정은 기존 수정본과 동일(세로, 여백 축소, 머리글/바닥글 등)
    """

    # 파일 경로 결정
    if file_path is None:
        if open_print_preview:
            # tempfile 대신 프로젝트 폴더의 data 디렉토리 사용
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            temp_dir = os.path.join(PROJECT_ROOT, 'data', 'temp')
            os.makedirs(temp_dir, exist_ok=True)
            file_path = os.path.join(temp_dir, f'preview_revised_{timestamp}.xlsx')
        else:
            initial_dir = get_excel_path()
            timestamped_filename = get_timestamped_filename(default_filename)
            chosen = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[["Excel files", "*.xlsx"]],
                initialdir=initial_dir,
                initialfile=timestamped_filename,
                title="생산처방(수정본) 저장"
            )
            if not chosen:
                return
            file_path = chosen
            save_excel_path(os.path.dirname(file_path))

    # 스타일
    thin = Side(style='thin', color='2C3E50')
    medium = Side(style='medium', color='2C3E50')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    medium_border = Border(left=medium, right=medium, top=medium, bottom=medium)
    title_font = Font(name='맑은 고딕', size=20, bold=True, color='1F4E78')
    header_font = Font(name='맑은 고딕', size=11, bold=True, color='2C3E50')
    header_font_white = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    label_font = Font(name='맑은 고딕', size=10, bold=True, color='34495E')
    default_font = Font(name='맑은 고딕', size=10, color='2C3E50')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    # 본문(내용목록) 셀은 위아래 정렬을 '센터'로 통일
    left_top = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    table_header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    label_fill = PatternFill(start_color="D5E8F0", end_color="D5E8F0", fill_type="solid")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "생산처방"

        # Get details first!
        details = production_data.get('details', {})
        
        # 1) a1:e2 (생산지시서) merged
        tcell = ws['A1']
        tcell.value = "생산지시서"
        tcell.font = title_font
        tcell.alignment = center
        # Set borders on all cells in A1:E2 first
        for row in [1,2]:
            for col in ['A','B','C','D','E']:
                ws[f"{col}{row}"].border = thin_border
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 22
        # Now merge
        ws.merge_cells('A1:E2')
        
        # 2) f1:f2 (결제방) merged
        f1_cell = ws['F1']
        f1f2_value = details.get('결제방','')
        if f1f2_value:
            f1_cell.value = "결제방: " + f1f2_value
        else:
            f1_cell.value = "결제방"
        f1_cell.font = label_font
        f1_cell.fill = label_fill
        f1_cell.alignment = center
        f1_cell.border = thin_border
        # Set borders on F1 and F2
        ws['F2'].border = thin_border
        # Now merge F1:F2
        ws.merge_cells('F1:F2')
        
        # 3) g1 (작성) g2 (빈칸)
        ws['G1'].value = '작성'; ws['G1'].font = label_font; ws['G1'].fill = label_fill; ws['G1'].alignment = center; ws['G1'].border = thin_border
        ws['G2'].border = thin_border
        # 4) h1 (검토) h2 (빈칸)
        ws['H1'].value = '검토'; ws['H1'].font = label_font; ws['H1'].fill = label_fill; ws['H1'].alignment = center; ws['H1'].border = thin_border
        ws['H2'].border = thin_border
        # 5) i1 (승인) i2 (빈칸)
        ws['I1'].value = '승인'; ws['I1'].font = label_font; ws['I1'].fill = label_fill; ws['I1'].alignment = center; ws['I1'].border = thin_border
        ws['I2'].border = thin_border
        
        # Row 3
        prod_name = details.get('제품명', '')
        # A3 라벨, B3:D3 병합 (제품명)
        ws['A3'].value = '제품명'; ws['A3'].font = label_font; ws['A3'].fill = label_fill; ws['A3'].alignment = center; ws['A3'].border = thin_border
        ws['B3'].value = prod_name
        ws['B3'].font = default_font; ws['B3'].alignment = center; ws['B3'].border = thin_border
        for col in ['C','D']:
            ws[f"{col}3"].border = thin_border
        ws.merge_cells('B3:D3')
        # 6. e3 (생산코드 label) f3:g3 (merged, 생산실제코드)
        ws['E3'].value = '생산코드'; ws['E3'].font = label_font; ws['E3'].fill = label_fill; ws['E3'].alignment = center; ws['E3'].border = thin_border
        ws['F3'].value = details.get('생산코드',''); ws['F3'].font = default_font; ws['F3'].alignment = center; ws['F3'].border = thin_border
        ws['G3'].border = thin_border
        ws.merge_cells('F3:G3')
        # 7. h3 (제조자 label) i3 (빈칸)
        ws['H3'].value = '제조자'; ws['H3'].font = label_font; ws['H3'].fill = label_fill; ws['H3'].alignment = center; ws['H3'].border = thin_border
        ws['I3'].border = thin_border
        
        ws.row_dimensions[3].height = 20
        
        # Row 4
        ws['A4'].value = '지시일'; ws['A4'].font = label_font; ws['A4'].fill = label_fill; ws['A4'].alignment = center; ws['A4'].border = thin_border
        instruction_date_raw = details.get('지시일', details.get('출력일시',''))
        if instruction_date_raw and ' ' in str(instruction_date_raw):
            instruction_date = str(instruction_date_raw).split(' ')[0] 
        else:
            instruction_date = instruction_date_raw
        ws['B4'].value = instruction_date; ws['B4'].font = default_font; ws['B4'].alignment = center; ws['B4'].border = thin_border
        ws['C4'].value = '제조일'; ws['C4'].font = label_font; ws['C4'].fill = label_fill; ws['C4'].alignment = center; ws['C4'].border = thin_border
        ws['D4'].value = details.get('제조일',''); ws['D4'].font = default_font; ws['D4'].alignment = center; ws['D4'].border = thin_border
        ws['E4'].value = '생산량(kg)'; ws['E4'].font = label_font; ws['E4'].fill = label_fill; ws['E4'].alignment = center; ws['E4'].border = thin_border
        ws['F4'].value = details.get('생산량(kg)',''); ws['F4'].font = default_font; ws['F4'].alignment = center; ws['F4'].border = thin_border
        ws['G4'].border = thin_border
        ws.merge_cells('F4:G4')
        # 8. h4 (수득량 label) i4 (빈칸)
        ws['H4'].value = '수득량'; ws['H4'].font = label_font; ws['H4'].fill = label_fill; ws['H4'].alignment = center; ws['H4'].border = thin_border
        ws['I4'].border = thin_border
        
        ws.row_dimensions[4].height = 20

        # Row 5: 비고 or empty
        ws.row_dimensions[5].height = 6

        # Row 6: 빈 줄(여백)
        ws.row_dimensions[6].height = 6

        # 4) 테이블 헤더 (UI 기준 A..J) - '계량량(kg)'을 생산량 옆에 추가
        headers = ["Ph", "구분", "코드", "원료명", "함량(%)", "생산량(kg)", "계량량(kg)", "제조공정", "공정검사"]
        header_row = 7
        ws.row_dimensions[header_row].height = 22
        for c_idx, h in enumerate(headers, 1):
            hc = ws.cell(row=header_row, column=c_idx, value=h)
            hc.font = header_font_white
            hc.fill = table_header_fill
            hc.alignment = center
            hc.border = medium_border

        # 5) 데이터 행 (Ph 구간 병합: A/G/H)
        current = header_row + 1
        total_ratio = 0.0
        items = production_data.get('items', [])

        def norm_phase(v):
            s = str(v).strip() if v is not None else ''
            return s.replace('Ph.', '').replace('PH', '').strip() if s else ''

        # 연속 구간으로 그룹화: 빈 Phase는 직전의 Phase에 속하도록 처리
        groups = []  # list of (phase_label, start_idx, end_idx) on items index (0-based)
        if items:
            def eff_phase(i: int, last_non_empty: str | None) -> tuple[str, str | None]:
                raw = norm_phase(items[i].get('Ph', items[i].get('Ph.', '')))
                if raw:
                    return raw, raw
                return (last_non_empty or ''), last_non_empty
            
            last_ph = None
            for idx, item in enumerate(items):
                curr_ph, last_ph = eff_phase(idx, last_ph)
                if not groups or groups[-1][0] != curr_ph:
                    groups.append([curr_ph, idx, idx])
                else:
                    groups[-1][2] = idx

        # Ph 구간별로 행 추가 + 병합
        for ph, s_idx, e_idx in groups:
            group_len = e_idx - s_idx + 1
            # 그룹 내 최대 텍스트를 기준으로 폭 기반 줄바꿈 적용 후 줄 수 산정
            proc_texts = []
            insp_texts = []
            for i in range(s_idx, e_idx+1):
                proc_texts.append(str(items[i].get('제조공정', '') or '').replace('"','').replace("'", ''))
                insp_texts.append(str(items[i].get('공정검사', '') or '').replace('"','').replace("'", ''))
            group_proc = max(proc_texts, key=len) if proc_texts else ''
            group_insp = max(insp_texts, key=len) if insp_texts else ''

            # 폭 기준 줄바꿈 (엑셀 컬럼폭 G=26, H=22 → 픽셀 근사치 7px/단위, 여백 12px 제외)
            def _wrap_by_px(txt: str, px: int) -> str:
                if not txt:
                    return ''
                try:
                    font = ImageFont.truetype("malgun.ttf", 10)
                except Exception:
                    try:
                        font = ImageFont.truetype("malgunbd.ttf", 10)
                    except Exception:
                        font = ImageFont.load_default()
                tmp = Image.new('RGB', (px, 20), 'white')
                d = ImageDraw.Draw(tmp)
                lines = []
                for para in str(txt).splitlines() or ['']:
                    if not para:
                        lines.append('')
                        continue
                    line = ''
                    tokens = para.split(' ')
                    if len(tokens) == 1:
                        for ch in para:
                            test = line + ch
                            w = d.textbbox((0,0), test, font=font)[2]
                            if w <= px or not line:
                                line = test
                            else:
                                lines.append(line)
                                line = ch
                        if line:
                            lines.append(line)
                    else:
                        for tok in tokens:
                            test = (line + (' ' if line else '') + tok)
                            w = d.textbbox((0,0), test, font=font)[2]
                            if w <= px or not line:
                                line = test
                            else:
                                lines.append(line)
                                line = tok
                        if line:
                            lines.append(line)
                return "\n".join(lines)

            g_px = max(10, int(26 * 7) - 12)
            h_px = max(10, int(22 * 7) - 12)
            group_proc_wrapped = _wrap_by_px(group_proc, g_px)
            group_insp_wrapped = _wrap_by_px(group_insp, h_px)
            wrapped_lines = max(group_proc_wrapped.count('\n')+1 if group_proc_wrapped else 1,
                                group_insp_wrapped.count('\n')+1 if group_insp_wrapped else 1)
            # tighter rows: minimal height 16pt and ~14pt per visual line across the group
            per_row_height = max(16, int(14 * wrapped_lines / group_len) + 2)

            # 각 아이템 행 작성
            for i in range(s_idx, e_idx+1):
                item = items[i]
                # 생산량(kg) 보정
                qty_kg = item.get('생산량(kg)')
                if qty_kg in (None, ""):
                    try:
                        g = float(item.get('기준중량(g)'))
                        qty_kg = g / 1000.0
                    except Exception:
                        qty_kg = item.get('생산량(kg)')

                # 계량량(kg) 값은 수기/데이터 입력 값 사용
                weigh_kg = item.get('계량량(kg)')
                vals = [
                    norm_phase(item.get('Ph', item.get('Ph.', ''))),
                    item.get('구분'), item.get('코드'), item.get('원료명'),
                    item.get('함량(%)'), qty_kg, weigh_kg,
                    '', ''  # 병합 예정(H,I)
                ]

                ws.row_dimensions[current].height = per_row_height
                for c_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=current, column=c_idx, value=val)
                    cell.border = thin_border
                    cell.font = default_font
                    if c_idx in (1,2,3):
                        cell.alignment = center
                    elif c_idx == 4:
                        cell.alignment = left
                    elif c_idx in (5,6,7):
                        try:
                            fval = float(val)
                            cell.value = fval
                            if c_idx == 5:
                                cell.number_format = '0.0000'
                                total_ratio += fval
                            else:
                                cell.number_format = '#,##0.0'
                        except Exception:
                            pass
                        cell.alignment = right
                    else:
                        # 내용목록 텍스트 셀도 세로 중앙 정렬
                        cell.alignment = left
                current += 1

            # A열, H열, I열을 그룹 단위로 병합 후 값/정렬 적용 (제조공정/공정검사 위치 이동)
            start_row = header_row + 1 + s_idx if groups and groups[0][1] == 0 else (current - group_len)
            # start_row 계산 보정: current는 그룹 끝 다음 행이므로...
            start_row = current - group_len
            end_row = current - 1
            if group_len >= 1:
                # A열(Ph) 병합: 다음 파스 전까지 항상 병합, 텍스트는 있을 때만 표시
                if start_row < end_row:
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                a_cell = ws.cell(row=start_row, column=1, value=(ph or ''))
                a_cell.alignment = center; a_cell.border = thin_border; a_cell.font = header_font

                # H열 병합 및 텍스트 (제조공정)
                if start_row < end_row:
                    ws.merge_cells(start_row=start_row, start_column=8, end_row=end_row, end_column=8)
                g_cell = ws.cell(row=start_row, column=8, value=group_proc_wrapped)
                g_cell.alignment = left; g_cell.border = thin_border; g_cell.font = default_font

                # I열 병합 및 텍스트 (공정검사)
                if start_row < end_row:
                    ws.merge_cells(start_row=start_row, start_column=9, end_row=end_row, end_column=9)
                h_cell = ws.cell(row=start_row, column=9, value=group_insp_wrapped)
                h_cell.alignment = left; h_cell.border = thin_border; h_cell.font = default_font

        # 6) 합계 행 (A:D 병합)
        sum_row = current
        ws.row_dimensions[sum_row].height = 22
        # 전체 셀 채움/테두리 적용
        for col in range(1, 10):  # A..I for table
            c = ws.cell(row=sum_row, column=col)
            c.fill = table_header_fill
            c.border = medium_border
        # A:D 병합 및 라벨 중앙 정렬
        ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=4)
        sl = ws.cell(row=sum_row, column=1, value="합계 (Total)")
        sl.font = header_font_white
        sl.alignment = center
        # 합계 값(E열)
        sr = ws.cell(row=sum_row, column=5, value=total_ratio)
        sr.font = header_font_white
        sr.number_format = '0.0000'
        sr.alignment = right

        # 7) 필터/고정
        ws.auto_filter.ref = f"A{header_row}:I{sum_row}"
        ws.freeze_panes = f"A{header_row+1}"

        # 8) 컬럼 폭 고정
        fixed_widths = {
            'A': 8, 'B': 10, 'C': 15, 'D': 45, 'E': 12, 'F': 12, 
            'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12, 'L': 12, 'M': 12
        }
        for col, w in fixed_widths.items():
            ws.column_dimensions[col].width = w
            
        # Define variables for later use
        use_com_header = False
        approval_img_width = None
        approval_img_height = None
        ap_img_path = None

        # 9) 인쇄 설정 (세로 + 여백 축소) + 머리글/바닥글 + 타이틀 행 반복/인쇄 영역
        try:
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.print_options.horizontalCentered = True
            ws.page_margins.left = 0.3; ws.page_margins.right = 0.3
            ws.page_margins.top = 1.5 / 2.54
            ws.page_margins.bottom = 0.4
            # 머리글/바닥글: 날짜/시간, 문서명, 페이지 번호
            try:
                ws.header_footer.left_header = "&D &T"
                title_text = f"생산지시서 - {details.get('제품명','')}"
                ws.header_footer.center_header = title_text
                ws.header_footer.right_footer = "Page &[Page] / &[Pages]"
            except Exception:
                pass
            # 각 페이지에 표 헤더 반복
            try:
                ws.print_title_rows = f"1:{header_row}"
            except Exception:
                pass
            # 인쇄 영역 지정
            try:
                ws.print_area = f"A1:I{sum_row}"
            except Exception:
                pass
        except Exception:
            pass

        # 파일 저장 시도 (파일이 열려있으면 처리)
        max_attempts = 2
        for attempt in range(max_attempts):
            try:
                wb.save(file_path)
                break
            except PermissionError:
                if attempt < max_attempts - 1:
                    # 파일이 열려있음
                    response = messagebox.askyesno(
                        "파일 사용 중",
                        f"파일이 다른 프로그램에서 사용 중입니다:\n{os.path.basename(file_path)}\n\n"
                        "파일을 닫고 다시 시도하시겠습니까?",
                        icon='warning'
                    )
                    if response:
                        # 사용자가 예를 선택 - 잠시 대기 후 재시도
                        time.sleep(1)
                        continue
                    else:
                        # 사용자가 아니오 선택 - 저장 취소
                        messagebox.showinfo("저장 취소", "파일 저장이 취소되었습니다.")
                        return
                else:
                    # 마지막 시도에서도 실패
                    messagebox.showerror(
                        "저장 실패",
                        f"파일을 저장할 수 없습니다.\n파일이 다른 프로그램에서 열려있는지 확인하세요."
                    )
                    return
            except Exception as e:
                messagebox.showerror("저장 오류", f"파일 저장 중 오류 발생:\n{e}")
                return

        # 저장 후, COM 헤더 이미지 삽입 시도 (머리글 오른쪽)
        if use_com_header and ap_img_path and os.path.exists(ap_img_path):
            try:
                import win32com.client as win32  # type: ignore
                excel = win32.Dispatch("Excel.Application")
                excel.Visible = bool(open_print_preview)
                wb_com = excel.Workbooks.Open(os.path.abspath(file_path))
                ws_com = wb_com.ActiveSheet
                # 헤더 오른쪽 그림 지정
                ws_com.PageSetup.RightHeader = "&G"
                ws_com.PageSetup.RightHeaderPicture.Filename = os.path.abspath(ap_img_path)
                # 폭을 대략 px->pt(0.75배)로 변환하여 지정
                try:
                    if approval_img_width:
                        ws_com.PageSetup.RightHeaderPicture.Width = int((approval_img_width // 4) * 0.75)
                except Exception:
                    pass
                wb_com.Save()
                if open_print_preview:
                    try:
                        ws_com.PrintPreview()
                    except Exception:
                        pass
                wb_com.Close(SaveChanges=False)
                excel.Quit()
            except Exception:
                pass
            finally:
                try:
                    if os.path.exists(ap_img_path):
                        os.remove(ap_img_path)
                except Exception:
                    pass

        # 미리보기/출력 처리
        if open_print_preview:
            try:
                import win32com.client as win32  # type: ignore
                excel = win32.Dispatch("Excel.Application")
                excel.Visible = True
                wb_com = excel.Workbooks.Open(os.path.abspath(file_path))
                try:
                    wb_com.ActiveSheet.PrintPreview()
                finally:
                    wb_com.Close(SaveChanges=False)
                    excel.Quit()
            except Exception:
                try:
                    os.startfile(os.path.abspath(file_path))
                except Exception:
                    try:
                        os.startfile(os.path.abspath(file_path), "print")
                    except Exception:
                        pass
        else:
            try:
                os.startfile(os.path.abspath(file_path))
            except Exception:
                pass
    except Exception as e:
        messagebox.showerror("내보내기 오류", f"파일 저장 중 오류가 발생했습니다: {e}")


def export_production_formulation_original_to_excel(
    production_data,
    default_filename: str = "production.xlsx",
    file_path: str | None = None,
    open_print_preview: bool = False,
):
    """원래 템플릿으로 생산처방을 내보냅니다.
    - 제목: A1:F2 병합, "생산처방서\nProduction Formulation Sheet"
    - 결재란: G-H 영역 이미지(우측 정렬, 약 1/3 확대), H열 우측 끝 정렬
    - 기본정보: 좌측 B–E 병합, 우측 G/H 라벨/값(병합 없음)
    - 비고: 2–8 병합
    - 본문: "계량량(kg)" 포함, Phase/H/I 병합, 지브라 스트라이프, 합계 행
    - 컬럼폭: H=26.25, I=17.25, D는 30~50, 기타 최대 20
    - 인쇄: 가로 모드, fitToWidth=1, 여백 0.3/0.5
    - 추가 시트: 제조공정, 원료 환산
    """

    # 파일 경로 결정
    if file_path is None:
        if open_print_preview:
            # tempfile 대신 프로젝트 폴더의 data 디렉토리 사용
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            temp_dir = os.path.join(PROJECT_ROOT, 'data', 'temp')
            os.makedirs(temp_dir, exist_ok=True)
            file_path = os.path.join(temp_dir, f'preview_production_{timestamp}.xlsx')
        else:
            initial_dir = get_excel_path()
            timestamped_filename = get_timestamped_filename(default_filename)
            chosen = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[["Excel files", "*.xlsx"]],
                initialdir=initial_dir,
                initialfile=timestamped_filename,
                title="생산처방 저장"
            )
            if not chosen:
                return
            file_path = chosen
            save_excel_path(os.path.dirname(file_path))

    # 스타일 정의
    thin = Side(style='thin', color='2C3E50')
    medium = Side(style='medium', color='2C3E50')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    medium_border = Border(left=medium, right=medium, top=medium, bottom=medium)
    title_font = Font(name='맑은 고딕', size=20, bold=True, color='1F4E78')
    header_font = Font(name='맑은 고딕', size=11, bold=True, color='2C3E50')
    header_font_white = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    label_font = Font(name='맑은 고딕', size=10, bold=True, color='34495E')
    default_font = Font(name='맑은 고딕', size=10, color='2C3E50')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    # 본문(내용목록) 셀은 위아래 정렬을 '센터'로 통일
    left_top = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')
    header_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    label_fill = PatternFill(start_color="D5E8F0", end_color="D5E8F0", fill_type="solid")
    table_header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")

    # 결재란 이미지 생성 함수
    def create_approval_image(canvas_width=None, scale=4/3):
        base_w, base_h = 300, 117
        width = int(canvas_width) if canvas_width else int(base_w * scale)
        content_w = min(int(base_w * scale), width)
        height = int(content_w * (base_h / base_w))
        img = Image.new('RGB', (width, height), 'white')
        draw = ImageDraw.Draw(img)
        try:
            font_title = ImageFont.truetype("malgunbd.ttf", 15)
            font_small = ImageFont.truetype("malgun.ttf", 10)
        except Exception:
            font_title = ImageFont.load_default(); font_small = font_title
        margin_left = width - content_w
        box_w = content_w // 3
        header_h = int(height * 0.28)
        for i in range(3):
            x0 = margin_left + i * box_w
            draw.rectangle([x0, 0, x0 + box_w, header_h], fill='#E8F4F8', outline='black', width=2)
        draw.rectangle([margin_left, 0, margin_left + content_w - 1, height - 1], outline='#2C3E50', width=3)
        draw.line([margin_left + box_w, 0, margin_left + box_w, height], fill='#2C3E50', width=2)
        draw.line([margin_left + box_w*2, 0, margin_left + box_w*2, height], fill='#2C3E50', width=2)
        draw.line([margin_left, header_h, margin_left + content_w, header_h], fill='#2C3E50', width=2)
        for i, label in enumerate(["작성","검토","승인"]):
            x0 = margin_left + i * box_w
            xr = x0 + box_w - 10
            draw.text((xr, header_h//2), label, fill='#2C3E50', font=font_title, anchor='rm')
            sy = header_h + (height - header_h)//2 + 15
            draw.text((xr, sy), "(인)", fill='#95A5A6', font=font_small, anchor='rm')
        # tempfile 대신 프로젝트 폴더 사용
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        temp_dir = os.path.join(PROJECT_ROOT, 'data', 'temp')
        os.makedirs(temp_dir, exist_ok=True)
        temp_img_path = os.path.join(temp_dir, f'approval_orig_{timestamp}.png')
        img.save(temp_img_path, 'PNG')
        return temp_img_path

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "생산처방"

        # 제목 & 결재란
        ws.merge_cells('A1:F2')
        tc = ws['A1']
        tc.value = "생산처방서\nProduction Formulation Sheet"
        tc.font = title_font; tc.alignment = center
        tc.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        # 결재 이미지는 컬럼 너비 산정 후 추가(오른쪽을 I열 경계에 맞춤)

        # 기본정보
        details = production_data.get('details', {})
        # '제조일'이 없을 경우 구버전 키인 '적용일'을 폴백으로 사용
        manufacture_date = details.get("제조일", details.get("적용일", ""))
        left_pairs = [
            ("제품명", details.get("제품명", "")),
            ("LAB NO.", details.get("LAB NO.", "")),
            ("거래처", details.get("거래처", "")),
            ("제조일", manufacture_date),
            ("승인자", details.get("승인자", "")),
        ]
        right_pairs = [
            ("생산코드", details.get("생산코드", "")),
            ("차수", details.get("차수", "")),
            ("생산량(kg)", details.get("생산량(kg)", "")),
            ("상태", details.get("상태", "")),
            ("결제방", details.get("결제방", "")),
            ("출력일시", details.get("출력일시", "")),
        ]
        r = 4
        for i in range(max(len(left_pairs), len(right_pairs))):
            ws.row_dimensions[r].height = 20
            if i < len(left_pairs):
                label, value = left_pairs[i]
                lc = ws.cell(row=r, column=1, value=label)
                lc.font = label_font; lc.fill = label_fill; lc.alignment = center; lc.border = thin_border
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
                vc = ws.cell(row=r, column=2, value=value)
                vc.font = default_font; vc.alignment = center; vc.border = thin_border
                for c in range(2,6): ws.cell(row=r, column=c).border = thin_border
            else:
                for c in range(1,6): ws.cell(row=r, column=c).border = thin_border
            if i < len(right_pairs):
                label, value = right_pairs[i]
                rc = ws.cell(row=r, column=7, value=label)
                rc.font = label_font; rc.fill = label_fill; rc.alignment = center; rc.border = thin_border
                rv = ws.cell(row=r, column=8, value=value)
                rv.font = default_font; rv.alignment = center; rv.border = thin_border
            else:
                for c in range(7,9): ws.cell(row=r, column=c).border = thin_border
            ws.cell(row=r, column=6).border = thin_border
            r += 1

        # 비고
        if details.get('비고'):
            ws.row_dimensions[r].height = 24
            nl = ws.cell(row=r, column=1, value="비고\nNote")
            nl.font = label_font; nl.fill = label_fill; nl.alignment = center; nl.border = thin_border
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
            nv = ws.cell(row=r, column=2, value=details.get('비고'))
            nv.font = default_font; nv.alignment = center; nv.border = thin_border
            for c in range(2,9): ws.cell(row=r, column=c).border = thin_border
            r += 1

        ws.row_dimensions[r].height = 8; r += 1

        # 테이블 헤더
        headers = ["Ph.", "구분", "코드", "원료명", "함량(%)", "생산량(kg)", "계량량(kg)", "제조공정", "공정검사"]
        header_row = r
        ws.row_dimensions[header_row].height = 22
        for idx, h in enumerate(headers, 1):
            hc = ws.cell(row=header_row, column=idx, value=h)
            hc.font = header_font_white; hc.fill = table_header_fill; hc.alignment = center; hc.border = medium_border

        # 데이터
        current = header_row + 1
        total_ratio = 0.0
        phase_merge_info = []
        current_phase = None
        phase_start_row = None
        phase_max_lines = 1
        for item in production_data.get('items', []):
            ph_val = str(item.get('Ph.', '')).strip()
            proc_text = str(item.get('제조공정', '')).strip()
            insp_text = str(item.get('공정검사', '')).strip()
            proc_lines = proc_text.count('\n') + 1 if proc_text else 1
            insp_lines = insp_text.count('\n') + 1 if insp_text else 1
            max_lines_row = max(proc_lines, insp_lines)
            if ph_val:
                if current_phase is not None and phase_start_row is not None:
                    phase_merge_info.append((current_phase, phase_start_row, current - 1, phase_max_lines))
                current_phase = ph_val; phase_start_row = current; phase_max_lines = max_lines_row
            else:
                phase_max_lines = max(phase_max_lines, max_lines_row)

            qty_kg = item.get('생산량(kg)')
            if qty_kg in (None, ""):
                try:
                    g = float(item.get('기준중량(g)')); qty_kg = g/1000.0
                except Exception:
                    qty_kg = item.get('생산량(kg)')
            weigh_kg = item.get('계량량(kg)')  # 수기 기입 전제(자동 채움 없음)

            vals = [ph_val if ph_val else "", item.get('구분'), item.get('코드'), item.get('원료명'), item.get('함량(%)'), qty_kg, weigh_kg, item.get('제조공정'), item.get('공정검사')]
            row_fill = PatternFill(start_color="F8FBFD", end_color="F8FBFD", fill_type="solid") if (current - header_row) % 2 == 0 else None
            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=current, column=c_idx, value=val)
                cell.border = thin_border; cell.font = default_font
                if row_fill and c_idx not in [1,8,9]: cell.fill = row_fill
                if c_idx in (1,2,3):
                    cell.alignment = center
                elif c_idx == 4:
                    cell.alignment = left
                elif c_idx in (5,6,7):
                    try:
                        fval = float(val); cell.value = fval
                        if c_idx == 5:
                            cell.number_format = '0.0000'; total_ratio += fval
                        else:
                            cell.number_format = '#,##0.0'
                    except Exception:
                        pass
                    cell.alignment = right
                else:
                    if val:
                        cleaned = str(val).replace('"','').replace("'", '')
                        cell.value = cleaned
                    # 내용목록 텍스트 셀도 세로 중앙 정렬
                    cell.alignment = left
            ws.row_dimensions[current].height = 16
            current += 1

        if current_phase is not None and phase_start_row is not None:
            phase_merge_info.append((current_phase, phase_start_row, current - 1, phase_max_lines))

        for phase_val, start_row, end_row, max_lines in phase_merge_info:
            num_rows = end_row - start_row + 1
            total_h = max(16, max_lines * 14 + 2)
            row_h = total_h / num_rows
            for rr in range(start_row, end_row + 1): ws.row_dimensions[rr].height = row_h
            if start_row < end_row:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                mc = ws.cell(row=start_row, column=1); mc.value = phase_val; mc.alignment = center; mc.border = thin_border; mc.font = header_font
                ws.merge_cells(start_row=start_row, start_column=8, end_row=end_row, end_column=8)
                pc = ws.cell(row=start_row, column=8); pc.alignment = left; pc.border = thin_border
                ws.merge_cells(start_row=start_row, start_column=9, end_row=end_row, end_column=9)
                ic = ws.cell(row=start_row, column=9); ic.alignment = left; ic.border = thin_border
            else:
                ws.row_dimensions[start_row].height = total_h

        # 합계 행
        sum_row = current
        ws.row_dimensions[sum_row].height = 22
        ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=4)
        sl = ws.cell(row=sum_row, column=1, value="합계 (Total)")
        sl.font = header_font_white; sl.fill = table_header_fill; sl.alignment = center; sl.border = medium_border
        sr = ws.cell(row=sum_row, column=5, value=total_ratio)
        sr.font = header_font_white; sr.fill = table_header_fill; sr.number_format = '0.0000'; sr.alignment = right; sr.border = medium_border
        for col in [1,2,3,4]: ws.cell(row=sum_row, column=col).border = medium_border
        for col in [6,7,8]:
            ec = ws.cell(row=sum_row, column=col); ec.fill = table_header_fill; ec.border = medium_border

        # 필터/고정
        ws.auto_filter.ref = f"A{header_row}:I{sum_row}"
        ws.freeze_panes = f"A{header_row+1}"

        # 컬럼 너비
        for col_idx, col_letter in enumerate(['A','B','C','D','E','F','G','H','I'], 1):
            max_w = 10
            try:
                hc = ws.cell(row=header_row, column=col_idx)
                if hc.value: max_w = max(max_w, len(str(hc.value)) * 1.2)
            except Exception: pass
            for row_idx in range(header_row + 1, sum_row):
                try:
                    cv = ws.cell(row=row_idx, column=col_idx).value
                    if cv:
                        s = str(cv)
                        if '\n' in s:
                            cell_len = max(len(line) for line in s.split('\n')) * 1.1
                        else:
                            cell_len = len(s) * 1.1
                        max_w = max(max_w, cell_len)
                except Exception: pass
            if col_letter == 'H':
                max_w = 26.25
            elif col_letter == 'I':
                max_w = 17.25
            elif col_letter == 'D':
                max_w = min(max_w, 50); max_w = max(max_w, 30)
            else:
                max_w = min(max_w, 20)
            ws.column_dimensions[col_letter].width = max_w

        # 결재 이미지: 기본은 페이지 머리글 '중앙'(Excel COM), COM 불가 시 워크시트 A1 폴백
        approval_img_path = None
        full_px = None
        use_com_header = False
        try:
            import win32com.client as _win32  # type: ignore
            use_com_header = True
        except Exception:
            use_com_header = False
        try:
            # 전체 열(A..I) 폭 기준으로 생성하여 헤더 중앙에 넣는다
            total_width_px = 0
            for col_idx in range(1, 10):
                col_letter = get_column_letter(col_idx)
                col_w = ws.column_dimensions[col_letter].width or 8
                total_width_px += col_w * 7
            full_px = int(total_width_px)
            # 결재란 이미지 생성 및 삽입
            approval_img_path = create_approval_image(full_px)
            if not use_com_header:
                img = XLImage(approval_img_path)
                img.anchor = 'A1'
                ws.add_image(img)
                ws.row_dimensions[1].height = 90
                ws.row_dimensions[2].height = 150
        except Exception:
            pass
        try:
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.print_options.horizontalCentered = True
            ws.page_margins.left = 0.3; ws.page_margins.right = 0.3
            # 위쪽 여백: 1.5cm 기본
            ws.page_margins.top = 1.5 / 2.54
            ws.page_margins.bottom = 0.5  # cm -> inch
        except Exception:
            pass

        # 요청에 따라 엑셀 내보내기 시 추가 시트(제조공정, 원료 환산)는 생성하지 않습니다.

        wb.save(file_path)

        # 저장 후, COM 헤더 이미지 삽입 시도 (머리글 중앙)
        if use_com_header and approval_img_path and os.path.exists(approval_img_path):
            try:
                import win32com.client as win32  # type: ignore
                excel = win32.Dispatch("Excel.Application")
                excel.Visible = bool(open_print_preview)
                wb_com = excel.Workbooks.Open(os.path.abspath(file_path))
                ws_com = wb_com.ActiveSheet
                ws_com.PageSetup.CenterHeader = "&G"
                ws_com.PageSetup.CenterHeaderPicture.Filename = os.path.abspath(approval_img_path)
                try:
                    if full_px:
                        ws_com.PageSetup.CenterHeaderPicture.Width = int(full_px * 0.75)
                except Exception:
                    pass
                wb_com.Save()
                if open_print_preview:
                    try:
                        ws_com.PrintPreview()
                    except Exception:
                        pass
                wb_com.Close(SaveChanges=False)
                excel.Quit()
            except Exception:
                pass
            finally:
                try:
                    if os.path.exists(approval_img_path):
                        os.remove(approval_img_path)
                except Exception:
                    pass

        if open_print_preview:
            try:
                import win32com.client as win32  # type: ignore
                excel = win32.Dispatch("Excel.Application"); excel.Visible = True
                wb_com = excel.Workbooks.Open(os.path.abspath(file_path))
                try:
                    wb_com.ActiveSheet.PrintPreview()
                finally:
                    wb_com.Close(SaveChanges=False); excel.Quit()
            except Exception:
                try:
                    os.startfile(os.path.abspath(file_path))
                except Exception:
                    pass
        else:
            try:
                os.startfile(os.path.abspath(file_path))
            except Exception:
                pass
    except Exception as e:
        messagebox.showerror("내보내기 오류", f"파일 저장 중 오류가 발생했습니다: {e}")

    

def export_data_to_excel(headers, data_rows, default_filename="export.xlsx"):
    """헤더와 데이터 행들을 단일 시트 엑셀 파일로 내보냅니다."""
    initial_dir = get_excel_path()
    timestamped_filename = get_timestamped_filename(default_filename)
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialdir=initial_dir,
        initialfile=timestamped_filename,
        title="데이터 저장"
    )
    if not file_path:
        return

    save_excel_path(os.path.dirname(file_path))

    # --- 스타일 정의 ---
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    header_font = Font(name='맑은 고딕', size=11, bold=True)
    default_font = Font(name='맑은 고딕', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    try:
        workbook = Workbook()
        sheet = workbook.active

        # 헤더 쓰기 및 스타일 적용
        for col_idx, header_text in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # 데이터 쓰기 및 스타일 적용
        for row_idx, row in enumerate(data_rows, 2):
            for col_idx, cell_value in enumerate(row, 1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.font = default_font
                cell.border = thin_border
                cell.alignment = left_align

        # 컬럼 너비 자동 조절
        for col in sheet.columns:
            max_length = 0
            safe_cell = next((c for c in col if not isinstance(c, openpyxl.cell.cell.MergedCell)), None)
            if not safe_cell:
                continue
            column_letter = safe_cell.column_letter
            for cell in col:
                if cell.value:
                    # _get_display_length 함수를 사용하여 한글/영문 길이를 고려
                    length = _get_display_length(cell.value)
                    if length > max_length:
                        max_length = length
            adjusted_width = max_length + 2
            if adjusted_width > 5:
                sheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(file_path)
        try:
            os.startfile(os.path.abspath(file_path))
        except Exception:
            pass
    except Exception as e:
        messagebox.showerror("오류", f"파일 저장 중 오류가 발생했습니다: {e}")

def export_multisheet_data_to_excel(sheets_data, default_filename="export.xlsx"):
    """데이터를 여러 시트를 가진 엑셀 파일로 내보냅니다."""
    initial_dir = get_excel_path()
    timestamped_filename = get_timestamped_filename(default_filename)
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialdir=initial_dir,
        initialfile=timestamped_filename,
        title="데이터 저장"
    )
    if not file_path:
        return

    save_excel_path(os.path.dirname(file_path))

    # --- 스타일 정의 ---
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    header_font = Font(name='맑은 고딕', size=11, bold=True)
    default_font = Font(name='맑은 고딕', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    try:
        workbook = Workbook()
        workbook.remove(workbook.active)
        for sheet_name, sheet_content in sheets_data.items():
            sheet = workbook.create_sheet(title=sheet_name)
            headers = sheet_content.get('headers', [])
            data_rows = sheet_content.get('data', [])
            apply_style = sheet_content.get('style', False)

            # 헤더 쓰기 및 스타일 적용
            for col_idx, header_text in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_idx, value=header_text)
                if apply_style:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border

            # 데이터 쓰기 및 스타일 적용
            for row_idx, row in enumerate(data_rows, 2):
                for col_idx, cell_value in enumerate(row, 1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    if apply_style:
                        cell.font = default_font
                        cell.border = thin_border
                        cell.alignment = left_align

            # 컬럼 너비 자동 조절
            for column_cells in sheet.columns:
                max_length = 0
                safe_cell = next((c for c in column_cells if not isinstance(c, openpyxl.cell.cell.MergedCell)), None)
                if not safe_cell:
                    continue
                column_letter = safe_cell.column_letter
                for cell in column_cells:
                    if cell.value is not None:
                        length = _get_display_length(cell.value)
                        if length > max_length:
                            max_length = length
                adjusted_width = max_length + 2
                if adjusted_width > 5: # 최소 너비
                    sheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(file_path)
        try:
            os.startfile(os.path.abspath(file_path))
        except Exception:
            pass
    except Exception as e:
        messagebox.showerror("오류", f"파일 저장 중 오류가 발생했습니다: {e}")


def export_all_change_logs(sheets, default_filename="change_logs.xlsx"):
    """
    이미 준비된 sheets 구조(엔티티별 headers/data)를 받아서 멀티시트 엑셀로 저장합니다.
    호출자는 DB에서 change_log를 모아 sheets dict를 만들어 전달하면 됩니다.
    """
    # 단순히 기존 export_multisheet_data_to_excel과 동일한 동작을 수행
    return export_multisheet_data_to_excel(sheets, default_filename=default_filename)

def export_formulation_template(formulation_data, default_filename="formulation.xlsx", is_lab_journal=False):
    """처방 정보를 특정 템플릿 형식의 엑셀 파일로 내보냅니다."""
    initial_dir = get_excel_path()
    timestamped_filename = get_timestamped_filename(default_filename)
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialdir=initial_dir,
        initialfile=timestamped_filename,
        title="처방 내보내기"
    )
    if not file_path:
        return

    save_excel_path(os.path.dirname(file_path))

    # --- 스타일 정의 ---
    # 테두리
    thin = Side(style='thin')
    medium = Side(style='medium')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    medium_border = Border(left=medium, right=medium, top=medium, bottom=medium)
    
    # 폰트
    title_font = Font(name='맑은 고딕', size=18, bold=True)
    header_font = Font(name='맑은 고딕', size=11, bold=True)
    default_font = Font(name='맑은 고딕', size=10)
    total_font = Font(name='맑은 고딕', size=10, bold=True)

    # 정렬
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')

    # 채우기
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    def set_column_widths_for_sheet(sheet):
        """시트의 컬럼 너비를 설정하는 헬퍼 함수"""
        sheet.column_dimensions['A'].width = 18
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 40
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 25 # 비고 열 추가

    try:
        wb = Workbook()
        wb.remove(wb.active) # 기본 시트 제거

        # --- 상세 정보 쓰기 ---
        def write_sheet_data(sheet, sheet_details, sheet_items, set_column_widths=True, is_target_sheet=False):
            """
            시트에 상세 정보와 아이템 목록을 쓰는 헬퍼 함수
            is_lab_journal 플래그에 따라 헤더와 내용을 다르게 처리합니다.
            """
            
            def apply_border_to_range(cell_range, border_style):
                """지정된 범위의 모든 셀에 테두리를 적용합니다."""
                rows = sheet[cell_range]
                for row in rows:
                    for cell in row:
                        cell.border = border_style

            # --- 문서 제목 ---
            if is_lab_journal:
                # 실험일지는 전체 너비(J열까지)로 병합
                sheet.merge_cells('A1:J2')
            else:
                # 처방전은 결재란을 제외하고 병합
                sheet.merge_cells('A1:C2')
            
            title_cell = sheet['A1'] # 병합된 셀의 첫 번째 셀
            if is_lab_journal:
                title_cell.value = "실험일지 (Lab Journal)"
            else:
                title_cell.value = "타겟 정보 (Target Information)" if is_target_sheet else "처방전 (Formulation Sheet)"
            title_cell.font = title_font; title_cell.alignment = center_align
            
            # --- 결재란 추가 (실험일지, 타겟 정보 시트에는 없음) ---
            if not is_target_sheet and not is_lab_journal:
                approval_labels = ["작성", "검토", "승인"]
                for i, label in enumerate(approval_labels):
                    col_idx = i + 4 # D, E, F 열로 수정
                    sheet.column_dimensions[chr(ord('A') + col_idx - 1)].width = 15 # 결재란 각 칸 너비 동일하게
                    # 직책 셀
                    label_cell = sheet.cell(row=1, column=col_idx)
                    label_cell.value = label
                    label_cell.font = header_font
                    label_cell.alignment = center_align
                    label_cell.border = thin_border
                    # 서명 셀
                    sign_cell = sheet.cell(row=2, column=col_idx)
                    sign_cell.border = thin_border
                    sheet.row_dimensions[2].height = 40 # 서명 공간 높이
            
                # 결재란 전체에 중간 테두리 적용
                approval_range = f"D1:F2"
                for row in sheet[approval_range]:
                    for cell in row:
                        cell.border = thin_border


            # --- 상단 정보 섹션 (실험일지에는 없음) ---
            row_idx = 4
            if is_target_sheet:
                info_layout = [
                    ("타겟 샘플명", sheet_details.get("타겟 샘플명")),
                    ("타겟 거래처", sheet_details.get("타겟 거래처")),
                ]
            else:
                info_layout = [
                    ("실험품명", sheet_details.get("실험품명")),
                    ("실험년월일", sheet_details.get("실험년월일")),
                    ("담당자", sheet_details.get("담당자")),
                    ("거래처", sheet_details.get("거래처")),
                    ("LAB NO.", str(sheet_details.get("LAB NO.") or "").upper()), ("차수", str(sheet_details.get("차수") or "").upper()),
                    ("담당번호", str(sheet_details.get("담당번호") or "").upper()), ("총 실험량", sheet_details.get("총 실험량")),
                ]

            if not is_lab_journal:
                for i in range(0, len(info_layout), 2):
                    sheet.row_dimensions[row_idx].height = 25
                    # 레이블
                    cell_a = sheet[f"A{row_idx}"]; cell_a.value = info_layout[i][0]; cell_a.font = header_font; cell_a.fill = header_fill; cell_a.alignment = center_align
                    if i + 1 < len(info_layout):
                        cell_d = sheet[f"D{row_idx}"]; cell_d.value = info_layout[i+1][0]; cell_d.font = header_font; cell_d.fill = header_fill; cell_d.alignment = center_align
                    # 값
                    sheet.merge_cells(f"B{row_idx}:C{row_idx}")
                    cell_b = sheet[f"B{row_idx}"]; cell_b.value = info_layout[i][1]; cell_b.alignment = left_align
                    if i + 1 < len(info_layout):
                        cell_e = sheet[f"E{row_idx}"]; cell_e.value = info_layout[i+1][1]; cell_e.alignment = left_align
                    row_idx += 1
                # 상단 정보 섹션 테두리
                apply_border_to_range(f"A4:F{row_idx-1}", thin_border)

            # --- 타겟 정보 시트 처리 ---
            if is_target_sheet:
                row_idx += 1 # 한 줄 띄우기
                result_start_row = row_idx

                # pH 행
                sheet.row_dimensions[row_idx].height = 25
                sheet[f"A{row_idx}"].value = "타겟 pH"; sheet[f"A{row_idx}"].font = header_font; sheet[f"A{row_idx}"].fill = header_fill; sheet[f"A{row_idx}"].alignment = center_align
                sheet[f"B{row_idx}"].value = f"당일: {sheet_details.get('타겟 pH (당일)', '')}"; sheet[f"B{row_idx}"].alignment = left_align
                sheet[f"C{row_idx}"].value = f"익일: {sheet_details.get('타겟 pH (익일)', '')}"; sheet[f"C{row_idx}"].alignment = left_align
                row_idx += 1

                # 점도 행
                sheet.row_dimensions[row_idx].height = 25
                sheet[f"A{row_idx}"].value = "타겟 점도"; sheet[f"A{row_idx}"].font = header_font; sheet[f"A{row_idx}"].fill = header_fill; sheet[f"A{row_idx}"].alignment = center_align
                sheet[f"B{row_idx}"].value = f"당일: {sheet_details.get('타겟 점도 (당일)', '')}"; sheet[f"B{row_idx}"].alignment = left_align
                sheet[f"C{row_idx}"].value = f"익일: {sheet_details.get('타겟 점도 (익일)', '')}"; sheet[f"C{row_idx}"].alignment = left_align
                row_idx += 1

                # 사용핀 및 기계 행
                sheet.row_dimensions[row_idx].height = 25
                sheet[f"A{row_idx}"].value = "사용핀 및 기계"; sheet[f"A{row_idx}"].font = header_font; sheet[f"A{row_idx}"].fill = header_fill; sheet[f"A{row_idx}"].alignment = center_align
                sheet.merge_cells(f"B{row_idx}:F{row_idx}")
                sheet[f"B{row_idx}"].value = sheet_details.get("사용핀 및 기계", ""); sheet[f"B{row_idx}"].alignment = left_align
                row_idx += 1

                # 타겟 결과 섹션 테두리 (F열까지)
                apply_border_to_range(f"A{result_start_row}:F{row_idx-1}", thin_border)
                
                if set_column_widths: set_column_widths_for_sheet(sheet)
                return # 타겟 시트는 여기서 종료

            # --- 처방 정보 시트 처리 (아래는 is_target_sheet=False 일 때만 실행) ---
            
            # --- 처방 내용 헤더 ---
            if is_lab_journal:
                item_headers = ["실험 날짜", "품명", "pH", "점도", "비중", "Pin", "실험번호", "업체", "샘플 전달", "기타"]
                item_header_row_num = 4 # 상단 정보가 없으므로 4행부터 시작
            else:
                item_headers = ["구분", "코드", "원료명", "함량(%)", "실험량(g)", "비고"]
                item_header_row_num = row_idx + 1
            sheet.row_dimensions[item_header_row_num].height = 25
            for col_idx, header_text in enumerate(item_headers, 1):
                cell = sheet.cell(row=item_header_row_num, column=col_idx, value=header_text) # noqa
                cell.border = thin_border; cell.font = header_font; cell.alignment = center_align; cell.fill = header_fill
                # 실험일지 너비 자동 조절을 위해 헤더 길이도 계산에 포함
                if is_lab_journal:
                    sheet.column_dimensions[cell.column_letter].max_len = len(header_text)

            # --- 처방 내용 데이터 ---
            current_item_row = item_header_row_num + 1
            total_ratio = 0.0; total_amount = 0.0 # noqa
            for item in sheet_items:
                sheet.row_dimensions[current_item_row].height = 20
                code_value = item.get("코드", "")
                # 줄내림 체크: 코드가 ---, -, --, ―, ㅡ 중 하나인지 확인
                is_separator = isinstance(code_value, str) and code_value.strip() in ["---", "-", "--", "―", "ㅡ"]

                if is_lab_journal:
                    item_values = [item.get(h) for h in item_headers]
                else:
                    # 줄내림일 경우 함량과 실험량을 빈 문자열로 처리
                    item_values = [item.get("구분"), code_value, item.get("원료명"), 
                                   "" if is_separator else try_convert_to_float(item.get("함량(%)")),
                                   "" if is_separator else try_convert_to_float(item.get("실험량(g)")), ""] # 비고 칸 추가

                for col_idx, value in enumerate(item_values, 1):
                    cell = sheet.cell(row=current_item_row, column=col_idx, value=value)
                    cell.border = thin_border; cell.font = default_font
                    if is_lab_journal:
                        # 실험일지는 모든 셀을 가운데 정렬
                        cell.alignment = center_align
                        # 컬럼 너비 자동 조절을 위해 최대 길이 계산
                        if not hasattr(sheet.column_dimensions[cell.column_letter], 'max_len'): # noqa
                            sheet.column_dimensions[cell.column_letter].max_len = 0
                        sheet.column_dimensions[cell.column_letter].max_len = max(sheet.column_dimensions[cell.column_letter].max_len, len(str(value)))
                    else:
                        if is_separator:
                            cell.fill = total_fill # 합계 행과 같은 회색 배경
                            cell.alignment = center_align
                        else:
                            if col_idx in [1, 2]: cell.alignment = center_align
                            elif col_idx == 3: cell.alignment = left_align
                            else: 
                                cell.alignment = right_align
                                if isinstance(value, (int, float)):
                                    cell.number_format = '0.0000'
                                    if col_idx == 4: total_ratio += value
                                    if col_idx == 5: total_amount += value
                current_item_row += 1

            # --- 합계 행 추가 ---
            total_row = current_item_row
            sheet.row_dimensions[total_row].height = 25
            sheet.merge_cells(f'A{total_row}:C{total_row}')
            total_label_cell = sheet[f'A{total_row}']
            total_label_cell.value = "합계 (Total)"; total_label_cell.font = total_font; total_label_cell.alignment = center_align; total_label_cell.fill = total_fill

            total_ratio_cell = sheet[f'D{total_row}']
            total_ratio_cell.value = total_ratio; total_ratio_cell.font = total_font; total_ratio_cell.alignment = right_align; total_ratio_cell.number_format = '0.0000'; total_ratio_cell.fill = total_fill

            total_amount_cell = sheet[f'E{total_row}']
            total_amount_cell.value = total_amount; total_amount_cell.font = total_font; total_amount_cell.alignment = right_align; total_amount_cell.number_format = '0.0000'; total_amount_cell.fill = total_fill

            for col_char in "ABCDEF": sheet[f'{col_char}{total_row}'].border = thin_border
            
            # 실험일지인 경우 합계 행과 실험 결과 섹션을 숨깁니다.
            if is_lab_journal:
                sheet.row_dimensions[total_row].hidden = True
            else:
                # --- 실험 결과 섹션 ---
                row_idx = total_row + 2
                result_start_row = row_idx

                # pH 행
                sheet.row_dimensions[row_idx].height = 25
                sheet[f"A{row_idx}"].value = "pH"; sheet[f"A{row_idx}"].font = header_font; sheet[f"A{row_idx}"].fill = header_fill; sheet[f"A{row_idx}"].alignment = center_align
                sheet[f"B{row_idx}"].value = f"당일: {sheet_details.get('pH (당일)', '')}"; sheet[f"B{row_idx}"].alignment = left_align
                sheet[f"C{row_idx}"].value = f"익일: {sheet_details.get('pH (익일)', '')}"; sheet[f"C{row_idx}"].alignment = left_align
                row_idx += 1

                # 점도 행
                sheet.row_dimensions[row_idx].height = 25
                sheet[f"A{row_idx}"].value = "점도"; sheet[f"A{row_idx}"].font = header_font; sheet[f"A{row_idx}"].fill = header_fill; sheet[f"A{row_idx}"].alignment = center_align
                sheet[f"B{row_idx}"].value = f"당일: {sheet_details.get('점도 (당일)', '')}"; sheet[f"B{row_idx}"].alignment = left_align
                sheet[f"C{row_idx}"].value = f"익일: {sheet_details.get('점도 (익일)', '')}"; sheet[f"C{row_idx}"].alignment = left_align
                row_idx += 1

                # 사용핀 및 기계 행
                sheet.row_dimensions[row_idx].height = 25
                sheet[f"A{row_idx}"].value = "사용핀 및 기계"; sheet[f"A{row_idx}"].font = header_font; sheet[f"A{row_idx}"].fill = header_fill; sheet[f"A{row_idx}"].alignment = center_align
                sheet.merge_cells(f"B{row_idx}:F{row_idx}")
                sheet[f"B{row_idx}"].value = sheet_details.get("사용핀 및 기계", ""); sheet[f"B{row_idx}"].alignment = left_align
                row_idx += 1

                # 품평결과 및 특이사항 행
                sheet.row_dimensions[row_idx].height = 80
                sheet[f"A{row_idx}"].value = "품평결과 및 특이사항"; sheet[f"A{row_idx}"].font = header_font; sheet[f"A{row_idx}"].fill = header_fill; sheet[f"A{row_idx}"].alignment = center_align
                sheet.merge_cells(f"B{row_idx}:F{row_idx}")
                sheet[f"B{row_idx}"].value = sheet_details.get("품평결과 및 특이사항", ""); sheet[f"B{row_idx}"].alignment = left_align
                row_idx += 1

                # 실험 결과 섹션 테두리 (F열까지)
                apply_border_to_range(f"A{result_start_row}:F{row_idx-1}", thin_border)

            # 처방 내용 섹션 테두리 (실험일지일 경우 F열이 아닌 마지막 열까지)
            end_col_char = 'J' if is_lab_journal else 'F'
            apply_border_to_range(f"A{item_header_row_num}:{end_col_char}{total_row}", thin_border)

            # --- 컬럼 너비 설정 ---
            if set_column_widths:
                if is_lab_journal:
                    for col_dim in sheet.column_dimensions.values():
                        if hasattr(col_dim, 'max_len'):
                            col_dim.width = col_dim.max_len + 5
                else:
                    # 컬럼 너비 자동 조절 (개선된 로직)
                    for column_cells in sheet.columns:
                        max_length = 0
                        safe_cell = next((c for c in column_cells if not isinstance(c, openpyxl.cell.cell.MergedCell)), None)
                        if not safe_cell:
                            continue
                        column_letter = safe_cell.column_letter
                        for cell in column_cells:
                            if cell.value is not None:
                                length = _get_display_length(cell.value)
                                if length > max_length:
                                    max_length = length
                        adjusted_width = max_length + 2
                        if adjusted_width > 2:
                            sheet.column_dimensions[column_letter].width = adjusted_width

        details = formulation_data.get("details", {}); items = formulation_data.get("items", [])
        target_details = formulation_data.get("target_details")
        if target_details:
            # 타겟 정보가 있을 경우: 2개 시트 생성
            ws_formulation = wb.create_sheet("처방 정보")
            write_sheet_data(ws_formulation, details, items)

            ws_target = wb.create_sheet("타겟 정보")
            # 타겟 정보 시트에는 처방 내용(items)이 없으므로 빈 리스트 전달
            write_sheet_data(ws_target, target_details, [], is_target_sheet=True)
        else:
            # 타겟 정보가 없을 경우: 기존 방식대로 1개 시트 생성
            ws_formulation = wb.create_sheet("처방 정보")
            write_sheet_data(ws_formulation, details, items)

        wb.save(file_path)
        try:
            os.startfile(os.path.abspath(file_path))
        except Exception:
            pass

    except Exception as e:
        messagebox.showerror("내보내기 오류", f"파일 저장 중 오류가 발생했습니다: {e}")
        # 오류 발생 시에도 파일은 생성되도록 복구 로직 추가
        details = formulation_data.get("details", {}); items = formulation_data.get("items", [])
        target_details = formulation_data.get("target_details")
        if target_details:
            # 타겟 정보가 있을 경우: 2개 시트 생성 (실험일지 모드에서는 실행되지 않음)
            ws_formulation = wb.create_sheet("처방 정보")
            write_sheet_data(ws_formulation, details, items)

            ws_target = wb.create_sheet("타겟 정보")
            # 타겟 정보 시트에는 처방 내용(items)이 없으므로 빈 리스트 전달
            write_sheet_data(ws_target, target_details, [], is_target_sheet=True)
        elif is_lab_journal:
            ws_journal = wb.create_sheet("실험일지")
            write_sheet_data(ws_journal, details, items, is_lab_journal=True)
        else:
            ws_formulation = wb.create_sheet("처방 정보")
            write_sheet_data(ws_formulation, details, items)
        wb.save(file_path)

def try_convert_to_float(value):
    """값을 float으로 변환 시도, 실패 시 원래 값 반환. 줄내림(---) 체크"""
    if value is None:
        return None
    # 줄내림 구분자 체크 (---, -, 등)
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in ["---", "-", "--", "―", "ㅡ"]:
            return "---"  # 줄내림으로 명시적 반환
    try:
        val = float(value)
        # 10자리 정도에서 반올림하여 부동소수점 오차 제거 (45.57 -> 45.569999...)
        # 안전하게 8자리 사용
        return round(val, 8)
    except (ValueError, TypeError):
        return value

def import_data(file_path: str | None = None) -> list[dict] | None:
    """단일 시트 엑셀 파일에서 데이터를 읽어 딕셔너리 리스트로 반환합니다."""
    if file_path is None:
        initial_dir = get_excel_path()
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=initial_dir,
            title="가져올 엑셀 파일 선택"
        )
        if not file_path:
            return None

    save_excel_path(os.path.dirname(file_path))

    def clean_cell(cell):
        if cell is None:
            return ""
        if isinstance(cell, str):
            cell = cell.strip()
            if cell == "-":
                return ""
        return cell

    try:
        workbook = _load_workbook_robust(file_path, data_only=True)
        sheet = workbook.active
        
        headers = [cell.value for cell in sheet[1]]
        headers = [str(h).strip() if h is not None else f"col_{idx}" for idx, h in enumerate(headers)]
        
        data_list = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(cell is None or str(cell).strip() in ("", "-") for cell in row):
                continue
            row = list(row) + [None] * (len(headers) - len(row))
            row = [clean_cell(cell) for cell in row]
            row_data = dict(zip(headers, row))
            data_list.append(row_data)
            
        return data_list
    except PermissionError:
        messagebox.showerror("오류", f"파일이 열려 있습니다.\n파일을 닫고 다시 시도해주세요.\n\n경로: {file_path}")
        return None
    except Exception as e:
        import traceback
        traceback.print_exc()
        messagebox.showerror("오류", f"파일을 읽는 중 오류가 발생했습니다: {e}")
        return None

def import_multisheet_data():
    """여러 시트를 가진 엑셀 파일에서 데이터를 가져옵니다."""
    initial_dir = get_excel_path()
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx")],
        initialdir=initial_dir,
        title="가져올 엑셀 파일 선택"
    )
    if not file_path:
        return None

    save_excel_path(os.path.dirname(file_path))

    def clean_cell(cell):
        """
        셀 값 정리:
        - None 또는 '-' → 빈 문자열
        - 문자열이면 앞뒤 공백 제거
        """
        if cell is None:
            return ""
        if isinstance(cell, str):
            cell = cell.strip()
            if cell == "-":
                return ""
        return cell

    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        result = {}
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # 헤더 행 읽기
            headers = [cell.value for cell in sheet[1]]
            
            # 데이터 행 읽기
            data_list = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # 전부 비어있으면 건너뜀
                if all(cell is None or str(cell).strip() in ("", "-") for cell in row):
                    continue
                # 헤더 개수에 맞춰 길이 조정
                row = list(row) + [None] * (len(headers) - len(row))
                # 각 셀 값 정리
                row = [clean_cell(cell) for cell in row]
                row_data = dict(zip(headers, row))
                data_list.append(row_data)
            
            result[sheet_name] = data_list
        
        return result
    except PermissionError:
        messagebox.showerror("오류", f"파일이 열려 있습니다.\n파일을 닫고 다시 시도해주세요.\n\n경로: {file_path}")
        return None
    except Exception as e:
        import traceback
        traceback.print_exc()
        messagebox.showerror("오류", f"파일을 읽는 중 오류가 발생했습니다: {e}")
        return None

def export_formulation_blank_template():
    """빈 처방 템플릿을 내보냅니다."""
    formulation_data = {
        "details": {},
        "items": [],
        "target_details": {}
    }
    export_formulation_template(formulation_data, default_filename="처방_템플릿.xlsx")

def import_formulation_template():
    """(최종 수정) 특정 템플릿 형식의 엑셀 파일에서 처방 정보를 안정적으로 읽어옵니다."""
    initial_dir = get_excel_path()
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx")],
        initialdir=initial_dir,
        title="가져올 처방 파일 선택"
    )
    if not file_path:
        return None

    save_excel_path(os.path.dirname(file_path))

    try:
        wb = _load_workbook_robust(file_path, data_only=True)
        ws = wb["처방 정보"] if "처방 정보" in wb.sheetnames else wb.active

        formulation_data = {"details": {}, "items": []}

        # 읽기 상태를 관리하는 변수 (상세정보, 처방내용, 실험결과)
        reading_state = "details" # 'details', 'items', 'results'

        # 헤더를 저장할 리스트
        item_headers = []

        for row in ws.iter_rows(values_only=True):
            key_cell = row[0]
            key = str(key_cell).strip() if key_cell is not None else ""

            # 상태 전환 로직
            # 헤더 행은 '구분', '코드', '원료명', '함량(%)' 등을 포함하므로 더 명확하게 체크
            row_str_values = [str(c).strip() for c in row if c is not None]
            if "함량(%)" in row_str_values and "원료명" in row_str_values:
                reading_state = "items"
                item_headers = [h for h in row_str_values if h]
                continue # 헤더 행은 건너뜀
            elif key == "합계 (Total)":
                reading_state = "results"
                continue # 합계 행은 건너뜀

            # 상태에 따른 데이터 처리
            if reading_state == "details":
                # A, B열 키-값 쌍 (row[0], row[1])
                if key:
                    formulation_data["details"][key] = row[1]
                # D, E열 키-값 쌍 (row[3], row[4])
                if len(row) > 3 and row[3]:
                    key_d = str(row[3]).strip()
                    if key_d:
                        formulation_data["details"][key_d] = row[4]

            elif reading_state == "items":
                if all(c is None or str(c).strip() == "" for c in row): continue

                row_list = list(row)
                item_data = dict(zip(item_headers, row_list))

                # 줄내림 체크: 코드가 줄내림이면 함량과 실험량을 빈 문자열로 처리
                code_val = item_data.get('코드', '')
                is_separator = isinstance(code_val, str) and code_val.strip() in ["---", "-", "--", "―", "ㅡ"]
                
                if is_separator:
                    # 줄내림인 경우 함량과 실험량을 빈 문자열로
                    item_data['함량(%)'] = ""
                    item_data['실험량(g)'] = ""
                else:
                    # 일반 원료인 경우 숫자 변환
                    if '함량(%)' in item_data:
                        item_data['함량(%)'] = try_convert_to_float(item_data['함량(%)'])
                    if '실험량(g)' in item_data:
                        item_data['실험량(g)'] = try_convert_to_float(item_data['실험량(g)'])
                
                formulation_data["items"].append(item_data)

            elif reading_state == "results":
                if key == "pH" and len(row) > 2:
                    formulation_data["details"]["pH (당일)"] = str(row[1]).replace("당일:", "").strip() if row[1] else ""
                    formulation_data["details"]["pH (익일)"] = str(row[2]).replace("익일:", "").strip() if row[2] else ""
                elif key == "점도" and len(row) > 2:
                    formulation_data["details"]["점도 (당일)"] = str(row[1]).replace("당일:", "").strip() if row[1] else ""
                    formulation_data["details"]["점도 (익일)"] = str(row[2]).replace("익일:", "").strip() if row[2] else ""
                elif key in ["사용핀 및 기계", "품평결과 및 특이사항"]:
                    formulation_data["details"][key] = row[1]

        # --- 타겟 정보 시트가 있으면 추가로 읽기 ---
        if "타겟 정보" in wb.sheetnames:
            ws_target = wb["타겟 정보"]
            target_details = {}
            for row in ws_target.iter_rows(max_col=5, values_only=True):
                if not row or not row[0]: continue
                key = str(row[0]).strip()
                if key == "타겟 샘플명" and len(row) > 1:
                    target_details["타겟 샘플명"] = row[1]
                elif key == "타겟 거래처" and len(row) > 1:
                    target_details["타겟 거래처"] = row[1]
                elif key == "타겟 pH" and len(row) > 2:
                    target_details["타겟 pH (당일)"] = str(row[1]).replace("당일:", "").strip() if row[1] else ""
                    target_details["타겟 pH (익일)"] = str(row[2]).replace("익일:", "").strip() if row[2] else ""
                elif key == "타겟 점도" and len(row) > 2:
                    target_details["타겟 점도 (당일)"] = str(row[1]).replace("당일:", "").strip() if row[1] else ""
                    target_details["타겟 점도 (익일)"] = str(row[2]).replace("익일:", "").strip() if row[2] else ""
                elif key == "사용핀 및 기계" and len(row) > 1:
                    target_details["사용핀 및 기계"] = row[1]
            formulation_data["target_details"] = target_details

        return formulation_data
    except Exception as e:
        # document_management 모듈에서 ClipboardErrorDialog를 동적으로 가져옵니다.
        from modules.document_management import ClipboardErrorDialog
        # 오류 발생 시 클립보드 복사 기능이 있는 다이얼로그를 사용합니다.
        ClipboardErrorDialog(None, title="가져오기 오류", error_message=f"파일을 읽는 중 오류가 발생했습니다:\n\n{e}")
        return None

def export_ingredient_report_multiple(products_data):
    """여러 제품의 원료목록 보고서 데이터를 템플릿 형식의 엑셀 파일로 내보냅니다."""
    if not products_data:
        messagebox.showwarning("데이터 없음", "내보낼 제품 데이터가 없습니다.")
        return
    
    default_filename = f"원료목록보고서_{len(products_data)}개제품.xlsx"
    initial_dir = get_excel_path()
    timestamped_filename = get_timestamped_filename(default_filename)
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialdir=initial_dir,
        initialfile=timestamped_filename,
        title="원료목록 보고서 저장"
    )
    if not file_path:
        return

    save_excel_path(os.path.dirname(file_path))

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "원료목록보고서"

        # --- 스타일 정의 ---
        header_font = Font(name='맑은 고딕', size=11, bold=True)
        default_font = Font(name='맑은 고딕', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # --- 헤더 작성 ---
        headers = [
            "일련번호", "제품명", "유형표시", "기능성화장품유형", "기능성화장품품목코드",
            "제조업자상호", "원료성분명", "용도(E:수출용)", "맞춤형내용물(C1:혼합용/C2:소분용)"
        ]
        for col_idx, header_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # --- 여러 제품의 데이터 작성 ---
        current_row = 2
        for product_data in products_data:
            items = product_data.get('items', [])
            for item in items:
                row_data = [
                    item.get("일련번호"),
                    product_data.get("제품명"),
                    product_data.get("유형표시"),
                    product_data.get("기능성화장품유형"),
                    product_data.get("기능성화장품품목코드"),
                    product_data.get("제조업자상호"),
                    item.get("원료성분명"),
                    item.get("용도(E:수출용)"),
                    item.get("맞춤형내용물(C1:혼합용/C2:소분용)")
                ]
                for col_idx, cell_value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
                    cell.font = default_font
                    cell.alignment = left_align
                current_row += 1

        # --- 컬럼 너비 자동 조절 ---
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        length = _get_display_length(cell.value)
                        if length > max_length:
                            max_length = length
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        wb.save(file_path)
        try:
            os.startfile(os.path.abspath(file_path))
        except Exception:
            pass
    except Exception as e:
        messagebox.showerror("오류", f"파일 저장 중 오류가 발생했습니다: {e}")

def export_ingredient_report(report_data):
    """원료목록 보고서 데이터를 템플릿 형식의 엑셀 파일로 내보냅니다."""
    default_filename = f"{report_data.get('제품명', '화장품')}_원료목록보고서.xlsx"
    initial_dir = get_excel_path()
    timestamped_filename = get_timestamped_filename(default_filename)
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialdir=initial_dir,
        initialfile=timestamped_filename,
        title="원료목록 보고서 저장"
    )
    if not file_path:
        return

    save_excel_path(os.path.dirname(file_path))

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "원료목록보고서"

        # --- 스타일 정의 ---
        header_font = Font(name='맑은 고딕', size=11, bold=True)
        default_font = Font(name='맑은 고딕', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # --- 헤더 작성 ---
        headers = [
            "일련번호", "제품명", "유형표시", "기능성화장품유형", "기능성화장품품목코드",
            "제조업자상호", "원료성분명", "용도(E:수출용)", "맞춤형내용물(C1:혼합용/C2:소분용)"
        ]
        for col_idx, header_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # --- 데이터 작성 ---
        items = report_data.get('items', [])
        for row_idx, item in enumerate(items, 2):
            row_data = [
                item.get("일련번호"),
                report_data.get("제품명"),
                report_data.get("유형표시"),
                report_data.get("기능성화장품유형"),
                report_data.get("기능성화장품품목코드"),
                report_data.get("제조업자상호"),
                item.get("원료성분명"),
                item.get("용도(E:수출용)"),
                item.get("맞춤형내용물(C1:혼합용/C2:소분용)")
            ]
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.font = default_font
                cell.alignment = left_align

        # --- 컬럼 너비 자동 조절 ---
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        length = _get_display_length(cell.value)
                        if length > max_length:
                            max_length = length
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        wb.save(file_path)
        try:
            os.startfile(os.path.abspath(file_path))
        except Exception:
            pass
    except Exception as e:
        messagebox.showerror("오류", f"파일 저장 중 오류가 발생했습니다: {e}")

def export_quotation_to_excel(quotation_data, default_filename="quotation.xlsx", lang="ko"):
    """견적서 데이터를 특정 템플릿 형식의 엑셀 파일로 내보냅니다. (한글 / 영문 다국어 지원)"""
    is_eng = (lang == "en")
    initial_dir = get_excel_path()
    timestamped_filename = get_timestamped_filename(default_filename)
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialdir=initial_dir,
        initialfile=timestamped_filename,
        title="Quotation Export" if is_eng else "견적서 저장"
    )
    if not file_path:
        return

    save_excel_path(os.path.dirname(file_path))

    try:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Quotation" if is_eng else "견적서"

        # --- 스타일 정의 ---
        thin = Side(style='thin')
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_font = Font(name='맑은 고딕', size=18, bold=True)
        header_font = Font(name='맑은 고딕', size=11, bold=True)
        default_font = Font(name='맑은 고딕', size=10)
        total_font = Font(name='맑은 고딕', size=10, bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        # --- 문서 제목 및 결재란 ---
        is_semi = quotation_data.get("is_semi", False)
        sheet.merge_cells('A1:C2')
        title_cell = sheet['A1']
        if is_eng:
            title_cell.value = "BULK PRICE QUOTATION" if is_semi else "FINISHED PRODUCT QUOTATION"
        else:
            title_cell.value = "반제품 견적서 (Bulk Quotation)" if is_semi else "완제품 견적서 (Quotation)"
        title_cell.font = title_font
        title_cell.alignment = center_align

        approval_labels = ["Prepared", "Reviewed", "Approved"] if is_eng else ["작성", "검토", "승인"]
        for i, label in enumerate(approval_labels):
            col_idx = i + 4 # D, E, F 열
            sheet.column_dimensions[chr(ord('A') + col_idx - 1)].width = 15
            sheet.cell(row=1, column=col_idx, value=label).font = header_font
            sheet.cell(row=1, column=col_idx).alignment = center_align
            sheet.cell(row=1, column=col_idx).border = thin_border
            sheet.cell(row=2, column=col_idx).border = thin_border
            sheet.row_dimensions[2].height = 40

        # --- 기본 정보 ---
        details = quotation_data.get("details", {})
        info_layout = []
        for k, v in details.items():
            lbl = k
            if is_eng:
                en_key_map = {"실험품명": "Product Name", "담당자": "Manager", "기준 중량": "Base Weight", "개당 용량": "Unit Capacity", "산출 수량": "Produced Units"}
                lbl = en_key_map.get(k, k)
            info_layout.append((lbl, v))
        row_idx = 4
        for label, value in info_layout:
            sheet.cell(row=row_idx, column=1, value=label).font = header_font
            sheet.cell(row=row_idx, column=2, value=value)
            row_idx += 1

        # --- 견적 항목 헤더 ---
        row_idx += 1
        if is_eng:
            item_headers = ["Phase", "Code", "Raw Material (INCI)", "Ratio (%)", "Unit Price (₩/kg)", "Cost (₩)"]
        else:
            item_headers = ["구분", "코드", "원료명", "함량(%)", "단가(원/kg)", "원가(원)"]
            
        for col_idx, header in enumerate(item_headers, 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=header)
            cell.font = header_font; cell.alignment = center_align; cell.fill = header_fill; cell.border = thin_border

        # --- 견적 항목 데이터 ---
        row_idx += 1
        for item_values in quotation_data.get("items", []):
            for col_idx, value in enumerate(item_values, 1):
                if isinstance(value, str):
                    try_val = try_convert_to_float(value.replace(",", ""))
                    value = try_val if isinstance(try_val, (int, float)) else value

                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.font = default_font; cell.border = thin_border
                if col_idx == 3:
                    cell.alignment = left_align
                else:
                    cell.alignment = right_align
                
                if isinstance(value, (int, float)):
                    if col_idx == 4: cell.number_format = '0.0000'
                    elif col_idx == 5: cell.number_format = '#,##0'
                    elif col_idx == 6: cell.number_format = '#,##0.00'
            row_idx += 1

        # --- 제조·부자재 비용 및 최종 견적 요약 ---
        row_idx += 1
        
        extra_expenses = quotation_data.get("extra_expenses", {})
        summary = quotation_data.get("summary", {})
        
        # 영문 매핑 딕셔너리
        en_extra_map = {
            "인력비": "Labor Cost", "제조비": "Manufacturing Cost",
            "운송비": "Shipping Cost", "용기": "Container/Packaging",
            "1kg당 총단가": "Total Cost per kg", "개당(EA) 총단가": "Total Cost per Unit (EA)"
        }
        en_summary_map = {
            "총 함량": "Total Ratio", "총 원료 원가": "Total Raw Material Cost",
            "총 제조 원가": "Total Manufacturing Cost",
            "최종가 (VAT 10% 포함)": "Total Final Price (Incl. VAT 10%)"
        }

        extra_items = []
        for k, v in extra_expenses.items():
            lbl = en_extra_map.get(k, k) if is_eng else k
            val = str(v).replace("원", "KRW").replace("개", "EA") if is_eng else v
            extra_items.append((lbl, val))

        summary_items = []
        for k, v in summary.items():
            lbl = k
            if is_eng:
                for ko_k, en_k in en_summary_map.items():
                    if ko_k in k:
                        lbl = k.replace(ko_k, en_k)
                        break
                if "이윤" in k:
                    lbl = k.replace("이윤", "Profit Margin").replace("포함 공급가", "Included Supply Price").replace("포함가", "Included Price")
            val = str(v).replace("원", "KRW") if is_eng else v
            summary_items.append((lbl, val))

        max_rows = max(len(extra_items), len(summary_items))
        
        for i in range(max_rows):
            # 좌측: 부가 비용
            if i < len(extra_items):
                ex_label, ex_val = extra_items[i]
                c_lbl = sheet.cell(row=row_idx, column=2, value=ex_label)
                c_lbl.font = default_font; c_lbl.alignment = left_align
                c_val = sheet.cell(row=row_idx, column=3, value=ex_val)
                is_highlight = ('단가' in ex_label) or ('Cost per' in ex_label)
                c_val.font = total_font if is_highlight else default_font
                c_val.alignment = right_align
                if is_highlight:
                    c_lbl.fill = header_fill
                    c_val.fill = header_fill
            
            # 우측: 견적 요약
            if i < len(summary_items):
                sm_label, sm_val = summary_items[i]
                label_cell = sheet.cell(row=row_idx, column=5, value=sm_label)
                label_cell.font = total_font; label_cell.alignment = right_align
                value_cell = sheet.cell(row=row_idx, column=6, value=sm_val)
                value_cell.font = total_font; value_cell.alignment = right_align
                if ('최종' in sm_label or '공급가' in sm_label or 'Final' in sm_label or 'Supply Price' in sm_label):
                    label_cell.fill = total_fill
                    value_cell.fill = total_fill
            
            row_idx += 1

        # --- 컬럼 너비 자동 조절 ---
        for column_cells in sheet.columns:
            max_length = 0
            safe_cell = next((c for c in column_cells if not isinstance(c, openpyxl.cell.cell.MergedCell)), None)
            if not safe_cell:
                continue
            column_letter = safe_cell.column_letter
            for cell in column_cells:
                if cell.value is not None:
                    length = _get_display_length(cell.value)
                    if length > max_length:
                        max_length = length
            adjusted_width = max_length + 2
            if adjusted_width > 2:
                sheet.column_dimensions[column_letter].width = adjusted_width

        wb.save(file_path)
        try:
            os.startfile(os.path.abspath(file_path))
        except Exception:
            pass
    except Exception as e:
        messagebox.showerror("내보내기 오류" if not is_eng else "Export Error", f"파일 저장 중 오류가 발생했습니다: {e}")

def export_ingredient_lists_to_excel(sheets_data, default_filename="ingredient_list.xlsx"):
    """
    다양한 형식의 전성분 목록 데이터를 여러 시트로 구성된 엑셀 파일로 내보냅니다.
    - sheets_data: {'시트명': {'type': 'table' or 'text', 'content': ...}}
    """
    initial_dir = get_excel_path()
    timestamped_filename = get_timestamped_filename(default_filename)
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialdir=initial_dir,
        initialfile=timestamped_filename,
        title="전성분 목록 저장"
    )
    if not file_path:
        return

    save_excel_path(os.path.dirname(file_path))

    try:
        wb = Workbook()
        wb.remove(wb.active)

        # 스타일 정의
        header_font = Font(name='맑은 고딕', size=11, bold=True)
        default_font = Font(name='맑은 고딕', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        for sheet_name, sheet_content in sheets_data.items():
            sheet = wb.create_sheet(title=sheet_name)
            content_type = sheet_content.get('type')
            content = sheet_content.get('content')

            if content_type == 'table':
                headers = content.get('headers', [])
                data_rows = content.get('data', [])
                number_formats = content.get('number_formats', {})
                
                # 헤더 쓰기
                for col_idx, header_text in enumerate(headers, 1): # noqa
                    cell = sheet.cell(row=1, column=col_idx, value=header_text.replace('\n', ' '))
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                
                # 데이터 쓰기
                # 먼저 'NO' 또는 '구분'이 포함된 첫 번째 열을 찾아, 그 열부터 모두 중앙 정렬합니다.
                # '함량' 또는 '%'가 포함된 열은 숫자이므로 오른쪽 정렬을 유지합니다.
                special_indices = []
                for idx, h in enumerate(headers, 1):
                    ht = str(h)
                    if ht.strip().upper() == 'NO' or '구분' in ht or '함량' in ht or '%' in ht:
                        special_indices.append(idx)
                last_special_idx = max(special_indices) if special_indices else 0

                # 'NO' 또는 '구분'이 포함된 첫 번째 열의 인덱스를 찾습니다.
                center_align_start_col = -1
                for idx, h in enumerate(headers, 1):
                    if str(h).strip().upper() == 'NO' or '구분' in str(h):
                        center_align_start_col = idx
                        break

                for row_idx, row_data in enumerate(data_rows, 2): # noqa
                    for col_idx, cell_value in enumerate(row_data, 1): # 1-based index
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        header_text = headers[col_idx - 1] # 0-based index for headers

                        # 숫자 형식 적용 대상인지 확인
                        num_format = number_formats.get(header_text)
                        if num_format:
                            # [수정] 빈 문자열이나 None일 경우를 안전하게 처리
                            if cell_value is None or cell_value == '':
                                cell.value = None # 빈 셀로 남겨둠
                            else:
                                try:
                                    cell.value = float(cell_value)
                                except (ValueError, TypeError):
                                    cell.value = cell_value # 변환 실패 시 원본 값 유지
                            cell.number_format = num_format
                        elif '함량' in header_text or '%' in header_text:
                            converted_value = try_convert_to_float(cell_value)
                            cell.value = converted_value
                        else:
                            cell.value = cell_value
                        cell.font = default_font

                        # 'NO' 또는 '구분' 열부터 시작하여 중앙 정렬 적용
                        if center_align_start_col != -1 and col_idx >= center_align_start_col:
                            cell.alignment = center_align
                        else:
                            # 숫자이면 오른쪽 정렬, 아니면 왼쪽 정렬을 기본으로 둡니다.
                            if isinstance(cell.value, (int, float)):
                                cell.alignment = right_align
                            else: # noqa
                                cell.alignment = left_align
                
                # 컬럼 너비 자동 조절 (개선된 로직)
                for column_cells in sheet.columns:
                    max_length = 0
                    safe_cell = next((c for c in column_cells if not isinstance(c, openpyxl.cell.cell.MergedCell)), None)
                    if not safe_cell:
                        continue
                    column_letter = safe_cell.column_letter
                    for cell in column_cells:
                        if cell.value is not None:
                            length = _get_display_length(cell.value)
                            if length > max_length:
                                max_length = length
                    
                    adjusted_width = (max_length + 2)
                    if adjusted_width > 2:
                        sheet.column_dimensions[column_letter].width = adjusted_width
                
                # --- 셀 병합 로직 추가 (원료별 목록 시트에만 적용) ---
                if sheet_name == "원료별 목록" or sheet_name == "By Raw Material":
                    try:
                        def get_col_idx(candidates):
                            for c in candidates:
                                if c in headers:
                                    return headers.index(c) + 1
                            return None

                        # 병합할 열의 인덱스 찾기
                        no_col_idx = get_col_idx(["NO"])
                        if not no_col_idx: raise ValueError("NO column not found")

                        material_name_col_idx = get_col_idx(["원료명", "Material Name"])
                        rm_ratio_col_idx = get_col_idx(["RM 함량(%)", "RM Ratio(%)"])
                        phase_col_idx = get_col_idx(["구분", "Phase"])
                        
                        # Material 레벨 컬럼들
                        hs_code_col_idx = get_col_idx(["HS CODE"])
                        origin_col_idx = get_col_idx(["원산지", "Origin"])
                        material_name_en_col_idx = get_col_idx(["영문원료명", "Material Name (EN)"])
                        nmpa_col_idx = get_col_idx(["NMPA"])
                        supplier_col_idx = get_col_idx(["거래처명", "Supplier"])

                        cols_to_merge = []
                        if material_name_col_idx: cols_to_merge.append(material_name_col_idx) # 원료명 필수
                        if rm_ratio_col_idx: cols_to_merge.append(rm_ratio_col_idx)
                        
                        # 항상 NO는 병합 대상 (그룹 식별자) - NO 자체를 병합할지는 선택이지만, 보통 병합함
                        cols_to_merge.append(no_col_idx)

                        if phase_col_idx: cols_to_merge.append(phase_col_idx)
                        if hs_code_col_idx: cols_to_merge.append(hs_code_col_idx)
                        if origin_col_idx: cols_to_merge.append(origin_col_idx)
                        if material_name_en_col_idx: cols_to_merge.append(material_name_en_col_idx)
                        if nmpa_col_idx: cols_to_merge.append(nmpa_col_idx)
                        if supplier_col_idx: cols_to_merge.append(supplier_col_idx)

                        merge_start_row = 2  # 데이터는 2행부터 시작
                        for current_row_idx, row_data in enumerate(data_rows, start=2):
                            # 새로운 원료 그룹 시작 (NO 컬럼에 값이 있음)
                            if row_data[no_col_idx - 1]:
                                # 이전 그룹에 대한 병합 처리
                                if current_row_idx > merge_start_row:
                                    for col_to_merge in cols_to_merge:
                                        sheet.merge_cells(start_row=merge_start_row, start_column=col_to_merge, end_row=current_row_idx - 1, end_column=col_to_merge)
                                # 새 그룹의 시작 행 업데이트
                                merge_start_row = current_row_idx

                        # 마지막 그룹에 대한 병합 처리
                        if len(data_rows) + 1 >= merge_start_row:
                            for col_to_merge in cols_to_merge:
                                sheet.merge_cells(start_row=merge_start_row, start_column=col_to_merge, end_row=len(data_rows) + 1, end_column=col_to_merge)

                        # 병합된 셀 정렬
                        for col_to_merge in cols_to_merge:
                            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_to_merge, max_col=col_to_merge):
                                for cell in row:
                                    cell.alignment = Alignment(vertical='center', horizontal=cell.alignment.horizontal, wrap_text=True)

                    except Exception as e:
                        print(f"'{sheet_name}' 시트 셀 병합 중 오류 발생 (필수 컬럼 부재): {e}")
                # --- 셀 병합 로직 종료 ---

        # '디자인용 전성분' 시트에 대한 특별 스타일링
        if "디자인용 전성분" in wb.sheetnames:
            sheet = wb["디자인용 전성분"]
            # 헤더 삭제
            sheet.delete_rows(1)
            # 셀 스타일 적용
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                row[0].font = header_font # 구분 (국문:, 영문:)
                row[0].alignment = left_align
                row[1].font = default_font # 전성분 목록
                row[1].alignment = Alignment(wrap_text=True, vertical='top')
            # 자동 너비 조절 적용
            for column_cells in sheet.columns:
                max_length = 0
                safe_cell = next((c for c in column_cells if not isinstance(c, openpyxl.cell.cell.MergedCell)), None)
                if not safe_cell:
                    continue
                column_letter = safe_cell.column_letter
                for cell in column_cells:
                    if cell.value is not None:
                        length = _get_display_length(cell.value)
                        if length > max_length: max_length = length
                sheet.column_dimensions[column_letter].width = max_length + 2
            sheet.column_dimensions['B'].width = 100

        wb.save(file_path)
        try:
            os.startfile(os.path.abspath(file_path))
        except Exception:
            pass
    except Exception as e:
        messagebox.showerror("내보내기 오류", f"파일 저장 중 오류가 발생했습니다: {e}")

def export_functional_cosmetics_report_template(report_data=None):
    """UI에서 작성된 내용을 바탕으로 기능성화장품 심사제외품목 보고서를 생성하고 내보냅니다."""
    if report_data is None:
        report_data = {}

    initial_dir = get_excel_path()
    product_name = report_data.get("제품명(국문)", "기능성보고서")
    default_filename = f"{product_name}_심사제외보고서.xlsx"
    timestamped_filename = get_timestamped_filename(default_filename)
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialdir=initial_dir,
        initialfile=timestamped_filename,
        title="심사제외 품목 보고서 저장"
    )
    if not file_path:
        return

    save_excel_path(os.path.dirname(file_path))

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "기능성심사제외보고서"

        # --- Styles ---
        title_font = Font(name='맑은 고딕', size=16, bold=True)
        header_font = Font(name='맑은 고딕', size=11, bold=True)
        default_font = Font(name='맑은 고딕', size=10)
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align_top_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # --- Column widths ---
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 50

        # --- Title ---
        ws.merge_cells("A1:C1")
        title_cell = ws['A1']
        title_cell.value = "기능성화장품 심사제외품목 보고서"
        title_cell.font = title_font
        title_cell.alignment = center_align
        ws.row_dimensions[1].height = 30

        row_idx = 3

        def write_row_and_style(r_idx, values, is_header=False):
            """Helper to write a row and apply styles"""
            cells = []
            for c_idx, value in enumerate(values, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cells.append(cell)
            
            for cell in cells:
                cell.border = thin_border
                cell.font = default_font
                cell.alignment = left_align_top_wrap
            
            if is_header and cells:
                cells[0].font = header_font
            
            # Auto row height for multiline content in the last cell
            if len(values) > 2 and values[2] and isinstance(values[2], str):
                num_lines = values[2].count('\n') + 1
                if num_lines > 1:
                    ws.row_dimensions[r_idx].height = 15 * num_lines

        # --- Data Parsing ---
        spf_pa_text = report_data.get("자외선 관련 (SPF / PA)", "")
        spf, pa = "", ""
        if "/" in spf_pa_text:
            parts = [p.strip() for p in spf_pa_text.split('/')]
            spf, pa = (parts[0], parts[1]) if len(parts) > 1 else (parts[0], "")
        else:
            if "spf" in spf_pa_text.lower(): spf = spf_pa_text
            elif "pa" in spf_pa_text.lower(): pa = spf_pa_text
        
        effects = report_data.get("효능·효과", "").replace(", ", "\n")

        # --- Report Layout ---
        layout = [
            ("보고정보", "제품명", report_data.get("제품명(국문)")),
            (None, "제품의 pH 기준치", report_data.get("pH (실측값)")),
            (None, "대상구분", report_data.get("제출유형")),
            ("제10조 제1항 제3호에 해당하는 경우", "이미 심사받은 품목", report_data.get("이미 심사받은 품목")),
            (None, "활성물질용량", report_data.get("활성물질용량")),
            (None, "자외선차단지수(SPF)", spf),
            (None, "자외선 A차단등급(PA)", pa),
            (None, "고시한 기준 및 시험방법", report_data.get("고시한 기준 및 시험방법")),
            (None, "효능효과", effects),
            (None, "용법용량", report_data.get("용법·용량")),
            (None, "사용할 때의 주의사항", report_data.get("사용할 때의 주의사항")),
            ("총량관리", "자동 입력", ""),
        ]

        # --- Write data and merge cells ---
        merge_start_row = 3
        for val_a, val_b, val_c in layout:
            if val_a is not None and row_idx > merge_start_row:
                ws.merge_cells(start_row=merge_start_row, start_column=1, end_row=row_idx - 1, end_column=1)
                ws.cell(merge_start_row, 1).alignment = left_align_top_wrap
                merge_start_row = row_idx
            
            write_row_and_style(row_idx, [val_a, val_b, val_c], is_header=(val_a is not None))
            row_idx += 1
        ws.merge_cells(start_row=merge_start_row, start_column=1, end_row=row_idx - 1, end_column=1)
        ws.cell(merge_start_row, 1).alignment = left_align_top_wrap

        # --- Section: 원료성분 및 배합 비율 ---
        start_row = row_idx
        
        ingredients_text = report_data.get("원료성분 및 배합비율", "")
        ingredients_list = []
        if ingredients_text and "예시:" not in ingredients_text:
            for line in ingredients_text.splitlines():
                if ":" in line:
                    parts = line.split(":", 1)
                    name = parts[0].strip()
                    amount = parts[1].strip()
                    ingredients_list.append((name, amount))
                elif line.strip():
                    ingredients_list.append((line.strip(), ""))
        
        if not ingredients_list:
            write_row_and_style(row_idx, ["원료성분 및 배합 비율", "(내용 없음)", ""], is_header=True)
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=3)
            row_idx += 1
        else:
            first_ing_name, first_ing_amount = ingredients_list[0]
            write_row_and_style(row_idx, ["원료성분 및 배합 비율", first_ing_name, first_ing_amount], is_header=True)
            row_idx += 1
            for name, amount in ingredients_list[1:]:
                write_row_and_style(row_idx, [None, name, amount])
                row_idx += 1

        ws.merge_cells(start_row=start_row, start_column=1, end_row=row_idx - 1, end_column=1)
        ws.cell(start_row, 1).alignment = left_align_top_wrap

        wb.save(file_path)
        try:
            os.startfile(os.path.abspath(file_path))
        except Exception:
            pass
    except Exception as e:
        messagebox.showerror("내보내기 오류", f"보고서 파일 저장 중 오류가 발생했습니다: {e}")

def export_production_formulation_to_excel(production_data, default_filename="production.xlsx", file_path: str | None = None, open_print_preview: bool = False, mode: str = "original"):
    """생산처방 정보를 엑셀로 내보냅니다.
    mode:
      - "original": 원래 템플릿(제목+영문부제, 결재란 이미지, 병합 포함, 추가 시트 포함)
      - "revised": 수정본 템플릿(결재란 텍스트 셀, 기본정보 병합 없음, 고정폭, 세로 인쇄, 단일 시트)
    """
    if (mode or "original").lower() == "revised":
        return export_production_formulation_revised_to_excel(
            production_data,
            default_filename=default_filename,
            file_path=file_path,
            open_print_preview=open_print_preview,
        )
    else:
        return export_production_formulation_original_to_excel(
            production_data,
            default_filename=default_filename,
            file_path=file_path,
            open_print_preview=open_print_preview,
        )

    