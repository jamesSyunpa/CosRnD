# modules/quality_management.py
import os
from datetime import datetime, date
import customtkinter as ctk
from tkinter import messagebox, ttk
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import tkinter.filedialog as fd
from modules import excel_handler
from utils.company_profile import get_company_profile
from modules.excel_handler import _get_display_length

from sqlalchemy.orm import joinedload
from modules.translation import get_texts
from database.db_manager import db_manager
from database.models import (
    IngredientReport, IngredientReportItem,
    SemiFinishedCOA, SemiFinishedCOAItem,
    FinishedProductCOA, FinishedProductCOAItem,
    MaterialInspectionReport, ProductStandard,
    BatchManufacturingRecord, StabilityTestReport,
    PackagingCompatibilityReport, MSDSReport, Material, Formulation
)

class QualityManagementFrame(ctk.CTkFrame):
    """품질 관리 관련 기능을 포함하는 프레임"""
    def __init__(self, master, user, app, texts):
        super().__init__(master)
        self.current_user = user
        self.app = app
        self.texts = texts

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # --- 메인 탭 뷰 ---
        self.tab_view = ctk.CTkTabview(
            self, border_width=1, border_color=("gray80", "gray30"),
            segmented_button_selected_color=('#3B8ED0', '#1F6AA5'),
            segmented_button_unselected_color=("gray92", "gray20"),
            text_color=("black", "white"),
            segmented_button_selected_hover_color=('#3671A8', '#144870'),
            segmented_button_unselected_hover_color=("gray85", "gray28")
        )
        self.tab_view.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        # --- 품질관리 8대 핵심 서류 탭 추가 (COA 바로 다음에 MSDS 배치) ---
        self.tab_labels = {
            'coa': self.texts.get('coa', 'COA (시험성적서)'),
            'msds': self.texts.get('msds', 'MSDS (물질안전보건자료)'),
            'ingredient_report': self.texts.get('ingredient_report', '원료목록보고'),
            'mat_inspection': self.texts.get('mat_inspection', '원료 입고검사'),
            'prod_standard': self.texts.get('prod_standard', '제품표준서'),
            'mfg_record': self.texts.get('mfg_record', '제조관리기록서'),
            'stability_test': self.texts.get('stability_test', '안정성 시험'),
            'compatibility_test': self.texts.get('compatibility_test', '용기 상용성 시험')
        }

        self.tab_view.add(self.tab_labels['coa'])
        self.tab_view.add(self.tab_labels['msds'])
        self.tab_view.add(self.tab_labels['ingredient_report'])
        self.tab_view.add(self.tab_labels['mat_inspection'])
        self.tab_view.add(self.tab_labels['prod_standard'])
        self.tab_view.add(self.tab_labels['mfg_record'])
        self.tab_view.add(self.tab_labels['stability_test'])
        self.tab_view.add(self.tab_labels['compatibility_test'])

        # --- 품질관리 8대 탭 지연 초기화(Lazy Tab Loading) 등록 ---
        self._initialized_tabs = set()
        self._tab_setup_handlers = {
            'coa': self.setup_coa_tab,
            'msds': self.setup_msds_tab,
            'ingredient_report': self.setup_ingredient_report_tab,
            'mat_inspection': self.setup_material_inspection_tab,
            'prod_standard': self.setup_product_standard_tab,
            'mfg_record': self.setup_batch_manufacturing_tab,
            'stability_test': self.setup_stability_test_tab,
            'compatibility_test': self.setup_packaging_compatibility_tab
        }

    def ensure_tab_initialized(self, tab_key):
        """요청된 탭이 아직 초기화되지 않았다면 즉시 빌드합니다 (초고속 지연 로딩)."""
        if tab_key in self._initialized_tabs:
            return
        handler = self._tab_setup_handlers.get(tab_key)
        tab_label = self.tab_labels.get(tab_key)
        if handler and tab_label:
            tab_frame = self.tab_view.tab(tab_label)
            handler(tab_frame)
            self._initialized_tabs.add(tab_key)

    def setup_ingredient_report_tab(self, tab_frame):
        """원료목록보고 탭의 UI를 설정합니다."""
        self.saved_products = []  # 저장된 제품 데이터 리스트 (엑셀용 버퍼)
        self.current_ingredient_report_id = None  # DB 로드 시 편집 중인 ID
        
        self.cosmetic_type_map = {
            "가": {"name": "만 3세 이하 영유아용 제품류", "items": {"가1": "영유아용 샴푸, 린스", "가2": "영유아용 로션, 크림", "가3": "영유아용 오일", "가4": "영유아용 인체 세정용 제품", "가5": "영유아용 목욕용 제품"}},
            "나": {"name": "목욕용 제품류", "items": {"나1": "목욕용 오일·정제·캡슐", "나2": "목욕용 소금류", "나3": "버블 배스", "나4": "그 밖의 목욕용 제품류"}},
            "다": {"name": "인체 세정용 제품류", "items": {"다1": "폼 클렌저", "다2": "바디 클렌저", "다3": "액체비누", "다3-1": "화장비누(고체형)", "다4": "외음부 세정제", "다5": "물휴지", "다6": "그 밖의 인체 세정용 제품류"}},
            "라": {"name": "눈화장용 제품류", "items": {"라1": "아이브로 제품", "라2": "아이라이너", "라3": "아이섀도", "라4": "마스카라", "라5": "아이 메이크업 리무버", "라6": "그 밖의 눈화장용 제품류"}},
            "마": {"name": "방향용 제품류", "items": {"마1": "향수", "마4": "코롱", "마5": "그 밖의 방향용 제품류"}},
            "바": {"name": "두발 염색용 제품류", "items": {"바1": "헤어 틴트", "바2": "헤어 칼라 스프레이", "바3": "그 밖의 염모용 제품류", "바4": "염모제", "바5": "탈염·탈색용 제품"}},
            "사": {"name": "색조 화장용 제품류", "items": {"사1": "볼연지", "사2": "페이스 파우더", "사3": "리퀴드·크림·케익 파운데이션", "사4": "메이크업 베이스", "사5": "메이크업 픽서티브", "사6": "립스틱·립라이너", "사7": "립글로스·립밤", "사8": "바디·페이스 페인팅·분장용 제품", "사9": "그 밖의 색조화장용 제품류"}},
            "아": {"name": "두발용 제품류", "items": {"아1": "헤어컨디셔너·트리트먼트·팩", "아2": "헤어토닉·헤어에센스", "아3": "헤어그루밍에이드", "아4": "헤어크림·로션", "아5": "헤어오일", "아6": "포마드", "아7": "헤어스프레이·무스·왁스·젤", "아8": "샴푸", "아9": "퍼머넌트 웨이브", "아10": "헤어스트레이트너", "아11": "그 밖의 두발용 제품류", "아12": "흑채"}},
            "자": {"name": "손발톱용 제품류", "items": {"자1": "베이스코트·언더코트", "자2": "네일폴리시·네일에나멜", "자3": "탑코트", "자4": "네일크림·로션·에센스·오일", "자5": "네일폴리시·네일에나멜 리무버", "자6": "그 밖의 손발톱용 제품류"}},
            "차": {"name": "면도용 제품류", "items": {"차1": "애프터셰이브 로션", "차3": "프리셰이브 로션", "차4": "세이빙 크림", "차5": "세이빙 폼", "차6": "그 밖의 면도용 제품류"}},
            "카": {"name": "기초화장용 제품류", "items": {"카1": "수렴·유연·영양화장수", "카2": "마사지 크림", "카3": "에센스·오일", "카4": "파우더", "카5": "바디 제품", "카6": "팩·마스크", "카7": "눈 주위 제품", "카8": "로션·크림", "카9": "손·발의 피부연화 제품", "카10": "클렌징워터·오일·로션·크림", "카11": "그 밖의 기초화장용 제품류"}},
            "타": {"name": "체취방지용 제품류", "items": {"타1": "데오도런트", "타2": "그 밖의 체취방지용 제품류"}},
            "파": {"name": "체모 제거용 제품류", "items": {"파1": "제모제", "파2": "그 밖의 체모 제거용 제품류", "파3": "제모 왁스"}},
        }
        self.functional_types = {
            "없음": "기능성화장품 아님",
            "F1": "미백", "F2": "주름개선", "F3": "자외선차단", "F4": "미백+주름개선",
            "F5": "미백+자외선차단", "F6": "주름개선+자외선차단", "F7": "미백+주름개선+자외선차단",
            "F8": "염모/탈염/탈색", "F9": "체모 제모", "F10": "탈모증상 완화",
            "F11": "여드름성 피부 완화", "F12": "피부 장벽 기능 회복", "F13": "튼살로 인한 붉은 선 완화",
            "F14": "기타 복합유형"
        }

        tab_frame.grid_columnconfigure(0, weight=1)
        tab_frame.grid_rowconfigure(0, weight=1)

        self.scrollable_frame = ctk.CTkScrollableFrame(tab_frame, label_text="화장품 원료목록 보고")
        self.scrollable_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        self.scrollable_frame.grid_columnconfigure(0, weight=1)

        self.report_entries = {}
        self.report_item_rows = []

        info_frame = ctk.CTkFrame(self.scrollable_frame)
        info_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 20))
        info_frame.grid_columnconfigure(1, weight=3)  # 제품명 필드가 더 넓게
        info_frame.grid_columnconfigure(3, weight=1)  # 제조업자상호 필드

        ctk.CTkLabel(info_frame, text="제품명", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.report_entries["제품명"] = ctk.CTkEntry(info_frame)
        self.report_entries["제품명"].grid(row=0, column=1, padx=10, pady=5, sticky="ew")

        ctk.CTkLabel(info_frame, text="제조업자상호", font=ctk.CTkFont(weight="bold")).grid(row=0, column=2, padx=10, pady=5, sticky="w")
        self.report_entries["제조업자상호"] = ctk.CTkEntry(info_frame)
        self.report_entries["제조업자상호"].grid(row=0, column=3, padx=10, pady=5, sticky="ew")

        ctk.CTkLabel(info_frame, text="유형표시", font=ctk.CTkFont(weight="bold")).grid(row=1, column=0, padx=10, pady=5, sticky="w")
        ctk.CTkLabel(info_frame, text="유형표시", font=ctk.CTkFont(weight="bold")).grid(row=1, column=0, padx=10, pady=5, sticky="w")
        category_main_options = [f"{k} : {v['name']}" for k, v in self.cosmetic_type_map.items()]
        self.report_entries["유형표시_대분류"] = ctk.CTkComboBox(info_frame, values=category_main_options, command=self._update_subcategory_combo, width=250)
        self.report_entries["유형표시_대분류"].grid(row=1, column=1, padx=10, pady=5, sticky="ew")
        
        self.report_entries["유형표시"] = ctk.CTkComboBox(info_frame, values=[], width=250)
        self.report_entries["유형표시"].grid(row=1, column=3, padx=10, pady=5, sticky="ew")
        self._update_subcategory_combo(category_main_options[0])

        ctk.CTkLabel(info_frame, text="기능성화장품유형", font=ctk.CTkFont(weight="bold")).grid(row=2, column=0, padx=10, pady=5, sticky="w")
        self.report_entries["기능성화장품유형"] = ctk.CTkComboBox(info_frame, values=[f"{k} : {v}" for k, v in self.functional_types.items()], width=250)
        self.report_entries["기능성화장품유형"].grid(row=2, column=1, padx=10, pady=5, sticky="ew")

        ctk.CTkLabel(info_frame, text="기능성화장품품목코드", font=ctk.CTkFont(weight="bold")).grid(row=2, column=2, padx=10, pady=5, sticky="w")
        self.report_entries["기능성화장품품목코드"] = ctk.CTkEntry(info_frame)
        self.report_entries["기능성화장품품목코드"].grid(row=2, column=3, padx=10, pady=5, sticky="ew")

        ctk.CTkLabel(info_frame, text="용도(수출전용 여부)", font=ctk.CTkFont(weight="bold")).grid(row=3, column=0, padx=10, pady=5, sticky="w")
        self.report_entries["용도"] = ctk.CTkComboBox(info_frame, values=["", "E"], state="readonly", width=200)
        self.report_entries["용도"].grid(row=3, column=1, padx=10, pady=5, sticky="w")

        ctk.CTkLabel(info_frame, text="맞춤형 내용물(C1/C2)", font=ctk.CTkFont(weight="bold")).grid(row=3, column=2, padx=10, pady=5, sticky="w")
        self.report_entries["맞춤형내용물"] = ctk.CTkComboBox(info_frame, values=["", "C1", "C2"], state="readonly", width=200)
        self.report_entries["맞춤형내용물"].grid(row=3, column=3, padx=10, pady=5, sticky="w")

        # 원료성분명 일괄 입력 프레임
        paste_frame = ctk.CTkFrame(self.scrollable_frame)
        paste_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=(10, 5))
        paste_frame.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(paste_frame, text="원료성분명(붙여넣기):", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.bulk_ingredient_entry = ctk.CTkTextbox(paste_frame, height=80)
        self.bulk_ingredient_entry.grid(row=0, column=1, padx=10, pady=5, sticky="ew")
        
        help_label = ctk.CTkLabel(paste_frame, text="※ 여러 성분을 ', ' (쉼표+공백) 또는 줄바꿈으로 구분하여 붙여넣으세요. 예: 물, 글리세린, 부틸렌글라이콜", 
                                 font=ctk.CTkFont(size=10), text_color="gray")
        help_label.grid(row=1, column=0, columnspan=2, padx=10, pady=(0, 5), sticky="w")

        self.report_table_frame = ctk.CTkFrame(self.scrollable_frame)
        self.report_table_frame.grid(row=2, column=0, sticky="ew", padx=10, pady=10)
        for i in range(9):  # 9개 열: 일련번호, 제품명, 유형표시, 기능성유형, 품목코드, 제조업자상호, 원료성분명, 용도, 맞춤형내용물
            self.report_table_frame.grid_columnconfigure(i, weight=1)

        # 테이블 헤더 생성은 _redraw_report_table에서 수행

        button_frame = ctk.CTkFrame(tab_frame, fg_color="transparent")
        button_frame.grid(row=1, column=0, padx=10, pady=10, sticky="e")
        
        ctk.CTkButton(button_frame, text="현재 제품 저장 후 계속", command=self.save_and_continue_report,
                     fg_color="#2FA572", hover_color="#106A43", width=150).pack(side="left", padx=5)
        ctk.CTkButton(button_frame, text="엑셀 보고서 생성", command=self.generate_ingredient_report,
                     fg_color="#3B8ED0", hover_color="#1F6AA5", width=150).pack(side="left", padx=5)
        ctk.CTkButton(button_frame, text="초기화", command=self.clear_ingredient_report_form, 
                     fg_color="gray50", hover_color="gray35").pack(side="left", padx=5)

        # DB 저장/불러오기 UI
        db_frame = ctk.CTkFrame(tab_frame, fg_color="transparent")
        db_frame.grid(row=2, column=0, padx=10, pady=(0, 10), sticky="ew")
        db_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkButton(db_frame, text="DB 저장", command=self.save_ingredient_report_to_db,
                      fg_color="#2FA572", hover_color="#106A43", width=120).grid(row=0, column=0, padx=(0, 8))
        self.ingredient_report_picker = ctk.CTkComboBox(db_frame, values=[], width=500)
        self.ingredient_report_picker.grid(row=0, column=1, padx=(0, 8), sticky="ew")
        ctk.CTkButton(db_frame, text="불러오기", command=self.load_selected_ingredient_report,
                      fg_color="#4C9AFF", hover_color="#1F6AA5", width=100).grid(row=0, column=2)
        ctk.CTkButton(db_frame, text="삭제", command=self.delete_selected_ingredient_report,
                      fg_color="#D32F2F", hover_color="#B71C1C", width=80).grid(row=0, column=3)

        self._redraw_report_table()
        # 최근 저장 목록 로드
        self.refresh_ingredient_report_list()

    def delete_selected_ingredient_report(self):
        """선택한 원료목록보고를 DB에서 삭제합니다."""
        if not getattr(self.current_user, 'can_delete', None) or not self.current_user.can_delete():
            messagebox.showwarning("권한 없음", "삭제 권한이 없습니다.", parent=self)
            return
        sel = self.ingredient_report_picker.get()
        if not sel or '|' not in sel:
            messagebox.showwarning("선택 필요", "삭제할 항목을 선택하세요.", parent=self)
            return
        try:
            report_id = int(sel.split('|')[0].strip())
        except Exception:
            messagebox.showwarning("선택 오류", "선택된 항목을 해석할 수 없습니다.", parent=self)
            return
        name_preview = ' | '.join([p.strip() for p in sel.split('|')[1:3]]) if '|' in sel else ''
        if not messagebox.askyesno("삭제 확인", f"선택한 원료목록보고를 삭제하시겠습니까?\n\n{sel}\n\n이 작업은 되돌릴 수 없습니다.", parent=self):
            return
        session = db_manager.get_session()
        try:
            header = session.query(IngredientReport).filter_by(id=report_id).first()
            if not header:
                messagebox.showwarning("오류", "선택한 데이터가 존재하지 않습니다.", parent=self)
                return
            session.delete(header)
            session.commit()
            if self.current_ingredient_report_id == report_id:
                self.current_ingredient_report_id = None
            self.refresh_ingredient_report_list()
            messagebox.showinfo("삭제 완료", "원료목록보고가 삭제되었습니다.", parent=self)
        except Exception as e:
            session.rollback()
            messagebox.showerror("오류", f"삭제 중 오류가 발생했습니다: {e}", parent=self)
        finally:
            session.close()

    def _update_subcategory_combo(self, selected_main_category):
        main_code = selected_main_category.split(" : ")[0]
        sub_items = self.cosmetic_type_map.get(main_code, {}).get("items", {})
        sub_options = [f"{k} : {v}" for k, v in sub_items.items()]
        self.report_entries["유형표시"].configure(values=sub_options)
        if sub_options:
            self.report_entries["유형표시"].set(sub_options[0])
        else:
            self.report_entries["유형표시"].set("")

    def _redraw_report_table(self):
        """테이블을 초기화하고 저장된 제품들을 표시합니다."""
        for row_widgets in self.report_item_rows:
            for key, widget in row_widgets.items():
                if hasattr(widget, 'destroy'):
                    widget.destroy()
        self.report_item_rows.clear()

        # 테이블 헤더
        headers = ["일련번호", "제품명", "유형표시", "기능성화장품유형", "기능성화장품품목코드", "제조업자상호", "원료성분명", "용도", "맞춤형내용물"]
        for i, h in enumerate(headers):
            header_label = ctk.CTkLabel(self.report_table_frame, text=h, font=ctk.CTkFont(weight="bold"), 
                                       fg_color=("gray85", "gray20"), corner_radius=0)
            header_label.grid(row=0, column=i, sticky="ew", padx=(1,0), pady=(1,0))

        # 저장된 제품들을 테이블에 표시
        if self.saved_products:
            row_number = 1
            for product in self.saved_products:
                for ingredient in product["원료성분명"]:
                    self._add_report_item_row_with_data(
                        row_number=row_number,
                        product_name=product["제품명"],
                        type_code=product["유형표시"],
                        functional_type=product["기능성화장품유형"],
                        functional_code=product["기능성화장품품목코드"],
                        manufacturer=product["제조업자상호"],
                        ingredient=ingredient,
                        usage=product["용도"],
                        custom_content=product["맞춤형내용물"]
                    )
                    row_number += 1

    def _add_report_item_row(self, ingredient_name=""):
        """테이블에 행을 추가합니다."""
        r = len(self.report_item_rows) + 1
        widgets = {}
        for i in range(9):
            e = ctk.CTkEntry(self.report_table_frame, corner_radius=0, border_width=0)
            if i == 0:  # 일련번호
                e.insert(0, str(r))
                e.configure(state="readonly", fg_color=("gray90", "gray25"))
            elif i == 6:  # 원료성분명
                e.insert(0, ingredient_name)
            e.grid(row=r, column=i, padx=(1,0), pady=(1,0), sticky="ew")
            widgets[f"col{i}"] = e
        self.report_item_rows.append(widgets)

    def _add_report_item_row_with_data(self, row_number, product_name, type_code, functional_type, functional_code, manufacturer, ingredient, usage, custom_content):
        """저장된 데이터로 테이블에 행을 추가합니다."""
        r = len(self.report_item_rows) + 1
        widgets = {}
        
        data = [str(row_number), product_name, type_code, functional_type, functional_code, manufacturer, ingredient, usage, custom_content]
        
        for i in range(9):
            e = ctk.CTkEntry(self.report_table_frame, corner_radius=0, border_width=0)
            e.insert(0, data[i])
            e.configure(state="readonly", fg_color=("gray90", "gray25"))
            e.grid(row=r, column=i, padx=(1,0), pady=(1,0), sticky="ew")
            widgets[f"col{i}"] = e
        self.report_item_rows.append(widgets)

    def clear_ingredient_report_form(self):
        """폼을 초기화합니다."""
        for entry in self.report_entries.values():
            if isinstance(entry, ctk.CTkComboBox):
                values = entry.cget("values")
                if values:
                    entry.set(values[0] if values else "")
            else:
                entry.delete(0, "end")
        self.bulk_ingredient_entry.delete("1.0", "end")
        self._redraw_report_table()
        self.saved_products.clear()  # 저장된 제품 목록도 초기화
        self.current_ingredient_report_id = None
        messagebox.showinfo("알림", "폼이 초기화되었습니다.", parent=self)

    def _collect_current_ingredient_report_data(self):
        """현재 화면 또는 저장 버퍼의 원료목록 데이터를 수집합니다.
        반환: 리스트[dict]: 각 제품 단위의 헤더 및 아이템 목록
        """
        collected = []
        if self.saved_products:
            # 여러 제품 버퍼를 각각 하나의 보고서로 저장
            for product in self.saved_products:
                collected.append({
                    'product_name': product.get("제품명", ""),
                    'type_code': product.get("유형표시", ""),
                    'functional_type_code': product.get("기능성화장품유형", ""),
                    'functional_code': product.get("기능성화장품품목코드", ""),
                    'manufacturer': product.get("제조업자상호", ""),
                    'usage': product.get("용도", ""),
                    'custom_content': product.get("맞춤형내용물", ""),
                    'ingredients': product.get("원료성분명", [])
                })
        else:
            # 현재 화면 데이터로 단일 저장
            pasted_text = self.bulk_ingredient_entry.get("1.0", "end-1c").strip()
            if ', ' in pasted_text:
                pasted_ingredients = pasted_text.split(', ')
            elif ',' in pasted_text:
                pasted_ingredients = pasted_text.split(',')
            else:
                pasted_ingredients = pasted_text.split('\n')
            pasted_ingredients = [line.strip() for line in pasted_ingredients if line.strip()]

            if not self.report_entries["제품명"].get() or not pasted_ingredients:
                return []

            type_selected = self.report_entries["유형표시"].get()
            type_code = type_selected.split(" : ")[0] if " : " in type_selected else type_selected
            functional_selected = self.report_entries["기능성화장품유형"].get()
            functional_type_code = functional_selected.split(" : ")[0] if " : " in functional_selected else ""
            
            # "없음"이 선택된 경우 기능성 유형과 품목코드를 빈칸으로 처리
            if functional_type_code == "없음":
                functional_type_code = ""
                functional_code = ""
            else:
                functional_code = self.report_entries["기능성화장품품목코드"].get()

            collected.append({
                'product_name': self.report_entries["제품명"].get(),
                'type_code': type_code,
                'functional_type_code': functional_type_code,
                'functional_code': functional_code,
                'manufacturer': self.report_entries["제조업자상호"].get(),
                'usage': self.report_entries.get("용도").get() if self.report_entries.get("용도") else "",
                'custom_content': self.report_entries.get("맞춤형내용물").get() if self.report_entries.get("맞춤형내용물") else "",
                'ingredients': pasted_ingredients
            })
        return collected

    def save_ingredient_report_to_db(self):
        """원료목록보고 현재 데이터를 DB에 저장하거나 업데이트합니다."""
        data_list = self._collect_current_ingredient_report_data()
        if not data_list:
            messagebox.showwarning("입력 오류", "제품명과 원료성분명을 확인해주세요.", parent=self)
            return

        session = db_manager.get_session()
        try:
            saved_count = 0
            # 편집 중인 단일 레코드가 있는 경우 업데이트, 아니면 새로 생성
            if self.current_ingredient_report_id and len(data_list) == 1:
                header = session.query(IngredientReport).filter_by(id=self.current_ingredient_report_id).first()
                if not header:
                    self.current_ingredient_report_id = None  # fallback to create
                else:
                    d = data_list[0]
                    header.product_name = d['product_name']
                    header.manufacturer = d['manufacturer']
                    header.type_code = d['type_code']
                    header.functional_type_code = d['functional_type_code']
                    header.functional_code = d['functional_code']
                    header.usage = d['usage']
                    header.custom_content = d['custom_content']
                    # replace items
                    header.items.clear()
                    for idx, ing in enumerate(d['ingredients'], start=1):
                        header.items.append(IngredientReportItem(row_no=idx, ingredient_name=ing))
                    session.commit()
                    saved_count = 1
            if saved_count == 0:
                # 새 레코드(들) 생성
                for d in data_list:
                    header = IngredientReport(
                        product_name=d['product_name'], manufacturer=d['manufacturer'],
                        type_code=d['type_code'], functional_type_code=d['functional_type_code'],
                        functional_code=d['functional_code'], usage=d['usage'], custom_content=d['custom_content']
                    )
                    for idx, ing in enumerate(d['ingredients'], start=1):
                        header.items.append(IngredientReportItem(row_no=idx, ingredient_name=ing))
                    session.add(header)
                    saved_count += 1
                session.commit()

            self.refresh_ingredient_report_list()
            messagebox.showinfo("저장 완료", f"원료목록보고 {saved_count}건 저장되었습니다.", parent=self)
        except Exception as e:
            session.rollback()
            messagebox.showerror("오류", f"DB 저장 중 오류가 발생했습니다: {e}", parent=self)
        finally:
            session.close()

    def refresh_ingredient_report_list(self):
        """최근 저장된 원료목록보고 목록을 콤보에 로드."""
        try:
            session = db_manager.get_session()
            recs = session.query(IngredientReport).order_by(IngredientReport.created_at.desc()).limit(50).all()
            values = [f"{r.id} | {r.product_name} | {r.manufacturer or ''} | {r.created_at.strftime('%Y-%m-%d %H:%M')}" for r in recs]
            self.ingredient_report_picker.configure(values=values)
            if values:
                self.ingredient_report_picker.set(values[0])
        except Exception as e:
            print(f"[경고] 원료목록보고 목록 로드 실패: {e}")
        finally:
            try:
                session.close()
            except Exception:
                pass

    def load_selected_ingredient_report(self):
        """콤보에서 선택된 원료목록보고를 폼에 로드합니다."""
        sel = self.ingredient_report_picker.get()
        if not sel or '|' not in sel:
            messagebox.showwarning("선택 필요", "불러올 항목을 선택하세요.", parent=self)
            return
        try:
            report_id = int(sel.split('|')[0].strip())
        except Exception:
            messagebox.showwarning("선택 오류", "선택된 항목을 해석할 수 없습니다.", parent=self)
            return
        session = db_manager.get_session()
        try:
            header = session.query(IngredientReport).filter_by(id=report_id).first()
            if not header:
                messagebox.showwarning("오류", "선택한 데이터가 존재하지 않습니다.", parent=self)
                return
            # 폼 채우기
            self.report_entries["제품명"].delete(0, 'end'); self.report_entries["제품명"].insert(0, header.product_name or '')
            self.report_entries["제조업자상호"].delete(0, 'end'); self.report_entries["제조업자상호"].insert(0, header.manufacturer or '')
            # 유형표시 콤보 값 구성과 세팅
            # 대분류는 유지, 소분류 코드만 있는 경우 그대로 세팅 시도
            try:
                # 소분류 전체 값 목록
                main_selected = self.report_entries["유형표시_대분류"].get()
                # 대분류 변경 없이 소분류 값 직접 세팅
                if header.type_code:
                    # value는 "코드 : 이름" 형식이지만 직접 코드만 세팅해도 표시됨
                    self.report_entries["유형표시"].set(header.type_code)
            except Exception:
                pass
            # 기능성 유형 설정 - 빈 문자열이면 "없음"으로 표시
            if header.functional_type_code:
                self.report_entries["기능성화장품유형"].set(header.functional_type_code)
            else:
                self.report_entries["기능성화장품유형"].set("없음 : 기능성화장품 아님")
            
            # 기능성 품목코드 설정
            self.report_entries["기능성화장품품목코드"].delete(0, 'end')
            if header.functional_type_code:  # 기능성 유형이 있을 때만 품목코드 표시
                self.report_entries["기능성화장품품목코드"].insert(0, header.functional_code or '')
            
            self.report_entries["용도"].set(header.usage or '')
            self.report_entries["맞춤형내용물"].set(header.custom_content or '')

            # 테이블
            for row_widgets in self.report_item_rows:
                for key, widget in row_widgets.items():
                    if hasattr(widget, 'destroy'):
                        widget.destroy()
            self.report_item_rows.clear()
            for idx, item in enumerate(sorted(header.items, key=lambda x: (x.row_no or 0, x.id)) , start=1):
                self._add_report_item_row_with_data(
                    row_number=idx,
                    product_name=header.product_name or '',
                    type_code=header.type_code or '',
                    functional_type=header.functional_type_code or '',
                    functional_code=header.functional_code or '',
                    manufacturer=header.manufacturer or '',
                    ingredient=item.ingredient_name or '',
                    usage=header.usage or '',
                    custom_content=header.custom_content or ''
                )
            self.current_ingredient_report_id = header.id
            self.saved_products.clear()
            messagebox.showinfo("불러오기 완료", "원료목록보고 데이터가 로드되었습니다.", parent=self)
        except Exception as e:
            messagebox.showerror("오류", f"불러오기 중 오류가 발생했습니다: {e}", parent=self)
        finally:
            session.close()

    def save_and_continue_report(self):
        """현재 제품 데이터를 저장하고 폼을 초기화하여 다음 제품을 입력받습니다."""
        # 필수 필드 확인
        if not self.report_entries["제품명"].get():
            messagebox.showwarning("입력 오류", "제품명을 입력해주세요.", parent=self)
            return

        # 원료성분명 붙여넣기 데이터 파싱
        pasted_text = self.bulk_ingredient_entry.get("1.0", "end-1c").strip()
        if ', ' in pasted_text:
            pasted_ingredients = pasted_text.split(', ')
        elif ',' in pasted_text:
            pasted_ingredients = pasted_text.split(',')
        else:
            pasted_ingredients = pasted_text.split('\n')
        pasted_ingredients = [line.strip() for line in pasted_ingredients if line.strip()]

        if not pasted_ingredients:
            messagebox.showwarning("입력 오류", "원료성분명을 입력해주세요.", parent=self)
            return

        # 제품 데이터 수집
        product_name = self.report_entries["제품명"].get()
        
        type_selected = self.report_entries["유형표시"].get()
        type_code = type_selected.split(" : ")[0] if " : " in type_selected else type_selected
        
        functional_selected = self.report_entries["기능성화장품유형"].get()
        functional_type_code = functional_selected.split(" : ")[0] if " : " in functional_selected else ""

        functional_code = self.report_entries["기능성화장품품목코드"].get()
        manufacturer = self.report_entries["제조업자상호"].get()
        usage = self.report_entries.get("용도", ctk.CTkComboBox(self, values=[])).get()
        custom_content = self.report_entries.get("맞춤형내용물", ctk.CTkComboBox(self, values=[])).get()

        # 제품 데이터 저장
        product_data = {
            "제품명": product_name,
            "유형표시": type_code,
            "기능성화장품유형": functional_type_code,
            "기능성화장품품목코드": functional_code,
            "제조업자상호": manufacturer,
            "용도": usage,
            "맞춤형내용물": custom_content,
            "원료성분명": pasted_ingredients
        }
        self.saved_products.append(product_data)

        # 폼 초기화 (제조업자상호는 유지)
        for key, entry in self.report_entries.items():
            if key == "제조업자상호":
                continue  # 제조업자상호는 유지
            elif isinstance(entry, ctk.CTkComboBox):
                values = entry.cget("values")
                if values:
                    entry.set(values[0] if values else "")
            else:
                entry.delete(0, "end")
        
        self.bulk_ingredient_entry.delete("1.0", "end")
        self._redraw_report_table()
        
        messagebox.showinfo("저장 완료", 
                          f"제품 '{product_name}'이(가) 저장되었습니다.\n"
                          f"현재 저장된 제품 수: {len(self.saved_products)}개\n\n"
                          f"계속해서 다음 제품을 입력하거나\n"
                          f"'엑셀 보고서 생성' 버튼을 눌러 모든 제품을 한 번에 출력하세요.", 
                          parent=self)

    def generate_ingredient_report(self):
        """원료목록 보고서를 엑셀로 생성합니다."""
        # 저장된 제품이 있는지 확인
        if not self.saved_products:
            # 저장된 제품이 없으면 현재 화면의 데이터를 사용
            if not self.report_entries["제품명"].get():
                messagebox.showwarning("입력 오류", "제품명을 입력해주세요.", parent=self)
                return

            # 원료성분명 붙여넣기 데이터 파싱
            pasted_text = self.bulk_ingredient_entry.get("1.0", "end-1c").strip()
            if ', ' in pasted_text:
                pasted_ingredients = pasted_text.split(', ')
            elif ',' in pasted_text:
                pasted_ingredients = pasted_text.split(',')
            else:
                pasted_ingredients = pasted_text.split('\n')
            pasted_ingredients = [line.strip() for line in pasted_ingredients if line.strip()]

            if not pasted_ingredients:
                messagebox.showwarning("입력 오류", "원료성분명을 입력해주세요.", parent=self)
                return

            # 단일 제품 처리
            self._generate_single_product_report(pasted_ingredients)
        else:
            # 여러 제품 처리
            self._generate_multiple_product_report()
    def _generate_single_product_report(self, pasted_ingredients):
        """단일 제품 보고서를 생성합니다."""
        # 테이블 재생성 및 데이터 채우기
        self._redraw_report_table()
        for ingredient in pasted_ingredients:
            self._add_report_item_row(ingredient_name=ingredient)

        # 상단 입력 데이터 수집
        product_name = self.report_entries["제품명"].get()
        
        type_selected = self.report_entries["유형표시"].get()
        type_code = type_selected.split(" : ")[0] if " : " in type_selected else type_selected
        
        functional_selected = self.report_entries["기능성화장품유형"].get()
        functional_type_code = functional_selected.split(" : ")[0] if " : " in functional_selected else ""

        functional_code = self.report_entries["기능성화장품품목코드"].get()
        manufacturer = self.report_entries["제조업자상호"].get()
        usage = self.report_entries.get("용도", ctk.CTkComboBox(self, values=[])).get()
        custom_content = self.report_entries.get("맞춤형내용물", ctk.CTkComboBox(self, values=[])).get()

        # 각 행에 데이터 채우기
        for row_widgets in self.report_item_rows:
            row_widgets["col1"].configure(state="normal")
            row_widgets["col1"].delete(0, "end")
            row_widgets["col1"].insert(0, product_name)
            row_widgets["col1"].configure(state="readonly", fg_color=("gray90", "gray25"))

            row_widgets["col2"].configure(state="normal")
            row_widgets["col2"].delete(0, "end")
            row_widgets["col2"].insert(0, type_code)
            row_widgets["col2"].configure(state="readonly", fg_color=("gray90", "gray25"))

            row_widgets["col3"].configure(state="normal")
            row_widgets["col3"].delete(0, "end")
            row_widgets["col3"].insert(0, functional_type_code)
            row_widgets["col3"].configure(state="readonly", fg_color=("gray90", "gray25"))

            row_widgets["col4"].configure(state="normal")
            row_widgets["col4"].delete(0, "end")
            row_widgets["col4"].insert(0, functional_code)
            row_widgets["col4"].configure(state="readonly", fg_color=("gray90", "gray25"))

            row_widgets["col5"].configure(state="normal")
            row_widgets["col5"].delete(0, "end")
            row_widgets["col5"].insert(0, manufacturer)
            row_widgets["col5"].configure(state="readonly", fg_color=("gray90", "gray25"))

            row_widgets["col7"].configure(state="normal")
            row_widgets["col7"].delete(0, "end")
            row_widgets["col7"].insert(0, usage)
            row_widgets["col7"].configure(state="readonly", fg_color=("gray90", "gray25"))

            row_widgets["col8"].configure(state="normal")
            row_widgets["col8"].delete(0, "end")
            row_widgets["col8"].insert(0, custom_content)
            row_widgets["col8"].configure(state="readonly", fg_color=("gray90", "gray25"))

        # 엑셀로 내보내기
        self._export_to_excel(usage, custom_content)

    def _generate_multiple_product_report(self):
        """여러 제품 보고서를 생성합니다."""
        wb = Workbook()
        ws = wb.active
        ws.title = "원료목록보고"
        
        # 스타일 정의
        header_font = Font(name='맑은 고딕', size=11, bold=True, color="FFFFFF")
        cell_font = Font(name='맑은 고딕', size=10)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # 헤더
        headers = ["일련번호", "제품명", "유형표시", "기능성화장품유형", "기능성화장품품목코드", 
                  "제조업자상호", "원료성분명", "용도", "맞춤형 내용물(혼합용'C1',소분용'C2')"]
        ws.append(headers)
        for col_idx, header_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # 모든 저장된 제품 데이터 추가
        row_number = 1
        for product in self.saved_products:
            for ingredient in product["원료성분명"]:
                excel_row = [
                    str(row_number),
                    product["제품명"],
                    product["유형표시"],
                    product["기능성화장품유형"],
                    product["기능성화장품품목코드"],
                    product["제조업자상호"],
                    ingredient,
                    product["용도"],
                    product["맞춤형내용물"]
                ]
                
                current_row_num = ws.max_row + 1
                for col_idx, value in enumerate(excel_row, 1):
                    cell = ws.cell(row=current_row_num, column=col_idx, value=value)
                    cell.font = cell_font
                    cell.border = thin_border
                    cell.alignment = left_align
                
                row_number += 1

        # 열 너비 자동 조절
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            
            # 헤더 길이 계산
            header_cell = ws[f"{column_letter}1"]
            if header_cell.value:
                header_lines = str(header_cell.value).split('\n')
                for line in header_lines:
                    if len(line) * 1.2 > max_length:
                        max_length = len(line) * 1.2
            
            # 데이터 길이 계산
            for cell in col:
                if cell.value:
                    length = sum(2 if '\uac00' <= char <= '\ud7a3' else 1 for char in str(cell.value))
                    if length > max_length:
                        max_length = length
            
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 파일 저장
        default_name = f"원료목록보고_{len(self.saved_products)}개제품.xlsx"
        file_path = fd.asksaveasfilename(defaultextension=".xlsx", 
                                        filetypes=[("Excel Files", "*.xlsx")], 
                                        initialfile=default_name, 
                                        title="엑셀로 저장")
        if file_path:
            wb.save(file_path)
            try:
                os.startfile(os.path.abspath(file_path))
            except Exception:
                pass

    def _export_to_excel(self, usage, custom_content):
        """테이블 데이터를 엑셀로 내보냅니다."""
        wb = Workbook()
        ws = wb.active
        ws.title = "원료목록보고"
        
        # 스타일 정의
        header_font = Font(name='맑은 고딕', size=11, bold=True, color="FFFFFF")
        cell_font = Font(name='맑은 고딕', size=10)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # 헤더
        headers = ["일련번호", "제품명", "유형표시", "기능성화장품유형", "기능성화장품품목코드", 
                  "제조업자상호", "원료성분명", "용도", "맞춤형 내용물(혼합용'C1',소분용'C2')"]
        ws.append(headers)
        for col_idx, header_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # 데이터 행
        for row_widgets in self.report_item_rows:
            excel_row = [row_widgets[f"col{i}"].get() for i in range(7)]
            excel_row.append(usage)  # 용도
            excel_row.append(custom_content)  # 맞춤형 내용물
            
            current_row_num = ws.max_row + 1
            for col_idx, value in enumerate(excel_row, 1):
                cell = ws.cell(row=current_row_num, column=col_idx, value=value)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = left_align

        # 열 너비 자동 조절
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            
            # 헤더 길이 계산
            header_cell = ws[f"{column_letter}1"]
            if header_cell.value:
                header_lines = str(header_cell.value).split('\n')
                for line in header_lines:
                    if len(line) * 1.2 > max_length:
                        max_length = len(line) * 1.2
            
            # 데이터 길이 계산
            for cell in col:
                if cell.value:
                    length = sum(2 if '\uac00' <= char <= '\ud7a3' else 1 for char in str(cell.value))
                    if length > max_length:
                        max_length = length
            
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 파일 저장
        default_name = f"원료목록보고_{self.report_entries['제품명'].get()}.xlsx"
        file_path = fd.asksaveasfilename(defaultextension=".xlsx", 
                                        filetypes=[("Excel Files", "*.xlsx")], 
                                        initialfile=default_name, 
                                        title="엑셀로 저장")
        if file_path:
            wb.save(file_path)
            try:
                os.startfile(os.path.abspath(file_path))
            except Exception:
                pass

    def setup_coa_tab(self, tab_frame):
        """COA 탭의 UI를 설정합니다. (서브탭 지연 로딩 적용으로 초고속 진입)"""
        tab_frame.grid_columnconfigure(0, weight=1)
        tab_frame.grid_rowconfigure(0, weight=1)

        self.coa_sub_tab_view = ctk.CTkTabview(
            tab_frame, 
            border_width=1, 
            border_color=("gray80", "gray30"),
            command=self._on_coa_subtab_changed
        )
        self.coa_sub_tab_view.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        self.coa_semi_tab_label = self.texts['semi_finished_product_report']
        self.coa_finished_tab_label = "완제품 시험성적서"

        self.coa_sub_tab_view.add(self.coa_semi_tab_label)
        self.coa_sub_tab_view.add(self.coa_finished_tab_label)
        
        self._initialized_coa_subtabs = set()
        # 기본 첫 탭인 반제품 시험성적서만 즉각 초기화 (약 0.15초)
        self.setup_semi_finished_product_tab(self.coa_sub_tab_view.tab(self.coa_semi_tab_label))
        self._initialized_coa_subtabs.add('semi')

    def _on_coa_subtab_changed(self):
        """COA 내부 서브탭 전환 시 완제품 탭 온디맨드 빌드"""
        current_tab = self.coa_sub_tab_view.get()
        if current_tab == self.coa_finished_tab_label and 'finished' not in self._initialized_coa_subtabs:
            self.setup_finished_product_tab(self.coa_sub_tab_view.tab(self.coa_finished_tab_label))
            self._initialized_coa_subtabs.add('finished')

    def setup_semi_finished_product_tab(self, tab_frame):
        """반제품 시험성적서 탭의 UI를 동적으로 재구성합니다."""
        tab_frame.grid_columnconfigure(0, weight=1)
        tab_frame.grid_rowconfigure(0, weight=1)

        scrollable_frame = ctk.CTkScrollableFrame(tab_frame, label_text=self.texts['semi_finished_product_report_title'])
        scrollable_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        scrollable_frame.grid_columnconfigure(0, weight=1)

        self.semi_product_entries = {}
        self.coa_item_rows = [] 
        self.current_semi_coa_id = None

        info_frame = ctk.CTkFrame(scrollable_frame)
        info_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 20))
        info_frame.grid_columnconfigure((1, 3), weight=1)

        info_fields = [("제 품 명", 0, 0), ("LOT", 0, 2), ("제조일자", 1, 0), ("시험일자", 1, 2)]
        for label, r, c in info_fields:
            ctk.CTkLabel(info_frame, text=label, font=ctk.CTkFont(weight="bold")).grid(row=r, column=c, padx=10, pady=5, sticky="w")
            entry = ctk.CTkEntry(info_frame)
            entry.grid(row=r, column=c+1, padx=10, pady=5, sticky="ew")
            self.semi_product_entries[label] = entry

        self.table_frame = ctk.CTkFrame(scrollable_frame)
        self.table_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=10)
        self.table_frame.grid_columnconfigure(2, weight=2) 
        self.table_frame.grid_columnconfigure(3, weight=3) 
        self.table_frame.grid_columnconfigure(4, weight=3) 
        self.table_frame.grid_columnconfigure(5, weight=1) 

        self.table_controls_frame = ctk.CTkFrame(scrollable_frame)
        self.table_controls_frame.grid(row=2, column=0, sticky="w", padx=10, pady=(0, 10))
        ctk.CTkButton(self.table_controls_frame, text="시험항목 추가", command=self._add_coa_item_row).pack(side="left")
        ctk.CTkButton(self.table_controls_frame, text="선택 항목 제거", command=self._remove_selected_coa_item_row).pack(side="left", padx=10)

        conclusion_frame = ctk.CTkFrame(scrollable_frame)
        conclusion_frame.grid(row=3, column=0, sticky="ew", padx=10, pady=(20, 10))
        conclusion_frame.grid_columnconfigure((1, 3, 5), weight=1)

        conclusion_fields = [("시험자", 0, 0), ("일자", 0, 2), ("종합판정", 0, 4)]
        for label, r, c in conclusion_fields:
            ctk.CTkLabel(conclusion_frame, text=label, font=ctk.CTkFont(weight="bold")).grid(row=r, column=c, padx=10, pady=5, sticky="w")
            entry = ctk.CTkEntry(conclusion_frame)
            entry.grid(row=r, column=c+1, padx=10, pady=5, sticky="ew")
            self.semi_product_entries[label] = entry

        button_frame = ctk.CTkFrame(tab_frame, fg_color="transparent")
        button_frame.grid(row=1, column=0, padx=10, pady=10, sticky="e")

        ctk.CTkButton(button_frame, text="📊 시험성적서 (KO)", command=lambda: self.generate_semi_product_report(lang="ko")).pack(side="left", padx=5)
        ctk.CTkButton(button_frame, text="🌐 영문 COA (EN)", fg_color="#1565C0", hover_color="#0D47A1", command=lambda: self.generate_semi_product_report(lang="en")).pack(side="left", padx=5)
        ctk.CTkButton(button_frame, text=self.texts['reset'], command=self.clear_semi_product_form, fg_color="gray50", hover_color="gray35").pack(side="left", padx=5)

        # DB 저장/불러오기 UI
        db_frame = ctk.CTkFrame(tab_frame, fg_color="transparent")
        db_frame.grid(row=2, column=0, padx=10, pady=(0, 10), sticky="ew")
        db_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkButton(db_frame, text="DB 저장", command=self.save_semi_coa_to_db,
                      fg_color="#2FA572", hover_color="#106A43", width=120).grid(row=0, column=0, padx=(0, 8))
        self.semi_coa_picker = ctk.CTkComboBox(db_frame, values=[], width=500)
        self.semi_coa_picker.grid(row=0, column=1, padx=(0, 8), sticky="ew")
        ctk.CTkButton(db_frame, text="불러오기", command=self.load_selected_semi_coa,
                      fg_color="#4C9AFF", hover_color="#1F6AA5", width=100).grid(row=0, column=2)
        ctk.CTkButton(db_frame, text="삭제", command=self.delete_selected_semi_coa,
                      fg_color="#D32F2F", hover_color="#B71C1C", width=80).grid(row=0, column=3)

        self._redraw_coa_table()
        self.refresh_semi_coa_list()

    def _redraw_coa_table(self, initial_items=None):
        for row_widgets in self.coa_item_rows:
            for widget in row_widgets.values():
                if isinstance(widget, (ctk.CTkEntry, ctk.CTkLabel, ctk.CTkCheckBox)):
                    widget.destroy()
        self.coa_item_rows.clear()

        headers = ["", "구분", "시험항목", "시험기준", "시험결과", "비고"]
        for i, h in enumerate(headers):
            header_label = ctk.CTkLabel(self.table_frame, text=h, font=ctk.CTkFont(weight="bold"), fg_color=("gray85", "gray20"), corner_radius=0)
            header_label.grid(row=0, column=i, sticky="ew", padx=(1,0), pady=(1,0))

        if initial_items:
            for item in initial_items:
                self._add_coa_item_row(item_data=item, redraw=False)
        elif not self.coa_item_rows: 
            initial_data = [
                {'num': "1", 'name': "성상", 'criteria': "표준품과 일치", 'result': "", 'remarks': ""},
                {'num': "2", 'name': "향취", 'criteria': "표준품과 일치", 'result': "", 'remarks': ""},
                {'num': "3", 'name': "사용감", 'criteria': "표준품과 일치", 'result': "", 'remarks': ""},
                {'num': "4", 'name': "pH (30℃)", 'criteria': "6.50 ± 1.00", 'result': "", 'remarks': ""},
                {'num': "5", 'name': "점도(30℃)", 'criteria': "33,000 ± 5,000", 'result': "", 'remarks': ""},
                {'num': "6", 'name': "비중(25℃)", 'criteria': "0.980 ± 0.02", 'result': "", 'remarks': ""},
                {'num': "7", 'name': "미생물(일반세균)", 'criteria': "100 cfu/ml 이하", 'result': "", 'remarks': ""},
                {'num': "", 'name': "미생물(효모/곰팡이)", 'criteria': "10 cfu/ml 이하", 'result': "", 'remarks': ""},
                {'num': "", 'name': "미생물(대장균)", 'criteria': "불검출", 'result': "", 'remarks': ""},
            ]
            for item in initial_data:
                self._add_coa_item_row(item_data=item, redraw=False)

    def _add_coa_item_row(self, item_data=None, redraw=True):
        if item_data is None:
            item_data = {'num': str(len(self.coa_item_rows) + 1), 'name': "", 'criteria': "", 'result': "", 'remarks': ""}

        row_index = len(self.coa_item_rows) + 1
        widgets = {'selected': ctk.BooleanVar()}

        chk = ctk.CTkCheckBox(self.table_frame, text="", variable=widgets['selected'])
        chk.grid(row=row_index, column=0, sticky="ew", padx=2)
        widgets['chk'] = chk

        widgets['num_label'] = ctk.CTkLabel(self.table_frame, text=item_data['num'])
        widgets['num_label'].grid(row=row_index, column=1, sticky="ew", padx=(1,0), pady=(1,0))

        widgets['name'] = ctk.CTkEntry(self.table_frame, corner_radius=0, border_width=0)
        widgets['name'].insert(0, item_data['name'])
        widgets['name'].grid(row=row_index, column=2, sticky="ew", padx=(1,0), pady=(1,0))

        widgets['criteria'] = ctk.CTkEntry(self.table_frame, corner_radius=0, border_width=0)
        widgets['criteria'].insert(0, item_data['criteria'])
        widgets['criteria'].grid(row=row_index, column=3, sticky="ew", padx=(1,0), pady=(1,0))

        widgets['result'] = ctk.CTkEntry(self.table_frame, corner_radius=0, border_width=0)
        widgets['result'].insert(0, item_data['result'])
        widgets['result'].grid(row=row_index, column=4, sticky="ew", padx=(1,0), pady=(1,0))

        widgets['remarks'] = ctk.CTkEntry(self.table_frame, corner_radius=0, border_width=0)
        widgets['remarks'].insert(0, item_data['remarks'])
        widgets['remarks'].grid(row=row_index, column=5, sticky="ew", padx=(1,0), pady=(1,0))

        self.coa_item_rows.append(widgets)
        if redraw:
            self._update_coa_row_numbers()

    def _remove_selected_coa_item_row(self):
        selected_rows = [i for i, row in enumerate(self.coa_item_rows) if row['selected'].get()]
        if not selected_rows:
            messagebox.showwarning("선택 오류", "삭제할 항목을 선택하세요.", parent=self)
            return

        for i in sorted(selected_rows, reverse=True):
            for widget in self.coa_item_rows[i].values():
                if isinstance(widget, (ctk.CTkEntry, ctk.CTkLabel, ctk.CTkCheckBox)):
                    widget.destroy()
            del self.coa_item_rows[i]
        self._update_coa_row_numbers()

    def _update_coa_row_numbers(self):
        for i, row_widgets in enumerate(self.coa_item_rows):
            row_widgets['num_label'].configure(text=str(i + 1))

    def clear_semi_product_form(self):
        """반제품 시험성적서 폼의 모든 입력 필드를 초기화합니다."""
        for entry in self.semi_product_entries.values():
            entry.delete(0, "end")
        
        self._redraw_coa_table(initial_items=None)

        self.current_semi_coa_id = None
        messagebox.showinfo(self.texts['notification'], self.texts['form_cleared'], parent=self)

    def _collect_semi_coa_data(self):
        return {
            'product_name': self.semi_product_entries.get("제 품 명").get(),
            'lot_no': self.semi_product_entries.get("LOT").get(),
            'manufacture_date': self.semi_product_entries.get("제조일자").get(),
            'test_date': self.semi_product_entries.get("시험일자").get(),
            'examiner': self.semi_product_entries.get("시험자").get(),
            'overall_result': self.semi_product_entries.get("종합판정").get(),
            'items': [
                {
                    'seq_no': i+1,
                    'item_name': row['name'].get(),
                    'spec': row['criteria'].get(),
                    'result': row['result'].get(),
                    'remark': row['remarks'].get(),
                }
                for i, row in enumerate(self.coa_item_rows)
            ]
        }

    def save_semi_coa_to_db(self):
        data = self._collect_semi_coa_data()
        if not data.get('product_name') or not data.get('overall_result'):
            messagebox.showwarning(self.texts['input_error'], self.texts['required_fields_missing'], parent=self)
            return
        from datetime import datetime
        def to_date(s):
            try:
                # 허용 가능한 몇 가지 포맷 시도
                for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"):
                    try:
                        return datetime.strptime(s.strip(), fmt).date()
                    except Exception:
                        continue
            except Exception:
                pass
            return None
        session = db_manager.get_session()
        try:
            if self.current_semi_coa_id:
                header = session.query(SemiFinishedCOA).filter_by(id=self.current_semi_coa_id).first()
                if not header:
                    self.current_semi_coa_id = None
                else:
                    header.product_name = data['product_name']
                    header.lot_no = data['lot_no']
                    header.manufacture_date = to_date(data['manufacture_date']) if data['manufacture_date'] else None
                    header.test_date = to_date(data['test_date']) if data['test_date'] else None
                    header.examiner = data['examiner']
                    header.overall_result = data['overall_result']
                    header.items.clear()
                    for it in data['items']:
                        header.items.append(SemiFinishedCOAItem(**it))
                    session.commit()
                    messagebox.showinfo(self.texts['success'], "DB에 업데이트되었습니다.", parent=self)
                    self.refresh_semi_coa_list()
                    return
            # create new
            header = SemiFinishedCOA(
                product_name=data['product_name'], lot_no=data['lot_no'],
                manufacture_date=to_date(data['manufacture_date']) if data['manufacture_date'] else None,
                test_date=to_date(data['test_date']) if data['test_date'] else None,
                examiner=data['examiner'], overall_result=data['overall_result']
            )
            for it in data['items']:
                header.items.append(SemiFinishedCOAItem(**it))
            session.add(header)
            session.commit()
            self.current_semi_coa_id = header.id
            self.refresh_semi_coa_list()
            messagebox.showinfo(self.texts['success'], "DB에 저장되었습니다.", parent=self)
        except Exception as e:
            session.rollback()
            messagebox.showerror(self.texts['error'], f"DB 저장 중 오류: {e}", parent=self)
        finally:
            session.close()

    def refresh_semi_coa_list(self):
        try:
            session = db_manager.get_session()
            recs = session.query(SemiFinishedCOA).order_by(SemiFinishedCOA.created_at.desc()).limit(50).all()
            values = [f"{r.id} | {r.product_name} | {r.lot_no or ''} | {r.created_at.strftime('%Y-%m-%d %H:%M')}" for r in recs]
            self.semi_coa_picker.configure(values=values)
            if values:
                self.semi_coa_picker.set(values[0])
        except Exception as e:
            print(f"[경고] 반제품 COA 목록 로드 실패: {e}")
        finally:
            try:
                session.close()
            except Exception:
                pass

    def load_selected_semi_coa(self):
        sel = self.semi_coa_picker.get()
        if not sel or '|' not in sel:
            messagebox.showwarning("선택 필요", "불러올 항목을 선택하세요.", parent=self)
            return
        try:
            header_id = int(sel.split('|')[0].strip())
        except Exception:
            messagebox.showwarning("선택 오류", "선택된 항목을 해석할 수 없습니다.", parent=self)
            return
        session = db_manager.get_session()
        try:
            header = session.query(SemiFinishedCOA).filter_by(id=header_id).first()
            if not header:
                messagebox.showwarning("오류", "선택한 데이터가 존재하지 않습니다.", parent=self)
                return
            # 기본 정보
            for key, entry in self.semi_product_entries.items():
                entry.delete(0, 'end')
            self.semi_product_entries["제 품 명"].insert(0, header.product_name or '')
            self.semi_product_entries["LOT"].insert(0, header.lot_no or '')
            self.semi_product_entries["제조일자"].insert(0, header.manufacture_date.strftime('%Y-%m-%d') if header.manufacture_date else '')
            self.semi_product_entries["시험일자"].insert(0, header.test_date.strftime('%Y-%m-%d') if header.test_date else '')
            self.semi_product_entries["시험자"].insert(0, header.examiner or '')
            self.semi_product_entries["종합판정"].insert(0, header.overall_result or '')

            # 테이블 재구성
            for row_widgets in self.coa_item_rows:
                for widget in row_widgets.values():
                    if isinstance(widget, (ctk.CTkEntry, ctk.CTkLabel, ctk.CTkCheckBox)):
                        widget.destroy()
            self.coa_item_rows.clear()
            for i, it in enumerate(header.items):
                self._add_coa_item_row({
                    'num': str(it.seq_no) if it.seq_no else '',
                    'name': it.item_name or '',
                    'criteria': it.spec or '',
                    'result': it.result or '',
                    'remarks': it.remark or ''
                }, redraw=False)
            self._update_coa_row_numbers()
            self.current_semi_coa_id = header.id
            messagebox.showinfo("불러오기 완료", "반제품 시험성적서가 로드되었습니다.", parent=self)
        except Exception as e:
            messagebox.showerror("오류", f"불러오기 중 오류가 발생했습니다: {e}", parent=self)
        finally:
            session.close()

    def delete_selected_semi_coa(self):
        """선택한 반제품 COA를 DB에서 삭제합니다."""
        if not getattr(self.current_user, 'can_delete', None) or not self.current_user.can_delete():
            messagebox.showwarning("권한 없음", "삭제 권한이 없습니다.", parent=self)
            return
        sel = self.semi_coa_picker.get()
        if not sel or '|' not in sel:
            messagebox.showwarning("선택 필요", "삭제할 항목을 선택하세요.", parent=self)
            return
        try:
            header_id = int(sel.split('|')[0].strip())
        except Exception:
            messagebox.showwarning("선택 오류", "선택된 항목을 해석할 수 없습니다.", parent=self)
            return
        if not messagebox.askyesno("삭제 확인", f"선택한 반제품 시험성적서를 삭제하시겠습니까?\n\n{sel}\n\n이 작업은 되돌릴 수 없습니다.", parent=self):
            return
        session = db_manager.get_session()
        try:
            header = session.query(SemiFinishedCOA).filter_by(id=header_id).first()
            if not header:
                messagebox.showwarning("오류", "선택한 데이터가 존재하지 않습니다.", parent=self)
                return
            session.delete(header)
            session.commit()
            if self.current_semi_coa_id == header_id:
                self.current_semi_coa_id = None
            self.refresh_semi_coa_list()
            messagebox.showinfo("삭제 완료", "반제품 시험성적서가 삭제되었습니다.", parent=self)
        except Exception as e:
            session.rollback()
            messagebox.showerror("오류", f"삭제 중 오류가 발생했습니다: {e}", parent=self)
        finally:
            session.close()

    def generate_semi_product_report(self, lang="ko"):
        """입력된 데이터를 기반으로 반제품 시험성적서(COA) 엑셀 파일을 생성합니다. (국문 / 영문 지원)"""
        is_eng = (lang == "en")
        try:
            kor_data = {key: entry.get() for key, entry in self.semi_product_entries.items()}
            
            # 빈 양식(폼) 상태에서도 제한 없이 엑셀 내보내기 허용
            p_name = kor_data.get("제 품 명", "").strip() or ("Semi_Product" if is_eng else "반제품")
            lot_no = kor_data.get("LOT", "").strip() or ("FORM" if is_eng else "양식")

            # 영문 항목명/기준 매핑 딕셔너리
            item_en_map = {
                "성상": ("Appearance", "Matches standard (Homogeneous cream)"),
                "향취": ("Odor", "Matches standard (Characteristic)"),
                "사용감": ("Texture/Feel", "Matches standard"),
                "pH": ("pH (30℃)", "6.50 ± 1.00"),
                "점도": ("Viscosity (30℃)", "33,000 ± 5,000 cps"),
                "비중": ("Specific Gravity (25℃)", "0.980 ± 0.020"),
                "미생물(일반세균)": ("Microbial Count (Total Aerobic)", "≤ 100 cfu/g (mL)"),
                "미생물(효모/곰팡이)": ("Microbial Count (Yeast/Mold)", "≤ 10 cfu/g (mL)"),
                "미생물(대장균)": ("Pathogens (E.coli)", "Negative (Not Detected)"),
            }

            dynamic_test_items = []
            for i, row_widgets in enumerate(self.coa_item_rows):
                raw_name = row_widgets['name'].get()
                raw_spec = row_widgets['criteria'].get()
                raw_res = row_widgets['result'].get()
                raw_rem = row_widgets['remarks'].get()
                
                if is_eng:
                    for k_key, (en_name, en_spec) in item_en_map.items():
                        if k_key in raw_name:
                            raw_name = en_name
                            if not raw_spec or raw_spec == "표준품과 일치":
                                raw_spec = en_spec
                            break
                    if "적합" in raw_res:
                        raw_res = raw_res.replace("적합", "Pass (Complies)")
                    elif "부적합" in raw_res:
                        raw_res = raw_res.replace("부적합", "Fail (Non-compliant)")

                item_data = (
                    row_widgets['num_label'].cget("text"),
                    raw_name,
                    raw_spec,
                    raw_res,
                    raw_rem
                )
                dynamic_test_items.append(item_data)

            wb = Workbook()
            ws1 = wb.active
            ws1.title = "COA (Semi-finished)" if is_eng else "반제품 시험성적서"
            
            title_font = Font(name='맑은 고딕', size=18, bold=True)
            header_font = Font(name='맑은 고딕', size=11, bold=True, color="FFFFFF")
            label_font = Font(name='맑은 고딕', size=10, bold=True)
            cell_font = Font(name='맑은 고딕', size=10)
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_align = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            label_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

            def apply_style_to_range(ws, cell_range, font=None, border=None, fill=None, alignment=None):
                rows = ws[cell_range]
                if not isinstance(rows, tuple):
                    rows = ((rows,),)
                for row in rows:
                    for cell in row:
                        if font: cell.font = font
                        if border: cell.border = border
                        if fill: cell.fill = fill
                        if alignment: cell.alignment = alignment

            ws1.merge_cells('A1:F2')
            ws1['A1'] = "CERTIFICATE OF ANALYSIS (Semi-Finished)" if is_eng else "반제품 시험성적서"
            ws1['A1'].font = title_font
            ws1['A1'].alignment = center_align
            
            ws1.merge_cells('B4:C4'); ws1.merge_cells('E4:F4')
            ws1.merge_cells('B5:C5'); ws1.merge_cells('E5:F5')
            ws1['A4'] = "Product Name" if is_eng else "제 품 명"; ws1['B4'] = kor_data.get("제 품 명")
            ws1['D4'] = "Batch / Lot No." if is_eng else "L O T"; ws1['E4'] = kor_data.get("LOT")
            ws1['A5'] = "Mfg. Date" if is_eng else "제조일자"; ws1['B5'] = kor_data.get("제조일자")
            ws1['D5'] = "Test Date" if is_eng else "시험일자"; ws1['E5'] = kor_data.get("시험일자")

            apply_style_to_range(ws1, 'A4:F5', border=thin_border)
            apply_style_to_range(ws1, 'A4:A5', font=label_font, fill=label_fill, alignment=center_align)
            apply_style_to_range(ws1, 'D4:D5', font=label_font, fill=label_fill, alignment=center_align)
            apply_style_to_range(ws1, 'B4:C5', font=cell_font, alignment=left_align)
            apply_style_to_range(ws1, 'E4:F5', font=cell_font, alignment=left_align)

            ws1.append([])
            table_start_row = ws1.max_row + 1
            
            headers = ["No.", "Test Parameters", "Specifications", "Results", "Remarks"] if is_eng else ["구분", "시험항목", "시험기준", "시험결과", "비고"]
            ws1.append(headers)
            ws1.merge_cells(start_row=table_start_row, start_column=5, end_row=table_start_row, end_column=6)

            for col_idx in range(1, 7):
                cell = ws1.cell(row=table_start_row, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            for item in dynamic_test_items:
                ws1.append(item)
                current_row = ws1.max_row
                ws1.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=6)
                apply_style_to_range(ws1, f'A{current_row}:F{current_row}', font=cell_font, border=thin_border, alignment=center_align)
                ws1[f'B{current_row}'].alignment = left_align
                ws1[f'C{current_row}'].alignment = left_align
                ws1[f'D{current_row}'].alignment = left_align
            
            ws1.append([])
            overall_val = kor_data.get("종합판정", "")
            if is_eng:
                if "적합" in overall_val: overall_val = "PASS / CONFORMS TO SPECIFICATION"
                elif "부적합" in overall_val: overall_val = "FAIL / NON-CONFORMING"
                
            ws1.append(["Overall Judgment" if is_eng else "종합판정", overall_val])
            current_row = ws1.max_row
            ws1.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)
            apply_style_to_range(ws1, f'A{current_row}:F{current_row}', border=thin_border)
            ws1[f'A{current_row}'].font = label_font; ws1[f'A{current_row}'].fill = label_fill; ws1[f'A{current_row}'].alignment = center_align
            ws1[f'B{current_row}'].font = Font(name='맑은 고딕', size=10, bold=True); ws1[f'B{current_row}'].alignment = center_align

            ws1.append(["Tester/Analyst" if is_eng else "시험자", kor_data.get("시험자"), "", "Date" if is_eng else "시험일자", kor_data.get("일자")])
            current_row = ws1.max_row
            ws1.merge_cells(f'B{current_row}:C{current_row}'); ws1.merge_cells(f'E{current_row}:F{current_row}')
            apply_style_to_range(ws1, f'A{current_row}:F{current_row}', border=thin_border)
            apply_style_to_range(ws1, f'A{current_row}', font=label_font, fill=label_fill, alignment=center_align)
            apply_style_to_range(ws1, f'D{current_row}', font=label_font, fill=label_fill, alignment=center_align)
            apply_style_to_range(ws1, f'B{current_row}:C{current_row}', font=cell_font, alignment=center_align)
            apply_style_to_range(ws1, f'E{current_row}:F{current_row}', font=cell_font, alignment=center_align)

            ws1.column_dimensions['A'].width = 16
            ws1.column_dimensions['B'].width = 28
            ws1.column_dimensions['C'].width = 28
            ws1.column_dimensions['D'].width = 16
            ws1.column_dimensions['E'].width = 25
            ws1.column_dimensions['F'].width = 12

            default_filename = f"{p_name}_{lot_no}_COA_EN.xlsx" if is_eng else f"{p_name}_{lot_no}_시험성적서.xlsx"
            file_path = fd.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel 통합 문서", "*.xlsx")],
                initialfile=default_filename,
                title="Save COA (English)" if is_eng else self.texts['save_report_as']
            )

            if file_path:
                wb.save(file_path)
                try:
                    os.startfile(os.path.abspath(file_path))
                except Exception:
                    pass

        except Exception as e:
            messagebox.showerror(self.texts['error'], f"{self.texts['report_generation_error']}:\n{e}", parent=self)

    def setup_finished_product_tab(self, tab_frame):
        """완제품 시험성적서 탭의 UI를 설정합니다."""
        tab_frame.grid_columnconfigure(0, weight=1)
        tab_frame.grid_rowconfigure(0, weight=1)

        scrollable_frame = ctk.CTkScrollableFrame(tab_frame, label_text="완제품 시험성적서")
        scrollable_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        scrollable_frame.grid_columnconfigure(0, weight=1)

        self.finished_product_entries = {}
        self.finished_item_rows = []
        self.current_finished_coa_id = None

        # 상단 기본 정보
        info_frame = ctk.CTkFrame(scrollable_frame)
        info_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 20))
        info_frame.grid_columnconfigure((1, 3), weight=1)

        info_items = [
            ("제 품 명", 0, 0, 3),
            ("반제품 제조일자", 1, 0, 1), ("반제품제조번호(LOT)", 1, 2, 1),
            ("포장 일자", 2, 0, 1), ("완제품제조번호(LOT)", 2, 2, 1),
            ("사용 기한", 3, 0, 1), ("단위 용량 (ml)", 3, 2, 1),
            ("샘플링 방법", 4, 0, 1), ("시험 일자", 4, 2, 1),
        ]

        for text, r, c, colspan in info_items:
            ctk.CTkLabel(info_frame, text=text, font=ctk.CTkFont(weight="bold")).grid(row=r, column=c, padx=10, pady=5, sticky="w")
            entry = ctk.CTkEntry(info_frame)
            entry.grid(row=r, column=c + 1, columnspan=colspan, padx=10, pady=5, sticky="ew")
            self.finished_product_entries[text] = entry

        # 시험항목 테이블
        self.finished_table_frame = ctk.CTkFrame(scrollable_frame)
        self.finished_table_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=10)
        self.finished_table_frame.grid_columnconfigure(0, weight=0)  # 체크박스 컬럼
        self.finished_table_frame.grid_columnconfigure(1, weight=1)  # 구분 컬럼
        self.finished_table_frame.grid_columnconfigure(2, weight=2)  # 시험항목 컬럼
        self.finished_table_frame.grid_columnconfigure(3, weight=3)  # 시험기준 컬럼
        self.finished_table_frame.grid_columnconfigure(4, weight=3)  # 시험결과 컬럼
        self.finished_table_frame.grid_columnconfigure(5, weight=1)  # 비고 컬럼

        # 테이블 헤더
        headers = ["선택", "구분", "시험항목", "시험기준", "시험결과", "비고"]
        for i, h in enumerate(headers):
            ctk.CTkLabel(self.finished_table_frame, text=h, font=ctk.CTkFont(weight="bold"), 
                        fg_color=("gray85", "gray20"), corner_radius=0).grid(row=0, column=i, sticky="ew", padx=2, pady=2)

        # 초기 행 생성
        self._create_initial_finished_product_rows()

        # 테이블 컨트롤 버튼 (판정/시험자 정보 바로 위에 배치)
        finished_table_controls = ctk.CTkFrame(scrollable_frame)
        finished_table_controls.grid(row=2, column=0, sticky="w", padx=10, pady=(10, 5))
        ctk.CTkButton(finished_table_controls, text="시험항목 추가", command=self._add_finished_item_row).pack(side="left")
        ctk.CTkButton(finished_table_controls, text="선택 항목 제거", command=self._remove_selected_finished_item_row).pack(side="left", padx=10)

        # 판정/시험자 정보 (특이사항)
        conclusion_frame = ctk.CTkFrame(scrollable_frame)
        conclusion_frame.grid(row=3, column=0, sticky="ew", padx=10, pady=(5, 10))
        conclusion_frame.grid_columnconfigure((1, 3, 5), weight=1)

        for idx, (label, key) in enumerate([("시험자", "시험자"), ("검토자", "검토자"), ("종합판정", "종합판정")]):
            ctk.CTkLabel(conclusion_frame, text=label, font=ctk.CTkFont(weight="bold")).grid(row=0, column=idx*2, padx=10, pady=5, sticky="w")
            self.finished_product_entries[key] = ctk.CTkEntry(conclusion_frame)
            self.finished_product_entries[key].grid(row=0, column=idx*2+1, padx=10, pady=5, sticky="ew")

        # 버튼 (스크롤 프레임 밖에 고정)
        button_frame = ctk.CTkFrame(tab_frame, fg_color="transparent")
        button_frame.grid(row=1, column=0, sticky="e", padx=10, pady=10)
        ctk.CTkButton(button_frame, text="📊 시험성적서 (KO)", command=lambda: self.generate_finished_product_report(lang="ko"), 
                      fg_color="#3B8ED0", hover_color="#1F6AA5").pack(side="left", padx=5)
        ctk.CTkButton(button_frame, text="🌐 영문 COA (EN)", command=lambda: self.generate_finished_product_report(lang="en"), 
                      fg_color="#1565C0", hover_color="#0D47A1").pack(side="left", padx=5)
        ctk.CTkButton(button_frame, text="초기화", command=self.clear_finished_product_form, 
                      fg_color="gray50", hover_color="gray35").pack(side="left", padx=5)

        # DB 저장/불러오기 UI
        db_frame = ctk.CTkFrame(tab_frame, fg_color="transparent")
        db_frame.grid(row=2, column=0, padx=10, pady=(0, 10), sticky="ew")
        db_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkButton(db_frame, text="DB 저장", command=self.save_finished_coa_to_db,
                      fg_color="#2FA572", hover_color="#106A43", width=120).grid(row=0, column=0, padx=(0, 8))
        self.finished_coa_picker = ctk.CTkComboBox(db_frame, values=[], width=500)
        self.finished_coa_picker.grid(row=0, column=1, padx=(0, 8), sticky="ew")
        ctk.CTkButton(db_frame, text="불러오기", command=self.load_selected_finished_coa,
                      fg_color="#4C9AFF", hover_color="#1F6AA5", width=100).grid(row=0, column=2)
        ctk.CTkButton(db_frame, text="삭제", command=self.delete_selected_finished_coa,
                      fg_color="#D32F2F", hover_color="#B71C1C", width=80).grid(row=0, column=3)

        self._redraw_finished_product_table()
        self.refresh_finished_coa_list()

    def _create_initial_finished_product_rows(self):
        """완제품 시험성적서의 초기 시험 항목들을 생성합니다."""
        initial_items = [
            {"id": "1", "item": "성 상", "spec": "유백색 크림상", "result": "", "note": "관능검사 / 표준품과 비교"},
            {"id": "2", "item": "향 취", "spec": "표준품과 일치", "result": "", "note": "관능검사 / 표준품과 비교"},
            {"id": "3", "item": "이물질", "spec": "미발견", "result": "", "note": "관능검사 / 표준품과 비교"},
            {"id": "4", "item": "내용량", "spec": "표시량의 97% 이상", "result": "", "note": "화장품 기준 및 시험방법 중 2.시험방법 1)내용량 가)용량으로 표시된 제품에 따른다"},
            {"id": "5", "item": "사용감", "spec": "표준품과 일치", "result": "", "note": "관능검사 / 표준품과 비교"},
            {"id": "6", "item": "pH(25℃)", "spec": "5.80 ± 1.00", "result": "", "note": "pH Meter 사용"},
            {"id": "7", "item": "점도(25℃)", "spec": "19,000 ± 4,000", "result": "", "note": "Helipath, 50rpm, Spindle-E, 1min"},
            {"id": "8", "item": "비중(25℃)", "spec": "1.010 ± 0.050", "result": "", "note": "비중병 사용"},
            {"id": "9", "item": "외관 및 용기상태", "spec": "표준품과 비교", "result": "", "note": "관능검사 / 표준품과 비교"},
            {"id": "10", "item": "인쇄상태", "spec": "육안 식별", "result": "", "note": "관능검사 / 표준품과 비교"},
            {"id": "11", "item": "법적 표시사항", "spec": "법적 규정에 적합", "result": "", "note": "화장품법 및 화장품법 시행규칙 참고"},
            {"id": "12", "item": "아데노신 함량 시험", "spec": "표시량의 90% 이상 (0.04%)", "result": "", "note": "기능성화장품 기준 및 시험방법 참고"},
            {"id": "13", "item": "아데노신 확인 시험", "spec": "검체와 표준액의 RT가 같다", "result": "", "note": "기능성화장품 기준 및 시험방법 참고"},
            {"id": "14", "item": "미생물(일반세균)", "spec": "100 cfu/ml 이하", "result": "", "note": "화장품의 미생물 한도 기준을 참고. 3M™ Petrifilm Plate법 사용"},
            {"id": "", "item": "미생물(진균(효모/곰팡이))", "spec": "10 cfu/ml 이하", "result": "", "note": ""},
            {"id": "", "item": "미생물(대장균)", "spec": "불검출", "result": "", "note": ""},
        ]
        for item_data in initial_items:
            self._add_finished_item_row_with_data(item_data)
        
        # 특이사항 줄을 맨 마지막에 추가
        self._add_special_remarks_row()

    def _redraw_finished_product_table(self):
        """완제품 시험성적서 테이블 헤더를 다시 그립니다."""
        # 헤더만 다시 그리기 (이미 _create_initial_finished_product_rows에서 행이 추가됨)
        pass

    def _add_finished_item_row(self, item_data=None, redraw=True):
        """완제품 시험항목 행을 추가합니다."""
        if item_data is None:
            item_data = {'id': '', 'item': '', 'spec': '', 'result': '', 'note': ''}

        # 특이사항 줄이 아닌 경우에만 번호를 부여
        if item_data.get('id') != "특이사항":
            # 현재 특이사항 줄을 제외한 행의 개수를 계산
            non_special_count = 0
            for row in self.finished_item_rows:
                if row.get('id_label') and row['id_label'].cget("text") != "특이사항":
                    non_special_count += 1
            item_data['id'] = str(non_special_count + 1)

        # 행 인덱스 계산 (특이사항은 항상 마지막에)
        if item_data.get('id') == "특이사항":
            row_index = len(self.finished_item_rows) + 1
        else:
            # 특이사항 줄이 있다면 그 바로 앞에 삽입
            special_index = None
            for i, row in enumerate(self.finished_item_rows):
                if row.get('id_label') and row['id_label'].cget("text") == "특이사항":
                    special_index = i
                    break
            
            if special_index is not None:
                row_index = special_index + 1
            else:
                row_index = len(self.finished_item_rows) + 1

        widgets = {'selected': ctk.BooleanVar()}

        # 체크박스 (column 0)
        chk = ctk.CTkCheckBox(self.finished_table_frame, text="", variable=widgets['selected'])
        chk.grid(row=row_index, column=0, sticky="w", padx=2, pady=2)
        widgets['chk'] = chk

        # 구분(ID) 라벨 (column 1)
        widgets['id_label'] = ctk.CTkLabel(self.finished_table_frame, text=item_data['id'])
        widgets['id_label'].grid(row=row_index, column=1, sticky="ew", padx=2, pady=2)

        # 시험항목, 시험기준, 시험결과, 비고 (columns 2-5)
        for i, key in enumerate(['item', 'spec', 'result', 'note'], start=2):
            entry = ctk.CTkEntry(self.finished_table_frame, corner_radius=0, border_width=0)
            entry.insert(0, item_data.get(key, ''))
            entry.grid(row=row_index, column=i, sticky="ew", padx=(1,0), pady=(1,0))
            widgets[key] = entry

        # 특이사항 줄이 아닌 경우, 특이사항 줄 바로 앞에 삽입
        if item_data.get('id') != "특이사항" and special_index is not None:
            self.finished_item_rows.insert(special_index, widgets)
            # 특이사항 줄과 그 뒤의 모든 행들을 한 칸씩 아래로 이동
            self._reposition_rows_from_index(special_index + 1)
        else:
            self.finished_item_rows.append(widgets)
            
        if redraw:
            self._update_finished_row_numbers()

    def _reposition_rows_from_index(self, start_index):
        """지정된 인덱스부터 모든 행들을 한 칸씩 아래로 이동시킵니다."""
        for i in range(start_index, len(self.finished_item_rows)):
            row = self.finished_item_rows[i]
            new_row_index = i + 2  # grid row는 1부터 시작하고 헤더가 있으므로 +2
            
            # 각 위젯을 새로운 행 위치로 이동
            if 'chk' in row:
                row['chk'].grid(row=new_row_index, column=0, sticky="w", padx=2, pady=2)
            if 'id_label' in row:
                row['id_label'].grid(row=new_row_index, column=1, sticky="ew", padx=2, pady=2)
            
            for j, key in enumerate(['item', 'spec', 'result', 'note'], start=2):
                if key in row:
                    row[key].grid(row=new_row_index, column=j, sticky="ew", padx=(1,0), pady=(1,0))

    def _add_finished_item_row_with_data(self, item_data):
        """완제품 시험항목 행을 데이터와 함께 추가합니다."""
        self._add_finished_item_row(item_data, redraw=False)

    def _add_special_remarks_row(self):
        """특이사항 줄을 맨 마지막에 추가합니다."""
        if not self._has_special_remarks_row():
            special_data = {"id": "특이사항", "item": "", "spec": "", "result": "", "note": ""}
            self._add_finished_item_row(special_data, redraw=False)

    def _has_special_remarks_row(self):
        """특이사항 줄이 있는지 확인합니다."""
        for row in self.finished_item_rows:
            if row.get('id_label') and row['id_label'].cget("text") == "특이사항":
                return True
        return False

    def _remove_selected_finished_item_row(self):
        """선택된 완제품 시험항목 행을 제거합니다. (특이사항 줄 제외)"""
        selected_rows = []
        for i, row in enumerate(self.finished_item_rows):
            # 특이사항 줄은 삭제 대상에서 제외
            if row['selected'].get() and row['id_label'].cget("text") != "특이사항":
                selected_rows.append(i)
        
        if not selected_rows:
            messagebox.showwarning("선택 오류", "삭제할 항목을 선택하세요. (특이사항 줄은 삭제할 수 없습니다)", parent=self)
            return

        for i in sorted(selected_rows, reverse=True):
            for widget in self.finished_item_rows[i].values():
                if hasattr(widget, 'destroy'):
                    widget.destroy()
            del self.finished_item_rows[i]
        
        # 삭제 후 모든 행을 다시 배치
        self._reposition_all_rows()
        self._update_finished_row_numbers()

    def _reposition_all_rows(self):
        """모든 행을 다시 배치합니다."""
        for i, row in enumerate(self.finished_item_rows):
            new_row_index = i + 2  # grid row는 1부터 시작하고 헤더가 있으므로 +2
            
            # 각 위젯을 새로운 행 위치로 이동
            if 'chk' in row:
                row['chk'].grid(row=new_row_index, column=0, sticky="w", padx=2, pady=2)
            if 'id_label' in row:
                row['id_label'].grid(row=new_row_index, column=1, sticky="ew", padx=2, pady=2)
            
            for j, key in enumerate(['item', 'spec', 'result', 'note'], start=2):
                if key in row:
                    row[key].grid(row=new_row_index, column=j, sticky="ew", padx=(1,0), pady=(1,0))

    def _update_finished_row_numbers(self):
        """완제품 시험항목의 일련번호를 업데이트합니다. (특이사항 줄 제외)"""
        number = 1
        for row_widgets in self.finished_item_rows:
            current_id = row_widgets['id_label'].cget("text")
            # 특이사항 줄이 아니고 기존에 숫자였던 경우만 번호 업데이트
            if current_id != "특이사항" and (current_id == "" or current_id.isdigit()):
                row_widgets['id_label'].configure(text=str(number))
                number += 1

    def clear_finished_product_form(self):
        """완제품 시험성적서 폼을 초기화합니다."""
        for entry in self.finished_product_entries.values():
            entry.delete(0, "end")

        for row_widgets in self.finished_item_rows[:]:
            for widget in row_widgets.values():
                if hasattr(widget, 'destroy'):
                    widget.destroy()
        self.finished_item_rows.clear()

        self._create_initial_finished_product_rows()
        self.current_finished_coa_id = None
        messagebox.showinfo("알림", "양식이 초기화되었습니다.", parent=self)

    def generate_finished_product_report(self, lang="ko"):
        """입력된 데이터를 기반으로 완제품 시험성적서(COA) 엑셀 파일을 생성합니다. (국문 / 영문 지원)"""
        is_eng = (lang == "en")
        try:
            # 데이터 수집
            info_data = {key: entry.get() for key, entry in self.finished_product_entries.items()}

            # 빈 양식(폼) 상태에서도 제한 없이 엑셀 내보내기 허용
            p_name = info_data.get("제 품 명", "").strip() or ("Finished_Product" if is_eng else "완제품")
            lot_no = info_data.get("완제품제조번호(LOT)", "").strip() or ("FORM" if is_eng else "양식")

            finished_item_en_map = {
                "성 상": ("Appearance", "Matches standard (Homogeneous cream)"),
                "향 취": ("Odor", "Matches standard (Characteristic)"),
                "이물질": ("Foreign Matter", "None detected"),
                "내용량": ("Net Contents", "≥ 97% of labeled volume"),
                "사용감": ("Application Feel", "Matches standard"),
                "pH": ("pH (25℃)", "5.80 ± 1.00"),
                "점도": ("Viscosity (25℃)", "19,000 ± 4,000 cps"),
                "비중": ("Specific Gravity (25℃)", "1.010 ± 0.050"),
                "외관 및 용기상태": ("Packaging / Container Integrity", "Matches standard"),
                "인쇄상태": ("Label Printing & Legibility", "Legible & Matches standard"),
                "법적 표시사항": ("Regulatory Compliance", "Complies with Cosmetics Act"),
                "아데노신 함량 시험": ("Assay (Adenosine)", "≥ 90.0% of labeled amount (0.04%)"),
                "아데노신 확인 시험": ("Identification (Adenosine)", "Retention time matches reference standard"),
                "미생물(일반세균)": ("Total Aerobic Microbial Count (TAMC)", "≤ 100 cfu/g (mL)"),
                "미생물(진균(효모/곰팡이))": ("Total Combined Yeasts/Molds (TYMC)", "≤ 10 cfu/g (mL)"),
                "미생물(대장균)": ("Pathogens (E.coli)", "Negative (Not Detected)"),
            }

            test_items = []
            for row_widgets in self.finished_item_rows:
                item_id = row_widgets["id_label"].cget("text")
                item = row_widgets["item"].get()
                spec = row_widgets["spec"].get()
                result = row_widgets["result"].get()
                note = row_widgets["note"].get()

                if is_eng:
                    for k_key, (en_name, en_spec) in finished_item_en_map.items():
                        if k_key in item:
                            item = en_name
                            if not spec or "표준품과 일치" in spec or "표준품과 비교" in spec:
                                spec = en_spec
                            break
                    if "적합" in result:
                        result = result.replace("적합", "Pass (Complies)")
                    elif "부적합" in result:
                        result = result.replace("부적합", "Fail (Non-compliant)")

                test_items.append({"id": item_id, "item": item, "spec": spec, "result": result, "note": note})

            # Excel 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "COA (Finished Product)" if is_eng else "완제품 시험성적서"

            # 스타일 정의
            title_font = Font(name='맑은 고딕', size=18, bold=True)
            header_font = Font(name='맑은 고딕', size=11, bold=True, color="FFFFFF")
            label_font = Font(name='맑은 고딕', size=10, bold=True)
            cell_font = Font(name='맑은 고딕', size=10)
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_align = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            label_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

            def apply_style(cell, font=None, border=None, fill=None, alignment=None):
                if font: cell.font = font
                if border: cell.border = border
                if fill: cell.fill = fill
                if alignment: cell.alignment = alignment

            # 문서 제목
            ws.merge_cells('A1:E2')
            ws['A1'] = "CERTIFICATE OF ANALYSIS (Finished Product)" if is_eng else "완제품 시험성적서"
            apply_style(ws['A1'], font=title_font, alignment=center_align)

            # 기본 정보
            if is_eng:
                info_layout = [
                    ("Product Name", info_data.get("제 품 명")), ("Bulk Mfg. Date", info_data.get("반제품 제조일자")), ("Bulk Lot No.", info_data.get("반제품제조번호(LOT)")),
                    ("Packaging Date", info_data.get("포장 일자")), ("Finished Lot No.", info_data.get("완제품제조번호(LOT)")), ("Expiry Date", info_data.get("사용 기한")),
                    ("Net Volume (ml)", info_data.get("단위 용량 (ml)")), ("Sampling Method", info_data.get("샘플링 방법")), ("Test Date", info_data.get("시험 일자"))
                ]
            else:
                info_layout = [
                    ("제 품 명", info_data.get("제 품 명")), ("반제품 제조일자", info_data.get("반제품 제조일자")), ("반제품제조번호(LOT)", info_data.get("반제품제조번호(LOT)")),
                    ("포장 일자", info_data.get("포장 일자")), ("완제품제조번호(LOT)", info_data.get("완제품제조번호(LOT)")), ("사용 기한", info_data.get("사용 기한")),
                    ("단위 용량 (ml)", info_data.get("단위 용량 (ml)")), ("샘플링 방법", info_data.get("샘플링 방법")), ("시험 일자", info_data.get("시험 일자"))
                ]
            row = 4
            # 제품명 행
            apply_style(ws.cell(row, 1, info_layout[0][0]), font=label_font, fill=label_fill, alignment=center_align, border=thin_border)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
            apply_style(ws.cell(row, 2, info_layout[0][1]), font=cell_font, alignment=left_align, border=thin_border)
            for c in range(3, 6): apply_style(ws.cell(row, c), border=thin_border)
            row += 1
            # 나머지 정보 행
            for i in range(1, len(info_layout), 2):
                apply_style(ws.cell(row, 1, info_layout[i][0]), font=label_font, fill=label_fill, alignment=center_align, border=thin_border)
                apply_style(ws.cell(row, 2, info_layout[i][1]), font=cell_font, alignment=left_align, border=thin_border)
                if i + 1 < len(info_layout):
                    apply_style(ws.cell(row, 3, info_layout[i+1][0]), font=label_font, fill=label_fill, alignment=center_align, border=thin_border)
                    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
                    apply_style(ws.cell(row, 4, info_layout[i+1][1]), font=cell_font, alignment=left_align, border=thin_border)
                    apply_style(ws.cell(row, 5), border=thin_border)
                row += 1

            # 시험 항목 테이블
            ws.append([])
            table_start_row = ws.max_row + 1
            headers = ["No.", "Test Parameters", "Specifications", "Results", "Remarks"] if is_eng else ["구분", "시험항목", "시험기준", "시험결과", "비고"]
            ws.append(headers)
            for col_idx, header_text in enumerate(headers, 1):
                apply_style(ws.cell(table_start_row, col_idx), font=header_font, fill=header_fill, alignment=center_align, border=thin_border)

            for item in test_items:
                ws.append([item["id"], item["item"], item["spec"], item["result"], item["note"]])
                for col in range(1, 6):
                    apply_style(ws.cell(ws.max_row, col), font=cell_font, border=thin_border, alignment=center_align)

            # 종합판정 및 시험자
            ws.append([])
            conclusion_row = ws.max_row + 1
            apply_style(ws.cell(conclusion_row, 1, "Tester" if is_eng else "시험자"), font=label_font, fill=label_fill, alignment=center_align, border=thin_border)
            apply_style(ws.cell(conclusion_row, 2, info_data.get("시험자")), font=cell_font, alignment=center_align, border=thin_border)
            apply_style(ws.cell(conclusion_row, 3, "Reviewer" if is_eng else "검토자"), font=label_font, fill=label_fill, alignment=center_align, border=thin_border)
            apply_style(ws.cell(conclusion_row, 4, info_data.get("검토자")), font=cell_font, alignment=center_align, border=thin_border)
            ws.cell(conclusion_row, 5).border = thin_border

            conclusion_row += 1
            overall_val = info_data.get("종합판정", "")
            if is_eng:
                if "적합" in overall_val: overall_val = "PASS / CONFORMS TO SPECIFICATION"
                elif "부적합" in overall_val: overall_val = "FAIL / NON-CONFORMING"

            apply_style(ws.cell(conclusion_row, 1, "Overall Judgment" if is_eng else "종합판정"), font=label_font, fill=label_fill, alignment=center_align, border=thin_border)
            ws.merge_cells(start_row=conclusion_row, start_column=2, end_row=conclusion_row, end_column=5)
            apply_style(ws.cell(conclusion_row, 2, overall_val), font=Font(name='맑은 고딕', size=10, bold=True), alignment=center_align, border=thin_border)
            for c in range(3, 6): apply_style(ws.cell(conclusion_row, c), border=thin_border)

            # 열 너비 조정
            ws.column_dimensions['A'].width = 16
            ws.column_dimensions['B'].width = 28
            ws.column_dimensions['C'].width = 32
            ws.column_dimensions['D'].width = 22
            ws.column_dimensions['E'].width = 35

            # 파일 저장
            default_filename = f"{p_name}_{lot_no}_COA_Finished_EN.xlsx" if is_eng else f"{p_name}_{lot_no}_완제품시험성적서.xlsx"
            file_path = fd.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel 통합 문서", "*.xlsx")],
                initialfile=default_filename,
                title="Save Finished Product COA (English)" if is_eng else "보고서 다른 이름으로 저장"
            )

            if file_path:
                wb.save(file_path)
                try:
                    os.startfile(os.path.abspath(file_path))
                except Exception:
                    pass

        except Exception as e:
            messagebox.showerror("오류" if not is_eng else "Error", f"보고서 생성 중 오류가 발생했습니다:\n{e}", parent=self)

    def _collect_finished_coa_data(self):
        info = {k: e.get() for k, e in self.finished_product_entries.items()}
        items = []
        for row in self.finished_item_rows:
            items.append({
                'item_id': row['id_label'].cget('text'),
                'item_name': row['item'].get(),
                'spec': row['spec'].get(),
                'result': row['result'].get(),
                'note': row['note'].get(),
            })
        return info, items

    def save_finished_coa_to_db(self):
        info, items = self._collect_finished_coa_data()
        if not info.get('제 품 명') or not info.get('완제품제조번호(LOT)') or not info.get('종합판정'):
            messagebox.showwarning("입력 오류", "제품명, 완제품 LOT, 종합판정은 필수입니다.", parent=self)
            return
        from datetime import datetime
        def to_date(s):
            try:
                for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"):
                    try:
                        return datetime.strptime(s.strip(), fmt).date()
                    except Exception:
                        continue
            except Exception:
                pass
            return None
        session = db_manager.get_session()
        try:
            if self.current_finished_coa_id:
                header = session.query(FinishedProductCOA).filter_by(id=self.current_finished_coa_id).first()
                if header:
                    header.product_name = info.get('제 품 명')
                    header.semi_mfg_date = to_date(info.get('반제품 제조일자') or '')
                    header.semi_lot_no = info.get('반제품제조번호(LOT)')
                    header.pack_date = to_date(info.get('포장 일자') or '')
                    header.finished_lot_no = info.get('완제품제조번호(LOT)')
                    header.expiry_date = to_date(info.get('사용 기한') or '')
                    try:
                        header.unit_volume_ml = float(info.get('단위 용량 (ml)') or 0)
                    except Exception:
                        header.unit_volume_ml = None
                    header.sampling_method = info.get('샘플링 방법')
                    header.test_date = to_date(info.get('시험 일자') or '')
                    header.examiner = info.get('시험자')
                    header.reviewer = info.get('검토자')
                    header.overall_result = info.get('종합판정')
                    header.items.clear()
                    for it in items:
                        header.items.append(FinishedProductCOAItem(**it))
                    session.commit()
                    self.refresh_finished_coa_list()
                    messagebox.showinfo("저장 완료", "DB에 업데이트되었습니다.", parent=self)
                    return
                else:
                    self.current_finished_coa_id = None

            header = FinishedProductCOA(
                product_name=info.get('제 품 명'),
                semi_mfg_date=to_date(info.get('반제품 제조일자') or ''),
                semi_lot_no=info.get('반제품제조번호(LOT)'),
                pack_date=to_date(info.get('포장 일자') or ''),
                finished_lot_no=info.get('완제품제조번호(LOT)'),
                expiry_date=to_date(info.get('사용 기한') or ''),
                unit_volume_ml=float(info.get('단위 용량 (ml)')) if (info.get('단위 용량 (ml)') or '').strip().replace('.', '', 1).isdigit() else None,
                sampling_method=info.get('샘플링 방법'),
                test_date=to_date(info.get('시험 일자') or ''),
                examiner=info.get('시험자'),
                reviewer=info.get('검토자'),
                overall_result=info.get('종합판정')
            )
            for it in items:
                header.items.append(FinishedProductCOAItem(**it))
            session.add(header)
            session.commit()
            self.current_finished_coa_id = header.id
            self.refresh_finished_coa_list()
            messagebox.showinfo("저장 완료", "DB에 저장되었습니다.", parent=self)
        except Exception as e:
            session.rollback()
            messagebox.showerror("오류", f"DB 저장 중 오류: {e}", parent=self)
        finally:
            session.close()

    def refresh_finished_coa_list(self):
        try:
            session = db_manager.get_session()
            recs = session.query(FinishedProductCOA).order_by(FinishedProductCOA.created_at.desc()).limit(50).all()
            values = [f"{r.id} | {r.product_name} | {r.finished_lot_no or ''} | {r.created_at.strftime('%Y-%m-%d %H:%M')}" for r in recs]
            self.finished_coa_picker.configure(values=values)
            if values:
                self.finished_coa_picker.set(values[0])
        except Exception as e:
            print(f"[경고] 완제품 COA 목록 로드 실패: {e}")
        finally:
            try:
                session.close()
            except Exception:
                pass

    def load_selected_finished_coa(self):
        sel = self.finished_coa_picker.get()
        if not sel or '|' not in sel:
            messagebox.showwarning("선택 필요", "불러올 항목을 선택하세요.", parent=self)
            return
        try:
            header_id = int(sel.split('|')[0].strip())
        except Exception:
            messagebox.showwarning("선택 오류", "선택된 항목을 해석할 수 없습니다.", parent=self)
            return
        session = db_manager.get_session()
        try:
            header = session.query(FinishedProductCOA).filter_by(id=header_id).first()
            if not header:
                messagebox.showwarning("오류", "선택한 데이터가 존재하지 않습니다.", parent=self)
                return
            # 입력 초기화
            for entry in self.finished_product_entries.values():
                entry.delete(0, 'end')
            # 헤더 채우기
            self.finished_product_entries['제 품 명'].insert(0, header.product_name or '')
            self.finished_product_entries['반제품 제조일자'].insert(0, header.semi_mfg_date.strftime('%Y-%m-%d') if header.semi_mfg_date else '')
            self.finished_product_entries['반제품제조번호(LOT)'].insert(0, header.semi_lot_no or '')
            self.finished_product_entries['포장 일자'].insert(0, header.pack_date.strftime('%Y-%m-%d') if header.pack_date else '')
            self.finished_product_entries['완제품제조번호(LOT)'].insert(0, header.finished_lot_no or '')
            self.finished_product_entries['사용 기한'].insert(0, header.expiry_date.strftime('%Y-%m-%d') if header.expiry_date else '')
            self.finished_product_entries['단위 용량 (ml)'].insert(0, str(header.unit_volume_ml or ''))
            self.finished_product_entries['샘플링 방법'].insert(0, header.sampling_method or '')
            self.finished_product_entries['시험 일자'].insert(0, header.test_date.strftime('%Y-%m-%d') if header.test_date else '')
            self.finished_product_entries['시험자'].insert(0, header.examiner or '')
            self.finished_product_entries['검토자'].insert(0, header.reviewer or '')
            self.finished_product_entries['종합판정'].insert(0, header.overall_result or '')

            # 테이블 재구성
            for row_widgets in self.finished_item_rows[:]:
                for widget in row_widgets.values():
                    if hasattr(widget, 'destroy'):
                        widget.destroy()
            self.finished_item_rows.clear()
            for it in header.items:
                self._add_finished_item_row_with_data({
                    'id': it.item_id or '',
                    'item': it.item_name or '',
                    'spec': it.spec or '',
                    'result': it.result or '',
                    'note': it.note or '',
                })
            if not any(r.get('id_label') and r['id_label'].cget('text') == '특이사항' for r in self.finished_item_rows):
                self._add_special_remarks_row()
            self.current_finished_coa_id = header.id
            messagebox.showinfo("불러오기 완료", "완제품 시험성적서가 로드되었습니다.", parent=self)
        except Exception as e:
            messagebox.showerror("오류", f"불러오기 중 오류가 발생했습니다: {e}", parent=self)
        finally:
            session.close()

    def delete_selected_finished_coa(self):
        """선택한 완제품 COA를 DB에서 삭제합니다."""
        # 권한 확인
        if not getattr(self.current_user, 'can_delete', None) or not self.current_user.can_delete():
            messagebox.showwarning("권한 없음", "삭제 권한이 없습니다.", parent=self)
            return
        # 선택값 파싱
        sel = self.finished_coa_picker.get()
        if not sel or '|' not in sel:
            messagebox.showwarning("선택 필요", "삭제할 항목을 선택하세요.", parent=self)
            return
        try:
            header_id = int(sel.split('|')[0].strip())
        except Exception:
            messagebox.showwarning("선택 오류", "선택된 항목을 해석할 수 없습니다.", parent=self)
            return
        # 확인 다이얼로그
        if not messagebox.askyesno("삭제 확인", f"선택한 완제품 시험성적서를 삭제하시겠습니까?\n\n{sel}\n\n이 작업은 되돌릴 수 없습니다.", parent=self):
            return
        # 삭제 실행
        session = db_manager.get_session()
        try:
            header = session.query(FinishedProductCOA).filter_by(id=header_id).first()
            if not header:
                messagebox.showwarning("오류", "선택한 데이터가 존재하지 않습니다.", parent=self)
                return
            session.delete(header)
            session.commit()
            if self.current_finished_coa_id == header_id:
                self.current_finished_coa_id = None
            self.refresh_finished_coa_list()
            messagebox.showinfo("삭제 완료", "완제품 시험성적서가 삭제되었습니다.", parent=self)
        except Exception as e:
            session.rollback()
            messagebox.showerror("오류", f"삭제 중 오류가 발생했습니다: {e}", parent=self)
        finally:
            session.close()

    # =========================================================================
    # 1. 원료 입고검사성적서 (Raw Material Incoming Inspection Report)
    # =========================================================================
    def setup_material_inspection_tab(self, tab_frame):
        tab_frame.grid_columnconfigure(0, weight=1)
        tab_frame.grid_rowconfigure(0, weight=1)

        scroll = ctk.CTkScrollableFrame(tab_frame, label_text="원료 입고검사성적서 (Raw Material Incoming Test)")
        scroll.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        scroll.grid_columnconfigure(0, weight=1)

        self.mat_insp_entries = {}
        self.current_mat_insp_id = None

        # 헤더 & 불러오기 툴바
        top_bar = ctk.CTkFrame(scroll, fg_color="transparent")
        top_bar.grid(row=0, column=0, sticky="ew", padx=10, pady=(5, 15))
        top_bar.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(top_bar, text="이력 선택:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=5, pady=5)
        self.mat_insp_picker = ctk.CTkComboBox(top_bar, values=["-- 저장된 검사성적서 선택 --"], width=350)
        self.mat_insp_picker.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ctk.CTkButton(top_bar, text="불러오기", width=80, command=self.load_selected_mat_insp).grid(row=0, column=2, padx=5, pady=5)
        ctk.CTkButton(top_bar, text="새로작성", width=80, fg_color="gray50", command=self.clear_mat_insp_form).grid(row=0, column=3, padx=5, pady=5)
        ctk.CTkButton(top_bar, text="삭제", width=70, fg_color="#D32F2F", hover_color="#B71C1C", command=self.delete_mat_insp).grid(row=0, column=4, padx=5, pady=5)

        # 기본 정보 프레임
        info_frame = ctk.CTkFrame(scroll)
        info_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=5)
        info_frame.grid_columnconfigure((1, 3, 5), weight=1)

        fields = [
            ("원료코드", 0, 0), ("원료명", 0, 2), ("공급업체", 0, 4),
            ("입고일자", 1, 0), ("입고LOT", 1, 2), ("입고수량(kg)", 1, 4),
            ("시험자", 2, 0), ("종합판정", 2, 2), ("포장상태", 2, 4)
        ]

        for label_text, r, c in fields:
            ctk.CTkLabel(info_frame, text=label_text, font=ctk.CTkFont(weight="bold")).grid(row=r, column=c, padx=10, pady=6, sticky="w")
            if label_text == "종합판정":
                cb = ctk.CTkComboBox(info_frame, values=["적합 (Pass)", "부적합 (Fail)", "조건부 적합"], width=160)
                cb.set("적합 (Pass)")
                cb.grid(row=r, column=c+1, padx=10, pady=6, sticky="ew")
                self.mat_insp_entries[label_text] = cb
            elif label_text == "포장상태":
                cb = ctk.CTkComboBox(info_frame, values=["정상 (밀봉양호)", "파손/오염", "라벨불량"], width=160)
                cb.set("정상 (밀봉양호)")
                cb.grid(row=r, column=c+1, padx=10, pady=6, sticky="ew")
                self.mat_insp_entries[label_text] = cb
            else:
                entry = ctk.CTkEntry(info_frame)
                if label_text == "입고일자":
                    entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
                entry.grid(row=r, column=c+1, padx=10, pady=6, sticky="ew")
                self.mat_insp_entries[label_text] = entry

        # 검사 항목 테이블 프레임 (동적 추가/제거 지원)
        self.mat_insp_test_frame = ctk.CTkFrame(scroll)
        self.mat_insp_test_frame.grid(row=2, column=0, sticky="ew", padx=10, pady=15)
        self.mat_insp_test_frame.grid_columnconfigure(1, weight=2)
        self.mat_insp_test_frame.grid_columnconfigure(2, weight=3)
        self.mat_insp_test_frame.grid_columnconfigure(3, weight=2)
        self.mat_insp_test_frame.grid_columnconfigure(4, weight=1)

        # 항목 관리 헤더 툴바
        test_header_bar = ctk.CTkFrame(self.mat_insp_test_frame, fg_color="transparent")
        test_header_bar.grid(row=0, column=0, columnspan=6, sticky="ew", padx=10, pady=(5, 5))
        test_header_bar.grid_columnconfigure(0, weight=1)
        
        ctk.CTkLabel(test_header_bar, text="🧪 시험 항목 및 판정 기준 (자유 추가/제거 가능)", font=ctk.CTkFont(weight="bold", size=13), text_color=("#1F497D", "#38BDF8")).grid(row=0, column=0, sticky="w")
        ctk.CTkButton(test_header_bar, text="➕ 시험 항목 추가", width=120, height=28, fg_color="#0284C7", hover_color="#0369A1", command=lambda: self.add_mat_insp_item_row()).grid(row=0, column=1, padx=5, sticky="e")

        # 테이블 헤더 라벨
        ctk.CTkLabel(self.mat_insp_test_frame, text="NO", font=ctk.CTkFont(weight="bold"), width=35).grid(row=1, column=0, padx=4, pady=5)
        ctk.CTkLabel(self.mat_insp_test_frame, text="시험 항목명", font=ctk.CTkFont(weight="bold")).grid(row=1, column=1, padx=4, pady=5)
        ctk.CTkLabel(self.mat_insp_test_frame, text="시험 기준 (Specification)", font=ctk.CTkFont(weight="bold")).grid(row=1, column=2, padx=4, pady=5)
        ctk.CTkLabel(self.mat_insp_test_frame, text="시험 결과 (Result)", font=ctk.CTkFont(weight="bold")).grid(row=1, column=3, padx=4, pady=5)
        ctk.CTkLabel(self.mat_insp_test_frame, text="판정", font=ctk.CTkFont(weight="bold"), width=70).grid(row=1, column=4, padx=4, pady=5)
        ctk.CTkLabel(self.mat_insp_test_frame, text="관리", font=ctk.CTkFont(weight="bold"), width=50).grid(row=1, column=5, padx=4, pady=5)

        self.mat_insp_item_rows = []
        self._init_default_mat_insp_items()

        # 비고 및 버튼
        note_frame = ctk.CTkFrame(scroll, fg_color="transparent")
        note_frame.grid(row=3, column=0, sticky="ew", padx=10, pady=10)
        note_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(note_frame, text="특이사항/비고:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10, pady=5, sticky="nw")
        self.mat_insp_note = ctk.CTkTextbox(note_frame, height=60)
        self.mat_insp_note.grid(row=0, column=1, padx=10, pady=5, sticky="ew")

        btn_bar = ctk.CTkFrame(scroll, fg_color="transparent")
        btn_bar.grid(row=4, column=0, sticky="e", padx=10, pady=15)

        ctk.CTkButton(btn_bar, text="💾 DB 저장/수정", width=120, command=self.save_mat_insp_to_db).pack(side="left", padx=5)
        ctk.CTkButton(btn_bar, text="📊 엑셀 성적서 출력", width=140, fg_color="#2E7D32", hover_color="#1B5E20", command=self.export_mat_insp_to_excel).pack(side="left", padx=5)

        self.refresh_mat_insp_list()

    def _init_default_mat_insp_items(self):
        """기본 표준 검사 항목들 로드"""
        self._clear_mat_insp_item_rows()
        default_items = [
            ("성상/외관", "고유의 성상을 띰", "적합", "적합"),
            ("색상", "고유의 색상", "적합", "적합"),
            ("향취", "고유의 취", "적합", "적합"),
            ("굴절률 (20℃)", "1.300 ~ 1.500", "1.334", "적합"),
            ("비중 (20℃)", "0.900 ~ 1.100", "1.002", "적합"),
            ("pH (10% soln)", "4.50 ~ 7.50", "6.20", "적합")
        ]
        for name, spec, res, judge in default_items:
            self.add_mat_insp_item_row(name, spec, res, judge)

    def _clear_mat_insp_item_rows(self):
        for row in self.mat_insp_item_rows:
            for w in row['widgets'].values():
                w.destroy()
        self.mat_insp_item_rows.clear()

    def add_mat_insp_item_row(self, name="", spec="", result="", judge="적합"):
        """새로운 시험 항목 행을 동적으로 추가합니다."""
        row_idx = len(self.mat_insp_item_rows) + 2  # 헤더가 0, 1행 차지
        
        no_lbl = ctk.CTkLabel(self.mat_insp_test_frame, text=str(len(self.mat_insp_item_rows) + 1), width=35)
        no_lbl.grid(row=row_idx, column=0, padx=4, pady=3)
        
        name_entry = ctk.CTkEntry(self.mat_insp_test_frame, placeholder_text="항목명 (예: 순도, 점도, 중금속 등)")
        name_entry.insert(0, name)
        name_entry.grid(row=row_idx, column=1, padx=4, pady=3, sticky="ew")

        spec_entry = ctk.CTkEntry(self.mat_insp_test_frame, placeholder_text="기준 규격")
        spec_entry.insert(0, spec)
        spec_entry.grid(row=row_idx, column=2, padx=4, pady=3, sticky="ew")

        result_entry = ctk.CTkEntry(self.mat_insp_test_frame, placeholder_text="시험 결과")
        result_entry.insert(0, result)
        result_entry.grid(row=row_idx, column=3, padx=4, pady=3, sticky="ew")

        judge_cb = ctk.CTkComboBox(self.mat_insp_test_frame, values=["적합", "부적합", "해당없음"], width=80)
        judge_cb.set(judge)
        judge_cb.grid(row=row_idx, column=4, padx=4, pady=3, sticky="ew")

        del_btn = ctk.CTkButton(
            self.mat_insp_test_frame, text="❌", width=36, height=28,
            fg_color="#EF4444", hover_color="#DC2626",
            command=lambda r_idx=len(self.mat_insp_item_rows): self.delete_mat_insp_item_row(r_idx)
        )
        del_btn.grid(row=row_idx, column=5, padx=4, pady=3)

        row_data = {
            "name": name_entry,
            "spec": spec_entry,
            "result": result_entry,
            "judge": judge_cb,
            "widgets": {
                "no": no_lbl,
                "name": name_entry,
                "spec": spec_entry,
                "result": result_entry,
                "judge": judge_cb,
                "del": del_btn
            }
        }
        self.mat_insp_item_rows.append(row_data)

    def delete_mat_insp_item_row(self, index):
        """특정 인덱스의 시험 항목 행을 삭제하고 재배치합니다."""
        if 0 <= index < len(self.mat_insp_item_rows):
            # 삭제할 행의 위젯 파괴
            for w in self.mat_insp_item_rows[index]['widgets'].values():
                w.destroy()
            self.mat_insp_item_rows.pop(index)
            
            # 남은 행들의 위치 및 번호 재정렬
            for new_idx, row in enumerate(self.mat_insp_item_rows):
                r_num = new_idx + 2
                row['widgets']['no'].configure(text=str(new_idx + 1))
                row['widgets']['no'].grid(row=r_num, column=0, padx=4, pady=3)
                row['widgets']['name'].grid(row=r_num, column=1, padx=4, pady=3, sticky="ew")
                row['widgets']['spec'].grid(row=r_num, column=2, padx=4, pady=3, sticky="ew")
                row['widgets']['result'].grid(row=r_num, column=3, padx=4, pady=3, sticky="ew")
                row['widgets']['judge'].grid(row=r_num, column=4, padx=4, pady=3, sticky="ew")
                row['widgets']['del'].configure(command=lambda i=new_idx: self.delete_mat_insp_item_row(i))
                row['widgets']['del'].grid(row=r_num, column=5, padx=4, pady=3)

    def refresh_mat_insp_list(self):
        try:
            session = db_manager.get_session()
            recs = session.query(MaterialInspectionReport).order_by(MaterialInspectionReport.created_at.desc()).limit(50).all()
            vals = [f"{r.id} | {r.material_name} ({r.material_code}) | LOT:{r.lot_no or ''} | {r.overall_result}" for r in recs]
            self.mat_insp_picker.configure(values=vals if vals else ["-- 저장된 검사성적서 없음 --"])
            if vals: self.mat_insp_picker.set(vals[0])
            session.close()
        except Exception as e:
            print(f"[경고] 원료 입고검사 목록 로드 실패: {e}")

    def load_selected_mat_insp(self):
        sel = self.mat_insp_picker.get()
        if not sel or '|' not in sel: return
        rec_id = int(sel.split('|')[0].strip())
        session = db_manager.get_session()
        try:
            r = session.query(MaterialInspectionReport).get(rec_id)
            if not r: return
            self.current_mat_insp_id = r.id
            self.mat_insp_entries["원료코드"].delete(0, "end"); self.mat_insp_entries["원료코드"].insert(0, r.material_code or "")
            self.mat_insp_entries["원료명"].delete(0, "end"); self.mat_insp_entries["원료명"].insert(0, r.material_name or "")
            self.mat_insp_entries["공급업체"].delete(0, "end"); self.mat_insp_entries["공급업체"].insert(0, r.supplier_name or "")
            self.mat_insp_entries["입고일자"].delete(0, "end"); self.mat_insp_entries["입고일자"].insert(0, str(r.incoming_date or ""))
            self.mat_insp_entries["입고LOT"].delete(0, "end"); self.mat_insp_entries["입고LOT"].insert(0, r.lot_no or "")
            self.mat_insp_entries["입고수량(kg)"].delete(0, "end"); self.mat_insp_entries["입고수량(kg)"].insert(0, str(r.incoming_amount_kg or ""))
            self.mat_insp_entries["시험자"].delete(0, "end"); self.mat_insp_entries["시험자"].insert(0, r.examiner or "")
            self.mat_insp_entries["종합판정"].set(r.overall_result or "적합 (Pass)")
            self.mat_insp_entries["포장상태"].set(r.packaging_status or "정상 (밀봉양호)")
            
            # 시험 항목들 로드
            self._clear_mat_insp_item_rows()
            if r.test_items_json:
                import json
                try:
                    items_list = json.loads(r.test_items_json)
                    for it in items_list:
                        self.add_mat_insp_item_row(it.get('name', ''), it.get('spec', ''), it.get('result', ''), it.get('judge', '적합'))
                except Exception:
                    self._init_default_mat_insp_items()
            else:
                # 레거시 단일 필드 호환
                self.add_mat_insp_item_row("성상/외관", r.appearance_spec or "고유의 성상을 띰", r.appearance_result or "적합")
                self.add_mat_insp_item_row("색상", r.color_spec or "고유의 색상", r.color_result or "적합")
                self.add_mat_insp_item_row("향취", r.odor_spec or "고유의 취", r.odor_result or "적합")
                self.add_mat_insp_item_row("굴절률 (20℃)", "1.300 ~ 1.500", r.refractive_index or "1.334")
                self.add_mat_insp_item_row("비중 (20℃)", "0.900 ~ 1.100", r.specific_gravity or "1.002")
                self.add_mat_insp_item_row("pH (10% soln)", "4.50 ~ 7.50", r.ph_val or "6.20")

            self.mat_insp_note.delete("1.0", "end")
            self.mat_insp_note.insert("1.0", r.notes or "")
            messagebox.showinfo("불러오기 완료", f"원료 입고검사성적서 '{r.material_name}' 데이터를 불러왔습니다.", parent=self)
        finally:
            session.close()

    def clear_mat_insp_form(self):
        self.current_mat_insp_id = None
        for k, e in self.mat_insp_entries.items():
            if isinstance(e, ctk.CTkComboBox): e.set(e.cget("values")[0])
            else:
                e.delete(0, "end")
                if k == "입고일자": e.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self._init_default_mat_insp_items()
        self.mat_insp_note.delete("1.0", "end")

    def save_mat_insp_to_db(self):
        mat_name = self.mat_insp_entries["원료명"].get().strip()
        mat_code = self.mat_insp_entries["원료코드"].get().strip()
        if not mat_name or not mat_code:
            messagebox.showwarning("입력 필요", "원료코드와 원료명은 필수입니다.", parent=self); return

        import json
        items_payload = []
        for row in self.mat_insp_item_rows:
            i_name = row['name'].get().strip()
            if i_name:
                items_payload.append({
                    "name": i_name,
                    "spec": row['spec'].get().strip(),
                    "result": row['result'].get().strip(),
                    "judge": row['judge'].get().strip()
                })

        session = db_manager.get_session()
        try:
            r = session.query(MaterialInspectionReport).get(self.current_mat_insp_id) if self.current_mat_insp_id else MaterialInspectionReport()
            r.material_code = mat_code
            r.material_name = mat_name
            r.supplier_name = self.mat_insp_entries["공급업체"].get().strip()
            
            try: r.incoming_date = datetime.strptime(self.mat_insp_entries["입고일자"].get().strip(), "%Y-%m-%d").date()
            except: r.incoming_date = None
            
            r.lot_no = self.mat_insp_entries["입고LOT"].get().strip()
            try: r.incoming_amount_kg = float(self.mat_insp_entries["입고수량(kg)"].get().strip())
            except: r.incoming_amount_kg = 0.0
            
            r.examiner = self.mat_insp_entries["시험자"].get().strip()
            r.overall_result = self.mat_insp_entries["종합판정"].get().strip()
            r.packaging_status = self.mat_insp_entries["포장상태"].get().strip()
            
            # 동적 항목 전체 JSON 저장
            r.test_items_json = json.dumps(items_payload, ensure_ascii=False)
            r.notes = self.mat_insp_note.get("1.0", "end-1c").strip()

            if not self.current_mat_insp_id:
                session.add(r)
            session.commit()
            self.current_mat_insp_id = r.id
            self.refresh_mat_insp_list()
            messagebox.showinfo("저장 완료", "원료 입고검사성적서가 DB에 성공적으로 저장되었습니다.", parent=self)
        except Exception as e:
            session.rollback()
            messagebox.showerror("오류", f"저장 실패: {e}", parent=self)
        finally:
            session.close()

    def delete_mat_insp(self):
        if not self.current_mat_insp_id:
            messagebox.showwarning("선택 오류", "삭제할 항목을 먼저 불러오세요.", parent=self); return
        if not messagebox.askyesno("삭제 확인", "선택한 원료 입고검사성적서를 삭제하시겠습니까?", parent=self): return
        session = db_manager.get_session()
        try:
            r = session.query(MaterialInspectionReport).get(self.current_mat_insp_id)
            if r:
                session.delete(r)
                session.commit()
                self.clear_mat_insp_form()
                self.refresh_mat_insp_list()
                messagebox.showinfo("완료", "삭제되었습니다.", parent=self)
        finally:
            session.close()

    def export_mat_insp_to_excel(self):
        mat_name = self.mat_insp_entries["원료명"].get().strip() or "원료"
        wb = Workbook()
        ws = wb.active
        ws.title = "원료입고검사성적서"
        
        ws.merge_cells("A1:E2")
        ws["A1"] = f"원료 입고검사성적서 ({mat_name})"
        ws["A1"].font = Font(name="맑은 고딕", size=16, bold=True, color="1F497D")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.append([])
        ws.append(["원료코드", self.mat_insp_entries["원료코드"].get(), "", "입고일자", self.mat_insp_entries["입고일자"].get()])
        ws.append(["원료명", mat_name, "", "입고LOT", self.mat_insp_entries["입고LOT"].get()])
        ws.append(["공급업체", self.mat_insp_entries["공급업체"].get(), "", "입고수량(kg)", self.mat_insp_entries["입고수량(kg)"].get()])
        ws.append(["포장상태", self.mat_insp_entries["포장상태"].get(), "", "종합판정", self.mat_insp_entries["종합판정"].get()])
        ws.append([])
        ws.append(["NO", "시험 항목", "시험 기준", "시험 결과", "판정"])
        
        for idx, row in enumerate(self.mat_insp_item_rows, 1):
            ws.append([idx, row['name'].get(), row['spec'].get(), row['result'].get(), row['judge'].get()])

        file_path = fd.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], initialfile=f"원료입고검사_{mat_name}.xlsx", title="성적서 엑셀 저장")
        if file_path:
            wb.save(file_path)
            messagebox.showinfo("완료", f"성적서가 저장되었습니다:\n{file_path}", parent=self)
            try: os.startfile(os.path.abspath(file_path))
            except: pass
    # =========================================================================
    # 2. 완제품/반제품 제품표준서 (Product Specification Standard) - CGMP 5대 심층 파트
    # =========================================================================
    def setup_product_standard_tab(self, tab_frame):
        """식약처 / CGMP 표준 [제품표준서] 공식 표지(Cover) + 12대 목차 초경량 데이터그리드(Treeview) 슬라이더 시스템"""
        tab_frame.grid_columnconfigure(0, weight=1)
        tab_frame.grid_rowconfigure(0, weight=1)

        scroll = ctk.CTkScrollableFrame(tab_frame, label_text="제품표준서 (Product Standard - 공식 표지 및 12대 전 목차 관리)")
        scroll.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        scroll.grid_columnconfigure(0, weight=1)

        self.prod_std_entries = {}
        self.current_prod_std_id = None
        self.current_prod_std_section_idx = 0

        # 상단 문서 이력 툴바
        top_bar = ctk.CTkFrame(scroll, fg_color="transparent")
        top_bar.grid(row=0, column=0, sticky="ew", padx=10, pady=(5, 10))
        top_bar.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(top_bar, text="표준서 이력:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=5, pady=5)
        self.prod_std_picker = ctk.CTkComboBox(top_bar, values=["-- 저장된 제품표준서 선택 --"], width=300)
        self.prod_std_picker.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ctk.CTkButton(top_bar, text="불러오기", width=80, command=self.load_selected_prod_std).grid(row=0, column=2, padx=5, pady=5)
        ctk.CTkButton(top_bar, text="새로작성", width=80, fg_color="gray50", command=self.clear_prod_std_form).grid(row=0, column=3, padx=5, pady=5)
        ctk.CTkButton(top_bar, text="삭제", width=70, fg_color="#D32F2F", command=self.delete_prod_std).grid(row=0, column=4, padx=5, pady=5)

        self.prod_std_confidential_var = ctk.BooleanVar(value=False)
        self.chk_prod_std_confidential = ctk.CTkCheckBox(
            top_bar, text="🔒 기밀사항 (외부제출용 함량 비공개)",
            variable=self.prod_std_confidential_var,
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=("#C62828", "#EF5350")
        )
        self.chk_prod_std_confidential.grid(row=0, column=5, padx=(12, 5), pady=5)

        # 13개 화면 슬라이드 네비게이터 바 (0. 표지 및 목차 + 1~12번 목차)
        nav_card = ctk.CTkFrame(scroll, fg_color=("gray90", "gray17"), corner_radius=8, border_width=1, border_color=("gray75", "gray30"))
        nav_card.grid(row=1, column=0, sticky="ew", padx=10, pady=5)
        nav_card.grid_columnconfigure(1, weight=1)

        self.prod_std_section_titles = [
            "0. 표지 및 목차 (Cover & Index)",
            "1. 변경이력 (목차 1/12)",
            "2. 제품참고사항 (목차 2/12)",
            "3. 공정규격 (목차 3/12)",
            "4. 원료성분 기준 및 시험방법 (목차 4/12)",
            "5. 계량 지시 및 기록서 (목차 5/12)",
            "6. 제조 지시 및 기록서 (목차 6/12)",
            "7. 제품사양 (포장재규격포함) (목차 7/12)",
            "8. 충진·포장 지시 및 기록서 (목차 8/12)",
            "9. 제품 규격서 (목차 9/12)",
            "10. 반제품 시험성적서 (목차 10/12)",
            "11. 완제품 시험성적서 (목차 11/12)",
            "12. 제조 및 품질관리 시설·기구 (목차 12/12)"
        ]

        self.btn_prev_ps = ctk.CTkButton(
            nav_card, text="◀ 이전", width=100, height=34,
            fg_color="#455A64", hover_color="#37474F",
            font=ctk.CTkFont(weight="bold", size=12),
            command=self.prev_prod_std_section
        )
        self.btn_prev_ps.grid(row=0, column=0, padx=8, pady=6)

        self.prod_std_section_selector = ctk.CTkComboBox(
            nav_card, values=self.prod_std_section_titles, height=34,
            font=ctk.CTkFont(weight="bold", size=13), dropdown_font=ctk.CTkFont(size=12),
            command=self.on_prod_std_section_selected
        )
        self.prod_std_section_selector.set(self.prod_std_section_titles[0])
        self.prod_std_section_selector.grid(row=0, column=1, sticky="ew", padx=8, pady=6)

        self.btn_next_ps = ctk.CTkButton(
            nav_card, text="다음 ▶", width=100, height=34,
            fg_color="#1565C0", hover_color="#0D47A1",
            font=ctk.CTkFont(weight="bold", size=12),
            command=self.next_prod_std_section
        )
        self.btn_next_ps.grid(row=0, column=2, padx=8, pady=6)

        # 본문 컨테이너 프레임
        self.prod_std_content_container = ctk.CTkFrame(scroll, fg_color="transparent")
        self.prod_std_content_container.grid(row=2, column=0, sticky="nsew", padx=10, pady=5)
        self.prod_std_content_container.grid_columnconfigure(0, weight=1)
        self.prod_std_content_container.grid_rowconfigure(0, weight=1)

        # 13개 온디맨드 렌더링 슬롯
        self.prod_std_section_frames = [None] * 13
        self.prod_std_section_builders = [
            self._build_ps_sec_cover, self._build_ps_sec_0, self._build_ps_sec_1,
            self._build_ps_sec_2, self._build_ps_sec_3, self._build_ps_sec_4,
            self._build_ps_sec_5, self._build_ps_sec_6, self._build_ps_sec_7,
            self._build_ps_sec_8, self._build_ps_sec_9, self._build_ps_sec_10,
            self._build_ps_sec_11
        ]

        # 첫 번째 화면(표지 및 목차)만 즉시 빌드 (0.05초 진입)
        self.show_prod_std_section(0)

        # 하단 저장 및 엑셀 출력 버튼 바
        btn_bar = ctk.CTkFrame(scroll, fg_color="transparent")
        btn_bar.grid(row=3, column=0, sticky="e", padx=10, pady=15)
        ctk.CTkButton(btn_bar, text="💾 제품표준서 DB 저장", width=140, command=self.save_prod_std_to_db).pack(side="left", padx=5)
        ctk.CTkButton(btn_bar, text="📊 국문 제품표준서 엑셀 (표지+12대 목차)", width=220, fg_color="#2E7D32", hover_color="#1B5E20", command=lambda: self.export_prod_std_to_excel("ko")).pack(side="left", padx=5)
        ctk.CTkButton(btn_bar, text="🌐 영문 제품표준서 엑셀 (Cover+12 Sections)", width=230, fg_color="#1565C0", hover_color="#0D47A1", command=lambda: self.export_prod_std_to_excel("en")).pack(side="left", padx=5)

        self.refresh_prod_std_list()

    def _create_ps_base_frame(self, sec_title_header):
        sec_frame = ctk.CTkScrollableFrame(
            self.prod_std_content_container, height=480, label_text=sec_title_header,
            label_font=ctk.CTkFont(weight="bold", size=13)
        )
        sec_frame.grid_columnconfigure(1, weight=1)
        sec_frame.grid_columnconfigure(3, weight=1)
        return sec_frame

    def _append_multi_ps(self, selected_val, target_entry):
        if not selected_val or selected_val.startswith("➕") or selected_val == "직접 입력": return
        cur = target_entry.get().strip()
        if not cur: target_entry.insert(0, selected_val)
        else:
            if selected_val in cur: return
            sep = " / " if ("/" in cur or "/" in selected_val or len(cur) > 25) else ", "
            target_entry.delete(0, "end"); target_entry.insert(0, f"{cur}{sep}{selected_val}")

    def _remove_multi_ps(self, target_entry):
        cur = target_entry.get().strip()
        if not cur: return
        for sep in [" / ", ", ", "/"]:
            if sep in cur:
                items = [it.strip() for it in cur.split(sep) if it.strip()]
                if len(items) > 1:
                    items.pop()
                    target_entry.delete(0, "end")
                    target_entry.insert(0, " / ".join(items) if " / " in cur or len(cur) > 30 else ", ".join(items))
                    return
        target_entry.delete(0, "end")

    def _setup_ps_fields(self, parent_frame, fields, combos_dict):
        for item in fields:
            if len(item) == 3: lbl, r, c = item; col_span = 1
            elif len(item) == 4: lbl, r, c, col_span = item
            else: continue

            ctk.CTkLabel(parent_frame, text=lbl, font=ctk.CTkFont(size=11, weight="bold")).grid(
                row=r, column=c, padx=(10, 6), pady=6, sticky="w"
            )
            if lbl in combos_dict:
                box_f = ctk.CTkFrame(parent_frame, fg_color="transparent")
                box_f.grid(row=r, column=c+1, columnspan=col_span, padx=6, pady=6, sticky="ew")
                box_f.grid_columnconfigure(0, weight=1)

                ent = ctk.CTkEntry(box_f, placeholder_text=f"{lbl} 입력 (우측 ➕에서 선택 추가)", height=30)
                ent.grid(row=0, column=0, sticky="ew", padx=(0, 4))
                self.prod_std_entries[lbl] = ent

                cb_vals = ["➕ 항목 추가 선택..."] + combos_dict[lbl]
                cb = ctk.CTkComboBox(
                    box_f, values=cb_vals, width=140, height=30,
                    command=lambda choice, target=ent: (
                        self._append_multi_ps(choice, target),
                        cb.set("➕ 항목 추가 선택...")
                    )
                )
                cb.set("➕ 항목 추가 선택...")
                cb.grid(row=0, column=1, padx=(0, 4))

                btn_del = ctk.CTkButton(
                    box_f, text="선택 제거", width=65, height=30,
                    fg_color="#C62828", hover_color="#B71C1C",
                    font=ctk.CTkFont(size=11, weight="bold"),
                    command=lambda target=ent: self._remove_multi_ps(target)
                )
                btn_del.grid(row=0, column=2)
            else:
                ent = ctk.CTkEntry(parent_frame, placeholder_text=f"{lbl} 입력", height=30)
                ent.grid(row=r, column=c+1, columnspan=col_span, padx=6, pady=6, sticky="ew")
                self.prod_std_entries[lbl] = ent

    # 헬퍼 함수: 초경량 Treeview 데이터그리드 생성
    def _create_treeview_grid(self, parent_frame, cols_dict, height=10):
        t_frame = ctk.CTkFrame(parent_frame, fg_color="transparent")
        t_frame.grid_columnconfigure(0, weight=1)
        t_frame.grid_rowconfigure(0, weight=1)

        tree = ttk.Treeview(t_frame, columns=list(cols_dict.keys()), show="headings", height=height, selectmode="extended")
        vsb = ttk.Scrollbar(t_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(t_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        for col_id, (col_text, col_w) in cols_dict.items():
            tree.heading(col_id, text=col_text)
            tree.column(col_id, width=col_w, anchor="center")

        return t_frame, tree

    # =========================================================================
    # [화면 0] 공식 표지 및 목차 (Cover Page & Index - PDF Page 1 기준)
    # =========================================================================
    def _build_ps_sec_cover(self):
        f = self._create_ps_base_frame("■ [공식 표지] 제품표준서 표지 및 결재란 (Cover & Approvals)")
        
        # 1. 상단 결재란 카드
        appr_card = ctk.CTkFrame(f, fg_color=("gray95", "gray18"), corner_radius=8, border_width=1, border_color=("gray80", "gray28"))
        appr_card.grid(row=0, column=0, columnspan=4, sticky="ew", padx=6, pady=6)
        appr_card.grid_columnconfigure((1, 3, 5), weight=1)

        ctk.CTkLabel(appr_card, text="[결재란 (Approvals)]", font=ctk.CTkFont(weight="bold", size=12), text_color=("#1F497D", "#38BDF8")).grid(row=0, column=0, columnspan=6, padx=8, pady=4, sticky="w")
        
        fields_appr = [
            ("작성자", 1, 0), ("작성일자", 1, 2), ("양식/관리번호", 1, 4),
            ("검토자", 2, 0), ("검토일자", 2, 2), ("개정번호", 2, 4),
            ("승인자", 3, 0), ("승인일자", 3, 2), ("연구소/회사명", 3, 4)
        ]
        for lbl, r, c in fields_appr:
            ctk.CTkLabel(appr_card, text=lbl, font=ctk.CTkFont(size=11, weight="bold")).grid(row=r, column=c, padx=6, pady=4, sticky="w")
            ent = ctk.CTkEntry(appr_card, placeholder_text=f"{lbl} 입력", height=28)
            ent.grid(row=r, column=c+1, padx=4, pady=4, sticky="ew")
            self.prod_std_entries[lbl] = ent

        # 설정에서 자회사 정보 가져와 자동 기본 채움
        cp = get_company_profile()
        self.prod_std_entries["양식/관리번호"].insert(0, cp.get("form_doc_no", "양0100-01"))
        self.prod_std_entries["개정번호"].insert(0, cp.get("form_rev_no", "Rev.0"))
        self.prod_std_entries["연구소/회사명"].insert(0, cp.get("company_name_ko", "(주)한국피부과학연구소"))
        if cp.get("manager_name"):
            self.prod_std_entries["작성자"].insert(0, cp.get("manager_name"))

        # 2. 제품 기본 식별 정보
        prod_card = ctk.CTkFrame(f, fg_color=("gray95", "gray18"), corner_radius=8, border_width=1, border_color=("gray80", "gray28"))
        prod_card.grid(row=1, column=0, columnspan=4, sticky="ew", padx=6, pady=6)
        prod_card.grid_columnconfigure((1, 3), weight=1)

        ctk.CTkLabel(prod_card, text="[제품 식별 정보 (Product Identification)]", font=ctk.CTkFont(weight="bold", size=12), text_color=("#1F497D", "#38BDF8")).grid(row=0, column=0, columnspan=4, padx=8, pady=4, sticky="w")
        p_fields = [
            ("제품명(국문)", 1, 0), ("제품명(영문)", 1, 2),
            ("제품표준서번호", 2, 0), ("제품 코드", 2, 2)
        ]
        for lbl, r, c in p_fields:
            ctk.CTkLabel(prod_card, text=lbl, font=ctk.CTkFont(size=11, weight="bold")).grid(row=r, column=c, padx=6, pady=4, sticky="w")
            ent = ctk.CTkEntry(prod_card, placeholder_text=f"{lbl} 입력", height=28)
            ent.grid(row=r, column=c+1, padx=4, pady=4, sticky="ew")
            self.prod_std_entries[lbl] = ent

        # 3. 12대 목차 인덱스 데이터그리드
        ctk.CTkLabel(f, text="📋 제품표준서 12대 목차 목록 (더블클릭 시 해당 목차로 즉시 이동):", font=ctk.CTkFont(weight="bold", size=12)).grid(row=2, column=0, columnspan=4, padx=8, pady=(8, 2), sticky="w")
        idx_cols = {
            "no": ("NO.", 60),
            "title_ko": ("목 차 명 (국문)", 260),
            "title_en": ("TABLE OF CONTENTS (EN)", 260),
            "remark": ("비고", 120)
        }
        f_tree, self.prod_std_index_tree = self._create_treeview_grid(f, idx_cols, height=8)
        f_tree.grid(row=3, column=0, columnspan=4, padx=6, pady=4, sticky="nsew")

        index_data = [
            ("1", "변경이력", "Revision History", "개정/변경 관리"),
            ("2", "제품참고사항", "General Information", "유형/성상/주의사항"),
            ("3", "공정규격", "Manufacturing Process Specifications", "공정조건/설비"),
            ("4", "원료성분 기준 및 시험방법 (100g당)", "Raw Material Specifications & Formulations", "Phase별 배합비"),
            ("5", "계량 지시 및 기록서", "Weighing Order & Record", "이론량/계량량/LOT"),
            ("6", "제조 지시 및 기록서", "Batch Manufacturing Record (BMR)", "제조 SOP/수율"),
            ("7", "제품사양 (포장재규격포함)", "Packaging Specifications", "용기/펌프/박스"),
            ("8", "충진·포장 지시 및 기록서", "Packaging Order & Record", "점검/수불/수율"),
            ("9", "제품 규격서", "Product Specifications", "14대 완제품 시험규격"),
            ("10", "반제품 시험성적서", "Semi-Finished Product COA", "반제품 성적서"),
            ("11", "완제품 시험성적서", "Finished Product COA", "완제품 성적서"),
            ("12", "제조 및 품질관리에 필요한 시설 및 기구", "Facilities and Equipment", "제조/QC 설비목록")
        ]
        for row_item in index_data:
            self.prod_std_index_tree.insert("", "end", values=row_item)

        # 목차 더블클릭 점프 이벤트
        def on_index_double_click(event):
            sel = self.prod_std_index_tree.selection()
            if sel:
                item_vals = self.prod_std_index_tree.item(sel[0], "values")
                if item_vals:
                    no_val = int(item_vals[0])
                    self.show_prod_std_section(no_val)

        self.prod_std_index_tree.bind("<Double-1>", on_index_double_click)
        return f

    # =========================================================================
    # [화면 1] 1. 변경이력 (Treeview 데이터그리드)
    # =========================================================================
    def _build_ps_sec_0(self):
        f = self._create_ps_base_frame("■ [목차 1] 변경이력 (Revision History)")
        
        # 입력 툴바
        in_bar = ctk.CTkFrame(f, fg_color="transparent")
        in_bar.grid(row=0, column=0, columnspan=4, sticky="ew", padx=6, pady=4)
        in_bar.grid_columnconfigure((1, 2), weight=1)

        e_date = ctk.CTkEntry(in_bar, placeholder_text="날짜 (YYYY-MM-DD)", width=120)
        e_date.grid(row=0, column=0, padx=2)
        e_content = ctk.CTkEntry(in_bar, placeholder_text="개정 및 변경 내용(사유)")
        e_content.grid(row=0, column=1, sticky="ew", padx=2)
        e_author = ctk.CTkEntry(in_bar, placeholder_text="작성자", width=90)
        e_author.grid(row=0, column=2, padx=2)
        e_appr = ctk.CTkEntry(in_bar, placeholder_text="승인자", width=90)
        e_appr.grid(row=0, column=3, padx=2)
        e_rem = ctk.CTkEntry(in_bar, placeholder_text="비고", width=90)
        e_rem.grid(row=0, column=4, padx=2)

        def add_rev():
            cnt = len(self.prod_std_rev_tree.get_children()) + 1
            self.prod_std_rev_tree.insert("", "end", values=(
                str(cnt), e_date.get().strip(), e_content.get().strip(),
                e_author.get().strip(), e_appr.get().strip(), e_rem.get().strip()
            ))
            e_content.delete(0, "end"); e_rem.delete(0, "end")

        def del_rev():
            for sel in self.prod_std_rev_tree.selection(): self.prod_std_rev_tree.delete(sel)

        ctk.CTkButton(in_bar, text="➕ 추가", width=65, fg_color="#0284C7", hover_color="#0369A1", command=add_rev).grid(row=0, column=5, padx=2)
        ctk.CTkButton(in_bar, text="선택 삭제", width=75, fg_color="#C62828", hover_color="#B71C1C", command=del_rev).grid(row=0, column=6, padx=2)

        cols = {"no": ("연번", 50), "date": ("날짜", 110), "content": ("내용 (변경사유)", 300), "author": ("작성자", 90), "appr": ("승인자", 90), "rem": ("비고", 90)}
        f_tree, self.prod_std_rev_tree = self._create_treeview_grid(f, cols, height=12)
        f_tree.grid(row=1, column=0, columnspan=4, padx=6, pady=4, sticky="nsew")
        return f

    # =========================================================================
    # [화면 2] 2. 제품참고사항 (General Information)
    # =========================================================================
    def _build_ps_sec_1(self):
        f = self._create_ps_base_frame("■ [목차 2] 제품참고사항 (General Information)")
        combos = {
            "유형 / 세부 유형": ["면도용 제품류 / 그 밖의 면도용 제품류", "기초화장용 제품류 / 크림·로션·에센스", "인체 세정용 제품류 / 클렌징 폼·바디워시", "두발용 제품류 / 샴푸·린스", "영유아용 제품류", "색조 화장용 제품류", "직접 입력"],
            "성 상": ["무색 투명 겔상", "유백색 점조성 크림 제형", "투명 점조성 액상", "미황색 에멀젼", "고체 스틱", "직접 입력"],
            "허가(보고) 취득일": ["해당사항없음 (일반화장품)", f"{datetime.now().strftime('%Y년 %m월 %d일')}", "직접 입력"],
            "허가(보고) 번호": ["해당사항없음", "제2026-보고-0001호", "직접 입력"],
            "사용기한 / 개봉 후 사용기간": ["제조일로부터 36개월 / 개봉후 12개월", "제조일로부터 24개월 / 개봉후 12개월", "제조일로부터 30개월 / 개봉후 6개월", "직접 입력"],
            "보 관 방 법": ["통풍이 잘되는 차광된 장소에서 상온(1~30℃)에서 보관", "직사광선을 피하고 서늘한 곳에 밀폐 보관 (1~30℃)", "직접 입력"],
            "효능효과": ["해당사항없음 (일반화장품)", "피부의 주름개선에 도움을 준다.", "피부의 미백에 도움을 준다.", "미백 및 주름개선 2중 기능성", "직접 입력"],
            "용법 용량": ["본품 적당량을 취해 피부에 골고루 펴 바릅니다.", "따뜻한 물로 적신 후 적당량을 펴 발라 사용합니다.", "직접 입력"],
            "사용할 때의 주의사항": ["1) 화장품 사용 시 또는 직사광선에 의한 이상 증상 시 전문의 상담 / 2) 상처 부위 자제 / 3) 어린이 손 닿지 않는 곳, 직사광선 피할 것", "직접 입력"]
        }
        fields = [
            ("작성 일자", 0, 0), ("작 성 자", 0, 2),
            ("제품용량 (ml/g)", 1, 0), ("성 상", 1, 2),
            ("허가(보고) 취득일", 2, 0), ("허가(보고) 번호", 2, 2),
            ("유형 / 세부 유형", 3, 0), ("사용기한 / 개봉 후 사용기간", 3, 2),
            ("보 관 방 법", 4, 0, 3),
            ("효능효과", 5, 0), ("용법 용량", 5, 2),
            ("화장품 성분 표기사항 (전성분)", 6, 0, 3),
            ("사용할 때의 주의사항", 7, 0, 3),
            ("기타 사항", 8, 0, 3)
        ]
        self._setup_ps_fields(f, fields, combos)
        return f

    # =========================================================================
    # [화면 3] 3. 공정규격 (Treeview 데이터그리드)
    # =========================================================================
    def _build_ps_sec_2(self):
        f = self._create_ps_base_frame("■ [목차 3] 공정규격 (Manufacturing Process Specifications)")
        
        in_bar = ctk.CTkFrame(f, fg_color="transparent")
        in_bar.grid(row=0, column=0, columnspan=4, sticky="ew", padx=6, pady=4)
        in_bar.grid_columnconfigure((2, 4), weight=1)

        e_room = ctk.CTkEntry(in_bar, placeholder_text="작업실(생산/품질)", width=110)
        e_room.grid(row=0, column=0, padx=2)
        e_name = ctk.CTkEntry(in_bar, placeholder_text="공정명", width=120)
        e_name.grid(row=0, column=1, padx=2)
        e_equip = ctk.CTkEntry(in_bar, placeholder_text="사용설비")
        e_equip.grid(row=0, column=2, sticky="ew", padx=2)
        e_cond = ctk.CTkEntry(in_bar, placeholder_text="제조(작업) 조건 (온도/RPM/시간)")
        e_cond.grid(row=0, column=3, sticky="ew", padx=2)
        e_insp = ctk.CTkEntry(in_bar, placeholder_text="공정검사 기준", width=120)
        e_insp.grid(row=0, column=4, padx=2)

        def add_proc():
            cnt = len(self.prod_std_proc_tree.get_children()) + 1
            self.prod_std_proc_tree.insert("", "end", values=(
                str(cnt), e_room.get().strip(), e_name.get().strip(),
                e_equip.get().strip(), e_cond.get().strip(), e_insp.get().strip()
            ))
            e_name.delete(0, "end"); e_cond.delete(0, "end")

        def del_proc():
            for sel in self.prod_std_proc_tree.selection(): self.prod_std_proc_tree.delete(sel)

        ctk.CTkButton(in_bar, text="➕ 추가", width=65, fg_color="#0284C7", hover_color="#0369A1", command=add_proc).grid(row=0, column=5, padx=2)
        ctk.CTkButton(in_bar, text="선택 삭제", width=75, fg_color="#C62828", hover_color="#B71C1C", command=del_proc).grid(row=0, column=6, padx=2)

        cols = {"no": ("No", 45), "room": ("작업실", 100), "name": ("공정명", 110), "equip": ("사용설비", 140), "cond": ("제조(작업) 조건", 280), "insp": ("공정검사", 120)}
        f_tree, self.prod_std_proc_tree = self._create_treeview_grid(f, cols, height=12)
        f_tree.grid(row=1, column=0, columnspan=4, padx=6, pady=4, sticky="nsew")
        return f

    # =========================================================================
    # [화면 4] 4. 원료성분 기준 및 함량 (Treeview 데이터그리드)
    # =========================================================================
    def _build_ps_sec_3(self):
        f = self._create_ps_base_frame("■ [목차 4] 원료 성분 기준 및 함량 (100g당)")
        
        in_bar = ctk.CTkFrame(f, fg_color="transparent")
        in_bar.grid(row=0, column=0, columnspan=4, sticky="ew", padx=6, pady=4)
        in_bar.grid_columnconfigure((2, 3), weight=1)

        e_phase = ctk.CTkEntry(in_bar, placeholder_text="Phase A", width=80)
        e_phase.grid(row=0, column=0, padx=2)
        e_code = ctk.CTkEntry(in_bar, placeholder_text="코드", width=90)
        e_code.grid(row=0, column=1, padx=2)
        e_name = ctk.CTkEntry(in_bar, placeholder_text="원료명")
        e_name.grid(row=0, column=2, sticky="ew", padx=2)
        e_inci = ctk.CTkEntry(in_bar, placeholder_text="허가명/INCI")
        e_inci.grid(row=0, column=3, sticky="ew", padx=2)
        e_spec = ctk.CTkEntry(in_bar, placeholder_text="시험기준", width=90); e_spec.insert(0, "자사규격")
        e_spec.grid(row=0, column=4, padx=2)
        e_ratio = ctk.CTkEntry(in_bar, placeholder_text="함량(%)", width=80)
        e_ratio.grid(row=0, column=5, padx=2)
        e_rem = ctk.CTkEntry(in_bar, placeholder_text="비고", width=80)
        e_rem.grid(row=0, column=6, padx=2)

        def add_ing():
            cnt = len(self.prod_std_ing_tree.get_children()) + 1
            self.prod_std_ing_tree.insert("", "end", values=(
                e_phase.get().strip() or "Phase A", str(cnt), e_code.get().strip() or "-",
                e_name.get().strip(), e_inci.get().strip(), e_spec.get().strip(),
                e_ratio.get().strip(), e_rem.get().strip()
            ))
            e_name.delete(0, "end"); e_inci.delete(0, "end"); e_ratio.delete(0, "end")

        def del_ing():
            for sel in self.prod_std_ing_tree.selection(): self.prod_std_ing_tree.delete(sel)

        ctk.CTkButton(in_bar, text="➕ 추가", width=65, fg_color="#0284C7", hover_color="#0369A1", command=add_ing).grid(row=0, column=7, padx=2)
        ctk.CTkButton(in_bar, text="선택 삭제", width=75, fg_color="#C62828", hover_color="#B71C1C", command=del_ing).grid(row=0, column=8, padx=2)

        cols = {"phase": ("Phase", 70), "no": ("No", 45), "code": ("코드", 80), "name": ("원료명", 160), "inci": ("허가명/INCI", 180), "spec": ("시험기준", 85), "ratio": ("함량(%)", 75), "rem": ("비고", 80)}
        f_tree, self.prod_std_ing_tree = self._create_treeview_grid(f, cols, height=12)
        f_tree.grid(row=1, column=0, columnspan=4, padx=6, pady=4, sticky="nsew")
        return f

    # =========================================================================
    # [화면 5] 5. 계량 지시 및 기록서 (Treeview 데이터그리드)
    # =========================================================================
    def _build_ps_sec_4(self):
        f = self._create_ps_base_frame("■ [목차 5] 계량 지시 및 기록서 (Weighing Order & Record)")
        combos = {"기준 제조량(kg)": ["100.0 kg", "200.0 kg", "500.0 kg", "1,000.0 kg", "직접 입력"]}
        fields = [("기준 제조량(kg)", 0, 0), ("계량 지시일", 0, 2), ("계량자", 1, 0), ("확인자(입회자)", 1, 2)]
        self._setup_ps_fields(f, fields, combos)

        in_bar = ctk.CTkFrame(f, fg_color="transparent")
        in_bar.grid(row=2, column=0, columnspan=4, sticky="ew", padx=6, pady=(10, 4))
        in_bar.grid_columnconfigure((1, 2), weight=1)

        e_vat = ctk.CTkEntry(in_bar, placeholder_text="Vat", width=50); e_vat.grid(row=0, column=0, padx=2)
        e_code = ctk.CTkEntry(in_bar, placeholder_text="코드", width=80); e_code.grid(row=0, column=1, padx=2)
        e_name = ctk.CTkEntry(in_bar, placeholder_text="원료명"); e_name.grid(row=0, column=2, sticky="ew", padx=2)
        e_ratio = ctk.CTkEntry(in_bar, placeholder_text="분량(%)", width=70); e_ratio.grid(row=0, column=3, padx=2)
        e_th = ctk.CTkEntry(in_bar, placeholder_text="이론량(kg)", width=80); e_th.grid(row=0, column=4, padx=2)
        e_wt = ctk.CTkEntry(in_bar, placeholder_text="계량량(kg)", width=80); e_wt.grid(row=0, column=5, padx=2)
        e_lot = ctk.CTkEntry(in_bar, placeholder_text="Lot No.", width=90); e_lot.grid(row=0, column=6, padx=2)

        def add_w():
            cnt = len(self.prod_std_weigh_tree.get_children()) + 1
            self.prod_std_weigh_tree.insert("", "end", values=(
                e_vat.get().strip() or "A", str(cnt), e_code.get().strip() or "-",
                e_name.get().strip(), e_ratio.get().strip(), e_th.get().strip(),
                e_wt.get().strip(), e_lot.get().strip(), "-", "-", "-"
            ))
            e_name.delete(0, "end"); e_th.delete(0, "end"); e_wt.delete(0, "end")

        def del_w():
            for sel in self.prod_std_weigh_tree.selection(): self.prod_std_weigh_tree.delete(sel)

        ctk.CTkButton(in_bar, text="➕ 추가", width=60, fg_color="#0284C7", hover_color="#0369A1", command=add_w).grid(row=0, column=7, padx=2)
        ctk.CTkButton(in_bar, text="선택 삭제", width=70, fg_color="#C62828", hover_color="#B71C1C", command=del_w).grid(row=0, column=8, padx=2)

        cols = {"vat": ("Vat", 45), "no": ("No", 40), "code": ("코드", 75), "name": ("원료명", 160), "ratio": ("분량(%)", 70), "th": ("이론량", 75), "wt": ("계량량", 75), "lot": ("Lot No.", 90), "by": ("칭량자", 70), "chk": ("확인자", 70), "rem": ("특이사항", 80)}
        f_tree, self.prod_std_weigh_tree = self._create_treeview_grid(f, cols, height=9)
        f_tree.grid(row=3, column=0, columnspan=4, padx=6, pady=4, sticky="nsew")
        return f

    # =========================================================================
    # [화면 6] 6. 제조 지시 및 기록서 (BMR SOP)
    # =========================================================================
    def _build_ps_sec_5(self):
        f = self._create_ps_base_frame("■ [목차 6] 제조 지시 및 기록서 (Batch Manufacturing SOP & Record)")
        combos = {
            "제조설비명": ["유화가마 1호기 (200kg)", "유화가마 2호기 (500kg)", "개방형 아지믹서 가마", "직접 입력"],
            "수득량 및 수율": ["수득량: 198.5 kg (수율: 99.25%)", "수득량: 197.0 kg (수율: 98.50%)", "직접 입력"]
        }
        fields = [("제조설비명", 0, 0), ("수득량 및 수율", 0, 2), ("제조작업자", 1, 0), ("제조책임자", 1, 2)]
        self._setup_ps_fields(f, fields, combos)

        ctk.CTkLabel(f, text="[세부 제조공정 지시사항 및 유화/혼합 조건 기록 (SOP)]:", font=ctk.CTkFont(weight="bold")).grid(row=2, column=0, columnspan=4, padx=10, pady=(10, 2), sticky="w")
        self.prod_std_mfg_summary = ctk.CTkTextbox(f, height=280)
        self.prod_std_mfg_summary.grid(row=3, column=0, columnspan=4, padx=6, pady=5, sticky="nsew")
        return f

    # =========================================================================
    # [화면 7] 7. 제품사양 (Treeview 데이터그리드)
    # =========================================================================
    def _build_ps_sec_6(self):
        f = self._create_ps_base_frame("■ [목차 7] 제품사양 및 포장재 규격 (Packaging Specifications)")
        
        in_bar = ctk.CTkFrame(f, fg_color="transparent")
        in_bar.grid(row=0, column=0, columnspan=4, sticky="ew", padx=6, pady=4)
        in_bar.grid_columnconfigure((1, 4), weight=1)

        e_code = ctk.CTkEntry(in_bar, placeholder_text="자재코드", width=100); e_code.grid(row=0, column=0, padx=2)
        e_name = ctk.CTkEntry(in_bar, placeholder_text="사양명(반제품/용기/펌프/단상자/박스)"); e_name.grid(row=0, column=1, sticky="ew", padx=2)
        e_qty = ctk.CTkEntry(in_bar, placeholder_text="수량", width=80); e_qty.grid(row=0, column=2, padx=2)
        e_unit = ctk.CTkEntry(in_bar, placeholder_text="단위(EA/ml)", width=90); e_unit.grid(row=0, column=3, padx=2)
        e_rem = ctk.CTkEntry(in_bar, placeholder_text="비고(재질/규격)"); e_rem.grid(row=0, column=4, sticky="ew", padx=2)

        def add_pkg():
            cnt = len(self.prod_std_pkg_tree.get_children()) + 1
            self.prod_std_pkg_tree.insert("", "end", values=(
                str(cnt), e_code.get().strip() or "-", e_name.get().strip(),
                e_qty.get().strip(), e_unit.get().strip() or "EA", e_rem.get().strip()
            ))
            e_name.delete(0, "end"); e_qty.delete(0, "end"); e_rem.delete(0, "end")

        def del_pkg():
            for sel in self.prod_std_pkg_tree.selection(): self.prod_std_pkg_tree.delete(sel)

        ctk.CTkButton(in_bar, text="➕ 추가", width=65, fg_color="#0284C7", hover_color="#0369A1", command=add_pkg).grid(row=0, column=5, padx=2)
        ctk.CTkButton(in_bar, text="선택 삭제", width=75, fg_color="#C62828", hover_color="#B71C1C", command=del_pkg).grid(row=0, column=6, padx=2)

        cols = {"no": ("No", 45), "code": ("자재코드", 110), "name": ("사양명 (자재명)", 220), "qty": ("수량", 80), "unit": ("단위", 75), "rem": ("비고 (재질/규격)", 220)}
        f_tree, self.prod_std_pkg_tree = self._create_treeview_grid(f, cols, height=12)
        f_tree.grid(row=1, column=0, columnspan=4, padx=6, pady=4, sticky="nsew")
        return f

    # =========================================================================
    # [화면 8] 8. 충진·포장 지시 및 기록서
    # =========================================================================
    def _build_ps_sec_7(self):
        f = self._create_ps_base_frame("■ [목차 8] 충진·포장 지시 및 기록서 (Packaging SOP & Yield)")
        combos = {
            "충진·포장 판정": ["적합 (양호)", "조건부 적합", "부적합", "직접 입력"],
            "포장 수율 기준": ["충진 포장 수율 95% 이상 적합", "수율 98% 이상", "직접 입력"]
        }
        fields = [
            ("표시 용량 (ml/g)", 0, 0), ("생산 계획량 (EA)", 0, 2),
            ("충진 작업조", 1, 0), ("포장 설비명", 1, 2),
            ("충진·포장 판정", 2, 0), ("포장 수율 기준", 2, 2)
        ]
        self._setup_ps_fields(f, fields, combos)

        ctk.CTkLabel(f, text="[충진·포장 공정 점검사항 및 자재/반제품 수불 종합 비고]:", font=ctk.CTkFont(weight="bold")).grid(row=3, column=0, columnspan=4, padx=10, pady=(10, 2), sticky="w")
        self.prod_std_pkg_log = ctk.CTkTextbox(f, height=240)
        self.prod_std_pkg_log.grid(row=4, column=0, columnspan=4, padx=6, pady=5, sticky="nsew")
        return f

    # =========================================================================
    # [화면 9] 9. 제품 규격서 (Treeview 데이터그리드)
    # =========================================================================
    def _build_ps_sec_8(self):
        f = self._create_ps_base_frame("■ [목차 9] 제품 규격서 (Product Specifications - 14대 품질규격)")
        
        in_bar = ctk.CTkFrame(f, fg_color="transparent")
        in_bar.grid(row=0, column=0, columnspan=4, sticky="ew", padx=6, pady=4)
        in_bar.grid_columnconfigure((1, 2), weight=1)

        e_item = ctk.CTkEntry(in_bar, placeholder_text="시험항목(성상/점도/pH/미생물 등)", width=140); e_item.grid(row=0, column=0, padx=2)
        e_spec = ctk.CTkEntry(in_bar, placeholder_text="규격(기준)"); e_spec.grid(row=0, column=1, sticky="ew", padx=2)
        e_meth = ctk.CTkEntry(in_bar, placeholder_text="시험방법"); e_meth.grid(row=0, column=2, sticky="ew", padx=2)
        e_rem = ctk.CTkEntry(in_bar, placeholder_text="비고", width=100); e_rem.grid(row=0, column=3, padx=2)

        def add_sp():
            cnt = len(self.prod_std_spec_tree.get_children()) + 1
            self.prod_std_spec_tree.insert("", "end", values=(
                str(cnt), e_item.get().strip(), e_spec.get().strip(),
                e_meth.get().strip(), e_rem.get().strip()
            ))
            e_item.delete(0, "end"); e_spec.delete(0, "end"); e_meth.delete(0, "end")

        def del_sp():
            for sel in self.prod_std_spec_tree.selection(): self.prod_std_spec_tree.delete(sel)

        ctk.CTkButton(in_bar, text="➕ 추가", width=65, fg_color="#0284C7", hover_color="#0369A1", command=add_sp).grid(row=0, column=4, padx=2)
        ctk.CTkButton(in_bar, text="선택 삭제", width=75, fg_color="#C62828", hover_color="#B71C1C", command=del_sp).grid(row=0, column=5, padx=2)

        cols = {"no": ("No", 45), "item": ("시험항목", 140), "spec": ("규격 (기준)", 240), "method": ("시험방법", 240), "rem": ("비고", 100)}
        f_tree, self.prod_std_spec_tree = self._create_treeview_grid(f, cols, height=12)
        f_tree.grid(row=1, column=0, columnspan=4, padx=6, pady=4, sticky="nsew")
        return f

    # =========================================================================
    # [화면 10] 10. 반제품 시험성적서 (Treeview 데이터그리드)
    # =========================================================================
    def _build_ps_sec_9(self):
        f = self._create_ps_base_frame("■ [목차 10] 반제품 시험성적서 (Semi-Finished COA)")
        combos = {"반제품 종합판정": ["적합 (Pass)", "부적합 (Fail)", "직접 입력"]}
        fields = [("반제품 제조일", 0, 0), ("반제품 LOT", 0, 2), ("반제품 시험일자", 1, 0), ("반제품 종합판정", 1, 2)]
        self._setup_ps_fields(f, fields, combos)

        in_bar = ctk.CTkFrame(f, fg_color="transparent")
        in_bar.grid(row=2, column=0, columnspan=4, sticky="ew", padx=6, pady=(10, 4))
        in_bar.grid_columnconfigure((1, 2), weight=1)

        e_item = ctk.CTkEntry(in_bar, placeholder_text="시험항목", width=120); e_item.grid(row=0, column=0, padx=2)
        e_spec = ctk.CTkEntry(in_bar, placeholder_text="시험기준"); e_spec.grid(row=0, column=1, sticky="ew", padx=2)
        e_res = ctk.CTkEntry(in_bar, placeholder_text="시험결과"); e_res.grid(row=0, column=2, sticky="ew", padx=2)
        e_rem = ctk.CTkEntry(in_bar, placeholder_text="비고", width=90); e_rem.grid(row=0, column=3, padx=2)

        def add_semi():
            cnt = len(self.prod_std_semi_tree.get_children()) + 1
            self.prod_std_semi_tree.insert("", "end", values=(
                str(cnt), e_item.get().strip(), e_spec.get().strip(), e_res.get().strip(), e_rem.get().strip()
            ))
            e_item.delete(0, "end"); e_res.delete(0, "end")

        def del_semi():
            for sel in self.prod_std_semi_tree.selection(): self.prod_std_semi_tree.delete(sel)

        ctk.CTkButton(in_bar, text="➕ 추가", width=60, fg_color="#0284C7", hover_color="#0369A1", command=add_semi).grid(row=0, column=4, padx=2)
        ctk.CTkButton(in_bar, text="선택 삭제", width=70, fg_color="#C62828", hover_color="#B71C1C", command=del_semi).grid(row=0, column=5, padx=2)

        cols = {"no": ("No", 45), "item": ("시험항목", 140), "spec": ("시험기준", 250), "res": ("시험결과", 160), "rem": ("비고", 100)}
        f_tree, self.prod_std_semi_tree = self._create_treeview_grid(f, cols, height=9)
        f_tree.grid(row=3, column=0, columnspan=4, padx=6, pady=4, sticky="nsew")
        return f

    # =========================================================================
    # [화면 11] 11. 완제품 시험성적서 (Treeview 데이터그리드)
    # =========================================================================
    def _build_ps_sec_10(self):
        f = self._create_ps_base_frame("■ [목차 11] 완제품 시험성적서 (Finished Product COA)")
        combos = {"완제품 종합판정": ["적합 (Pass)", "부적합 (Fail)", "직접 입력"]}
        fields = [("완제품 포장일자", 0, 0), ("완제품 LOT", 0, 2), ("완제품 시험일자", 1, 0), ("완제품 종합판정", 1, 2)]
        self._setup_ps_fields(f, fields, combos)

        in_bar = ctk.CTkFrame(f, fg_color="transparent")
        in_bar.grid(row=2, column=0, columnspan=4, sticky="ew", padx=6, pady=(10, 4))
        in_bar.grid_columnconfigure((1, 2), weight=1)

        e_item = ctk.CTkEntry(in_bar, placeholder_text="시험항목", width=120); e_item.grid(row=0, column=0, padx=2)
        e_spec = ctk.CTkEntry(in_bar, placeholder_text="시험기준"); e_spec.grid(row=0, column=1, sticky="ew", padx=2)
        e_res = ctk.CTkEntry(in_bar, placeholder_text="시험결과"); e_res.grid(row=0, column=2, sticky="ew", padx=2)
        e_rem = ctk.CTkEntry(in_bar, placeholder_text="비고", width=90); e_rem.grid(row=0, column=3, padx=2)

        def add_fin():
            cnt = len(self.prod_std_fin_tree.get_children()) + 1
            self.prod_std_fin_tree.insert("", "end", values=(
                str(cnt), e_item.get().strip(), e_spec.get().strip(), e_res.get().strip(), e_rem.get().strip()
            ))
            e_item.delete(0, "end"); e_res.delete(0, "end")

        def del_fin():
            for sel in self.prod_std_fin_tree.selection(): self.prod_std_fin_tree.delete(sel)

        ctk.CTkButton(in_bar, text="➕ 추가", width=60, fg_color="#0284C7", hover_color="#0369A1", command=add_fin).grid(row=0, column=4, padx=2)
        ctk.CTkButton(in_bar, text="선택 삭제", width=70, fg_color="#C62828", hover_color="#B71C1C", command=del_fin).grid(row=0, column=5, padx=2)

        cols = {"no": ("No", 45), "item": ("시험항목", 140), "spec": ("시험기준", 250), "res": ("시험결과", 160), "rem": ("비고", 100)}
        f_tree, self.prod_std_fin_tree = self._create_treeview_grid(f, cols, height=9)
        f_tree.grid(row=3, column=0, columnspan=4, padx=6, pady=4, sticky="nsew")
        return f

    # =========================================================================
    # [화면 12] 12. 제조 및 품질관리 시설·기구 (공장 기본 설비 영구 보존 & Treeview)
    # =========================================================================
    def _build_ps_sec_11(self):
        f = self._create_ps_base_frame("■ [목차 12] 제조 및 품질관리에 필요한 시설 및 기구 (Facilities & Equipment)")
        
        # 상단 공장 설비 프리셋 관리 툴바
        preset_bar = ctk.CTkFrame(f, fg_color=("gray95", "gray18"), corner_radius=6, border_width=1, border_color=("gray80", "gray28"))
        preset_bar.grid(row=0, column=0, columnspan=4, sticky="ew", padx=6, pady=(2, 6))
        preset_bar.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(preset_bar, text="⚙️ 우리 공장 기본 설비 인프라 (한 번 저장해두면 매번 자동으로 유지됩니다)", font=ctk.CTkFont(weight="bold", size=11), text_color=("#1F497D", "#38BDF8")).grid(row=0, column=0, padx=8, pady=4, sticky="w")
        
        p_btns = ctk.CTkFrame(preset_bar, fg_color="transparent")
        p_btns.grid(row=0, column=1, padx=6, pady=4, sticky="e")
        ctk.CTkButton(p_btns, text="💾 우리 공장 기본설비로 저장", width=170, height=28, fg_color="#2E7D32", hover_color="#1B5E20", font=ctk.CTkFont(weight="bold", size=11), command=self.save_default_facility_preset).pack(side="left", padx=3)
        ctk.CTkButton(p_btns, text="🔄 기본설비 불러오기", width=140, height=28, fg_color="#1565C0", hover_color="#0D47A1", font=ctk.CTkFont(weight="bold", size=11), command=self.load_default_facility_preset).pack(side="left", padx=3)
        ctk.CTkButton(p_btns, text="초기화(표준 복원)", width=110, height=28, fg_color="gray50", font=ctk.CTkFont(size=11), command=self.reset_standard_facility_preset).pack(side="left", padx=3)

        # 항목 입력 툴바
        in_bar = ctk.CTkFrame(f, fg_color="transparent")
        in_bar.grid(row=1, column=0, columnspan=4, sticky="ew", padx=6, pady=4)
        in_bar.grid_columnconfigure((0, 1, 2), weight=1)

        e_mfg = ctk.CTkEntry(in_bar, placeholder_text="제조시설 기구명"); e_mfg.grid(row=0, column=0, sticky="ew", padx=2)
        e_pkg = ctk.CTkEntry(in_bar, placeholder_text="충진 및 포장 설비명"); e_pkg.grid(row=0, column=1, sticky="ew", padx=2)
        e_qc = ctk.CTkEntry(in_bar, placeholder_text="품질관리 시험기구명"); e_qc.grid(row=0, column=2, sticky="ew", padx=2)

        def add_fac():
            cnt = len(self.prod_std_fac_tree.get_children()) + 1
            self.prod_std_fac_tree.insert("", "end", values=(
                str(cnt), e_mfg.get().strip() or "-", e_pkg.get().strip() or "-", e_qc.get().strip() or "-"
            ))
            e_mfg.delete(0, "end"); e_pkg.delete(0, "end"); e_qc.delete(0, "end")

        def del_fac():
            for sel in self.prod_std_fac_tree.selection(): self.prod_std_fac_tree.delete(sel)

        ctk.CTkButton(in_bar, text="➕ 추가", width=65, fg_color="#0284C7", hover_color="#0369A1", command=add_fac).grid(row=0, column=3, padx=2)
        ctk.CTkButton(in_bar, text="선택 삭제", width=75, fg_color="#C62828", hover_color="#B71C1C", command=del_fac).grid(row=0, column=4, padx=2)

        cols = {"no": ("No", 45), "mfg": ("제조시설 기구명", 240), "pkg": ("충진 및 포장 설비명", 220), "qc": ("품질관리 시험기구명", 240)}
        f_tree, self.prod_std_fac_tree = self._create_treeview_grid(f, cols, height=12)
        f_tree.grid(row=2, column=0, columnspan=4, padx=6, pady=4, sticky="nsew")

        # 저장된 기본 설비가 있으면 즉시 자동 로드
        self.load_default_facility_preset(silent=True)
        return f

    def _get_default_facility_list(self):
        """식약처/CGMP 표준 20대 시설 기구 마스터 목록"""
        return [
            ("1", "제조탱크 (1000L)", "-", "pH 메타"),
            ("2", "제조탱크 (300L)", "-", "현미경"),
            ("3", "제조탱크 (60L)", "-", "점도계"),
            ("4", "정제수 제조장치", "-", "인큐베이터-1"),
            ("5", "디스퍼 믹서-1", "-", "인큐베이터-2"),
            ("6", "디스퍼 믹서-2", "-", "인큐베이터-3"),
            ("7", "여과장치-1", "-", "수욕조"),
            ("8", "여과장치-2", "-", "호모믹서-1"),
            ("9", "소형 집진기", "-", "호모믹서-2"),
            ("10", "이동식 핸드 파렛트 리프트", "-", "디스퍼믹서-1"),
            ("11", "전자저울 (250kg)", "-", "디스퍼믹서-2"),
            ("12", "전자저울 (150kg)", "-", "아지믹서"),
            ("13", "소형 전자저울-1", "-", "클린벤치"),
            ("14", "모터식 드럼펌프-1", "-", "핫플레이트-1"),
            ("15", "모터식 드럼펌프-2", "-", "핫플레이트-2"),
            ("16", "-", "자동 충진기", "소형 전자저울-1"),
            ("17", "-", "라벨러", "소형 전자저울-2"),
            ("18", "-", "중량 선별기", "Vortexer"),
            ("19", "-", "수축 포장기", "시약 냉장고"),
            ("20", "-", "테이핑기 / 박스제함기", "비중병")
        ]

    def _get_facility_preset_file_path(self):
        cfg_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "config")
        os.makedirs(cfg_dir, exist_ok=True)
        return os.path.join(cfg_dir, "facility_preset.json")

    def save_default_facility_preset(self):
        """현재 Treeview에 있는 시설 목록을 우리 공장 기본값으로 영구 저장"""
        if not hasattr(self, 'prod_std_fac_tree'): return
        rows = []
        for item_id in self.prod_std_fac_tree.get_children():
            rows.append(list(self.prod_std_fac_tree.item(item_id, "values")))
        
        try:
            fpath = self._get_facility_preset_file_path()
            with open(fpath, "w", encoding="utf-8") as f:
                json.dump(rows, f, ensure_ascii=False, indent=2)
            messagebox.showinfo("저장 완료", f"우리 공장 기본 시설 및 기구 목록({len(rows)}개)이 기본 프리셋으로 안전하게 저장되었습니다. 다음번 작성 시에도 자동으로 로드됩니다.", parent=self)
        except Exception as e:
            messagebox.showerror("오류", f"기본설비 저장 실패: {e}", parent=self)

    def load_default_facility_preset(self, silent=False):
        """저장된 공장 기본 설비 로드"""
        if not hasattr(self, 'prod_std_fac_tree'): return
        fpath = self._get_facility_preset_file_path()
        rows = None
        if os.path.exists(fpath):
            try:
                with open(fpath, "r", encoding="utf-8") as f:
                    rows = json.load(f)
            except Exception: pass

        if not rows:
            rows = self._get_default_facility_list()

        for item in self.prod_std_fac_tree.get_children():
            self.prod_std_fac_tree.delete(item)

        for r in rows:
            self.prod_std_fac_tree.insert("", "end", values=r)

        if not silent:
            messagebox.showinfo("불러오기 완료", f"공장 기본 시설 및 기구 목록({len(rows)}개)을 성공적으로 불러왔습니다.", parent=self)

    def reset_standard_facility_preset(self):
        """식약처 표준 20대 설비로 리셋"""
        if not hasattr(self, 'prod_std_fac_tree'): return
        if not messagebox.askyesno("초기화 확인", "공장 설비 목록을 식약처/CGMP 표준 20대 기본 설비로 초기화하시겠습니까?", parent=self): return
        rows = self._get_default_facility_list()
        for item in self.prod_std_fac_tree.get_children(): self.prod_std_fac_tree.delete(item)
        for r in rows: self.prod_std_fac_tree.insert("", "end", values=r)
        self.save_default_facility_preset()

    def show_prod_std_section(self, idx):
        idx = max(0, min(idx, len(self.prod_std_section_titles) - 1))
        self.current_prod_std_section_idx = idx

        if self.prod_std_section_frames[idx] is None:
            self.prod_std_section_frames[idx] = self.prod_std_section_builders[idx]()

        for i, frame in enumerate(self.prod_std_section_frames):
            if frame is not None:
                if i == idx: frame.grid(row=0, column=0, sticky="nsew")
                else: frame.grid_forget()

        if hasattr(self, 'prod_std_section_selector'):
            self.prod_std_section_selector.set(self.prod_std_section_titles[idx])

        if hasattr(self, 'btn_prev_ps'):
            if idx == 0: self.btn_prev_ps.configure(state="disabled", fg_color="gray40")
            else: self.btn_prev_ps.configure(state="normal", fg_color="#455A64")

        if hasattr(self, 'btn_next_ps'):
            if idx == len(self.prod_std_section_titles) - 1: self.btn_next_ps.configure(state="disabled", fg_color="gray40")
            else: self.btn_next_ps.configure(state="normal", fg_color="#1565C0")

    def prev_prod_std_section(self):
        if self.current_prod_std_section_idx > 0:
            self.show_prod_std_section(self.current_prod_std_section_idx - 1)

    def next_prod_std_section(self):
        if self.current_prod_std_section_idx < len(self.prod_std_section_titles) - 1:
            self.show_prod_std_section(self.current_prod_std_section_idx + 1)

    def on_prod_std_section_selected(self, choice):
        if choice in self.prod_std_section_titles:
            idx = self.prod_std_section_titles.index(choice)
            self.show_prod_std_section(idx)

    def _ensure_all_prod_std_sections_built(self):
        """저장/로드/엑셀 출력 시 13개 전체 화면 생성 보장"""
        for i in range(13):
            if self.prod_std_section_frames[i] is None:
                self.prod_std_section_frames[i] = self.prod_std_section_builders[i]()

    def clear_prod_std_form(self):
        """완전 공란(Blank Form)으로 초기화"""
        self.current_prod_std_id = None
        for k, e in self.prod_std_entries.items():
            if isinstance(e, ctk.CTkComboBox):
                vals = e.cget("values")
                if vals: e.set(vals[0])
            else: e.delete(0, "end")
        
        # 설정에서 자회사 정보 복원
        cp = get_company_profile()
        if "양식/관리번호" in self.prod_std_entries: self.prod_std_entries["양식/관리번호"].insert(0, cp.get("form_doc_no", "양0100-01"))
        if "개정번호" in self.prod_std_entries: self.prod_std_entries["개정번호"].insert(0, cp.get("form_rev_no", "Rev.0"))
        if "연구소/회사명" in self.prod_std_entries: self.prod_std_entries["연구소/회사명"].insert(0, cp.get("company_name_ko", "(주)한국피부과학연구소"))

        trees = ['prod_std_rev_tree', 'prod_std_proc_tree', 'prod_std_ing_tree', 'prod_std_weigh_tree', 'prod_std_pkg_tree', 'prod_std_spec_tree', 'prod_std_semi_tree', 'prod_std_fin_tree']
        for tr_name in trees:
            if hasattr(self, tr_name):
                tr = getattr(self, tr_name)
                for item in tr.get_children(): tr.delete(item)

        # 12번 시설·기구는 사용자가 매번 다시 입력하지 않도록 공장 기본 설비 자동 유지
        if hasattr(self, 'load_default_facility_preset'):
            self.load_default_facility_preset(silent=True)

        if hasattr(self, 'prod_std_mfg_summary'): self.prod_std_mfg_summary.delete("1.0", "end")
        if hasattr(self, 'prod_std_pkg_log'): self.prod_std_pkg_log.delete("1.0", "end")

    def refresh_prod_std_list(self):
        try:
            session = db_manager.get_session()
            recs = session.query(ProductStandard).order_by(ProductStandard.created_at.desc()).limit(50).all()
            vals = [f"{r.id} | {r.product_name} | 코드:{r.product_code or ''}" for r in recs]
            self.prod_std_picker.configure(values=vals if vals else ["-- 저장된 제품표준서 없음 --"])
            if vals: self.prod_std_picker.set(vals[0])
            session.close()
        except Exception as e:
            print(f"[경고] 제품표준서 목록 로드 실패: {e}")

    def load_selected_prod_std(self):
        sel = self.prod_std_picker.get()
        if not sel or "--" in sel:
            messagebox.showwarning("선택 오류", "불러올 제품표준서를 선택해주세요.", parent=self)
            return

        ps_id = int(sel.split("|")[0].strip())
        session = db_manager.get_session()
        r = session.query(ProductStandard).filter(ProductStandard.id == ps_id).first()
        if not r:
            session.close()
            messagebox.showerror("오류", "선택한 제품표준서를 찾을 수 없습니다.", parent=self)
            return

        self.current_prod_std_id = r.id
        self._ensure_all_prod_std_sections_built()
        self.clear_prod_std_form()

        col_mappings = {
            "제품명(국문)": r.product_name,
            "제품명(영문)": r.product_name_en,
            "제품 코드": r.product_code,
            "유형 / 세부 유형": r.cosmetic_type,
            "제품용량 (ml/g)": r.package_volume,
            "성 상": r.appearance_criteria,
            "보 관 방 법": r.storage_condition,
            "사용기한 / 개봉 후 사용기간": r.expiry_period
        }
        for k, v in col_mappings.items():
            if k in self.prod_std_entries and v:
                e = self.prod_std_entries[k]
                e.delete(0, "end"); e.insert(0, str(v))

        if r.packaging_specs_json:
            try:
                bundle = json.loads(r.packaging_specs_json)
                if "entries" in bundle:
                    for k, v in bundle["entries"].items():
                        if k in self.prod_std_entries and v:
                            e = self.prod_std_entries[k]
                            e.delete(0, "end"); e.insert(0, str(v))
                
                # Treeview 데이터 복원
                tree_load_map = [
                    ("rev_data", 'prod_std_rev_tree'),
                    ("proc_data", 'prod_std_proc_tree'),
                    ("ing_data", 'prod_std_ing_tree'),
                    ("weigh_data", 'prod_std_weigh_tree'),
                    ("pkg_data", 'prod_std_pkg_tree'),
                    ("spec_data", 'prod_std_spec_tree'),
                    ("semi_data", 'prod_std_semi_tree'),
                    ("fin_data", 'prod_std_fin_tree'),
                    ("fac_data", 'prod_std_fac_tree')
                ]
                for key_name, tr_attr in tree_load_map:
                    if key_name in bundle and hasattr(self, tr_attr):
                        tr = getattr(self, tr_attr)
                        for r_vals in bundle[key_name]:
                            tr.insert("", "end", values=r_vals)

                if "mfg_summary" in bundle and hasattr(self, 'prod_std_mfg_summary'):
                    self.prod_std_mfg_summary.delete("1.0", "end")
                    self.prod_std_mfg_summary.insert("1.0", str(bundle["mfg_summary"]))
                if "pkg_log" in bundle and hasattr(self, 'prod_std_pkg_log'):
                    self.prod_std_pkg_log.delete("1.0", "end")
                    self.prod_std_pkg_log.insert("1.0", str(bundle["pkg_log"]))
            except Exception as e:
                print(f"[경고] 제품표준서 JSON 파싱 오류: {e}")

        session.close()
        messagebox.showinfo("불러오기 완료", f"[{r.product_name}] 제품표준서의 표지 및 12대 전 목차 데이터가 로드되었습니다.", parent=self)

    def save_prod_std_to_db(self):
        p_name = self.prod_std_entries.get("제품명(국문)", ctk.CTkEntry(self)).get().strip()
        if not p_name:
            messagebox.showwarning("필수 입력 누락", "제품명(국문)은 필수 입력 항목입니다.", parent=self)
            return

        session = db_manager.get_session()
        try:
            if self.current_prod_std_id:
                r = session.query(ProductStandard).filter(ProductStandard.id == self.current_prod_std_id).first()
                if not r: r = ProductStandard(); session.add(r)
            else:
                r = ProductStandard(); session.add(r)

            r.product_name = p_name
            r.product_name_en = self.prod_std_entries.get("제품명(영문)", ctk.CTkEntry(self)).get().strip()
            r.product_code = self.prod_std_entries.get("제품 코드", ctk.CTkEntry(self)).get().strip()
            r.cosmetic_type = self.prod_std_entries.get("유형 / 세부 유형", ctk.CTkEntry(self)).get().strip()
            r.package_volume = self.prod_std_entries.get("제품용량 (ml/g)", ctk.CTkEntry(self)).get().strip()
            r.appearance_criteria = self.prod_std_entries.get("성 상", ctk.CTkEntry(self)).get().strip()
            r.storage_condition = self.prod_std_entries.get("보 관 방 법", ctk.CTkEntry(self)).get().strip()
            r.expiry_period = self.prod_std_entries.get("사용기한 / 개봉 후 사용기간", ctk.CTkEntry(self)).get().strip()

            def get_tree_rows(tr_attr):
                res = []
                if hasattr(self, tr_attr):
                    tr = getattr(self, tr_attr)
                    for item_id in tr.get_children():
                        res.append(list(tr.item(item_id, "values")))
                return res

            all_entries = {k: w.get().strip() for k, w in self.prod_std_entries.items()}
            mfg_sum = self.prod_std_mfg_summary.get("1.0", "end-1c").strip() if hasattr(self, 'prod_std_mfg_summary') else ""
            pkg_lg = self.prod_std_pkg_log.get("1.0", "end-1c").strip() if hasattr(self, 'prod_std_pkg_log') else ""

            bundle = {
                "entries": all_entries,
                "rev_data": get_tree_rows('prod_std_rev_tree'),
                "proc_data": get_tree_rows('prod_std_proc_tree'),
                "ing_data": get_tree_rows('prod_std_ing_tree'),
                "weigh_data": get_tree_rows('prod_std_weigh_tree'),
                "pkg_data": get_tree_rows('prod_std_pkg_tree'),
                "spec_data": get_tree_rows('prod_std_spec_tree'),
                "semi_data": get_tree_rows('prod_std_semi_tree'),
                "fin_data": get_tree_rows('prod_std_fin_tree'),
                "fac_data": get_tree_rows('prod_std_fac_tree'),
                "mfg_summary": mfg_sum,
                "pkg_log": pkg_lg
            }
            r.packaging_specs_json = json.dumps(bundle, ensure_ascii=False)

            session.commit()
            self.current_prod_std_id = r.id
            messagebox.showinfo("저장 완료", f"[{p_name}] 제품표준서의 표지 및 12대 전 목차 정보가 DB에 저장되었습니다.", parent=self)
            self.refresh_prod_std_list()
        except Exception as e:
            session.rollback()
            messagebox.showerror("저장 오류", f"제품표준서 저장 실패: {e}", parent=self)
        finally:
            session.close()

    def delete_prod_std(self):
        if not self.current_prod_std_id:
            messagebox.showwarning("선택 오류", "삭제할 제품표준서를 먼저 불러와주세요.", parent=self)
            return
        if not messagebox.askyesno("삭제 확인", "선택한 제품표준서를 영구히 삭제하시겠습니까?", parent=self): return
        session = db_manager.get_session()
        try:
            r = session.query(ProductStandard).filter(ProductStandard.id == self.current_prod_std_id).first()
            if r:
                session.delete(r); session.commit()
                self.clear_prod_std_form()
                self.refresh_prod_std_list()
                messagebox.showinfo("완료", "제품표준서가 성공적으로 삭제되었습니다.", parent=self)
        finally:
            session.close()

    def export_prod_std_to_excel(self, lang="ko"):
        """식약처 / CGMP 표준 [제품표준서] 공식 표지 + 12대 목차 = 총 13개 시트 전수 100% 국문/영문 엑셀 출력 엔진"""
        import re
        self._ensure_all_prod_std_sections_built()
        is_en = (lang == "en")
        cp = get_company_profile()
        is_conf = getattr(self, 'prod_std_confidential_var', ctk.BooleanVar(value=False)).get()

        # 영문/국문 번역 사전
        dict_ps_en = {
            "적합": "Pass (Complies)", "부적합": "Fail (Non-compliant)", "조건부 적합": "Conditional Pass",
            "자사규격": "In-house Specification", "공정규격": "In-process Standard",
            "무색 투명 겔상": "Colorless transparent gel",
            "유백색 점조성 크림 제형": "Milky-white viscous cream formulation",
            "투명 점조성 액상": "Transparent viscous liquid",
            "미황색 에멀젼": "Pale-yellow emulsion",
            "고체 스틱": "Solid stick",
            "해당사항없음": "Not Applicable",
            "해당사항없음 (일반화장품)": "Not Applicable (General Cosmetics)",
            "제조일로부터 36개월 / 개봉후 12개월": "36 months from mfg date / 12 months after opening",
            "제조일로부터 24개월 / 개봉후 12개월": "24 months from mfg date / 12 months after opening",
            "제조일로부터 30개월 / 개봉후 6개월": "30 months from mfg date / 6 months after opening",
            "통풍이 잘되는 차광된 장소에서 상온(1~30℃)에서 보관": "Store in a well-ventilated, shaded place at room temp (1~30℃)",
            "직사광선을 피하고 서늘한 곳에 밀폐 보관 (1~30℃)": "Keep tightly closed in a cool place out of direct sunlight (1~30℃)",
            "피부의 주름개선에 도움을 준다.": "Helps improve skin wrinkles.",
            "피부의 미백에 도움을 준다.": "Helps brighten the skin.",
            "미백 및 주름개선 2중 기능성": "Dual functional for whitening and wrinkle improvement",
            "본품 적당량을 취해 피부에 골고루 펴 바릅니다.": "Take an appropriate amount and spread evenly over the skin.",
            "따뜻한 물로 적신 후 적당량을 펴 발라 사용합니다.": "Wet with warm water, take an appropriate amount and spread evenly.",
            "1) 화장품 사용 시 또는 직사광선에 의한 이상 증상 시 전문의 상담 / 2) 상처 부위 자제 / 3) 어린이 손 닿지 않는 곳, 직사광선 피할 것": "1) Consult a doctor if abnormal symptoms occur / 2) Avoid wounded areas / 3) Keep out of reach of children, avoid direct sunlight",
            "적합 (양호)": "Pass (Good)",
            "충진 포장 수율 95% 이상 적합": "Packaging yield 95% or higher is acceptable",
            "수율 98% 이상": "Yield 98% or higher",
            "적합 (Pass)": "Pass",
            "부적합 (Fail)": "Fail"
        }

        def tr_ps(text):
            if not text: return ""
            t = str(text).strip()
            if not is_en: return t
            if not re.search(r'[가-힣]', t): return t
            if t in dict_ps_en: return dict_ps_en[t]
            for k_d, v_d in dict_ps_en.items():
                if k_d in t: t = t.replace(k_d, v_d)
            return t

        p_name_ko = self.prod_std_entries.get("제품명(국문)", ctk.CTkEntry(self)).get().strip() or "제품표준서_양식"
        p_name_en = self.prod_std_entries.get("제품명(영문)", ctk.CTkEntry(self)).get().strip() or "Product_Standard_Form"
        p_name_disp = p_name_en if is_en else p_name_ko
        doc_no = self.prod_std_entries.get("제품표준서번호", ctk.CTkEntry(self)).get().strip() or self.prod_std_entries.get("양식/관리번호", ctk.CTkEntry(self)).get().strip() or cp.get("form_doc_no", "양0100-01")
        rev_no = self.prod_std_entries.get("개정번호", ctk.CTkEntry(self)).get().strip() or cp.get("form_rev_no", "Rev.0")
        corp_name = cp.get("company_name_en") if is_en else cp.get("company_name_ko", "(주)한국피부과학연구소")

        author_val = (cp.get("manager_name_en") if is_en else self.prod_std_entries.get("작성자", ctk.CTkEntry(self)).get().strip()) or "-"
        author_dt = self.prod_std_entries.get("작성일자", ctk.CTkEntry(self)).get().strip() or "-"
        review_val = (cp.get("manager_name_en") if is_en else self.prod_std_entries.get("검토자", ctk.CTkEntry(self)).get().strip()) or "-"
        review_dt = self.prod_std_entries.get("검토일자", ctk.CTkEntry(self)).get().strip() or "-"
        appr_val = (cp.get("representative") if is_en else self.prod_std_entries.get("승인자", ctk.CTkEntry(self)).get().strip()) or "-"
        appr_dt = self.prod_std_entries.get("승인일자", ctk.CTkEntry(self)).get().strip() or "-"

        wb = Workbook()

        font_title = Font(name="맑은 고딕", size=17, bold=True)
        font_sec_hdr = Font(name="맑은 고딕", size=12, bold=True, color="1F497D")
        font_tbl_hdr = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
        font_lbl = Font(name="맑은 고딕", size=10, bold=True)
        font_cell = Font(name="맑은 고딕", size=9.5)
        font_foot = Font(name="맑은 고딕", size=8.5, color="7F7F7F")

        align_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_l = Alignment(horizontal="left", vertical="center", wrap_text=True)

        fill_sec = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")
        fill_tbl_hdr = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin = Side(style="thin", color="A6B9D0")
        thin_dark = Side(style="thin", color="595959")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        table_border = Border(left=thin_dark, right=thin_dark, top=thin_dark, bottom=thin_dark)

        def style_r(ws_obj, cell_range, font=None, fill=None, align=None, border=None):
            for row in ws_obj[cell_range]:
                for c in row:
                    if font: c.font = font
                    if fill: c.fill = fill
                    if align: c.alignment = align
                    if border: c.border = border

        def add_sheet_header(ws_obj, sec_num, sec_title_ko, sec_title_en):
            ws_obj.views.sheetView[0].showGridLines = True
            ws_obj.merge_cells("A1:F2")
            ws_obj["A1"] = f"{sec_num}. {sec_title_en}" if is_en else f"{sec_num}. {sec_title_ko}"
            ws_obj["A1"].font = font_title; ws_obj["A1"].alignment = align_c
            ws_obj["E3"] = f"Doc No.: {doc_no}" if is_en else f"표준서번호: {doc_no}"
            ws_obj["E3"].font = font_lbl

            ws_obj.merge_cells("A4:B4"); ws_obj["A4"] = "Product Name" if is_en else "제 품 명"; ws_obj["A4"].font = font_lbl; ws_obj["A4"].alignment = align_c; ws_obj["A4"].fill = fill_sec
            ws_obj.merge_cells("C4:F4"); ws_obj["C4"] = p_name_disp; ws_obj["C4"].font = font_lbl; ws_obj["C4"].alignment = align_l
            style_r(ws_obj, "A4:F4", border=thin_border)

        def add_sheet_footer(ws_obj, end_r):
            curr = end_r + 2
            appr_lbl = "Approval" if is_en else "결재"
            prep_lbl = f"Prepared by: {author_val}" if is_en else f"작 성: {author_val}"
            rev_lbl = f"Reviewed by: {review_val}" if is_en else f"검 토: {review_val}"
            app_lbl = f"Approved by: {appr_val}" if is_en else f"승 인: {appr_val}"

            ws_obj.merge_cells(f"C{curr}:C{curr+1}"); ws_obj[f"C{curr}"] = appr_lbl; ws_obj[f"C{curr}"].font = font_lbl; ws_obj[f"C{curr}"].alignment = align_c; ws_obj[f"C{curr}"].fill = fill_sec
            ws_obj[f"D{curr}"] = prep_lbl; ws_obj[f"E{curr}"] = rev_lbl; ws_obj[f"F{curr}"] = app_lbl
            style_r(ws_obj, f"D{curr}:F{curr}", font=font_lbl, align=align_c, fill=fill_sec, border=thin_border)
            ws_obj[f"D{curr+1}"] = author_dt; ws_obj[f"E{curr+1}"] = review_dt; ws_obj[f"F{curr+1}"] = appr_dt
            style_r(ws_obj, f"C{curr}:F{curr+1}", border=thin_border, font=font_cell, align=align_c)

            curr += 3
            ws_obj[f"A{curr}"] = f"{doc_no}({rev_no})"; ws_obj[f"C{curr}"] = corp_name; ws_obj[f"F{curr}"] = "A4(210X297)"
            style_r(ws_obj, f"A{curr}:F{curr}", font=font_foot)

        # =============================================================
        # Sheet 0: [0. 표지 및 목차] (Cover & Index)
        # =============================================================
        ws0 = wb.active
        ws0.title = "Cover & Index" if is_en else "표지 및 목차"
        ws0.views.sheetView[0].showGridLines = True

        ws0.merge_cells("A1:F2")
        ws0["A1"] = "PRODUCT SPECIFICATION STANDARD" if is_en else "제 품 표 준 서"
        ws0["A1"].font = font_title; ws0["A1"].alignment = align_c
        
        ws0["A3"] = "[CONFIDENTIAL / TRADE SECRET]" if (is_en and is_conf) else ("[영업비밀 보호문서 / 기밀사항]" if is_conf else "")
        if is_conf: ws0["A3"].font = Font(name="맑은 고딕", size=10, bold=True, color="C00000")

        ws0["E3"] = f"Standard No.: {doc_no}" if is_en else f"표준서번호: {doc_no}"
        ws0["E3"].font = font_lbl

        ws0.merge_cells("A4:B4"); ws0["A4"] = "Product Name" if is_en else "제 품 명"; ws0["A4"].font = font_lbl; ws0["A4"].alignment = align_c; ws0["A4"].fill = fill_sec
        ws0.merge_cells("C4:F4"); ws0["C4"] = p_name_disp; ws0["C4"].font = font_lbl; ws0["C4"].alignment = align_l
        style_r(ws0, "A4:F4", border=thin_border)

        ws0.append([])
        hdr_toc = ["NO.", "TABLE OF CONTENTS", "", "", "", "Remarks"] if is_en else ["NO.", "목  차", "", "", "", "비고"]
        ws0.append(hdr_toc)
        ws0.merge_cells("B6:E6")
        style_r(ws0, "A6:F6", font=font_tbl_hdr, fill=fill_tbl_hdr, align=align_c, border=thin_border)

        toc_items = [
            ("1", "변경이력", "Revision History"),
            ("2", "제품참고사항", "General Information"),
            ("3", "공정규격", "Manufacturing Process Specifications"),
            ("4", "원료성분 기준 및 시험방법 (100g당)", "Raw Material Specifications & Formulations"),
            ("5", "계량 지시 및 기록서", "Weighing Order & Record"),
            ("6", "제조 지시 및 기록서", "Batch Manufacturing Record (BMR)"),
            ("7", "제품사양 (포장재규격포함)", "Packaging Specifications"),
            ("8", "충진·포장 지시 및 기록서", "Packaging Order & Record"),
            ("9", "제품 규격서", "Product Specifications"),
            ("10", "반제품 시험성적서", "Semi-Finished Product COA"),
            ("11", "완제품 시험성적서", "Finished Product COA"),
            ("12", "제조 및 품질관리에 필요한 시설 및 기구", "Facilities and Equipment")
        ]

        curr_r = 7
        for c_no, c_ko, c_en in toc_items:
            ws0[f"A{curr_r}"] = c_no
            ws0.merge_cells(f"B{curr_r}:E{curr_r}")
            ws0[f"B{curr_r}"] = c_en if is_en else c_ko
            ws0[f"F{curr_r}"] = "-"
            style_r(ws0, f"A{curr_r}:F{curr_r}", font=font_cell, align=align_c, border=thin_border)
            ws0[f"B{curr_r}"].alignment = align_l
            curr_r += 1

        add_sheet_footer(ws0, curr_r)
        for col_l, w in [('A', 8), ('B', 20), ('C', 20), ('D', 18), ('E', 18), ('F', 18)]: ws0.column_dimensions[col_l].width = w

        # =============================================================
        # Sheet 1: [1. 변경이력] (Revision History)
        # =============================================================
        ws1 = wb.create_sheet(title="1. Revision History" if is_en else "1. 변경이력")
        add_sheet_header(ws1, "1", "변경이력", "Revision History")
        
        ws1.append([])
        hdr1 = ["No.", "Date", "Revision & Change Description", "Author", "Approved by", "Remarks"] if is_en else ["연번", "날짜", "개정 및 변경 내용 (사유)", "작성자", "승인자", "비고"]
        ws1.append(hdr1)
        style_r(ws1, "A6:F6", font=font_tbl_hdr, fill=fill_tbl_hdr, align=align_c, border=thin_border)

        r_cnt1 = 7
        if hasattr(self, 'prod_std_rev_tree'):
            for item in self.prod_std_rev_tree.get_children():
                v = self.prod_std_rev_tree.item(item, "values")
                if v:
                    ws1.append([v[0], v[1], tr_ps(v[2]), tr_ps(v[3]), tr_ps(v[4]), tr_ps(v[5])])
                    style_r(ws1, f"A{r_cnt1}:F{r_cnt1}", font=font_cell, align=align_c, border=thin_border)
                    ws1[f"C{r_cnt1}"].alignment = align_l
                    r_cnt1 += 1
        if r_cnt1 == 7:
            ws1.append(["1", author_dt if author_dt != "-" else datetime.now().strftime("%Y-%m-%d"), "First Creation (Rev.0)" if is_en else "최초 제정 (Rev.0)", author_val, appr_val, "-"])
            style_r(ws1, f"A{r_cnt1}:F{r_cnt1}", font=font_cell, align=align_c, border=thin_border)
            r_cnt1 += 1

        add_sheet_footer(ws1, r_cnt1)
        for col_l, w in [('A', 8), ('B', 15), ('C', 35), ('D', 15), ('E', 15), ('F', 15)]: ws1.column_dimensions[col_l].width = w

        # =============================================================
        # Sheet 2: [2. 제품참고사항] (General Information)
        # =============================================================
        ws2 = wb.create_sheet(title="2. General Info" if is_en else "2. 제품참고사항")
        add_sheet_header(ws2, "2", "제품참고사항", "General Information")

        gen_fields = [
            ("Product Name" if is_en else "제 품 명", p_name_disp),
            ("Product Code" if is_en else "제품 코드", self.prod_std_entries.get("제품 코드", ctk.CTkEntry(self)).get().strip() or "-"),
            ("Approval Date" if is_en else "허가(보고) 취득일", tr_ps(self.prod_std_entries.get("허가(보고) 취득일", ctk.CTkEntry(self)).get().strip())),
            ("Approval No." if is_en else "허가(보고) 번호", tr_ps(self.prod_std_entries.get("허가(보고) 번호", ctk.CTkEntry(self)).get().strip())),
            ("Date of Creation" if is_en else "작성 일자", self.prod_std_entries.get("작성 일자", ctk.CTkEntry(self)).get().strip()),
            ("Author" if is_en else "작 성 자", author_val),
            ("Net Volume / Weight" if is_en else "제품용량 (ml/g)", self.prod_std_entries.get("제품용량 (ml/g)", ctk.CTkEntry(self)).get().strip()),
            ("Appearance" if is_en else "성 상", tr_ps(self.prod_std_entries.get("성 상", ctk.CTkEntry(self)).get().strip())),
            ("Cosmetic Category" if is_en else "유형 / 세부 유형", tr_ps(self.prod_std_entries.get("유형 / 세부 유형", ctk.CTkEntry(self)).get().strip())),
            ("Period After Opening" if is_en else "사용기한 / 개봉후", tr_ps(self.prod_std_entries.get("사용기한 / 개봉 후 사용기간", ctk.CTkEntry(self)).get().strip())),
            ("Storage Condition" if is_en else "보 관 방 법", tr_ps(self.prod_std_entries.get("보 관 방 법", ctk.CTkEntry(self)).get().strip())),
            ("Efficacy & Effects" if is_en else "효능효과", tr_ps(self.prod_std_entries.get("효능효과", ctk.CTkEntry(self)).get().strip())),
            ("Directions for Use" if is_en else "용법 용량", tr_ps(self.prod_std_entries.get("용법 용량", ctk.CTkEntry(self)).get().strip())),
            ("Full Ingredients List" if is_en else "전성분 표기사항", tr_ps(self.prod_std_entries.get("화장품 성분 표기사항 (전성분)", ctk.CTkEntry(self)).get().strip())),
            ("Precautions for Use" if is_en else "사용할 때의 주의사항", tr_ps(self.prod_std_entries.get("사용할 때의 주의사항", ctk.CTkEntry(self)).get().strip())),
            ("Other Remarks" if is_en else "기타 사항", tr_ps(self.prod_std_entries.get("기타 사항", ctk.CTkEntry(self)).get().strip()))
        ]

        r_cnt2 = 6
        for k_lbl, v_val in gen_fields:
            ws2[f"A{r_cnt2}"] = k_lbl; ws2[f"A{r_cnt2}"].font = font_lbl; ws2[f"A{r_cnt2}"].fill = fill_sec
            ws2.merge_cells(f"B{r_cnt2}:F{r_cnt2}")
            ws2[f"B{r_cnt2}"] = v_val or "-"; ws2[f"B{r_cnt2}"].font = font_cell; ws2[f"B{r_cnt2}"].alignment = align_l
            style_r(ws2, f"A{r_cnt2}:F{r_cnt2}", border=thin_border)
            r_cnt2 += 1

        add_sheet_footer(ws2, r_cnt2)
        for col_l, w in [('A', 22), ('B', 20), ('C', 20), ('D', 20), ('E', 20), ('F', 20)]: ws2.column_dimensions[col_l].width = w

        # =============================================================
        # Sheet 3: [3. 공정규격] (Process Specifications)
        # =============================================================
        ws3 = wb.create_sheet(title="3. Process Specs" if is_en else "3. 공정규격")
        add_sheet_header(ws3, "3", "공정규격", "Manufacturing Process Specifications")
        ws3.append([])
        hdr3 = ["No.", "Room", "Process Name", "Equipment", "Process Conditions (Temp/RPM/Time)", "In-Process Control"] if is_en else ["No", "작업실", "공정명", "사용설비", "제조(작업) 조건", "공정검사"]
        ws3.append(hdr3)
        style_r(ws3, "A6:F6", font=font_tbl_hdr, fill=fill_tbl_hdr, align=align_c, border=thin_border)

        r_cnt3 = 7
        if hasattr(self, 'prod_std_proc_tree'):
            for item in self.prod_std_proc_tree.get_children():
                v = self.prod_std_proc_tree.item(item, "values")
                if v:
                    ws3.append([v[0], tr_ps(v[1]), tr_ps(v[2]), tr_ps(v[3]), tr_ps(v[4]), tr_ps(v[5])])
                    style_r(ws3, f"A{r_cnt3}:F{r_cnt3}", font=font_cell, align=align_c, border=thin_border)
                    ws3[f"C{r_cnt3}"].alignment = align_l; ws3[f"E{r_cnt3}"].alignment = align_l
                    r_cnt3 += 1
        if r_cnt3 == 7:
            ws3.append(["1", "Mfg Room" if is_en else "제조실", "Emulsification / Mixing" if is_en else "유화 및 혼합", "Main Tank" if is_en else "메인 유화가마", "75~80℃ / 3000 RPM / 20 min", "Appearance / Viscosity" if is_en else "외관 및 점도 확인"])
            style_r(ws3, f"A{r_cnt3}:F{r_cnt3}", font=font_cell, align=align_c, border=thin_border)
            r_cnt3 += 1

        add_sheet_footer(ws3, r_cnt3)
        for col_l, w in [('A', 8), ('B', 15), ('C', 20), ('D', 20), ('E', 30), ('F', 20)]: ws3.column_dimensions[col_l].width = w

        # =============================================================
        # Sheet 4: [4. 원료성분 기준 및 함량] (Formulations)
        # =============================================================
        ws4 = wb.create_sheet(title="4. Formulations" if is_en else "4. 원료성분기준")
        add_sheet_header(ws4, "4", "원료성분 기준 및 배합비 (100g당)", "Raw Material Specifications & Formulations")
        ws4.append([])
        hdr4 = ["Phase", "No.", "Code", "Raw Material Name", "INCI / Standard Name", "Test Spec", "Ratio (%)", "Remarks"] if is_en else ["Phase", "No", "코드", "원료명", "허가명/INCI", "시험기준", "함량(%)", "비고"]
        ws4.append(hdr4)
        style_r(ws4, "A6:H6", font=font_tbl_hdr, fill=fill_tbl_hdr, align=align_c, border=thin_border)

        r_cnt4 = 7
        if hasattr(self, 'prod_std_ing_tree'):
            for item in self.prod_std_ing_tree.get_children():
                v = self.prod_std_ing_tree.item(item, "values")
                if v:
                    ratio_display = ("Trade Secret" if is_en else "영업비밀") if is_conf else v[6]
                    ws4.append([v[0], v[1], v[2], tr_ps(v[3]), tr_ps(v[4]), tr_ps(v[5]), ratio_display, tr_ps(v[7])])
                    style_r(ws4, f"A{r_cnt4}:H{r_cnt4}", font=font_cell, align=align_c, border=thin_border)
                    ws4[f"D{r_cnt4}"].alignment = align_l; ws4[f"E{r_cnt4}"].alignment = align_l
                    r_cnt4 += 1
        if r_cnt4 == 7:
            def_ratio = ("Trade Secret" if is_en else "영업비밀") if is_conf else "100.0%"
            ws4.append(["A", "1", "RM001", "Water" if is_en else "정제수", "Water / Aqua", "In-House" if is_en else "자사규격", def_ratio, "-"])
            style_r(ws4, f"A{r_cnt4}:H{r_cnt4}", font=font_cell, align=align_c, border=thin_border)
            r_cnt4 += 1

        add_sheet_footer(ws4, r_cnt4)
        for col_l, w in [('A', 8), ('B', 8), ('C', 12), ('D', 22), ('E', 25), ('F', 12), ('G', 12), ('H', 15)]: ws4.column_dimensions[col_l].width = w

        # =============================================================
        # Sheet 5: [5. 계량 지시 및 기록서] (Weighing Order & Record)
        # =============================================================
        ws5 = wb.create_sheet(title="5. Weighing Record" if is_en else "5. 계량지시기록")
        add_sheet_header(ws5, "5", "계량 지시 및 기록서", "Weighing Order & Record")
        ws5.append([])
        
        batch_wt = self.prod_std_entries.get("기준 제조량(kg)", ctk.CTkEntry(self)).get().strip() or "100.0 kg"
        w_date = self.prod_std_entries.get("계량 지시일", ctk.CTkEntry(self)).get().strip() or "-"
        w_man = tr_ps(self.prod_std_entries.get("계량자", ctk.CTkEntry(self)).get().strip()) or author_val
        w_chk = tr_ps(self.prod_std_entries.get("확인자(입회자)", ctk.CTkEntry(self)).get().strip()) or review_val

        ws5.append(["Batch Size" if is_en else "기준제조량", batch_wt, "Order Date" if is_en else "계량지시일", w_date, "Weighed by" if is_en else "계량자", w_man])
        style_r(ws5, "A6:F6", font=font_lbl, align=align_c, fill=fill_sec, border=thin_border)

        ws5.append([])
        hdr5 = ["Vat", "No.", "Code", "Raw Material", "Ratio (%)", "Theoretical (kg)", "Actual (kg)", "Lot No.", "Weighed by", "Checked by", "Remarks"] if is_en else ["Vat", "No", "코드", "원료명", "분량(%)", "이론량(kg)", "계량량(kg)", "Lot No.", "칭량자", "확인자", "특이사항"]
        ws5.append(hdr5)
        style_r(ws5, "A8:K8", font=font_tbl_hdr, fill=fill_tbl_hdr, align=align_c, border=thin_border)

        r_cnt5 = 9
        if hasattr(self, 'prod_std_weigh_tree'):
            for item in self.prod_std_weigh_tree.get_children():
                v = self.prod_std_weigh_tree.item(item, "values")
                if v:
                    w_ratio = ("Trade Secret" if is_en else "영업비밀") if is_conf else v[4]
                    w_th = ("***" if is_en else "비공개") if is_conf else v[5]
                    w_act = ("***" if is_en else "비공개") if is_conf else v[6]
                    ws5.append([v[0], v[1], v[2], tr_ps(v[3]), w_ratio, w_th, w_act, v[7], tr_ps(v[8]), tr_ps(v[9]), tr_ps(v[10])])
                    style_r(ws5, f"A{r_cnt5}:K{r_cnt5}", font=font_cell, align=align_c, border=thin_border)
                    ws5[f"D{r_cnt5}"].alignment = align_l
                    r_cnt5 += 1
        if r_cnt5 == 9:
            ws5.append(["A", "1", "RM001", "Water" if is_en else "정제수", "100.0%", "100.0", "100.0", "LOT-01", w_man, w_chk, "-"])
            style_r(ws5, f"A{r_cnt5}:K{r_cnt5}", font=font_cell, align=align_c, border=thin_border)
            r_cnt5 += 1

        add_sheet_footer(ws5, r_cnt5)
        for col_l, w in [('A', 8), ('B', 8), ('C', 12), ('D', 20), ('E', 10), ('F', 12), ('G', 12), ('H', 15), ('I', 12), ('J', 12), ('K', 15)]: ws5.column_dimensions[col_l].width = w

        # =============================================================
        # Sheet 6: [6. 제조 지시 및 기록서] (BMR Manufacturing SOP)
        # =============================================================
        ws6 = wb.create_sheet(title="6. Manufacturing SOP" if is_en else "6. 제조지시기록서")
        add_sheet_header(ws6, "6", "제조 지시 및 기록서 (BMR)", "Batch Manufacturing Record (BMR SOP)")
        ws6.append([])

        mfg_eq = tr_ps(self.prod_std_entries.get("제조설비명", ctk.CTkEntry(self)).get().strip()) or "Main Tank"
        mfg_yd = tr_ps(self.prod_std_entries.get("수득량 및 수율", ctk.CTkEntry(self)).get().strip()) or "Yield 99.0%"
        mfg_op = tr_ps(self.prod_std_entries.get("제조작업자", ctk.CTkEntry(self)).get().strip()) or author_val
        mfg_mg = tr_ps(self.prod_std_entries.get("제조책임자", ctk.CTkEntry(self)).get().strip()) or appr_val

        ws6.append(["Equipment" if is_en else "제조설비명", mfg_eq, "Yield" if is_en else "수득량/수율", mfg_yd, "Operator" if is_en else "제조작업자", mfg_op])
        style_r(ws6, "A6:F6", font=font_lbl, align=align_c, fill=fill_sec, border=thin_border)

        ws6.append([])
        ws6.merge_cells("A8:F8"); ws6["A8"] = "Standard Manufacturing Operating Procedure (SOP)" if is_en else "세부 제조공정 지시사항 및 유화/혼합 조건 기록 (SOP)"
        ws6["A8"].font = font_lbl; ws6["A8"].fill = fill_sec; style_r(ws6, "A8:F8", border=thin_border)

        mfg_text = self.prod_std_mfg_summary.get("1.0", "end-1c").strip() if hasattr(self, 'prod_std_mfg_summary') else ""
        if not mfg_text: mfg_text = "1. Phase A ingredients are weighed and heated to 75~80℃.\n2. Homogenize at 3,000 RPM for 15 minutes.\n3. Cool down to 35℃ and check in-process quality standards." if is_en else "1. Phase A 원료를 계량하여 75~80℃로 가온 용해한다.\n2. 호모믹서 3,000 RPM으로 15분간 균질 유화한다.\n3. 35℃로 냉각 후 공정 품질 규격을 검사한다."

        ws6.merge_cells("A9:F18"); ws6["A9"] = tr_ps(mfg_text); ws6["A9"].font = font_cell; ws6["A9"].alignment = align_l
        style_r(ws6, "A9:F18", border=thin_border)

        add_sheet_footer(ws6, 19)
        for col_l, w in [('A', 15), ('B', 20), ('C', 15), ('D', 20), ('E', 15), ('F', 20)]: ws6.column_dimensions[col_l].width = w

        # =============================================================
        # Sheet 7: [7. 제품사양 (포장재규격)] (Packaging Specifications)
        # =============================================================
        ws7 = wb.create_sheet(title="7. Packaging Specs" if is_en else "7. 제품사양")
        add_sheet_header(ws7, "7", "제품사양 (포장재규격 포함)", "Packaging Specifications")
        ws7.append([])
        hdr7 = ["No.", "Material Code", "Component / Part Name", "Quantity", "Unit", "Material & Spec Remarks"] if is_en else ["No", "자재코드", "사양명 (자재명)", "수량", "단위", "비고 (재질/규격)"]
        ws7.append(hdr7)
        style_r(ws7, "A6:F6", font=font_tbl_hdr, fill=fill_tbl_hdr, align=align_c, border=thin_border)

        r_cnt7 = 7
        if hasattr(self, 'prod_std_pkg_tree'):
            for item in self.prod_std_pkg_tree.get_children():
                v = self.prod_std_pkg_tree.item(item, "values")
                if v:
                    ws7.append([v[0], v[1], tr_ps(v[2]), v[3], v[4], tr_ps(v[5])])
                    style_r(ws7, f"A{r_cnt7}:F{r_cnt7}", font=font_cell, align=align_c, border=thin_border)
                    ws7[f"C{r_cnt7}"].alignment = align_l; ws7[f"F{r_cnt7}"].alignment = align_l
                    r_cnt7 += 1
        if r_cnt7 == 7:
            ws7.append(["1", "PKG001", "Container / Bottle" if is_en else "본품 용기", "1", "EA", "PET / 100ml"])
            style_r(ws7, f"A{r_cnt7}:F{r_cnt7}", font=font_cell, align=align_c, border=thin_border)
            r_cnt7 += 1

        add_sheet_footer(ws7, r_cnt7)
        for col_l, w in [('A', 8), ('B', 15), ('C', 25), ('D', 10), ('E', 10), ('F', 25)]: ws7.column_dimensions[col_l].width = w

        # =============================================================
        # Sheet 8: [8. 충진·포장 지시 및 기록서] (Packaging Order & Record)
        # =============================================================
        ws8 = wb.create_sheet(title="8. Packaging Record" if is_en else "8. 충진포장기록")
        add_sheet_header(ws8, "8", "충진·포장 지시 및 기록서", "Packaging Order & Record")
        ws8.append([])

        p_vol = self.prod_std_entries.get("표시 용량 (ml/g)", ctk.CTkEntry(self)).get().strip() or "100 ml"
        p_plan = self.prod_std_entries.get("생산 계획량 (EA)", ctk.CTkEntry(self)).get().strip() or "1,000 EA"
        p_crew = tr_ps(self.prod_std_entries.get("충진 작업조", ctk.CTkEntry(self)).get().strip()) or "Packaging Line 1"
        p_eq = tr_ps(self.prod_std_entries.get("포장 설비명", ctk.CTkEntry(self)).get().strip()) or "Auto Filler"
        p_judg = tr_ps(self.prod_std_entries.get("충진·포장 판정", ctk.CTkEntry(self)).get().strip()) or "Pass"
        p_std = tr_ps(self.prod_std_entries.get("포장 수율 기준", ctk.CTkEntry(self)).get().strip()) or "Yield >= 95%"

        ws8.append(["Volume" if is_en else "표시용량", p_vol, "Plan Qty" if is_en else "계획수량", p_plan, "Packaging Line" if is_en else "작업조", p_crew])
        ws8.append(["Equipment" if is_en else "포장설비", p_eq, "Judgement" if is_en else "판정", p_judg, "Yield Criteria" if is_en else "수율기준", p_std])
        style_r(ws8, "A6:F7", font=font_lbl, align=align_c, fill=fill_sec, border=thin_border)

        ws8.append([])
        ws8.merge_cells("A9:F9"); ws8["A9"] = "Packaging Process Log & Reconciliation Summary" if is_en else "충진·포장 공정 점검사항 및 자재/반제품 수불 종합 비고"
        ws8["A9"].font = font_lbl; ws8["A9"].fill = fill_sec; style_r(ws8, "A9:F9", border=thin_border)

        pkg_text = self.prod_std_pkg_log.get("1.0", "end-1c").strip() if hasattr(self, 'prod_std_pkg_log') else ""
        if not pkg_text: pkg_text = "Packaging inspection completed: Container torque, labeling position, and carton seal passed all CGMP criteria." if is_en else "포장 공정 점검 완료: 용기 캡핑 토크, 라벨 부착 상태, 단상자 봉합 상태 모두 CGMP 기준 적합."

        ws8.merge_cells("A10:F16"); ws8["A10"] = tr_ps(pkg_text); ws8["A10"].font = font_cell; ws8["A10"].alignment = align_l
        style_r(ws8, "A10:F16", border=thin_border)

        add_sheet_footer(ws8, 17)
        for col_l, w in [('A', 15), ('B', 20), ('C', 15), ('D', 20), ('E', 15), ('F', 20)]: ws8.column_dimensions[col_l].width = w

        # =============================================================
        # Sheet 9: [9. 제품 규격서] (Product Specifications)
        # =============================================================
        ws9 = wb.create_sheet(title="9. Product Specs" if is_en else "9. 제품규격서")
        add_sheet_header(ws9, "9", "제품 규격서 (14대 완제품 시험기준)", "Product Specifications (14 QC Criteria)")
        ws9.append([])
        hdr9 = ["No.", "Test Parameter", "Specification Criteria", "Test Method", "Remarks"] if is_en else ["No", "시험항목", "규격 (기준)", "시험방법", "비고"]
        ws9.append(hdr9)
        style_r(ws9, "A6:E6", font=font_tbl_hdr, fill=fill_tbl_hdr, align=align_c, border=thin_border)

        r_cnt9 = 7
        if hasattr(self, 'prod_std_spec_tree'):
            for item in self.prod_std_spec_tree.get_children():
                v = self.prod_std_spec_tree.item(item, "values")
                if v:
                    ws9.append([v[0], tr_ps(v[1]), tr_ps(v[2]), tr_ps(v[3]), tr_ps(v[4])])
                    style_r(ws9, f"A{r_cnt9}:E{r_cnt9}", font=font_cell, align=align_c, border=thin_border)
                    ws9[f"B{r_cnt9}"].alignment = align_l; ws9[f"C{r_cnt9}"].alignment = align_l
                    r_cnt9 += 1
        if r_cnt9 == 7:
            spec_defaults = [
                ("1", "Appearance" if is_en else "성상", "Clear gel" if is_en else "무색 투명 겔상", "Visual" if is_en else "육안검사", "-"),
                ("2", "pH (at 25℃)", "5.50 ~ 7.00", "pH Meter", "-"),
                ("3", "Viscosity (at 25℃)", "15,000 ~ 25,000 cPs", "Brookfield LVT", "-"),
                ("4", "Microbiology (Total Colony)", "<= 100 CFU/g", "MFDS Method" if is_en else "식약처 고시", "-")
            ]
            for s_row in spec_defaults:
                ws9.append(list(s_row))
                style_r(ws9, f"A{r_cnt9}:E{r_cnt9}", font=font_cell, align=align_c, border=thin_border)
                r_cnt9 += 1

        add_sheet_footer(ws9, r_cnt9)
        for col_l, w in [('A', 8), ('B', 22), ('C', 28), ('D', 22), ('E', 15)]: ws9.column_dimensions[col_l].width = w

        # =============================================================
        # Sheet 10: [10. 반제품 시험성적서] (Semi-Finished COA)
        # =============================================================
        ws10 = wb.create_sheet(title="10. Semi-Finished COA" if is_en else "10. 반제품성적서")
        add_sheet_header(ws10, "10", "반제품 시험성적서", "Semi-Finished Product Certificate of Analysis")
        ws10.append([])

        semi_mfg_dt = self.prod_std_entries.get("반제품 제조일", ctk.CTkEntry(self)).get().strip() or "-"
        semi_lot = self.prod_std_entries.get("반제품 LOT", ctk.CTkEntry(self)).get().strip() or "-"
        semi_t_dt = self.prod_std_entries.get("반제품 시험일자", ctk.CTkEntry(self)).get().strip() or "-"
        semi_judg = tr_ps(self.prod_std_entries.get("반제품 종합판정", ctk.CTkEntry(self)).get().strip()) or "Pass"

        ws10.append(["Mfg Date" if is_en else "제조일자", semi_mfg_dt, "Semi-Lot No." if is_en else "반제품 LOT", semi_lot, "Judgement" if is_en else "판정", semi_judg])
        style_r(ws10, "A6:F6", font=font_lbl, align=align_c, fill=fill_sec, border=thin_border)

        ws10.append([])
        hdr10 = ["No.", "Test Parameter", "Specification Criteria", "Test Result", "Remarks"] if is_en else ["No", "시험항목", "시험기준", "시험결과", "비고"]
        ws10.append(hdr10)
        style_r(ws10, "A8:E8", font=font_tbl_hdr, fill=fill_tbl_hdr, align=align_c, border=thin_border)

        r_cnt10 = 9
        if hasattr(self, 'prod_std_semi_tree'):
            for item in self.prod_std_semi_tree.get_children():
                v = self.prod_std_semi_tree.item(item, "values")
                if v:
                    ws10.append([v[0], tr_ps(v[1]), tr_ps(v[2]), tr_ps(v[3]), tr_ps(v[4])])
                    style_r(ws10, f"A{r_cnt10}:E{r_cnt10}", font=font_cell, align=align_c, border=thin_border)
                    r_cnt10 += 1
        if r_cnt10 == 9:
            ws10.append(["1", "Appearance" if is_en else "성상", "Clear gel" if is_en else "무색 투명 겔상", "Complies" if is_en else "적합", "-"])
            style_r(ws10, f"A{r_cnt10}:E{r_cnt10}", font=font_cell, align=align_c, border=thin_border)
            r_cnt10 += 1

        add_sheet_footer(ws10, r_cnt10)
        for col_l, w in [('A', 8), ('B', 22), ('C', 28), ('D', 22), ('E', 15)]: ws10.column_dimensions[col_l].width = w

        # =============================================================
        # Sheet 11: [11. 완제품 시험성적서] (Finished Product COA)
        # =============================================================
        ws11 = wb.create_sheet(title="11. Finished COA" if is_en else "11. 완제품성적서")
        add_sheet_header(ws11, "11", "완제품 시험성적서", "Finished Product Certificate of Analysis")
        ws11.append([])

        fin_pkg_dt = self.prod_std_entries.get("완제품 포장일자", ctk.CTkEntry(self)).get().strip() or "-"
        fin_lot = self.prod_std_entries.get("완제품 LOT", ctk.CTkEntry(self)).get().strip() or "-"
        fin_t_dt = self.prod_std_entries.get("완제품 시험일자", ctk.CTkEntry(self)).get().strip() or "-"
        fin_judg = tr_ps(self.prod_std_entries.get("완제품 종합판정", ctk.CTkEntry(self)).get().strip()) or "Pass"

        ws11.append(["Pkg Date" if is_en else "포장일자", fin_pkg_dt, "Finished Lot No." if is_en else "완제품 LOT", fin_lot, "Judgement" if is_en else "판정", fin_judg])
        style_r(ws11, "A6:F6", font=font_lbl, align=align_c, fill=fill_sec, border=thin_border)

        ws11.append([])
        hdr11 = ["No.", "Test Parameter", "Specification Criteria", "Test Result", "Remarks"] if is_en else ["No", "시험항목", "시험기준", "시험결과", "비고"]
        ws11.append(hdr11)
        style_r(ws11, "A8:E8", font=font_tbl_hdr, fill=fill_tbl_hdr, align=align_c, border=thin_border)

        r_cnt11 = 9
        if hasattr(self, 'prod_std_fin_tree'):
            for item in self.prod_std_fin_tree.get_children():
                v = self.prod_std_fin_tree.item(item, "values")
                if v:
                    ws11.append([v[0], tr_ps(v[1]), tr_ps(v[2]), tr_ps(v[3]), tr_ps(v[4])])
                    style_r(ws11, f"A{r_cnt11}:E{r_cnt11}", font=font_cell, align=align_c, border=thin_border)
                    r_cnt11 += 1
        if r_cnt11 == 9:
            ws11.append(["1", "Appearance" if is_en else "성상", "Clear gel" if is_en else "무색 투명 겔상", "Complies" if is_en else "적합", "-"])
            style_r(ws11, f"A{r_cnt11}:E{r_cnt11}", font=font_cell, align=align_c, border=thin_border)
            r_cnt11 += 1

        add_sheet_footer(ws11, r_cnt11)
        for col_l, w in [('A', 8), ('B', 22), ('C', 28), ('D', 22), ('E', 15)]: ws11.column_dimensions[col_l].width = w

        # =============================================================
        # Sheet 12: [12. 제조 및 품질관리 시설·기구] (Facilities & Equipment)
        # =============================================================
        ws12 = wb.create_sheet(title="12. Facilities" if is_en else "12. 시설기구")
        add_sheet_header(ws12, "12", "제조 및 품질관리에 필요한 시설 및 기구", "Facilities and Equipment for Manufacturing & Quality Control")
        ws12.append([])
        hdr12 = ["No.", "Manufacturing Facility & Machinery", "Filling & Packaging Equipment", "QC Testing Instruments"] if is_en else ["No", "제조시설 기구명", "충진 및 포장 설비명", "품질관리 시험기구명"]
        ws12.append(hdr12)
        style_r(ws12, "A6:D6", font=font_tbl_hdr, fill=fill_tbl_hdr, align=align_c, border=thin_border)

        r_cnt12 = 7
        if hasattr(self, 'prod_std_fac_tree'):
            for item in self.prod_std_fac_tree.get_children():
                v = self.prod_std_fac_tree.item(item, "values")
                if v:
                    ws12.append([v[0], tr_ps(v[1]), tr_ps(v[2]), tr_ps(v[3])])
                    style_r(ws12, f"A{r_cnt12}:D{r_cnt12}", font=font_cell, align=align_c, border=thin_border)
                    r_cnt12 += 1

        add_sheet_footer(ws12, r_cnt12)
        for col_l, w in [('A', 8), ('B', 30), ('C', 30), ('D', 30)]: ws12.column_dimensions[col_l].width = w

        # 저장 다이얼로그
        f_prefix = f"Product_Standard_Cover_12Sections_{p_name_en}" if is_en else f"제품표준서_공식표지및12대목차_{p_name_ko}"
        file_path_dlg = fd.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"{f_prefix}.xlsx",
            title="Save Product Standard (Cover + 12 Sections Full Edition)" if is_en else "제품표준서 (공식 표지 + 12대 전 목차 표준 엑셀) 저장"
        )
        if file_path_dlg:
            wb.save(file_path_dlg)
            msg_ok = f"Product Standard (Cover + 12 Sections Full Edition) exported successfully:\n{file_path_dlg}" if is_en else f"식약처/CGMP 공식 표지 및 12대 목차 전수(13개 시트)가 포함된 제품표준서가 생성되었습니다:\n{file_path_dlg}"
            messagebox.showinfo("Export Complete" if is_en else "저장 완료", msg_ok, parent=self)
            try: os.startfile(os.path.abspath(file_path_dlg))
            except Exception: pass

    def setup_batch_manufacturing_tab(self, tab_frame):
        tab_frame.grid_columnconfigure(0, weight=1)
        tab_frame.grid_rowconfigure(0, weight=1)

        scroll = ctk.CTkScrollableFrame(tab_frame, label_text="제조지시 및 기록서 (Batch Manufacturing Record - CGMP BMR 심층 기록)")
        scroll.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        scroll.grid_columnconfigure(0, weight=1)

        self.bmr_entries = {}
        self.current_bmr_id = None

        top_bar = ctk.CTkFrame(scroll, fg_color="transparent")
        top_bar.grid(row=0, column=0, sticky="ew", padx=10, pady=(5, 10))
        top_bar.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(top_bar, text="제조기록 이력:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=5, pady=5)
        self.bmr_picker = ctk.CTkComboBox(top_bar, values=["-- 저장된 제조기록서 선택 --"], width=300)
        self.bmr_picker.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ctk.CTkButton(top_bar, text="불러오기", width=80, command=self.load_selected_bmr).grid(row=0, column=2, padx=5, pady=5)
        ctk.CTkButton(top_bar, text="새로작성", width=80, fg_color="gray50", command=self.clear_bmr_form).grid(row=0, column=3, padx=5, pady=5)
        ctk.CTkButton(top_bar, text="삭제", width=70, fg_color="#D32F2F", command=self.delete_bmr).grid(row=0, column=4, padx=5, pady=5)

        # BMR 심층 서브 탭뷰 (지시서 / 칭량검증표 / 공정로그 / IPC검사 / 수율정산)
        self.bmr_subtabs = ctk.CTkTabview(scroll, height=450)
        self.bmr_subtabs.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)

        tab_b1 = self.bmr_subtabs.add("1. 제조지시 및 배치개요")
        tab_b2 = self.bmr_subtabs.add("2. 원료 칭량 및 교차검증표")
        tab_b3 = self.bmr_subtabs.add("3. 단계별 제조공정 일지")
        tab_b4 = self.bmr_subtabs.add("4. 공정검사(IPC) 및 수율정산")

        # BMR 1: 제조지시 및 배치개요
        tab_b1.grid_columnconfigure((1, 3, 5), weight=1)
        b1_fields = [
            ("제품명", 0, 0), ("제조번호(LOT)", 0, 2), ("제조일자", 0, 4),
            ("제조지시량(kg)", 1, 0), ("제조설비(Tank)", 1, 2), ("제조지시자", 1, 4),
            ("제조작업자", 2, 0), ("제조책임자(승인)", 2, 2), ("배치 완료상태", 2, 4)
        ]
        for lbl, r, c in b1_fields:
            ctk.CTkLabel(tab_b1, text=lbl, font=ctk.CTkFont(weight="bold")).grid(row=r, column=c, padx=8, pady=8, sticky="w")
            if lbl == "배치 완료상태":
                cb = ctk.CTkComboBox(tab_b1, values=["제조 진행중", "공정완료 (적합)", "보류 / 재작업", "불합격 (폐기)"])
                cb.set("공정완료 (적합)")
                cb.grid(row=r, column=c+1, padx=8, pady=8, sticky="ew")
                self.bmr_entries[lbl] = cb
            else:
                ent = ctk.CTkEntry(tab_b1)
                if lbl == "제조일자": ent.insert(0, datetime.now().strftime("%Y-%m-%d"))
                elif lbl == "제조지시량(kg)": ent.insert(0, "500.00")
                elif lbl == "제조설비(Tank)": ent.insert(0, "MAIN-TANK-01 (1,000L 진공유화기)")
                ent.grid(row=r, column=c+1, padx=8, pady=8, sticky="ew")
                self.bmr_entries[lbl] = ent

        # BMR 2: 원료 칭량 및 교차검증표
        tab_b2.grid_columnconfigure(0, weight=1)
        tab_b2.grid_rowconfigure(1, weight=1)
        ctk.CTkLabel(tab_b2, text="원료별 처방배합비(%) -> 지시량(kg) vs 실측량(kg) vs 칭량 오차율(%) vs 칭량자/확인자 2중 서명:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.bmr_weighing_textbox = ctk.CTkTextbox(tab_b2, height=350)
        self.bmr_weighing_textbox.grid(row=1, column=0, padx=10, pady=5, sticky="nsew")
        self.bmr_weighing_textbox.insert("1.0", "• [Phase A] 1. 정제수: 지시 350.00kg | 실측 350.02kg (오차 +0.01% - 적합) | 칭량자: 홍길동, 확인자: 김책임\n• [Phase A] 2. 글리세린: 지시 25.00kg | 실측 25.00kg (오차 0.00% - 적합) | 칭량자: 홍길동, 확인자: 김책임\n• [Phase A] 3. 1,2-헥산다이올: 지시 10.00kg | 실측 10.00kg (적합)\n• [Phase B] 4. 세테아릴알코올: 지시 15.00kg | 실측 15.01kg (오차 +0.07% - 적합)\n• [Phase B] 5. 카프릴릭/카프릭트라이글리세라이드: 지시 40.00kg | 실측 40.00kg (적합)\n• [Phase B] 6. 폴리솔베이트60: 지시 10.00kg | 실측 10.00kg (적합)\n• [Phase C] 7. 아데노신(주성분): 지시 0.200kg (200.0g) | 정밀저울 실측 200.1g (적합) | 칭량자: 홍길동, 확인자: 김책임\n• [Phase C] 8. 향료 및 보존보조제: 지시 2.50kg | 실측 2.50kg (적합)\n=======================================================\n[총 칭량 합계] 지시량: 500.00kg | 실칭량 합계: 500.03kg (칭량 공정 적합 완료)")

        # BMR 3: 단계별 제조공정 일지
        tab_b3.grid_columnconfigure(0, weight=1)
        tab_b3.grid_rowconfigure(1, weight=1)
        ctk.CTkLabel(tab_b3, text="공정별 실시간 작업 로그 (투입시각, 설정온도/실측온도, 교반 RPM, 진공도, 냉각, 여과):", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.bmr_log_textbox = ctk.CTkTextbox(tab_b3, height=350)
        self.bmr_log_textbox.grid(row=1, column=0, padx=10, pady=5, sticky="nsew")
        self.bmr_log_textbox.insert("1.0", "[09:00~09:30] 제조 가마 CIP/SIP 세척 및 청결 상태 확인 (검사자: 홍길동 - 적합)\n[09:30~10:15] Phase A 수상 탱크 투입 및 승온 (설정: 80℃, 실측: 78.5℃ 도달 완전 용해)\n[10:00~10:20] Phase B 유상 서브탱크 투입 및 가열 용해 (78℃ 투명 액상 확인)\n[10:30~10:45] 유화 공정: 메인 호모믹서 3,600 RPM 가동 (5분간 고속 유화) / 아지믹서 25 RPM\n[10:45~11:00] 진공 탈포: 진공 펌프 가동 (진공도 -0.082 MPa, 기포 완전 제거 확인)\n[11:00~11:40] 냉각 공정: 냉각수 순환 (45℃ 도달 시 Phase C 주성분 및 첨가제 투입 균일 교반)\n[11:40~12:00] 최종 여과: 100 mesh 여과망 통과 후 완제품 드럼 이송 (이물 불검출 확인)")

        # BMR 4: 공정검사(IPC) 및 수율정산
        tab_b4.grid_columnconfigure((1, 3), weight=1)
        b4_fields = [
            ("실생산량(kg)", 0, 0), ("공정수율(%)", 0, 2),
            ("제조중 pH(25℃)", 1, 0), ("제조중 점도(cps)", 1, 2),
            ("제조중 비중", 2, 0), ("외관/유화입도 IPC", 2, 2),
            ("손실량(Loss kg)", 3, 0), ("종합 품질 판정", 3, 2)
        ]
        b4_defaults = {
            "실생산량(kg)": "496.50",
            "공정수율(%)": "99.30%",
            "제조중 pH(25℃)": "6.12 (적합)",
            "제조중 점도(cps)": "38,500 cps (적합)",
            "제조중 비중": "0.998 (적합)",
            "외관/유화입도 IPC": "현미경 입도 균일, 층분리 없음 (적합)",
            "손실량(Loss kg)": "3.50 kg (배관 잔류 및 샘플링 손실)",
            "종합 품질 판정": "적합 (포장 공정 이관 승인)"
        }
        for lbl, r, c in b4_fields:
            ctk.CTkLabel(tab_b4, text=lbl, font=ctk.CTkFont(weight="bold")).grid(row=r, column=c, padx=10, pady=8, sticky="w")
            ent = ctk.CTkEntry(tab_b4)
            if lbl in b4_defaults: ent.insert(0, b4_defaults[lbl])
            ent.grid(row=r, column=c+1, padx=10, pady=8, sticky="ew")
            self.bmr_entries[lbl] = ent

        btn_bar = ctk.CTkFrame(scroll, fg_color="transparent")
        btn_bar.grid(row=2, column=0, sticky="e", padx=10, pady=15)
        ctk.CTkButton(btn_bar, text="💾 제조기록 DB 저장", width=120, command=self.save_bmr_to_db).pack(side="left", padx=5)
        ctk.CTkButton(btn_bar, text="📊 국문 BMR 엑셀 (4대 시트)", width=170, fg_color="#2E7D32", hover_color="#1B5E20", command=lambda: self.export_bmr_to_excel("ko")).pack(side="left", padx=5)
        ctk.CTkButton(btn_bar, text="🌐 영문 BMR 엑셀 (4 Sheets)", width=170, fg_color="#1565C0", hover_color="#0D47A1", command=lambda: self.export_bmr_to_excel("en")).pack(side="left", padx=5)

        self.refresh_bmr_list()

    def refresh_bmr_list(self):
        try:
            session = db_manager.get_session()
            recs = session.query(BatchManufacturingRecord).order_by(BatchManufacturingRecord.created_at.desc()).limit(50).all()
            vals = [f"{r.id} | {r.product_name} | LOT:{r.batch_no} | {r.batch_size_kg or 0}kg" for r in recs]
            self.bmr_picker.configure(values=vals if vals else ["-- 저장된 제조기록서 없음 --"])
            if vals: self.bmr_picker.set(vals[0])
            session.close()
        except Exception as e:
            print(f"[경고] 제조기록서 목록 로드 실패: {e}")

    def clear_bmr_form(self):
        self.current_bmr_id = None
        for k, e in self.bmr_entries.items():
            if isinstance(e, ctk.CTkComboBox):
                vals = e.cget("values")
                if vals: e.set(vals[0])
            else:
                e.delete(0, "end")
        if "제조일자" in self.bmr_entries: self.bmr_entries["제조일자"].insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.bmr_log_textbox.delete("1.0", "end")
        self.bmr_weighing_textbox.delete("1.0", "end")

    def load_selected_bmr(self):
        sel = self.bmr_picker.get()
        if not sel or '|' not in sel: return
        rec_id = int(sel.split('|')[0].strip())
        session = db_manager.get_session()
        try:
            r = session.query(BatchManufacturingRecord).get(rec_id)
            if not r: return
            self.current_bmr_id = r.id
            if "제품명" in self.bmr_entries: self.bmr_entries["제품명"].delete(0, "end"); self.bmr_entries["제품명"].insert(0, r.product_name or "")
            if "제조번호(LOT)" in self.bmr_entries: self.bmr_entries["제조번호(LOT)"].delete(0, "end"); self.bmr_entries["제조번호(LOT)"].insert(0, r.batch_no or "")
            if "제조일자" in self.bmr_entries: self.bmr_entries["제조일자"].delete(0, "end"); self.bmr_entries["제조일자"].insert(0, str(r.manufacture_date or ""))
            if "제조지시량(kg)" in self.bmr_entries: self.bmr_entries["제조지시량(kg)"].delete(0, "end"); self.bmr_entries["제조지시량(kg)"].insert(0, str(r.batch_size_kg or ""))
            if "실생산량(kg)" in self.bmr_entries: self.bmr_entries["실생산량(kg)"].delete(0, "end"); self.bmr_entries["실생산량(kg)"].insert(0, str(r.actual_yield_kg or ""))
            if "공정수율(%)" in self.bmr_entries: self.bmr_entries["공정수율(%)"].delete(0, "end"); self.bmr_entries["공정수율(%)"].insert(0, f"{r.yield_rate or 99.0}%")
            if "제조설비(Tank)" in self.bmr_entries: self.bmr_entries["제조설비(Tank)"].delete(0, "end"); self.bmr_entries["제조설비(Tank)"].insert(0, r.tank_no or "")
            if "제조작업자" in self.bmr_entries: self.bmr_entries["제조작업자"].delete(0, "end"); self.bmr_entries["제조작업자"].insert(0, r.operator_name or "")
            if "제조책임자(승인)" in self.bmr_entries: self.bmr_entries["제조책임자(승인)"].delete(0, "end"); self.bmr_entries["제조책임자(승인)"].insert(0, r.supervisor_name or "")
            if "제조중 pH(25℃)" in self.bmr_entries: self.bmr_entries["제조중 pH(25℃)"].delete(0, "end"); self.bmr_entries["제조중 pH(25℃)"].insert(0, r.ph_result or "")
            if "제조중 점도(cps)" in self.bmr_entries: self.bmr_entries["제조중 점도(cps)"].delete(0, "end"); self.bmr_entries["제조중 점도(cps)"].insert(0, r.viscosity_result or "")
            
            self.bmr_log_textbox.delete("1.0", "end"); self.bmr_log_textbox.insert("1.0", r.process_log or "")
            self.bmr_weighing_textbox.delete("1.0", "end"); self.bmr_weighing_textbox.insert("1.0", r.weighing_records_json or "")
            messagebox.showinfo("불러오기 완료", f"제조기록서 '{r.product_name}'를 불러왔습니다.", parent=self)
        finally:
            session.close()

    def save_bmr_to_db(self):
        p_name = self.bmr_entries.get("제품명", ctk.CTkEntry(self)).get().strip()
        b_no = self.bmr_entries.get("제조번호(LOT)", ctk.CTkEntry(self)).get().strip()
        if not p_name or not b_no:
            messagebox.showwarning("입력 필요", "제품명과 제조번호(LOT)는 필수입니다.", parent=self); return
        session = db_manager.get_session()
        try:
            r = session.query(BatchManufacturingRecord).get(self.current_bmr_id) if self.current_bmr_id else BatchManufacturingRecord()
            r.product_name = p_name
            r.batch_no = b_no
            try: r.manufacture_date = datetime.strptime(self.bmr_entries.get("제조일자", ctk.CTkEntry(self)).get().strip(), "%Y-%m-%d").date()
            except: r.manufacture_date = None
            try: r.batch_size_kg = float(self.bmr_entries.get("제조지시량(kg)", ctk.CTkEntry(self)).get().strip())
            except: r.batch_size_kg = 0.0
            try: r.actual_yield_kg = float(self.bmr_entries.get("실생산량(kg)", ctk.CTkEntry(self)).get().strip())
            except: r.actual_yield_kg = 0.0
            
            if r.batch_size_kg > 0 and r.actual_yield_kg > 0:
                r.yield_rate = round((r.actual_yield_kg / r.batch_size_kg) * 100, 2)
            else:
                r.yield_rate = 100.0

            r.tank_no = self.bmr_entries.get("제조설비(Tank)", ctk.CTkEntry(self)).get().strip()
            r.operator_name = self.bmr_entries.get("제조작업자", ctk.CTkEntry(self)).get().strip()
            r.supervisor_name = self.bmr_entries.get("제조책임자(승인)", ctk.CTkEntry(self)).get().strip()
            r.ph_result = self.bmr_entries.get("제조중 pH(25℃)", ctk.CTkEntry(self)).get().strip()
            r.viscosity_result = self.bmr_entries.get("제조중 점도(cps)", ctk.CTkEntry(self)).get().strip()
            if "배치 완료상태" in self.bmr_entries: r.overall_status = self.bmr_entries["배치 완료상태"].get().strip()
            r.process_log = self.bmr_log_textbox.get("1.0", "end-1c").strip()
            r.weighing_records_json = self.bmr_weighing_textbox.get("1.0", "end-1c").strip()
            
            if not self.current_bmr_id: session.add(r)
            session.commit()
            self.current_bmr_id = r.id
            self.refresh_bmr_list()
            messagebox.showinfo("저장 완료", "제조지시 및 기록서가 DB에 성공적으로 저장되었습니다.", parent=self)
        except Exception as e:
            session.rollback(); messagebox.showerror("오류", f"저장 실패: {e}", parent=self)
        finally:
            session.close()

    def delete_bmr(self):
        if not self.current_bmr_id: return
        if not messagebox.askyesno("삭제 확인", "선택한 제조기록서를 삭제하시겠습니까?", parent=self): return
        session = db_manager.get_session()
        try:
            r = session.query(BatchManufacturingRecord).get(self.current_bmr_id)
            if r:
                session.delete(r); session.commit()
                self.clear_bmr_form(); self.refresh_bmr_list()
                messagebox.showinfo("완료", "삭제되었습니다.", parent=self)
        finally:
            session.close()

    def export_bmr_to_excel(self, lang="ko"):
        is_en = (lang == "en")
        p_name = self.bmr_entries.get("제품명", ctk.CTkEntry(self)).get().strip() or ("BMR_Product" if is_en else "제조기록서")
        wb = Workbook()
        
        # 시트 1: 배치 제조 지시 및 총괄 (Order & Summary)
        ws1 = wb.active
        ws1.title = "1.Batch_Order" if is_en else "1.제조지시및총괄"
        ws1.merge_cells("A1:E2")
        ws1["A1"] = f"BATCH MANUFACTURING RECORD (BMR) - {p_name}" if is_en else f"제조지시 및 기록서 [제1부 배치개요] - {p_name}"
        ws1["A1"].font = Font(name="맑은 고딕", size=15, bold=True, color="1F497D")
        ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws1.append([])
        for k in ["제품명", "제조번호(LOT)", "제조일자", "제조지시량(kg)", "제조설비(Tank)", "제조지시자", "제조작업자", "제조책임자(승인)", "배치 완료상태"]:
            if k in self.bmr_entries: ws1.append([k, self.bmr_entries[k].get(), "", ""])

        # 시트 2: 원료 칭량 검증표 (Weighing Verification)
        ws2 = wb.create_sheet(title="2.Weighing_Sheet" if is_en else "2.원료칭량검증표")
        ws2.append(["[원료 칭량 지시량 vs 실측량 vs 교차 검증표 / Weighing Verification]"])
        for line in self.bmr_weighing_textbox.get("1.0", "end-1c").split('\n'): ws2.append([line])

        # 시트 3: 공정 작업 일지 (Processing Log)
        ws3 = wb.create_sheet(title="3.Process_Log" if is_en else "3.공정작업일지")
        ws3.append(["[제조 공정 실시간 작업 일지 / Real-time Processing Log]"])
        for line in self.bmr_log_textbox.get("1.0", "end-1c").split('\n'): ws3.append([line])

        # 시트 4: 공정검사(IPC) 및 수율 정산 (IPC & Yield)
        ws4 = wb.create_sheet(title="4.IPC_Yield" if is_en else "4.공정검사및수율")
        ws4.append(["검사 및 정산 항목", "결과값 / 측정치", "기준 규격", "판정"])
        for k in ["실생산량(kg)", "공정수율(%)", "제조중 pH(25℃)", "제조중 점도(cps)", "제조중 비중", "외관/유화입도 IPC", "손실량(Loss kg)", "종합 품질 판정"]:
            if k in self.bmr_entries: ws4.append([k, self.bmr_entries[k].get(), "BMR 기준치", "적합"])

        f_prefix = f"BMR_EN_4Sheets_{p_name}" if is_en else f"제조기록서_BMR4대시트_{p_name}"
        file_path = fd.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], initialfile=f"{f_prefix}.xlsx", title=f"제조기록서 (4대 시트) 엑셀 저장")
        if file_path:
            wb.save(file_path)
            messagebox.showinfo("완료", f"제조기록서 (4개 시트)가 저장되었습니다:\n{file_path}", parent=self)
            try: os.startfile(os.path.abspath(file_path))
            except: pass

    # =========================================================================
    # 0. 물질안전보건자료 (Material Safety Data Sheet - MSDS) 16대 표준 파트 전수 심층 구조
    # =========================================================================
    def setup_msds_tab(self, tab_frame):
        """고용노동부고시 제2023-9호 16대 법정 MSDS 초고속 온디맨드(Lazy Loading) 슬라이더 시스템"""
        tab_frame.grid_columnconfigure(0, weight=1)
        tab_frame.grid_rowconfigure(0, weight=1)

        scroll = ctk.CTkScrollableFrame(tab_frame, label_text="물질안전보건자료 (MSDS - 고용노동부고시 제2023-9호 16대 법정 기준 전수 관리)")
        scroll.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        scroll.grid_columnconfigure(0, weight=1)

        self.msds_entries = {}
        self.current_msds_id = None
        self.current_msds_section_idx = 0

        # 상단 문서 이력 툴바
        top_bar = ctk.CTkFrame(scroll, fg_color="transparent")
        top_bar.grid(row=0, column=0, sticky="ew", padx=10, pady=(5, 10))
        top_bar.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(top_bar, text="MSDS 이력:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=5, pady=5)
        self.msds_picker = ctk.CTkComboBox(top_bar, values=["-- 저장된 MSDS 선택 --"], width=300)
        self.msds_picker.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ctk.CTkButton(top_bar, text="불러오기", width=80, command=self.load_selected_msds).grid(row=0, column=2, padx=5, pady=5)
        ctk.CTkButton(top_bar, text="새로작성", width=80, fg_color="gray50", command=self.clear_msds_form).grid(row=0, column=3, padx=5, pady=5)
        ctk.CTkButton(top_bar, text="삭제", width=70, fg_color="#D32F2F", command=self.delete_msds).grid(row=0, column=4, padx=5, pady=5)

        # 16대 섹션 슬라이드 네비게이터 바
        nav_card = ctk.CTkFrame(scroll, fg_color=("gray90", "gray17"), corner_radius=8, border_width=1, border_color=("gray75", "gray30"))
        nav_card.grid(row=1, column=0, sticky="ew", padx=10, pady=5)
        nav_card.grid_columnconfigure(1, weight=1)

        self.msds_section_titles = [
            "1. 화학제품과 회사에 관한 정보 (Section 1/16)",
            "2. 유해성·위험성 (Section 2/16)",
            "3. 구성성분의 명칭 및 함유량 (Section 3/16)",
            "4. 응급조치 요령 (Section 4/16)",
            "5. 폭발·화재시 대처방법 (Section 5/16)",
            "6. 누출 사고 시 대처방법 (Section 6/16)",
            "7. 취급 및 저장방법 (Section 7/16)",
            "8. 노출방지 및 개인보호구 (Section 8/16)",
            "9. 물리화학적 특성 (Section 9/16)",
            "10. 안정성 및 반응성 (Section 10/16)",
            "11. 독성에 관한 정보 (Section 11/16)",
            "12. 환경에 미치는 영향 (Section 12/16)",
            "13. 폐기시 주의사항 (Section 13/16)",
            "14. 운송에 필요한 정보 (Section 14/16)",
            "15. 법적 규제현황 (Section 15/16)",
            "16. 그 밖의 참고사항 (Section 16/16)"
        ]

        self.btn_prev_sec = ctk.CTkButton(
            nav_card, text="◀ 이전 섹션", width=110, height=34,
            fg_color="#455A64", hover_color="#37474F",
            font=ctk.CTkFont(weight="bold", size=12),
            command=self.prev_msds_section
        )
        self.btn_prev_sec.grid(row=0, column=0, padx=8, pady=6)

        self.msds_section_selector = ctk.CTkComboBox(
            nav_card, values=self.msds_section_titles, height=34,
            font=ctk.CTkFont(weight="bold", size=13), dropdown_font=ctk.CTkFont(size=12),
            command=self.on_msds_section_selected
        )
        self.msds_section_selector.set(self.msds_section_titles[0])
        self.msds_section_selector.grid(row=0, column=1, sticky="ew", padx=8, pady=6)

        self.btn_next_sec = ctk.CTkButton(
            nav_card, text="다음 섹션 ▶", width=110, height=34,
            fg_color="#1565C0", hover_color="#0D47A1",
            font=ctk.CTkFont(weight="bold", size=12),
            command=self.next_msds_section
        )
        self.btn_next_sec.grid(row=0, column=2, padx=8, pady=6)

        # 16대 섹션 본문 컨테이너 프레임
        self.msds_content_container = ctk.CTkFrame(scroll, fg_color="transparent")
        self.msds_content_container.grid(row=2, column=0, sticky="nsew", padx=10, pady=5)
        self.msds_content_container.grid_columnconfigure(0, weight=1)
        self.msds_content_container.grid_rowconfigure(0, weight=1)

        # 온디맨드 렌더링용 16개 슬롯
        self.msds_section_frames = [None] * 16
        self.msds_section_builders = [
            self._build_msds_sec_0, self._build_msds_sec_1, self._build_msds_sec_2,
            self._build_msds_sec_3, self._build_msds_sec_4, self._build_msds_sec_5,
            self._build_msds_sec_6, self._build_msds_sec_7, self._build_msds_sec_8,
            self._build_msds_sec_9, self._build_msds_sec_10, self._build_msds_sec_11,
            self._build_msds_sec_12, self._build_msds_sec_13, self._build_msds_sec_14,
            self._build_msds_sec_15
        ]

        # 첫 번째 섹션만 즉시 빌드 (0.05초 초고속 렌더링)
        self.show_msds_section(0)

        # 하단 저장 및 엑셀 출력 버튼 바
        btn_bar = ctk.CTkFrame(scroll, fg_color="transparent")
        btn_bar.grid(row=3, column=0, sticky="e", padx=10, pady=15)
        ctk.CTkButton(btn_bar, text="💾 MSDS DB 저장", width=140, command=self.save_msds_to_db).pack(side="left", padx=5)
        ctk.CTkButton(btn_bar, text="📊 국문 MSDS 엑셀 (16대 전 섹션)", width=200, fg_color="#2E7D32", hover_color="#1B5E20", command=lambda: self.export_msds_to_excel("ko")).pack(side="left", padx=5)
        ctk.CTkButton(btn_bar, text="🌐 영문 MSDS 엑셀 (16 Sections)", width=200, fg_color="#1565C0", hover_color="#0D47A1", command=lambda: self.export_msds_to_excel("en")).pack(side="left", padx=5)

        self.refresh_msds_list()

    def _create_msds_base_frame(self, sec_title_header):
        sec_frame = ctk.CTkScrollableFrame(
            self.msds_content_container, height=480, label_text=sec_title_header,
            label_font=ctk.CTkFont(weight="bold", size=13)
        )
        sec_frame.grid_columnconfigure(1, weight=1)
        sec_frame.grid_columnconfigure(3, weight=1)
        return sec_frame

    def _append_multi_msds(self, selected_val, target_entry):
        if not selected_val or selected_val.startswith("➕") or selected_val == "직접 입력": return
        cur = target_entry.get().strip()
        if not cur: target_entry.insert(0, selected_val)
        else:
            if selected_val in cur: return
            sep = " / " if ("/" in cur or "/" in selected_val or len(cur) > 25) else ", "
            target_entry.delete(0, "end"); target_entry.insert(0, f"{cur}{sep}{selected_val}")

    def _remove_multi_msds(self, target_entry):
        cur = target_entry.get().strip()
        if not cur: return
        for sep in [" / ", ", ", "/"]:
            if sep in cur:
                items = [it.strip() for it in cur.split(sep) if it.strip()]
                if len(items) > 1:
                    items.pop()
                    target_entry.delete(0, "end")
                    target_entry.insert(0, " / ".join(items) if " / " in cur or len(cur) > 30 else ", ".join(items))
                    return
        target_entry.delete(0, "end")

    def _setup_msds_fields(self, parent_frame, fields, combos_dict):
        for item in fields:
            if len(item) == 3: lbl, r, c = item; col_span = 1
            elif len(item) == 4: lbl, r, c, col_span = item
            else: continue

            ctk.CTkLabel(parent_frame, text=lbl, font=ctk.CTkFont(size=11, weight="bold")).grid(
                row=r, column=c, padx=(10, 6), pady=6, sticky="w"
            )
            if lbl in combos_dict:
                box_f = ctk.CTkFrame(parent_frame, fg_color="transparent")
                box_f.grid(row=r, column=c+1, columnspan=col_span, padx=6, pady=6, sticky="ew")
                box_f.grid_columnconfigure(0, weight=1)

                ent = ctk.CTkEntry(box_f, placeholder_text=f"{lbl} 입력 (우측 ➕에서 선택 추가)", height=30)
                ent.grid(row=0, column=0, sticky="ew", padx=(0, 4))
                self.msds_entries[lbl] = ent

                cb_vals = ["➕ 항목 추가 선택..."] + combos_dict[lbl]
                cb = ctk.CTkComboBox(
                    box_f, values=cb_vals, width=140, height=30,
                    command=lambda choice, target=ent: (
                        self._append_multi_msds(choice, target),
                        cb.set("➕ 항목 추가 선택...")
                    )
                )
                cb.set("➕ 항목 추가 선택...")
                cb.grid(row=0, column=1, padx=(0, 4))

                btn_del = ctk.CTkButton(
                    box_f, text="선택 제거", width=65, height=30,
                    fg_color="#C62828", hover_color="#B71C1C",
                    font=ctk.CTkFont(size=11, weight="bold"),
                    command=lambda target=ent: self._remove_multi_msds(target)
                )
                btn_del.grid(row=0, column=2)
            else:
                ent = ctk.CTkEntry(parent_frame, placeholder_text=f"{lbl} 입력", height=30)
                ent.grid(row=r, column=c+1, columnspan=col_span, padx=6, pady=6, sticky="ew")
                self.msds_entries[lbl] = ent

    # 16개 온디맨드 빌더 메서드
    def _build_msds_sec_0(self):
        f = self._create_msds_base_frame("■ [Section 1] 화학제품과 회사에 관한 정보 (Chemical Product & Company)")
        combos = {
            "권장용도 [별표5]": [
                "36. 화장품 및 개인위생용품", "1. 원료 및 중간체", "4. 방향제 및 탈취제", "25. 향수 및 향료 제품",
                "32. 세정제 및 세척제품", "37. 추출용제 및 가공보조제", "48. 기타", "직접 입력"
            ],
            "사용상의 제한": ["화장품 제조 및 배합 용도 외 사용 금지", "산업용 전용 원료로서 식품/의약품 용도 사용 금지", "직접 입력"]
        }
        fields = [
            ("제품명(국문)", 0, 0), ("제품명(영문)", 0, 2),
            ("제품코드/식별자", 1, 0), ("권장용도 [별표5]", 1, 2),
            ("사용상의 제한", 2, 0), ("공급업체(회사명)", 2, 2),
            ("담당부서", 3, 0), ("담당자", 3, 2),
            ("주소", 4, 0, 3),
            ("긴급연락전화번호", 5, 0), ("팩스번호", 5, 2)
        ]
        self._setup_msds_fields(f, fields, combos)
        return f

    def _build_msds_sec_1(self):
        f = self._create_msds_base_frame("■ [Section 2] 유해성·위험성 (Hazards Identification - 고용노동부고시 별표 2 기준)")
        combos = {
            "GHS 유해성분류": [
                "해당사항없음 (GHS 유해성 미분류)", "피부 부식성/자극성 : 구분 2 (Skin Irrit. 2)",
                "심한 눈 손상성/자극성 : 구분 2 (Eye Irrit. 2)", "인화성 액체 : 구분 3 (Flam. Liq. 3)",
                "인화성 액체 : 구분 4 (Flam. Liq. 4)", "급성 독성(경구) : 구분 4 (Acute Tox. 4)",
                "피부 과민성 : 구분 1 (Skin Sens. 1)", "특정표적장기 독성(1회 노출) : 구분 3 (STOT SE 3)",
                "수생환경 유해성(만성) : 구분 3 (Aquatic Chronic 3)", "직접 입력"
            ],
            "신호어": ["해당사항없음", "경고 (Warning)", "위험 (Danger)", "직접 입력"],
            "그림문자(픽토그램)": [
                "해당사항없음 (GHS 비분류)", "GHS07 (느낌표/감탄사 - 유해성)", "GHS02 (불꽃 - 인화성)",
                "GHS05 (부식성 - 피부/눈 손상)", "GHS08 (건강유해성 - 호흡기/발암성)",
                "GHS09 (환경 - 수생환경 유해성)", "GHS06 (해골 - 급성독성)", "GHS01 (폭발성)", "직접 입력"
            ],
            "유해위험문구 (H코드)": [
                "해당사항없음", "H315 : 피부에 자극을 일으킴", "H319 : 눈에 심한 자극을 일으킴",
                "H226 : 인화성 액체 및 증기", "H227 : 가연성 액체", "H302 : 삼키면 유해함",
                "H317 : 알레르기성 피부 반응을 일으킬 수 있음", "H335 : 호흡기계 자극을 일으킬 수 있음",
                "H412 : 장기적인 영향에 의해 수생생물에게 유해함", "직접 입력"
            ],
            "예방조치(예방)": [
                "P102 : 어린이의 손이 닿지 않는 곳에 보관하시오.",
                "P210 : 열·스파크·화염·고열로부터 멀리하시오. - 금연.",
                "P280 : 보호장갑·보안경·보호면을 착용하시오.",
                "P261 : 분진·흄·가스·미스트·증기·스프레이의 흡입을 피하시오.",
                "P264 : 취급 후에는 취급 부위를 철저히 씻으시오.", "직접 입력"
            ],
            "예방조치(대응)": [
                "P302+P352 : 피부에 묻으면 다량의 물과 비누로 씻으시오.",
                "P305+P351+P338 : 눈에 묻으면 몇 분간 물로 조심해서 씻으시오. 가능하면 콘택트렌즈를 제거하고 계속 씻으시오.",
                "P337+P313 : 눈에 자극이 지속되면 의학적인 조치·조언을 구하시오.",
                "P301+P312 : 삼켜서 불편함을 느끼면 의료기관(의사)의 진찰을 받으시오.", "직접 입력"
            ],
            "예방조치(저장)": [
                "P403+P235 : 환기가 잘 되는 곳에 보관하고 저온으로 유지하시오.",
                "P403+P233 : 환기가 잘 되는 곳에 밀폐하여 보관하시오.",
                "P405 : 잠금장치가 있는 곳에 보관하시오.", "직접 입력"
            ],
            "예방조치(폐기)": [
                "P501 : 폐기물관리법 및 관계 법령에 따라 내용물 및 용기를 폐기하시오.",
                "P502 : 재활용 또는 회수에 대한 정보는 제조자 또는 공급자를 참조하시오.", "직접 입력"
            ],
            "기타 유해위험성 (NFPA)": [
                "보건: 1, 화재: 0, 반응성: 0 (안전한 화장품 제형)",
                "보건: 1, 화재: 1, 반응성: 0", "보건: 2, 화재: 2, 반응성: 0",
                "보건: 0, 화재: 0, 반응성: 0", "직접 입력"
            ]
        }
        fields = [
            ("GHS 유해성분류", 0, 0, 3),
            ("신호어", 1, 0), ("그림문자(픽토그램)", 1, 2),
            ("유해위험문구 (H코드)", 2, 0, 3),
            ("예방조치(예방)", 3, 0, 3),
            ("예방조치(대응)", 4, 0, 3),
            ("예방조치(저장)", 5, 0, 3),
            ("예방조치(폐기)", 6, 0, 3),
            ("기타 유해위험성 (NFPA)", 7, 0, 3)
        ]
        self._setup_msds_fields(f, fields, combos)
        return f

    def _build_msds_sec_2(self):
        f = self._create_msds_base_frame("■ [Section 3] 구성성분의 명칭 및 함유량 (Composition / Ingredients)")
        combos = {
            "비공개 승인번호 (영업비밀)": [
                "해당사항없음", "산업안전보건법 제112조에 따른 대체자료 승인번호: 2026-승인-0001호",
                "대체자료 승인 유효기간: 5년 이내", "직접 입력"
            ]
        }
        fields = [("비공개 승인번호 (영업비밀)", 0, 0, 3)]
        self._setup_msds_fields(f, fields, combos)

        # 간편 입력 툴바
        in_bar = ctk.CTkFrame(f, fg_color="transparent")
        in_bar.grid(row=1, column=0, columnspan=4, sticky="ew", padx=6, pady=(10, 4))
        in_bar.grid_columnconfigure((0, 1), weight=1)

        e_chem = ctk.CTkEntry(in_bar, placeholder_text="화학물질명 (국문 또는 INCI)")
        e_chem.grid(row=0, column=0, sticky="ew", padx=2)
        e_syn = ctk.CTkEntry(in_bar, placeholder_text="관용명 및 이명 (Synonyms)")
        e_syn.grid(row=0, column=1, sticky="ew", padx=2)
        e_cas = ctk.CTkEntry(in_bar, placeholder_text="CAS 번호", width=110)
        e_cas.grid(row=0, column=2, padx=2)
        e_pct = ctk.CTkEntry(in_bar, placeholder_text="함유량(%)", width=90)
        e_pct.grid(row=0, column=3, padx=2)
        e_iden = ctk.CTkEntry(in_bar, placeholder_text="식별/승인번호", width=100)
        e_iden.grid(row=0, column=4, padx=2)

        def add_ing_grid():
            cnt = len(self.msds_ing_tree.get_children()) + 1
            c_val = e_chem.get().strip()
            s_val = e_syn.get().strip()
            cas_val = e_cas.get().strip() or "-"
            pct_val = e_pct.get().strip() or "-"
            id_val = e_iden.get().strip() or "-"
            if c_val or s_val or cas_val != "-":
                self.msds_ing_tree.insert("", "end", values=(str(cnt), c_val, s_val, cas_val, pct_val, id_val))
                e_chem.delete(0, "end"); e_syn.delete(0, "end"); e_cas.delete(0, "end"); e_pct.delete(0, "end"); e_iden.delete(0, "end")

        def del_ing_grid():
            for sel in self.msds_ing_tree.selection():
                self.msds_ing_tree.delete(sel)

        ctk.CTkButton(in_bar, text="➕ 추가", width=65, fg_color="#0284C7", hover_color="#0369A1", command=add_ing_grid).grid(row=0, column=5, padx=2)
        ctk.CTkButton(in_bar, text="선택 삭제", width=75, fg_color="#C62828", hover_color="#B71C1C", command=del_ing_grid).grid(row=0, column=6, padx=2)

        cols = {
            "no": ("No", 45),
            "chem": ("화학물질명 (국문/INCI)", 220),
            "syn": ("관용명 및 이명 (Synonyms)", 200),
            "cas": ("CAS 번호", 110),
            "pct": ("함유량 (%)", 90),
            "iden": ("식별/승인번호", 110)
        }
        f_tree, self.msds_ing_tree = self._create_treeview_grid(f, cols, height=10)
        f_tree.grid(row=2, column=0, columnspan=4, padx=6, pady=4, sticky="nsew")
        return f

    def _build_msds_sec_3(self):
        f = self._create_msds_base_frame("■ [Section 4] 응급조치 요령 (First Aid Measures)")
        combos = {
            "눈에 들어갔을 때": ["다량의 흐르는 물로 15분 이상 충분히 씻어내고 의사의 진료를 받을 것", "콘택트렌즈 제거 후 물로 세척", "직접 입력"],
            "피부에 접촉했을 때": ["오염된 의복을 벗고 다량의 물과 비누로 피부를 깨끗이 씻을 것", "자극이 발생하면 피부과 의사의 진료를 받을 것", "직접 입력"],
            "흡입했을 때": ["신선한 공기가 있는 곳으로 이동시키고 호흡이 곤란하면 인공호흡을 실시하고 의사의 진료를 받을 것", "자료없음", "직접 입력"],
            "먹었을 때": ["억지로 구토하게 하지 말고 즉시 물로 입을 헹궈낸 후 의사의 진료를 받을 것", "다량의 물을 마시게 할 것", "직접 입력"],
            "기타 의사의 주의사항": ["노출 정도에 따른 대증 치료를 실시할 것", "자료없음", "직접 입력"]
        }
        fields = [
            ("눈에 들어갔을 때", 0, 0, 3), ("피부에 접촉했을 때", 1, 0, 3),
            ("흡입했을 때", 2, 0, 3), ("먹었을 때", 3, 0, 3),
            ("기타 의사의 주의사항", 4, 0, 3)
        ]
        self._setup_msds_fields(f, fields, combos)
        return f

    def _build_msds_sec_4(self):
        f = self._create_msds_base_frame("■ [Section 5] 폭발·화재시 대처방법 (Fire-Fighting Measures)")
        combos = {
            "적절한 소화제": ["물분무, 내알코올포말, 분말소화제, 이산화탄소(CO2), 건사(모래)", "직접 입력"],
            "부적절한 소화제": ["직사주수 (물줄기 직접 분사 금지)", "자료없음", "직접 입력"],
            "화재시 발생 유해물질": ["열분해 또는 연소 시 일산화탄소(CO), 이산화탄소(CO2), 자극성 가스 및 유독성 흄이 발생할 수 있음", "직접 입력"],
            "소방관 착용 보호구": ["공기호흡기 및 전신 방화복, 내화학 보호의 착용", "직접 입력"],
            "화재진압 요령": ["위험하지 않은 경우 용기를 화재 지역 밖으로 이동시킬 것. 주변 화재에 맞는 소화 방법 사용", "직접 입력"]
        }
        fields = [
            ("적절한 소화제", 0, 0, 3), ("부적절한 소화제", 1, 0, 3),
            ("화재시 발생 유해물질", 2, 0, 3), ("소방관 착용 보호구", 3, 0, 3),
            ("화재진압 요령", 4, 0, 3)
        ]
        self._setup_msds_fields(f, fields, combos)
        return f

    def _build_msds_sec_5(self):
        f = self._create_msds_base_frame("■ [Section 6] 누출 사고 시 대처방법 (Accidental Release Measures)")
        combos = {
            "작업자 조치 및 보호구": ["보호장갑, 보안경 등 적절한 개인보호구를 착용하고 접근할 것. 누출 지역 환기", "직접 입력"],
            "환경 보호 조치": ["누출물이 하수도, 하천, 토양 및 수계로 유입되지 않도록 방제둑 설치", "직접 입력"],
            "정화 및 제거 방법": ["소량: 모래나 불활성 흡착제(버미큘라이트 등)로 흡수 후 폐기용기에 수거 / 대량: 방제둑 구축 후 펌프로 회수", "직접 입력"]
        }
        fields = [
            ("작업자 조치 및 보호구", 0, 0, 3),
            ("환경 보호 조치", 1, 0, 3),
            ("정화 및 제거 방법", 2, 0, 3)
        ]
        self._setup_msds_fields(f, fields, combos)
        return f

    def _build_msds_sec_6(self):
        f = self._create_msds_base_frame("■ [Section 7] 취급 및 저장방법 (Handling & Storage)")
        combos = {
            "안전취급요령": ["국소배기장치를 가동하고 취급할 것. 눈과 피부 접촉을 피하고 취급 후 손을 깨끗이 씻을 것", "직접 입력"],
            "안전저장조건": ["직사광선을 피하고 통풍이 잘되는 서늘하고 건조한 장소에 밀폐 보관할 것 (상온 1~30℃)", "직접 입력"],
            "피해야 할 조건 및 물질": ["고열, 스파크, 화염 및 강산화제, 강산 접촉 피함", "직접 입력"]
        }
        fields = [
            ("안전취급요령", 0, 0, 3),
            ("안전저장조건", 1, 0, 3),
            ("피해야 할 조건 및 물질", 2, 0, 3)
        ]
        self._setup_msds_fields(f, fields, combos)
        return f

    def _build_msds_sec_7(self):
        f = self._create_msds_base_frame("■ [Section 8] 노출방지 및 개인보호구 (Exposure Controls / PPE)")
        combos = {
            "국내 노출기준 (고용노동부)": ["고용노동부고시 화학물질 노출기준 미설정", "TWA: 해당없음 / STEL: 해당없음", "직접 입력"],
            "ACGIH / OSHA 노출기준": ["ACGIH TLV: 자료없음 / OSHA PEL: 자료없음", "직접 입력"],
            "공학적 관리 (환기)": ["공정 격리 또는 국소배기장치 설치로 작업장 공기 중 농도를 허용기준 이하로 유지할 것", "직접 입력"],
            "호흡기 보호구": ["방진/방독 마스크 착용 (분진 또는 미스트 발생 시)", "필요시 한국산업안전보건공단 인증 방독마스크 착용", "직접 입력"],
            "눈 보호구": ["보안경 또는 안면보호구 착용", "직접 입력"],
            "손 보호구": ["내화학성 고무장갑 또는 니트릴 보호장갑 착용", "직접 입력"],
            "신체 보호구": ["내화학성 작업복 및 안전화 착용", "직접 입력"]
        }
        fields = [
            ("국내 노출기준 (고용노동부)", 0, 0), ("ACGIH / OSHA 노출기준", 0, 2),
            ("공학적 관리 (환기)", 1, 0, 3),
            ("호흡기 보호구", 2, 0), ("눈 보호구", 2, 2),
            ("손 보호구", 3, 0), ("신체 보호구", 3, 2)
        ]
        self._setup_msds_fields(f, fields, combos)
        return f

    def _build_msds_sec_8(self):
        f = self._create_msds_base_frame("■ [Section 9] 물리화학적 특성 (Physical & Chemical Properties - 19대 지표)")
        combos = {
            "외관 (성상/색상)": ["무색 투명 겔상", "유백색 점조성 액체", "투명 점조성 액상", "미황색 액체", "백색 크림", "직접 입력"],
            "냄새": ["특이취", "무취", "원료 고유의 향", "향료향", "직접 입력"],
            "냄새역치": ["자료없음", "해당사항없음", "직접 입력"],
            "pH": ["5.50 ± 1.00", "6.50 ± 1.00", "5.0 ~ 7.0", "자료없음", "직접 입력"],
            "녹는점/어는점": ["자료없음", "0 ℃ 부근 (수용성 제형)", "직접 입력"],
            "초기 끓는점 및 끓는점 범위": ["100 ℃ 이상", "자료없음", "직접 입력"],
            "인화점": ["해당사항없음 (비인화성 제형)", "> 93 ℃", "자료없음", "직접 입력"],
            "증발속도": ["자료없음", "직접 입력"],
            "인화성 (고체, 기체)": ["해당사항없음 (비인화성)", "비해당", "직접 입력"],
            "인화 또는 폭발 한계의 상한/하한": ["해당사항없음", "자료없음", "직접 입력"],
            "증기압": ["자료없음", "직접 입력"],
            "용해도 (물에 대한 용해도)": ["수용성 (물에 쉽게 용해됨)", "물에 분산됨", "불용성", "직접 입력"],
            "증기밀도": ["자료없음", "직접 입력"],
            "비중 / 밀도": ["1.010 ± 0.050", "0.980 ± 0.050", "1.000 ± 0.050", "직접 입력"],
            "n-옥탄올/물 분배계수": ["자료없음", "직접 입력"],
            "자연발화온도": ["해당사항없음", "자료없음", "직접 입력"],
            "분해온도": ["자료없음", "직접 입력"],
            "점도": ["12,000 ± 3,000 cps", "25,000 ± 5,000 cps", "30,000 ± 5,000 cps", "자료없음", "직접 입력"],
            "분자량": ["해당사항없음 (혼합물)", "혼합물", "직접 입력"]
        }
        fields = [
            ("외관 (성상/색상)", 0, 0), ("냄새", 0, 2),
            ("냄새역치", 1, 0), ("pH", 1, 2),
            ("녹는점/어는점", 2, 0), ("초기 끓는점 및 끓는점 범위", 2, 2),
            ("인화점", 3, 0), ("증발속도", 3, 2),
            ("인화성 (고체, 기체)", 4, 0), ("인화 또는 폭발 한계의 상한/하한", 4, 2),
            ("증기압", 5, 0), ("용해도 (물에 대한 용해도)", 5, 2),
            ("증기밀도", 6, 0), ("비중 / 밀도", 6, 2),
            ("n-옥탄올/물 분배계수", 7, 0), ("자연발화온도", 7, 2),
            ("분해온도", 8, 0), ("점도", 8, 2),
            ("분자량", 9, 0, 3)
        ]
        self._setup_msds_fields(f, fields, combos)
        return f

    def _build_msds_sec_9(self):
        f = self._create_msds_base_frame("■ [Section 10] 안정성 및 반응성 (Stability & Reactivity)")
        combos = {
            "화학적 안정성": ["상온, 상압 및 권장 보관 조건에서 매우 안정함", "직접 입력"],
            "유해 반응 가능성": ["통상적인 보관 및 취급 시 유해 중합 또는 위험한 반응 없음", "직접 입력"],
            "피해야 할 조건": ["고열, 스파크, 화염, 직사광선 및 동결", "직접 입력"],
            "피해야 할 물질": ["강산, 강염기, 강산화제, 환원제", "직접 입력"],
            "분해시 생성되는 유해물질": ["열분해 시 일산화탄소, 이산화탄소 등 유해가스 발생 가능", "직접 입력"]
        }
        fields = [
            ("화학적 안정성", 0, 0, 3), ("유해 반응 가능성", 1, 0, 3),
            ("피해야 할 조건", 2, 0, 3), ("피해야 할 물질", 3, 0, 3),
            ("분해시 생성되는 유해물질", 4, 0, 3)
        ]
        self._setup_msds_fields(f, fields, combos)
        return f

    def _build_msds_sec_10(self):
        f = self._create_msds_base_frame("■ [Section 11] 독성에 관한 정보 (Toxicological Information - 14대 지표)")
        combos = {
            "급성 독성 (경구)": ["LD50 > 2,000 mg/kg (Rat, 혼합물 기준 비독성)", "자료없음", "직접 입력"],
            "급성 독성 (경피)": ["LD50 > 2,000 mg/kg (Rabbit)", "자료없음", "직접 입력"],
            "급성 독성 (흡입)": ["LC50 > 5.0 mg/L (4h, Rat)", "자료없음", "직접 입력"],
            "피부 부식성 또는 자극성": ["자극 없음 또는 경미한 자극 (Human Patch Test 음성)", "경미한 자극성", "직접 입력"],
            "심한 눈 손상 또는 자극성": ["자극 없음 또는 미약한 자극", "경미한 일시적 자극", "직접 입력"],
            "호흡기 과민성": ["자료없음 (과민성 유발 사례 없음)", "직접 입력"],
            "피부 과민성": ["피부 감작성 없음 (Non-sensitizer)", "자료없음", "직접 입력"],
            "발암성 (IARC / NTP / 고시)": ["IARC: 미분류 (Not Classifiable) / 발암성 물질 비해당", "직접 입력"],
            "생식세포 변이원성": ["음성 (Ames Test 음성)", "자료없음", "직접 입력"],
            "생식독성": ["생식독성 없음 또는 해당사항 없음", "자료없음", "직접 입력"],
            "특정표적장기독성 (1회 노출)": ["자료없음 (유의미한 표적장기 독성 없음)", "직접 입력"],
            "특정표적장기독성 (반복 노출)": ["자료없음 (반복 투여 독성 없음)", "직접 입력"],
            "흡인 유해성": ["흡인 유해성 없음", "자료없음", "직접 입력"],
            "기타 독성학적 정보": ["화장품 안전기준 등에 관한 규정 적합 제품", "자료없음", "직접 입력"]
        }
        fields = [
            ("급성 독성 (경구)", 0, 0), ("급성 독성 (경피)", 0, 2),
            ("급성 독성 (흡입)", 1, 0), ("피부 부식성 또는 자극성", 1, 2),
            ("심한 눈 손상 또는 자극성", 2, 0), ("호흡기 과민성", 2, 2),
            ("피부 과민성", 3, 0), ("발암성 (IARC / NTP / 고시)", 3, 2),
            ("생식세포 변이원성", 4, 0), ("생식독성", 4, 2),
            ("특정표적장기독성 (1회 노출)", 5, 0), ("특정표적장기독성 (반복 노출)", 5, 2),
            ("흡인 유해성", 6, 0), ("기타 독성학적 정보", 6, 2)
        ]
        self._setup_msds_fields(f, fields, combos)
        return f

    def _build_msds_sec_11(self):
        f = self._create_msds_base_frame("■ [Section 12] 환경에 미치는 영향 (Ecological Information)")
        combos = {
            "수생태 독성 (어류)": ["LC50 > 100 mg/L (96h, Fish)", "자료없음", "직접 입력"],
            "수생태 독성 (갑각류)": ["EC50 > 100 mg/L (48h, Daphnia magna)", "자료없음", "직접 입력"],
            "수생태 독성 (조류)": ["ErC50 > 100 mg/L (72h, Algae)", "자료없음", "직접 입력"],
            "잔류성 및 분해성": ["생분해성 우수 (Readily Biodegradable)", "자료없음", "직접 입력"],
            "생물 농축성": ["생물 축적 가능성 낮음 (Log Kow < 3)", "자료없음", "직접 입력"],
            "토양 이동성": ["자료없음", "직접 입력"],
            "기타 유해 영향": ["오존층 파괴 물질 및 환경호르몬 물질 비해당", "직접 입력"]
        }
        fields = [
            ("수생태 독성 (어류)", 0, 0), ("수생태 독성 (갑각류)", 0, 2),
            ("수생태 독성 (조류)", 1, 0), ("잔류성 및 분해성", 1, 2),
            ("생물 농축성", 2, 0), ("토양 이동성", 2, 2),
            ("기타 유해 영향", 3, 0, 3)
        ]
        self._setup_msds_fields(f, fields, combos)
        return f

    def _build_msds_sec_12(self):
        f = self._create_msds_base_frame("■ [Section 13] 폐기시 주의사항 (Disposal Considerations)")
        combos = {
            "폐기방법": ["폐기물관리법 규정에 따라 허가받은 폐기물 처리업체에 위탁 처리할 것", "직접 입력"],
            "폐기시 주의사항": ["하수구, 하천, 토양 등에 직접 방류를 엄격히 금지함. 폐기물 관련 법령 준수", "직접 입력"]
        }
        fields = [("폐기방법", 0, 0, 3), ("폐기시 주의사항", 1, 0, 3)]
        self._setup_msds_fields(f, fields, combos)
        return f

    def _build_msds_sec_13(self):
        f = self._create_msds_base_frame("■ [Section 14] 운송에 필요한 정보 (Transport Information - UN 규정)")
        combos = {
            "유엔번호 (UN No.)": ["해당사항없음 (비위험물)", "UN 1993", "UN 1170", "직접 입력"],
            "유엔 적정 선적명": ["해당사항없음 (Not Regulated as Dangerous Goods)", "비위험물", "직접 입력"],
            "운송에서의 위험성 등급": ["해당사항없음 (Class Non-hazardous)", "Class 3", "직접 입력"],
            "용기등급": ["해당사항없음", "PG III", "PG II", "직접 입력"],
            "해양오염물질": ["해당사항없음 (Not Marine Pollutant)", "해당", "직접 입력"],
            "특별 안전대책": ["운송 시 용기 파손, 누출 및 낙하 주의. 취급 주의 및 밀폐 상태 유지", "직접 입력"]
        }
        fields = [
            ("유엔번호 (UN No.)", 0, 0), ("유엔 적정 선적명", 0, 2),
            ("운송에서의 위험성 등급", 1, 0), ("용기등급", 1, 2),
            ("해양오염물질", 2, 0), ("특별 안전대책", 2, 2)
        ]
        self._setup_msds_fields(f, fields, combos)
        return f

    def _build_msds_sec_14(self):
        f = self._create_msds_base_frame("■ [Section 15] 법적 규제현황 (Regulatory Information)")
        combos = {
            "산업안전보건법에 의한 규제": ["작업환경측정 대상물질: 비해당, 관리대상 유해물질: 비해당, 특수건강진단 대상: 비해당", "직접 입력"],
            "화학물질관리법에 의한 규제": ["유독물질: 비해당, 제한물질: 비해당, 사고대비물질: 비해당", "직접 입력"],
            "위험물안전관리법에 의한 규제": ["비위험물 (위험물안전관리법상 위험물 비해당)", "제4류 인화성액체", "직접 입력"],
            "폐기물관리법에 의한 규제": ["지정폐기물 비해당 (일반 사업장폐기물로 처리)", "직접 입력"],
            "기타 국내 및 외국법에 의한 규제": ["화장품법 준수 / 미국 TSCA, 유럽 REACH 규정 준수", "직접 입력"]
        }
        fields = [
            ("산업안전보건법에 의한 규제", 0, 0, 3),
            ("화학물질관리법에 의한 규제", 1, 0, 3),
            ("위험물안전관리법에 의한 규제", 2, 0, 3),
            ("폐기물관리법에 의한 규제", 3, 0, 3),
            ("기타 국내 및 외국법에 의한 규제", 4, 0, 3)
        ]
        self._setup_msds_fields(f, fields, combos)
        return f

    def _build_msds_sec_15(self):
        f = self._create_msds_base_frame("■ [Section 16] 그 밖의 참고사항 (Other Information)")
        combos = {
            "자료출처": ["한국산업안전보건공단(KOSHA), 국립환경과학원(NCIS), 화학물질정보시스템, 원료 공급사 MSDS", "직접 입력"],
            "최초 작성일자": [f"{datetime.now().strftime('%Y년 %m월 %d일')}", "직접 입력"],
            "개정횟수 및 최종 개정일자": [f"개정횟수: 0회 (최종 개정일: {datetime.now().strftime('%Y년 %m월 %d일')})", "직접 입력"],
            "기타": ["본 MSDS는 고용노동부고시 제2023-9호 및 산업안전보건법 제110조에 따라 작성되었습니다.", "직접 입력"]
        }
        fields = [
            ("자료출처", 0, 0, 3),
            ("최초 작성일자", 1, 0), ("개정횟수 및 최종 개정일자", 1, 2),
            ("기타", 2, 0, 3)
        ]
        self._setup_msds_fields(f, fields, combos)
        return f

    def show_msds_section(self, idx):
        idx = max(0, min(idx, len(self.msds_section_titles) - 1))
        self.current_msds_section_idx = idx

        # 온디맨드 빌드: 아직 생성되지 않은 섹션이면 0.01초 만에 즉시 빌드
        if self.msds_section_frames[idx] is None:
            self.msds_section_frames[idx] = self.msds_section_builders[idx]()

        for i, frame in enumerate(self.msds_section_frames):
            if frame is not None:
                if i == idx: frame.grid(row=0, column=0, sticky="nsew")
                else: frame.grid_forget()

        if hasattr(self, 'msds_section_selector'):
            self.msds_section_selector.set(self.msds_section_titles[idx])

        if hasattr(self, 'btn_prev_sec'):
            if idx == 0: self.btn_prev_sec.configure(state="disabled", fg_color="gray40")
            else: self.btn_prev_sec.configure(state="normal", fg_color="#455A64")

        if hasattr(self, 'btn_next_sec'):
            if idx == len(self.msds_section_titles) - 1: self.btn_next_sec.configure(state="disabled", fg_color="gray40")
            else: self.btn_next_sec.configure(state="normal", fg_color="#1565C0")

    def prev_msds_section(self):
        if self.current_msds_section_idx > 0:
            self.show_msds_section(self.current_msds_section_idx - 1)

    def next_msds_section(self):
        if self.current_msds_section_idx < len(self.msds_section_titles) - 1:
            self.show_msds_section(self.current_msds_section_idx + 1)

    def on_msds_section_selected(self, choice):
        if choice in self.msds_section_titles:
            idx = self.msds_section_titles.index(choice)
            self.show_msds_section(idx)

    # -------------------------------------------------------------
    def add_msds_ing_row(self, data=None):
        self._ensure_all_msds_sections_built()
        if not hasattr(self, 'msds_ing_tree'): return
        cnt = len(self.msds_ing_tree.get_children()) + 1
        if data and isinstance(data, dict):
            c_val = str(data.get('name', ''))
            s_val = str(data.get('syn', ''))
            cas_val = str(data.get('cas', '-'))
            pct_val = str(data.get('pct', '-'))
            id_val = str(data.get('iden', '-'))
            self.msds_ing_tree.insert("", "end", values=(str(cnt), c_val, s_val, cas_val, pct_val, id_val))
        elif data and isinstance(data, (list, tuple)):
            self.msds_ing_tree.insert("", "end", values=data)

    def remove_selected_msds_ing_row(self):
        if hasattr(self, 'msds_ing_tree'):
            for sel in self.msds_ing_tree.selection():
                self.msds_ing_tree.delete(sel)

    def refresh_msds_list(self):
        try:
            session = db_manager.get_session()
            recs = session.query(MSDSDocument).order_by(MSDSDocument.created_at.desc()).limit(50).all()
            vals = [f"{r.id} | {r.product_name} | {r.doc_no or ''}" for r in recs]
            self.msds_picker.configure(values=vals if vals else ["-- 저장된 MSDS 없음 --"])
            if vals: self.msds_picker.set(vals[0])
            session.close()
        except Exception as e:
            print(f"[경고] MSDS 목록 로드 실패: {e}")

    def clear_msds_form(self):
        """완전 공란(Blank Form)으로 초기화"""
        self.current_msds_id = None
        for k, e in self.msds_entries.items():
            if isinstance(e, ctk.CTkComboBox):
                vals = e.cget("values")
                if vals: e.set(vals[0])
            else: e.delete(0, "end")
        if hasattr(self, 'msds_ing_tree'):
            for item in self.msds_ing_tree.get_children():
                self.msds_ing_tree.delete(item)

    def _ensure_all_msds_sections_built(self):
        """저장/로드 시 모든 섹션이 빌드되도록 보장"""
        for i in range(16):
            if self.msds_section_frames[i] is None:
                self.msds_section_frames[i] = self.msds_section_builders[i]()

    def load_selected_msds(self):
        sel = self.msds_picker.get()
        if not sel or "--" in sel:
            messagebox.showwarning("선택 오류", "불러올 MSDS를 선택해주세요.", parent=self)
            return

        msds_id = int(sel.split("|")[0].strip())
        session = db_manager.get_session()
        r = session.query(MSDSDocument).filter(MSDSDocument.id == msds_id).first()
        if not r:
            session.close()
            messagebox.showerror("오류", "선택한 MSDS 문서를 찾을 수 없습니다.", parent=self)
            return

        self.current_msds_id = r.id
        self._ensure_all_msds_sections_built()
        self.clear_msds_form()

        if r.sections_json:
            try:
                bundle = json.loads(r.sections_json)
                if "entries" in bundle:
                    for k, v in bundle["entries"].items():
                        if k in self.msds_entries and v:
                            e = self.msds_entries[k]
                            e.delete(0, "end"); e.insert(0, str(v))
                if "ingredients" in bundle and isinstance(bundle["ingredients"], list):
                    for ing in bundle["ingredients"]:
                        self.add_msds_ing_row(ing)
            except Exception as e:
                print(f"[경고] MSDS JSON 파싱 오류: {e}")

        session.close()
        messagebox.showinfo("불러오기 완료", f"[{r.product_name}] MSDS 16대 전 섹션 데이터가 성공적으로 로드되었습니다.", parent=self)

    def save_msds_to_db(self):
        p_name = self.msds_entries.get("제품명(국문)", ctk.CTkEntry(self)).get().strip()
        if not p_name:
            messagebox.showwarning("필수 입력 누락", "제품명(국문)은 필수 입력 항목입니다.", parent=self)
            return

        session = db_manager.get_session()
        try:
            if self.current_msds_id:
                r = session.query(MSDSDocument).filter(MSDSDocument.id == self.current_msds_id).first()
                if not r: r = MSDSDocument(); session.add(r)
            else:
                r = MSDSDocument(); session.add(r)

            r.product_name = p_name
            r.product_name_en = self.msds_entries.get("제품명(영문)", ctk.CTkEntry(self)).get().strip()
            r.doc_no = self.msds_entries.get("제품코드/식별자", ctk.CTkEntry(self)).get().strip()
            r.ghs_classification = self.msds_entries.get("GHS 유해성분류", ctk.CTkEntry(self)).get().strip()
            r.signal_word = self.msds_entries.get("신호어", ctk.CTkEntry(self)).get().strip()
            r.hazard_statements = self.msds_entries.get("유해위험문구 (H코드)", ctk.CTkEntry(self)).get().strip()
            r.precautionary_statements = self.msds_entries.get("예방조치(예방)", ctk.CTkEntry(self)).get().strip()

            all_entries = {k: w.get().strip() for k, w in self.msds_entries.items()}
            ing_list = []
            if hasattr(self, 'msds_ing_rows'):
                for row in self.msds_ing_rows:
                    nm = row['name'].get().strip()
                    if nm:
                        ing_list.append({
                            'name': nm, 'syn': row['syn'].get().strip(),
                            'cas': row['cas'].get().strip(), 'pct': row['pct'].get().strip(),
                            'iden': row['iden'].get().strip()
                        })

            bundle = {"entries": all_entries, "ingredients": ing_list}
            r.sections_json = json.dumps(bundle, ensure_ascii=False)

            session.commit()
            self.current_msds_id = r.id
            messagebox.showinfo("저장 완료", f"[{p_name}] MSDS 16대 전 섹션 정보가 DB에 안전하게 저장되었습니다.", parent=self)
            self.refresh_msds_list()
        except Exception as e:
            session.rollback()
            messagebox.showerror("저장 오류", f"MSDS 저장 실패: {e}", parent=self)
        finally:
            session.close()

    def delete_msds(self):
        if not self.current_msds_id:
            messagebox.showwarning("선택 오류", "삭제할 MSDS를 먼저 불러와주세요.", parent=self)
            return
        if not messagebox.askyesno("삭제 확인", "선택한 MSDS 문서를 영구히 삭제하시겠습니까?", parent=self): return
        session = db_manager.get_session()
        try:
            r = session.query(MSDSDocument).filter(MSDSDocument.id == self.current_msds_id).first()
            if r:
                session.delete(r); session.commit()
                self.clear_msds_form()
                self.refresh_msds_list()
                messagebox.showinfo("완료", "MSDS 문서가 성공적으로 삭제되었습니다.", parent=self)
        finally:
            session.close()

    def export_msds_to_excel(self, lang="ko"):
        """GHS 국제 표준 16대 전 섹션 및 고용노동부고시 제2023-9호 기준 1:1 국문/영문 전문 해석 엑셀 엔진"""
        import re
        self._ensure_all_msds_sections_built()
        is_en = (lang == "en")

        # =====================================================================
        # 고용노동부고시 제2023-9호 & GHS Rev. 국제 표준 1:1 영문화 대조 매트릭스
        # =====================================================================
        dict_ko_to_en = {
            # 직책 및 부서
            "품질책임자": "Quality Assurance Manager",
            "품질관리자": "QA Manager",
            "품질관리팀": "Quality Assurance Team",
            "품질보증팀": "Quality Assurance Team",
            "중앙기술연구소": "Central R&D Center",
            "중앙연구소": "Central R&D Center",
            "연구소": "R&D Center",
            "연구원": "Researcher",
            "선임연구원": "Senior Researcher",
            "책임연구원": "Lead Researcher",
            "팀장": "Team Leader",
            "태성연구원": "Taesung Kim (QA Researcher)",

            # 기본 상태
            "해당사항없음": "Not Applicable",
            "자료없음": "No data available",
            "해당없음": "Not Applicable",
            "비해당": "Not Applicable",
            "미설정": "Not established",
            "적합": "Pass (Complies)",
            "부적합": "Fail (Non-compliant)",
            "경고": "Warning",
            "경고 (Warning)": "Warning",
            "위험": "Danger",
            "위험 (Danger)": "Danger",

            # GHS 유해성분류
            "해당사항없음 (GHS 유해성 미분류)": "Not Classified (Complies with cosmetic safety standards)",
            "피부 부식성/자극성 : 구분 2 (Skin Irrit. 2)": "Skin corrosion/irritation - Category 2 (H315)",
            "심한 눈 손상성/자극성 : 구분 2 (Eye Irrit. 2)": "Serious eye damage/eye irritation - Category 2 (H319)",
            "심한 눈 손상성 : 구분 1 (Eye Dam. 1)": "Serious eye damage/eye irritation - Category 1 (H318)",
            "인화성 액체 : 구분 3 (Flam. Liq. 3)": "Flammable liquids - Category 3 (H226)",
            "인화성 액체 : 구분 4 (Flam. Liq. 4)": "Flammable liquids - Category 4 (H227)",
            "급성 독성(경구) : 구분 4 (Acute Tox. 4)": "Acute toxicity (Oral) - Category 4 (H302)",
            "피부 과민성 : 구분 1 (Skin Sens. 1)": "Skin sensitization - Category 1 (H317)",
            "특정표적장기 독성(1회 노출) : 구분 3 (STOT SE 3)": "Specific target organ toxicity (Single exposure) - Category 3 (H335)",
            "수생환경 유해성(만성) : 구분 3 (Aquatic Chronic 3)": "Hazardous to the aquatic environment (Chronic) - Category 3 (H412)",

            # 그림문자
            "해당사항없음 (GHS 비분류)": "None (Non-hazardous)",
            "GHS07 (느낌표/감탄사 - 유해성)": "GHS07 (Exclamation mark - Harmful)",
            "GHS02 (불꽃 - 인화성)": "GHS02 (Flame - Flammable)",
            "GHS05 (부식성 - 피부/눈 손상)": "GHS05 (Corrosion - Corrosive)",
            "GHS08 (건강유해성 - 호흡기/발암성)": "GHS08 (Health hazard)",
            "GHS09 (환경 - 수생환경 유해성)": "GHS09 (Environment - Aquatic toxicity)",
            "GHS06 (해골 - 급성독성)": "GHS06 (Skull and crossbones - Toxic)",
            "GHS01 (폭발성)": "GHS01 (Exploding bomb)",

            # 유해위험문구 (H-Codes)
            "H315 : 피부에 자극을 일으킴": "H315: Causes skin irritation.",
            "H319 : 눈에 심한 자극을 일으킴": "H319: Causes serious eye irritation.",
            "H318 : 눈에 심한 손상을 일으킴": "H318: Causes serious eye damage.",
            "H226 : 인화성 액체 및 증기": "H226: Flammable liquid and vapour.",
            "H227 : 가연성 액체": "H227: Combustible liquid.",
            "H302 : 삼키면 유해함": "H302: Harmful if swallowed.",
            "H317 : 알레르기성 피부 반응을 일으킬 수 있음": "H317: May cause an allergic skin reaction.",
            "H335 : 호흡기계 자극을 일으킬 수 있음": "H335: May cause respiratory irritation.",
            "H412 : 장기적인 영향에 의해 수생생물에게 유해함": "H412: Harmful to aquatic life with long lasting effects.",

            # 예방조치문구 (P-Codes)
            "P102 : 어린이의 손이 닿지 않는 곳에 보관하시오.": "P102: Keep out of reach of children.",
            "P210 : 열·스파크·화염·고열로부터 멀리하시오. - 금연.": "P210: Keep away from heat, sparks, open flames, hot surfaces. - No smoking.",
            "P280 : 보호장갑·보안경·보호면을 착용하시오.": "P280: Wear protective gloves and eye protection.",
            "P261 : 분진·흄·가스·미스트·증기·스프레이의 흡입을 피하시오.": "P261: Avoid breathing dust/fume/gas/mist/vapours/spray.",
            "P264 : 취급 후에는 취급 부위를 철저히 씻으시오.": "P264: Wash hands and exposed areas thoroughly after handling.",
            "P302+P352 : 피부에 묻으면 다량의 물과 비누로 씻으시오.": "P302+P352: IF ON SKIN: Wash with plenty of soap and water.",
            "P305+P351+P338 : 눈에 묻으면 몇 분간 물로 조심해서 씻으시오. 가능하면 콘택트렌즈를 제거하고 계속 씻으시오.": "P305+P351+P338: IF IN EYES: Rinse cautiously with water for several minutes. Remove contact lenses, if present and easy to do. Continue rinsing.",
            "P337+P313 : 눈에 자극이 지속되면 의학적인 조치·조언을 구하시오.": "P337+P313: If eye irritation persists: Get medical advice/attention.",
            "P301+P312 : 삼켜서 불편함을 느끼면 의료기관(의사)의 진찰을 받으시오.": "P301+P312: IF SWALLOWED: Call a POISON CENTER or doctor/physician if you feel unwell.",
            "P403+P235 : 환기가 잘 되는 곳에 보관하고 저온으로 유지하시오.": "P403+P235: Store in a well-ventilated place. Keep cool.",
            "P403+P233 : 환기가 잘 되는 곳에 밀폐하여 보관하시오.": "P403+P233: Store in a well-ventilated place. Keep container tightly closed.",
            "P405 : 잠금장치가 있는 곳에 보관하시오.": "P405: Store locked up.",
            "P501 : 폐기물관리법 및 관계 법령에 따라 내용물 및 용기를 폐기하시오.": "P501: Dispose of contents/container in accordance with local and national regulations.",
            "P502 : 재활용 또는 회수에 대한 정보는 제조자 또는 공급자를 참조하시오.": "P502: Refer to manufacturer or supplier for information on recovery or recycling."
        }

        # 스마트 복합 번역 엔진 (한글이 없거나 이미 영문인 문자열은 100% 원본 영문 그대로 보존)
        def tr(text):
            if not text:
                return ""
            t_str = str(text).strip()
            if not is_en:
                return t_str

            # 핵심: 한글이 전혀 없고 이미 영문/숫자/기호로만 구성되어 있다면 원래 영문 그대로 보존!
            if not re.search(r'[가-힣]', t_str):
                return t_str

            # 1. 완전 일치 매핑
            if t_str in dict_ko_to_en:
                return dict_ko_to_en[t_str]

            # 2. 복합 문장 ( / 또는 , 로 연결된 다중 항목) 분할 번역
            for sep in [" / ", ", ", "/"]:
                if sep in t_str:
                    parts = [p.strip() for p in t_str.split(sep) if p.strip()]
                    trans_parts = []
                    for part in parts:
                        if not re.search(r'[가-힣]', part):
                            trans_parts.append(part)  # 이미 영문인 부분은 그대로 유지
                        elif part in dict_ko_to_en:
                            trans_parts.append(dict_ko_to_en[part])
                        else:
                            matched_code = False
                            for k_dict, v_dict in dict_ko_to_en.items():
                                if (part.startswith("H") or part.startswith("P")) and part[:4] in k_dict:
                                    trans_parts.append(v_dict)
                                    matched_code = True
                                    break
                            if not matched_code:
                                trans_parts.append(part)
                    return (" / " if " / " in t_str or len(t_str) > 25 else ", ").join(trans_parts)

            # 3. H-코드 및 P-코드 패턴 매칭
            for k_dict, v_dict in dict_ko_to_en.items():
                if (t_str.startswith("H") or t_str.startswith("P")) and t_str[:4] in k_dict:
                    return v_dict
                if k_dict in t_str:
                    t_str = t_str.replace(k_dict, v_dict)

            return t_str

        # 제품 식별 및 공급자 정보
        p_ko = self.msds_entries.get("제품명(국문)", ctk.CTkEntry(self)).get().strip()
        p_en = self.msds_entries.get("제품명(영문)", ctk.CTkEntry(self)).get().strip()
        p_name = (p_en if p_en else tr(p_ko)) if is_en else (p_ko if p_ko else p_en)
        if not p_name: p_name = "Product_Safety_Data_Sheet" if is_en else "물질안전보건자료"

        p_code = self.msds_entries.get("제품코드/식별자", ctk.CTkEntry(self)).get().strip() or self.msds_entries.get("제품코드", ctk.CTkEntry(self)).get().strip() or "-"
        cp = get_company_profile()
        if is_en:
            comp_name = cp.get("company_name_en") or tr(self.msds_entries.get("공급업체(회사명)", ctk.CTkEntry(self)).get().strip()) or "ROCPOMA COSMETIC CO., LTD."
            dept_name = cp.get("department_en") or tr(self.msds_entries.get("담당부서", ctk.CTkEntry(self)).get().strip()) or "Quality Assurance Team"
            manager_name = cp.get("manager_name_en") or tr(self.msds_entries.get("담당자", ctk.CTkEntry(self)).get().strip()) or "Taesung Kim (QA Manager)"
            address_txt = cp.get("address_en") or tr(self.msds_entries.get("주소", ctk.CTkEntry(self)).get().strip()) or "Headquarters & R&D Center"
            emer_tel = self.msds_entries.get("긴급연락전화번호", ctk.CTkEntry(self)).get().strip() or cp.get("emergency_phone") or cp.get("phone", "-")
            fax_no = self.msds_entries.get("팩스번호", ctk.CTkEntry(self)).get().strip() or cp.get("fax", "-")
        else:
            comp_name = self.msds_entries.get("공급업체(회사명)", ctk.CTkEntry(self)).get().strip() or cp.get("company_name_ko", "(주)한국피부과학연구소")
            dept_name = self.msds_entries.get("담당부서", ctk.CTkEntry(self)).get().strip() or cp.get("department_ko", "품질관리팀")
            manager_name = self.msds_entries.get("담당자", ctk.CTkEntry(self)).get().strip() or cp.get("manager_name", "품질관리자")
            address_txt = self.msds_entries.get("주소", ctk.CTkEntry(self)).get().strip() or cp.get("address_ko", "본사 및 연구소")
            emer_tel = self.msds_entries.get("긴급연락전화번호", ctk.CTkEntry(self)).get().strip() or cp.get("emergency_phone") or cp.get("phone", "-")
            fax_no = self.msds_entries.get("팩스번호", ctk.CTkEntry(self)).get().strip() or cp.get("fax", "-")
        
        usage_txt = tr(self.msds_entries.get("권장용도 [별표5]", ctk.CTkEntry(self)).get().strip()) or ("36. Cosmetics and personal care products" if is_en else "36. 화장품 및 개인위생용품")
        usage_limit = tr(self.msds_entries.get("사용상의 제한", ctk.CTkEntry(self)).get().strip()) or ("Do not use for purposes other than cosmetic use." if is_en else "화장품 제조 및 배합 용도 외 사용 금지")
        ghs_class = tr(self.msds_entries.get("GHS 유해성분류", ctk.CTkEntry(self)).get().strip()) or ("Not Classified (Complies with cosmetic safety standards)" if is_en else "해당사항없음 (GHS 유해성 미분류)")
        sig_word = tr(self.msds_entries.get("신호어", ctk.CTkEntry(self)).get().strip()) or ("None" if is_en else "해당사항없음")
        picto = tr(self.msds_entries.get("그림문자(픽토그램)", ctk.CTkEntry(self)).get().strip()) or ("None" if is_en else "해당사항없음")
        ghs_hazard = tr(self.msds_entries.get("유해위험문구 (H코드)", ctk.CTkEntry(self)).get().strip()) or ("No significant hazards under normal conditions of intended use." if is_en else "해당사항없음")

        wb = Workbook()
        ws = wb.active
        ws.title = "MSDS (16 Sections)" if is_en else "물질안전보건자료"
        ws.views.sheetView[0].showGridLines = True

        # 스타일 정의
        font_title = Font(name="맑은 고딕", size=16, bold=True)
        font_sec_header = Font(name="맑은 고딕", size=11, bold=True, color="1F497D")
        font_tbl_header = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
        font_label = Font(name="맑은 고딕", size=10, bold=True)
        font_bold = Font(name="맑은 고딕", size=10, bold=True)
        font_body = Font(name="맑은 고딕", size=9.5)

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        fill_sec = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")
        fill_tbl_header = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin = Side(style="thin", color="A6B9D0")
        thin_dark = Side(style="thin", color="595959")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        table_border = Border(left=thin_dark, right=thin_dark, top=thin_dark, bottom=thin_dark)

        def style_range(cell_range, font=None, alignment=None, fill=None, border=None):
            for r in ws[cell_range]:
                for c in r:
                    if font: c.font = font
                    if alignment: c.alignment = alignment
                    if fill: c.fill = fill
                    if border: c.border = border

        # 문서 제목
        ws.merge_cells("A1:F2")
        ws["A1"] = "SAFETY DATA SHEET (SDS - GHS 16 SECTIONS)" if is_en else "물 질 안 전 보 건 자 료 (MSDS - 16대 법정 기준)"
        ws["A1"].font = font_title
        ws["A1"].alignment = align_center

        curr_row = 4

        def add_section_header(title_ko, title_en):
            nonlocal curr_row
            ws.merge_cells(f"A{curr_row}:F{curr_row}")
            ws[f"A{curr_row}"] = title_en if is_en else title_ko
            ws[f"A{curr_row}"].font = font_sec_header
            ws[f"A{curr_row}"].alignment = align_left
            style_range(f"A{curr_row}:F{curr_row}", fill=fill_sec, border=thin_border)
            curr_row += 1

        def add_sub_item(lbl_ko, lbl_en, val, sub_tag=""):
            nonlocal curr_row
            ws[f"A{curr_row}"] = sub_tag
            ws[f"A{curr_row}"].font = font_bold
            ws[f"A{curr_row}"].alignment = align_left

            ws[f"B{curr_row}"] = lbl_en if is_en else lbl_ko
            ws[f"B{curr_row}"].font = font_label
            ws[f"B{curr_row}"].alignment = align_left

            ws.merge_cells(f"C{curr_row}:F{curr_row}")
            ws[f"C{curr_row}"] = tr(val) if val else "-"
            ws[f"C{curr_row}"].font = font_body
            ws[f"C{curr_row}"].alignment = align_left
            style_range(f"A{curr_row}:F{curr_row}", border=thin_border)
            curr_row += 1

        # Section 1
        add_section_header("1. 화학제품과 회사에 관한 정보", "1. IDENTIFICATION OF THE SUBSTANCE/MIXTURE AND OF THE COMPANY")
        add_sub_item("가. 제품명", "A. Product Name", p_name)
        add_sub_item("나. 제품코드/식별자", "B. Product Code / Identifier", p_code)
        add_sub_item("다. 권장용도", "C. Recommended Use", usage_txt)
        add_sub_item("라. 사용상의 제한", "D. Restrictions on Use", usage_limit)
        add_sub_item("마. 공급업체(회사명)", "E. Supplier / Manufacturer Name", comp_name)
        add_sub_item("바. 주소", "F. Address", address_txt)
        add_sub_item("사. 담당부서 및 담당자", "G. Department & Contact Person", f"{dept_name} / {manager_name}")
        add_sub_item("아. 긴급연락전화번호", "H. Emergency Phone Number", emer_tel)
        add_sub_item("자. 팩스번호", "I. Fax Number", fax_no)
        curr_row += 1

        # Section 2
        add_section_header("2. 유해성·위험성", "2. HAZARDS IDENTIFICATION")
        add_sub_item("가. 유해성·위험성 분류", "A. GHS Hazard Classification", ghs_class)
        add_sub_item("나. 신호어", "B. Signal Word", sig_word)
        add_sub_item("다. 그림문자(픽토그램)", "C. Pictograms", picto)
        add_sub_item("라. 유해위험문구 (H코드)", "D. Hazard Statements (H-Codes)", ghs_hazard)
        add_sub_item("마. 예방조치(예방)", "E. Precautionary - Prevention", self.msds_entries.get("예방조치(예방)", ctk.CTkEntry(self)).get().strip())
        add_sub_item("바. 예방조치(대응)", "F. Precautionary - Response", self.msds_entries.get("예방조치(대응)", ctk.CTkEntry(self)).get().strip())
        add_sub_item("사. 예방조치(저장)", "G. Precautionary - Storage", self.msds_entries.get("예방조치(저장)", ctk.CTkEntry(self)).get().strip())
        add_sub_item("아. 예방조치(폐기)", "H. Precautionary - Disposal", self.msds_entries.get("예방조치(폐기)", ctk.CTkEntry(self)).get().strip())
        add_sub_item("자. 기타 유해위험성 (NFPA)", "I. Other Hazards (NFPA Rating)", self.msds_entries.get("기타 유해위험성 (NFPA)", ctk.CTkEntry(self)).get().strip())
        curr_row += 1

        # Section 3
        add_section_header("3. 구성성분의 명칭 및 함유량", "3. COMPOSITION / INFORMATION ON INGREDIENTS")
        trade_sec = self.msds_entries.get("비공개 승인번호 (영업비밀)", ctk.CTkEntry(self)).get().strip()
        if trade_sec:
            add_sub_item("비공개 대체자료 승인번호", "Trade Secret Approval No.", trade_sec)
        
        ws.merge_cells(f"A{curr_row}:A{curr_row+1}"); ws[f"A{curr_row}"] = "No"
        ws.merge_cells(f"B{curr_row}:B{curr_row+1}"); ws[f"B{curr_row}"] = "Chemical Name" if is_en else "화학물질명"
        ws.merge_cells(f"C{curr_row}:C{curr_row+1}"); ws[f"C{curr_row}"] = "INCI / Synonyms" if is_en else "관용명 및 이명 (INCI)"
        ws.merge_cells(f"D{curr_row}:D{curr_row+1}"); ws[f"D{curr_row}"] = "CAS No."
        ws.merge_cells(f"E{curr_row}:E{curr_row+1}"); ws[f"E{curr_row}"] = "Concentration (%)" if is_en else "함유량 (%)"
        ws.merge_cells(f"F{curr_row}:F{curr_row+1}"); ws[f"F{curr_row}"] = "Approval / Ident No." if is_en else "식별번호/승인번호"

        style_range(f"A{curr_row}:F{curr_row+1}", font=font_tbl_header, fill=fill_tbl_header, alignment=align_center, border=table_border)
        curr_row += 2

        parsed_rows = []
        if hasattr(self, 'msds_ing_tree'):
            for r_i, item_id in enumerate(self.msds_ing_tree.get_children(), start=1):
                v = self.msds_ing_tree.item(item_id, "values")
                if len(v) >= 6:
                    c_name = v[1].strip()
                    syn_name = v[2].strip()
                    cas = v[3].strip() or ("No data" if is_en else "자료없음")
                    pct_val = v[4].strip()
                    ident = v[5].strip() or ("N/A" if is_en else "해당없음")
                    display_chem = tr(c_name) if c_name else "-"
                    display_syn = tr(syn_name) if syn_name else "-"
                    parsed_rows.append((str(r_i), display_chem, display_syn, cas, pct_val, ident))

        if not parsed_rows:
            parsed_rows = [
                ("1", p_name, "Cosmetic Formulation Mixture", "N/A" if is_en else "혼합물", "100.0%", "N/A" if is_en else "해당없음")
            ]

        for r_no, r_chem, r_inci, r_cas, r_pct, r_ident in parsed_rows:
            ws[f"A{curr_row}"] = r_no
            ws[f"B{curr_row}"] = r_chem
            ws[f"C{curr_row}"] = r_inci
            ws[f"D{curr_row}"] = r_cas
            ws[f"E{curr_row}"] = r_pct
            ws[f"F{curr_row}"] = r_ident
            style_range(f"A{curr_row}:F{curr_row}", font=font_body, alignment=align_center, border=table_border)
            ws[f"B{curr_row}"].alignment = align_left
            ws[f"C{curr_row}"].alignment = align_left
            curr_row += 1
        curr_row += 1

        # Section 4
        add_section_header("4. 응급조치 요령", "4. FIRST-AID MEASURES")
        add_sub_item("가. 눈에 들어갔을 때", "A. Eye Contact", self.msds_entries.get("눈에 들어갔을 때", ctk.CTkEntry(self)).get().strip())
        add_sub_item("나. 피부에 접촉했을 때", "B. Skin Contact", self.msds_entries.get("피부에 접촉했을 때", ctk.CTkEntry(self)).get().strip())
        add_sub_item("다. 흡입했을 때", "C. Inhalation", self.msds_entries.get("흡입했을 때", ctk.CTkEntry(self)).get().strip())
        add_sub_item("라. 먹었을 때", "D. Ingestion", self.msds_entries.get("먹었을 때", ctk.CTkEntry(self)).get().strip())
        add_sub_item("마. 기타 의사의 주의사항", "E. Notes to Physician", self.msds_entries.get("기타 의사의 주의사항", ctk.CTkEntry(self)).get().strip())
        curr_row += 1

        # Section 5
        add_section_header("5. 폭발·화재시 대처방법", "5. FIRE-FIGHTING MEASURES")
        add_sub_item("가. 적절한 소화제", "A. Suitable Extinguishing Media", self.msds_entries.get("적절한 소화제", ctk.CTkEntry(self)).get().strip())
        add_sub_item("나. 부적절한 소화제", "B. Unsuitable Extinguishing Media", self.msds_entries.get("부적절한 소화제", ctk.CTkEntry(self)).get().strip())
        add_sub_item("다. 화재시 발생 유해물질", "C. Specific Hazards from Chemical", self.msds_entries.get("화재시 발생 유해물질", ctk.CTkEntry(self)).get().strip())
        add_sub_item("라. 소방관 착용 보호구", "D. Protective Equipment for Firefighters", self.msds_entries.get("소방관 착용 보호구", ctk.CTkEntry(self)).get().strip())
        add_sub_item("마. 화재진압 요령", "E. Fire Fighting Instructions", self.msds_entries.get("화재진압 요령", ctk.CTkEntry(self)).get().strip())
        curr_row += 1

        # Section 6
        add_section_header("6. 누출 사고 시 대처방법", "6. ACCIDENTAL RELEASE MEASURES")
        add_sub_item("가. 작업자 조치 및 보호구", "A. Personal Precautions & PPE", self.msds_entries.get("작업자 조치 및 보호구", ctk.CTkEntry(self)).get().strip())
        add_sub_item("나. 환경 보호 조치", "B. Environmental Precautions", self.msds_entries.get("환경 보호 조치", ctk.CTkEntry(self)).get().strip())
        add_sub_item("다. 정화 및 제거 방법", "C. Clean-up & Containment", self.msds_entries.get("정화 및 제거 방법", ctk.CTkEntry(self)).get().strip())
        curr_row += 1

        # Section 7
        add_section_header("7. 취급 및 저장방법", "7. HANDLING AND STORAGE")
        add_sub_item("가. 안전취급요령", "A. Safe Handling Precautions", self.msds_entries.get("안전취급요령", ctk.CTkEntry(self)).get().strip())
        add_sub_item("나. 안전저장조건", "B. Safe Storage Conditions", self.msds_entries.get("안전저장조건", ctk.CTkEntry(self)).get().strip())
        add_sub_item("다. 피해야 할 조건 및 물질", "C. Incompatible Conditions & Materials", self.msds_entries.get("피해야 할 조건 및 물질", ctk.CTkEntry(self)).get().strip())
        curr_row += 1

        # Section 8
        add_section_header("8. 노출방지 및 개인보호구", "8. EXPOSURE CONTROLS / PERSONAL PROTECTION")
        add_sub_item("가. 국내 노출기준 (고용노동부)", "A. Occupational Exposure Limits (Domestic)", self.msds_entries.get("국내 노출기준 (고용노동부)", ctk.CTkEntry(self)).get().strip())
        add_sub_item("나. ACGIH / OSHA 노출기준", "B. ACGIH / OSHA Exposure Limits", self.msds_entries.get("ACGIH / OSHA 노출기준", ctk.CTkEntry(self)).get().strip())
        add_sub_item("다. 공학적 관리 (환기)", "C. Engineering Controls", self.msds_entries.get("공학적 관리 (환기)", ctk.CTkEntry(self)).get().strip())
        add_sub_item("라. 호흡기 보호구", "D. Respiratory Protection", self.msds_entries.get("호흡기 보호구", ctk.CTkEntry(self)).get().strip())
        add_sub_item("마. 눈 보호구", "E. Eye / Face Protection", self.msds_entries.get("눈 보호구", ctk.CTkEntry(self)).get().strip())
        add_sub_item("바. 손 보호구", "F. Hand Protection", self.msds_entries.get("손 보호구", ctk.CTkEntry(self)).get().strip())
        add_sub_item("사. 신체 보호구", "G. Body Protection", self.msds_entries.get("신체 보호구", ctk.CTkEntry(self)).get().strip())
        curr_row += 1

        # Section 9
        add_section_header("9. 물리화학적 특성", "9. PHYSICAL AND CHEMICAL PROPERTIES")
        p9_map = [
            ("가. 외관 (성상/색상)", "A. Appearance", "외관 (성상/색상)"),
            ("나. 냄새", "B. Odour", "냄새"),
            ("다. 냄새역치", "C. Odour Threshold", "냄새역치"),
            ("라. pH", "D. pH", "pH"),
            ("마. 녹는점/어는점", "E. Melting / Freezing Point", "녹는점/어는점"),
            ("바. 초기 끓는점 및 범위", "F. Initial Boiling Point & Range", "초기 끓는점 및 끓는점 범위"),
            ("사. 인화점", "G. Flash Point", "인화점"),
            ("아. 증발속도", "H. Evaporation Rate", "증발속도"),
            ("자. 인화성 (고체, 기체)", "I. Flammability (Solid, Gas)", "인화성 (고체, 기체)"),
            ("차. 인화/폭발 한계 상한/하한", "J. Flammability / Explosive Limits", "인화 또는 폭발 한계의 상한/하한"),
            ("카. 증기압", "K. Vapour Pressure", "증기압"),
            ("타. 용해도 (물에 대한 용해도)", "L. Solubility", "용해도 (물에 대한 용해도)"),
            ("파. 증기밀도", "M. Vapour Density", "증기밀도"),
            ("하. 비중 / 밀도", "N. Specific Gravity / Density", "비중 / 밀도"),
            ("거. n-옥탄올/물 분배계수", "O. Partition Coefficient (n-octanol/water)", "n-옥탄올/물 분배계수"),
            ("너. 자연발화온도", "P. Auto-ignition Temperature", "자연발화온도"),
            ("더. 분해온도", "Q. Decomposition Temperature", "분해온도"),
            ("러. 점도", "R. Viscosity", "점도"),
            ("머. 분자량", "S. Molecular Weight", "분자량")
        ]
        for l_ko, l_en, k_ent in p9_map:
            add_sub_item(l_ko, l_en, self.msds_entries.get(k_ent, ctk.CTkEntry(self)).get().strip())
        curr_row += 1

        # Section 10
        add_section_header("10. 안정성 및 반응성", "10. STABILITY AND REACTIVITY")
        add_sub_item("가. 화학적 안정성", "A. Chemical Stability", self.msds_entries.get("화학적 안정성", ctk.CTkEntry(self)).get().strip())
        add_sub_item("나. 유해 반응 가능성", "B. Possibility of Hazardous Reactions", self.msds_entries.get("유해 반응 가능성", ctk.CTkEntry(self)).get().strip())
        add_sub_item("다. 피해야 할 조건", "C. Conditions to Avoid", self.msds_entries.get("피해야 할 조건", ctk.CTkEntry(self)).get().strip())
        add_sub_item("라. 피해야 할 물질", "D. Incompatible Materials", self.msds_entries.get("피해야 할 물질", ctk.CTkEntry(self)).get().strip())
        add_sub_item("마. 분해시 생성되는 유해물질", "E. Hazardous Decomposition Products", self.msds_entries.get("분해시 생성되는 유해물질", ctk.CTkEntry(self)).get().strip())
        curr_row += 1

        # Section 11
        add_section_header("11. 독성에 관한 정보", "11. TOXICOLOGICAL INFORMATION")
        p11_map = [
            ("가. 급성 독성 (경구)", "A. Acute Toxicity (Oral)", "급성 독성 (경구)"),
            ("나. 급성 독성 (경피)", "B. Acute Toxicity (Dermal)", "급성 독성 (경피)"),
            ("다. 급성 독성 (흡입)", "C. Acute Toxicity (Inhalation)", "급성 독성 (흡입)"),
            ("라. 피부 부식성 또는 자극성", "D. Skin Corrosion / Irritation", "피부 부식성 또는 자극성"),
            ("마. 심한 눈 손상 또는 자극성", "E. Serious Eye Damage / Irritation", "심한 눈 손상 또는 자극성"),
            ("바. 호흡기 과민성", "F. Respiratory Sensitization", "호흡기 과민성"),
            ("사. 피부 과민성", "G. Skin Sensitization", "피부 과민성"),
            ("아. 발암성 (IARC/NTP/고시)", "H. Carcinogenicity", "발암성 (IARC / NTP / 고시)"),
            ("자. 생식세포 변이원성", "I. Germ Cell Mutagenicity", "생식세포 변이원성"),
            ("차. 생식독성", "J. Reproductive Toxicity", "생식독성"),
            ("카. 특정표적장기독성 (1회 노출)", "K. STOT - Single Exposure", "특정표적장기독성 (1회 노출)"),
            ("타. 특정표적장기독성 (반복 노출)", "L. STOT - Repeated Exposure", "특정표적장기독성 (반복 노출)"),
            ("파. 흡인 유해성", "M. Aspiration Hazard", "흡인 유해성"),
            ("하. 기타 독성학적 정보", "N. Other Toxicological Info", "기타 독성학적 정보")
        ]
        for l_ko, l_en, k_ent in p11_map:
            add_sub_item(l_ko, l_en, self.msds_entries.get(k_ent, ctk.CTkEntry(self)).get().strip())
        curr_row += 1

        # Section 12
        add_section_header("12. 환경에 미치는 영향", "12. ECOLOGICAL INFORMATION")
        p12_map = [
            ("가. 수생태 독성 (어류)", "A. Aquatic Toxicity (Fish)", "수생태 독성 (어류)"),
            ("나. 수생태 독성 (갑각류)", "B. Aquatic Toxicity (Crustacea)", "수생태 독성 (갑각류)"),
            ("다. 수생태 독성 (조류)", "C. Aquatic Toxicity (Algae)", "수생태 독성 (조류)"),
            ("라. 잔류성 및 분해성", "D. Persistence & Degradability", "잔류성 및 분해성"),
            ("마. 생물 농축성", "E. Bioaccumulative Potential", "생물 농축성"),
            ("바. 토양 이동성", "F. Mobility in Soil", "토양 이동성"),
            ("사. 기타 유해 영향", "G. Other Adverse Effects", "기타 유해 영향")
        ]
        for l_ko, l_en, k_ent in p12_map:
            add_sub_item(l_ko, l_en, self.msds_entries.get(k_ent, ctk.CTkEntry(self)).get().strip())
        curr_row += 1

        # Section 13
        add_section_header("13. 폐기시 주의사항", "13. DISPOSAL CONSIDERATIONS")
        add_sub_item("가. 폐기방법", "A. Waste Treatment Methods", self.msds_entries.get("폐기방법", ctk.CTkEntry(self)).get().strip())
        add_sub_item("나. 폐기시 주의사항", "B. Disposal Precautions", self.msds_entries.get("폐기시 주의사항", ctk.CTkEntry(self)).get().strip())
        curr_row += 1

        # Section 14
        add_section_header("14. 운송에 필요한 정보", "14. TRANSPORT INFORMATION")
        add_sub_item("가. 유엔번호 (UN No.)", "A. UN Number", self.msds_entries.get("유엔번호 (UN No.)", ctk.CTkEntry(self)).get().strip())
        add_sub_item("나. 유엔 적정 선적명", "B. UN Proper Shipping Name", self.msds_entries.get("유엔 적정 선적명", ctk.CTkEntry(self)).get().strip())
        add_sub_item("다. 운송에서의 위험성 등급", "C. Transport Hazard Class", self.msds_entries.get("운송에서의 위험성 등급", ctk.CTkEntry(self)).get().strip())
        add_sub_item("라. 용기등급", "D. Packing Group", self.msds_entries.get("용기등급", ctk.CTkEntry(self)).get().strip())
        add_sub_item("마. 해양오염물질", "E. Marine Pollutant", self.msds_entries.get("해양오염물질", ctk.CTkEntry(self)).get().strip())
        add_sub_item("바. 특별 안전대책", "F. Special Precautions for User", self.msds_entries.get("특별 안전대책", ctk.CTkEntry(self)).get().strip())
        curr_row += 1

        # Section 15
        add_section_header("15. 법적 규제현황", "15. REGULATORY INFORMATION")
        add_sub_item("가. 산업안전보건법에 의한 규제", "A. Occupational Safety & Health Act", self.msds_entries.get("산업안전보건법에 의한 규제", ctk.CTkEntry(self)).get().strip())
        add_sub_item("나. 화학물질관리법에 의한 규제", "B. Chemicals Control Act", self.msds_entries.get("화학물질관리법에 의한 규제", ctk.CTkEntry(self)).get().strip())
        add_sub_item("다. 위험물안전관리법에 의한 규제", "C. Dangerous Goods Safety Management Act", self.msds_entries.get("위험물안전관리법에 의한 규제", ctk.CTkEntry(self)).get().strip())
        add_sub_item("라. 폐기물관리법에 의한 규제", "D. Wastes Control Act", self.msds_entries.get("폐기물관리법에 의한 규제", ctk.CTkEntry(self)).get().strip())
        add_sub_item("마. 기타 국내 및 외국법에 의한 규제", "E. Other Domestic & International Acts", self.msds_entries.get("기타 국내 및 외국법에 의한 규제", ctk.CTkEntry(self)).get().strip())
        curr_row += 1

        # Section 16
        add_section_header("16. 그 밖의 참고사항", "16. OTHER INFORMATION")
        add_sub_item("가. 자료출처", "A. References & Data Sources", self.msds_entries.get("자료출처", ctk.CTkEntry(self)).get().strip())
        add_sub_item("나. 최초 작성일자", "B. First Creation Date", self.msds_entries.get("최초 작성일자", ctk.CTkEntry(self)).get().strip() or datetime.now().strftime("%Y-%m-%d"))
        add_sub_item("다. 개정횟수 및 최종 개정일자", "C. Revision Number & Latest Date", self.msds_entries.get("개정횟수 및 최종 개정일자", ctk.CTkEntry(self)).get().strip() or f"{datetime.now().strftime('%Y-%m-%d')} (Rev. 0)")
        add_sub_item("라. 기타", "D. Other Notes", self.msds_entries.get("기타", ctk.CTkEntry(self)).get().strip())

        # 열 너비 설정
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 28
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 20

        f_prefix = f"MSDS_EN_{p_name}" if is_en else f"MSDS_국문16대섹션_{p_name}"
        file_path_dlg = fd.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"{f_prefix}.xlsx",
            title="Save MSDS (English 16 Sections)" if is_en else "MSDS (16대 전 섹션 표준 엑셀) 저장"
        )
        if file_path_dlg:
            wb.save(file_path_dlg)
            msg_txt = f"공식 16대 표준 서식의 MSDS가 성공적으로 생성되었습니다: {file_path_dlg}" if not is_en else f"MSDS (16 Sections Full English Edition) exported successfully: {file_path_dlg}"
            messagebox.showinfo("저장 완료" if not is_en else "Export Complete", msg_txt, parent=self)
            try:
                os.startfile(os.path.abspath(file_path_dlg))
            except Exception:
                pass

    def setup_stability_test_tab(self, tab_frame):
        tab_frame.grid_columnconfigure(0, weight=1)
        tab_frame.grid_rowconfigure(0, weight=1)

        scroll = ctk.CTkScrollableFrame(tab_frame, label_text="안정성(경시변화) 시험보고서 (Stability Test Report)")
        scroll.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        scroll.grid_columnconfigure(0, weight=1)

        self.stab_entries = {}
        self.current_stab_id = None

        top_bar = ctk.CTkFrame(scroll, fg_color="transparent")
        top_bar.grid(row=0, column=0, sticky="ew", padx=10, pady=(5, 15))
        top_bar.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(top_bar, text="안정성 이력:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=5, pady=5)
        self.stab_picker = ctk.CTkComboBox(top_bar, values=["-- 저장된 안정성보고서 선택 --"], width=300)
        self.stab_picker.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ctk.CTkButton(top_bar, text="불러오기", width=80, command=self.load_selected_stab).grid(row=0, column=2, padx=5, pady=5)
        ctk.CTkButton(top_bar, text="새로작성", width=80, fg_color="gray50", command=self.clear_stab_form).grid(row=0, column=3, padx=5, pady=5)
        ctk.CTkButton(top_bar, text="삭제", width=70, fg_color="#D32F2F", command=self.delete_stab).grid(row=0, column=4, padx=5, pady=5)

        info_frame = ctk.CTkFrame(scroll)
        info_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=5)
        info_frame.grid_columnconfigure((1, 3, 5), weight=1)

        fields = [
            ("제품명", 0, 0), ("제조번호(LOT)", 0, 2), ("시험시작일", 0, 4),
            ("시험담당자", 1, 0), ("종합평가", 1, 2)
        ]

        for label_text, r, c in fields:
            ctk.CTkLabel(info_frame, text=label_text, font=ctk.CTkFont(weight="bold")).grid(row=r, column=c, padx=8, pady=5, sticky="w")
            if label_text == "종합평가":
                cb = ctk.CTkComboBox(info_frame, values=["안정 (양호)", "경미한 변화(허용범위)", "불안정(변색/분리)"], width=160)
                cb.set("안정 (양호)")
                cb.grid(row=r, column=c+1, padx=8, pady=5, sticky="ew")
                self.stab_entries[label_text] = cb
            else:
                entry = ctk.CTkEntry(info_frame)
                if label_text == "시험시작일": entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
                entry.grid(row=r, column=c+1, padx=8, pady=5, sticky="ew")
                self.stab_entries[label_text] = entry

        # 가혹 조건별 경시변화 매트릭스 테이블
        cond_frame = ctk.CTkFrame(scroll)
        cond_frame.grid(row=2, column=0, sticky="ew", padx=10, pady=10)
        cond_frame.grid_columnconfigure((1, 2, 3, 4, 5), weight=1)

        headers = ["보관 조건", "초기 (0주)", "2주 경과", "4주 경과", "8주 경과", "12주 경과"]
        for idx, h in enumerate(headers):
            ctk.CTkLabel(cond_frame, text=h, font=ctk.CTkFont(weight="bold")).grid(row=0, column=idx, padx=5, pady=5)

        conditions = [
            "실온 (25℃)",
            "고온 (45℃)",
            "극고온 (50℃)",
            "저온 (4℃)",
            "Cycle (-10~40℃)",
            "광안정성 (UV)"
        ]

        self.stab_matrix_entries = {}
        for r_idx, cond in enumerate(conditions, 1):
            ctk.CTkLabel(cond_frame, text=cond, font=ctk.CTkFont(weight="bold")).grid(row=r_idx, column=0, padx=5, pady=4, sticky="w")
            self.stab_matrix_entries[cond] = []
            for c_idx in range(1, 6):
                e = ctk.CTkEntry(cond_frame, width=90)
                e.insert(0, "이상없음" if c_idx < 3 else "안정")
                e.grid(row=r_idx, column=c_idx, padx=4, pady=4, sticky="ew")
                self.stab_matrix_entries[cond].append(e)

        btn_bar = ctk.CTkFrame(scroll, fg_color="transparent")
        btn_bar.grid(row=3, column=0, sticky="e", padx=10, pady=15)
        ctk.CTkButton(btn_bar, text="💾 안정성보고서 DB 저장", width=140, command=self.save_stab_to_db).pack(side="left", padx=5)
        ctk.CTkButton(btn_bar, text="📊 안정성 엑셀 출력", width=140, fg_color="#2E7D32", hover_color="#1B5E20", command=self.export_stab_to_excel).pack(side="left", padx=5)

        self.refresh_stab_list()

    def refresh_stab_list(self):
        try:
            session = db_manager.get_session()
            recs = session.query(StabilityTestReport).order_by(StabilityTestReport.created_at.desc()).limit(50).all()
            vals = [f"{r.id} | {r.product_name} | LOT:{r.lot_no or ''} | {r.overall_evaluation or ''}" for r in recs]
            self.stab_picker.configure(values=vals if vals else ["-- 저장된 안정성보고서 없음 --"])
            if vals: self.stab_picker.set(vals[0])
            session.close()
        except Exception as e:
            print(f"[경고] 안정성보고서 목록 로드 실패: {e}")

    def clear_stab_form(self):
        self.current_stab_id = None
        for k, e in self.stab_entries.items():
            if isinstance(e, ctk.CTkComboBox): e.set(e.cget("values")[0])
            else: e.delete(0, "end")
        self.stab_entries["시험시작일"].insert(0, datetime.now().strftime("%Y-%m-%d"))

    def load_selected_stab(self):
        sel = self.stab_picker.get()
        if not sel or '|' not in sel: return
        rec_id = int(sel.split('|')[0].strip())
        session = db_manager.get_session()
        try:
            r = session.query(StabilityTestReport).get(rec_id)
            if not r: return
            self.current_stab_id = r.id
            self.stab_entries["제품명"].delete(0, "end"); self.stab_entries["제품명"].insert(0, r.product_name or "")
            self.stab_entries["제조번호(LOT)"].delete(0, "end"); self.stab_entries["제조번호(LOT)"].insert(0, r.lot_no or "")
            self.stab_entries["시험시작일"].delete(0, "end"); self.stab_entries["시험시작일"].insert(0, str(r.test_start_date or ""))
            self.stab_entries["시험담당자"].delete(0, "end"); self.stab_entries["시험담당자"].insert(0, r.examiner or "")
            self.stab_entries["종합평가"].set(r.overall_evaluation or "안정 (양호)")
            
            if r.test_results_json:
                import json
                matrix = json.loads(r.test_results_json)
                for cond, vals in matrix.items():
                    if cond in self.stab_matrix_entries:
                        for i, v in enumerate(vals):
                            if i < len(self.stab_matrix_entries[cond]):
                                self.stab_matrix_entries[cond][i].delete(0, "end")
                                self.stab_matrix_entries[cond][i].insert(0, str(v))
            messagebox.showinfo("불러오기 완료", f"안정성보고서 '{r.product_name}'를 불러왔습니다.", parent=self)
        finally:
            session.close()

    def save_stab_to_db(self):
        p_name = self.stab_entries["제품명"].get().strip()
        if not p_name:
            messagebox.showwarning("입력 필요", "제품명은 필수 항목입니다.", parent=self); return
        import json
        matrix_data = {}
        for cond, entries in self.stab_matrix_entries.items():
            matrix_data[cond] = [e.get().strip() for e in entries]

        session = db_manager.get_session()
        try:
            r = session.query(StabilityTestReport).get(self.current_stab_id) if self.current_stab_id else StabilityTestReport()
            r.product_name = p_name
            r.lot_no = self.stab_entries["제조번호(LOT)"].get().strip()
            try: r.test_start_date = datetime.strptime(self.stab_entries["시험시작일"].get().strip(), "%Y-%m-%d").date()
            except: r.test_start_date = None
            r.examiner = self.stab_entries["시험담당자"].get().strip()
            r.overall_evaluation = self.stab_entries["종합평가"].get().strip()
            r.test_results_json = json.dumps(matrix_data, ensure_ascii=False)

            if not self.current_stab_id: session.add(r)
            session.commit()
            self.current_stab_id = r.id
            self.refresh_stab_list()
            messagebox.showinfo("저장 완료", "안정성 시험보고서가 DB에 성공적으로 저장되었습니다.", parent=self)
        except Exception as e:
            session.rollback(); messagebox.showerror("오류", f"저장 실패: {e}", parent=self)
        finally:
            session.close()

    def delete_stab(self):
        if not self.current_stab_id: return
        if not messagebox.askyesno("삭제 확인", "선택한 안정성보고서를 삭제하시겠습니까?", parent=self): return
        session = db_manager.get_session()
        try:
            r = session.query(StabilityTestReport).get(self.current_stab_id)
            if r:
                session.delete(r); session.commit()
                self.clear_stab_form(); self.refresh_stab_list()
                messagebox.showinfo("완료", "삭제되었습니다.", parent=self)
        finally:
            session.close()

    def export_stab_to_excel(self):
        p_name = self.stab_entries["제품명"].get().strip() or "안정성보고서"
        wb = Workbook(); ws = wb.active; ws.title = "안정성보고서"
        ws.merge_cells("A1:F2")
        ws["A1"] = f"화장품 안정성(경시변화) 시험보고서 - {p_name}"
        ws["A1"].font = Font(name="맑은 고딕", size=16, bold=True, color="1F497D")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.append([])
        ws.append(["제품명", p_name, "", "제조번호(LOT)", self.stab_entries["제조번호(LOT)"].get(), ""])
        ws.append(["시험시작일", self.stab_entries["시험시작일"].get(), "", "시험담당자", self.stab_entries["시험담당자"].get(), ""])
        ws.append(["종합평가", self.stab_entries["종합평가"].get(), "", "", "", ""])
        ws.append([])
        ws.append(["보관 조건", "초기 (0주)", "2주 경과", "4주 경과", "8주 경과", "12주 경과"])
        for cond, entries in self.stab_matrix_entries.items():
            row = [cond] + [e.get().strip() for e in entries]
            ws.append(row)

        file_path = fd.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], initialfile=f"안정성보고서_{p_name}.xlsx", title="안정성보고서 엑셀 저장")
        if file_path:
            wb.save(file_path)
            messagebox.showinfo("완료", f"안정성보고서가 저장되었습니다:\n{file_path}", parent=self)
            try: os.startfile(os.path.abspath(file_path))
            except: pass

    # =========================================================================
    # 5. 내용물-용기 상용성(적합성) 시험보고서 (Packaging Compatibility Report)
    # =========================================================================
    def setup_packaging_compatibility_tab(self, tab_frame):
        tab_frame.grid_columnconfigure(0, weight=1)
        tab_frame.grid_rowconfigure(0, weight=1)

        scroll = ctk.CTkScrollableFrame(tab_frame, label_text="용기-내용물 상용성(적합성) 시험보고서 (Packaging Compatibility)")
        scroll.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        scroll.grid_columnconfigure(0, weight=1)

        self.comp_entries = {}
        self.current_comp_id = None

        top_bar = ctk.CTkFrame(scroll, fg_color="transparent")
        top_bar.grid(row=0, column=0, sticky="ew", padx=10, pady=(5, 15))
        top_bar.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(top_bar, text="상용성 이력:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=5, pady=5)
        self.comp_picker = ctk.CTkComboBox(top_bar, values=["-- 저장된 상용성보고서 선택 --"], width=300)
        self.comp_picker.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ctk.CTkButton(top_bar, text="불러오기", width=80, command=self.load_selected_comp).grid(row=0, column=2, padx=5, pady=5)
        ctk.CTkButton(top_bar, text="새로작성", width=80, fg_color="gray50", command=self.clear_comp_form).grid(row=0, column=3, padx=5, pady=5)
        ctk.CTkButton(top_bar, text="삭제", width=70, fg_color="#D32F2F", command=self.delete_comp).grid(row=0, column=4, padx=5, pady=5)

        info_frame = ctk.CTkFrame(scroll)
        info_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=5)
        info_frame.grid_columnconfigure((1, 3, 5), weight=1)

        fields = [
            ("제품명", 0, 0), ("LAB NO", 0, 2), ("의뢰업체", 0, 4),
            ("용기명", 1, 0), ("용기재질", 1, 2), ("캡/펌프사양", 1, 4),
            ("코팅/인쇄", 2, 0), ("시험담당자", 2, 2), ("종합판정", 2, 4)
        ]

        for label_text, r, c in fields:
            ctk.CTkLabel(info_frame, text=label_text, font=ctk.CTkFont(weight="bold")).grid(row=r, column=c, padx=8, pady=5, sticky="w")
            if label_text == "종합판정":
                cb = ctk.CTkComboBox(info_frame, values=["적합 (Pass)", "부적합 (Fail)", "조건부 적합"], width=160)
                cb.set("적합 (Pass)")
                cb.grid(row=r, column=c+1, padx=8, pady=5, sticky="ew")
                self.comp_entries[label_text] = cb
            elif label_text == "용기재질":
                cb = ctk.CTkComboBox(info_frame, values=["PET", "PP", "PE", "HDPE", "아크릴", "유리", "알루미늄", "복합튜브"], width=160)
                cb.set("PET")
                cb.grid(row=r, column=c+1, padx=8, pady=5, sticky="ew")
                self.comp_entries[label_text] = cb
            else:
                entry = ctk.CTkEntry(info_frame)
                entry.grid(row=r, column=c+1, padx=8, pady=5, sticky="ew")
                self.comp_entries[label_text] = entry

        # 핵심 평가 6대 검사 항목
        eval_frame = ctk.CTkFrame(scroll)
        eval_frame.grid(row=2, column=0, sticky="ew", padx=10, pady=10)
        eval_frame.grid_columnconfigure((1, 3), weight=1)

        eval_fields = [
            ("감량율 (45℃ 4주, %)", 0, 0, "0.32% (기준 1.5% 이하 적합)"),
            ("용기외관 변형여부", 0, 2, "함몰/팽창/크랙 없음 (정상)"),
            ("인쇄/코팅 박리여부", 1, 0, "박리 및 번짐 없음 (이상무)"),
            ("펌프/토출 작동성", 1, 2, "토출량 편차 없음 / 누액 없음"),
            ("진공감압 누액시험", 2, 0, "-0.08MPa 10분 누액 없음 (적합)"),
            ("내용물 변질여부", 2, 2, "변색/취변/분리 없음 (정상)")
        ]

        for label_text, r, c, default_val in eval_fields:
            ctk.CTkLabel(eval_frame, text=label_text, font=ctk.CTkFont(weight="bold")).grid(row=r, column=c, padx=8, pady=6, sticky="w")
            entry = ctk.CTkEntry(eval_frame)
            entry.insert(0, default_val)
            entry.grid(row=r, column=c+1, padx=8, pady=6, sticky="ew")
            self.comp_entries[label_text] = entry

        # 권고사항 및 비고
        note_frame = ctk.CTkFrame(scroll)
        note_frame.grid(row=3, column=0, sticky="ew", padx=10, pady=10)
        note_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(note_frame, text="개선 권고사항 및 종합 소견:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10, pady=5, sticky="nw")
        self.comp_recom_textbox = ctk.CTkTextbox(note_frame, height=70)
        self.comp_recom_textbox.grid(row=0, column=1, padx=10, pady=5, sticky="ew")
        self.comp_recom_textbox.insert("1.0", "내용물과 용기 간 감량율 및 토출 작동성이 매우 우수하여 양산 적합 판정함.")

        btn_bar = ctk.CTkFrame(scroll, fg_color="transparent")
        btn_bar.grid(row=4, column=0, sticky="e", padx=10, pady=15)
        ctk.CTkButton(btn_bar, text="💾 상용성보고서 DB 저장", width=150, command=self.save_comp_to_db).pack(side="left", padx=5)
        ctk.CTkButton(btn_bar, text="📊 상용성 엑셀 출력", width=150, fg_color="#2E7D32", hover_color="#1B5E20", command=self.export_comp_to_excel).pack(side="left", padx=5)

        self.refresh_comp_list()

    def refresh_comp_list(self):
        try:
            session = db_manager.get_session()
            recs = session.query(PackagingCompatibilityReport).order_by(PackagingCompatibilityReport.created_at.desc()).limit(50).all()
            vals = [f"{r.id} | {r.product_name} | 용기:{r.container_name or ''} | {r.overall_result}" for r in recs]
            self.comp_picker.configure(values=vals if vals else ["-- 저장된 상용성보고서 없음 --"])
            if vals: self.comp_picker.set(vals[0])
            session.close()
        except Exception as e:
            print(f"[경고] 상용성보고서 목록 로드 실패: {e}")

    def clear_comp_form(self):
        self.current_comp_id = None
        for k, e in self.comp_entries.items():
            if isinstance(e, ctk.CTkComboBox): e.set(e.cget("values")[0])
            else: e.delete(0, "end")
        self.comp_recom_textbox.delete("1.0", "end")

    def load_selected_comp(self):
        sel = self.comp_picker.get()
        if not sel or '|' not in sel: return
        rec_id = int(sel.split('|')[0].strip())
        session = db_manager.get_session()
        try:
            r = session.query(PackagingCompatibilityReport).get(rec_id)
            if not r: return
            self.current_comp_id = r.id
            self.comp_entries["제품명"].delete(0, "end"); self.comp_entries["제품명"].insert(0, r.product_name or "")
            self.comp_entries["LAB NO"].delete(0, "end"); self.comp_entries["LAB NO"].insert(0, r.lab_no or "")
            self.comp_entries["의뢰업체"].delete(0, "end"); self.comp_entries["의뢰업체"].insert(0, r.client_name or "")
            self.comp_entries["용기명"].delete(0, "end"); self.comp_entries["용기명"].insert(0, r.container_name or "")
            self.comp_entries["용기재질"].set(r.container_material or "PET")
            self.comp_entries["캡/펌프사양"].delete(0, "end"); self.comp_entries["캡/펌프사양"].insert(0, r.pump_specs or "")
            self.comp_entries["코팅/인쇄"].delete(0, "end"); self.comp_entries["코팅/인쇄"].insert(0, r.coating_printing or "")
            self.comp_entries["시험담당자"].delete(0, "end"); self.comp_entries["시험담당자"].insert(0, r.examiner or "")
            self.comp_entries["종합판정"].set(r.overall_result or "적합 (Pass)")
            
            self.comp_entries["감량율 (45℃ 4주, %)"].delete(0, "end"); self.comp_entries["감량율 (45℃ 4주, %)"].insert(0, r.weight_loss_rate or "")
            self.comp_entries["용기외관 변형여부"].delete(0, "end"); self.comp_entries["용기외관 변형여부"].insert(0, r.container_deformation or "")
            self.comp_entries["인쇄/코팅 박리여부"].delete(0, "end"); self.comp_entries["인쇄/코팅 박리여부"].insert(0, r.coating_peeling or "")
            self.comp_entries["펌프/토출 작동성"].delete(0, "end"); self.comp_entries["펌프/토출 작동성"].insert(0, r.pump_operability or "")
            self.comp_entries["진공감압 누액시험"].delete(0, "end"); self.comp_entries["진공감압 누액시험"].insert(0, r.leakage_test or "")
            self.comp_entries["내용물 변질여부"].delete(0, "end"); self.comp_entries["내용물 변질여부"].insert(0, r.contents_change or "")

            self.comp_recom_textbox.delete("1.0", "end"); self.comp_recom_textbox.insert("1.0", r.recommendations or "")
            messagebox.showinfo("불러오기 완료", f"상용성보고서 '{r.product_name}'를 불러왔습니다.", parent=self)
        finally:
            session.close()

    def save_comp_to_db(self):
        p_name = self.comp_entries["제품명"].get().strip()
        if not p_name:
            messagebox.showwarning("입력 필요", "제품명은 필수 항목입니다.", parent=self); return
        session = db_manager.get_session()
        try:
            r = session.query(PackagingCompatibilityReport).get(self.current_comp_id) if self.current_comp_id else PackagingCompatibilityReport()
            r.product_name = p_name
            r.lab_no = self.comp_entries["LAB NO"].get().strip()
            r.client_name = self.comp_entries["의뢰업체"].get().strip()
            r.container_name = self.comp_entries["용기명"].get().strip()
            r.container_material = self.comp_entries["용기재질"].get().strip()
            r.pump_specs = self.comp_entries["캡/펌프사양"].get().strip()
            r.coating_printing = self.comp_entries["코팅/인쇄"].get().strip()
            r.examiner = self.comp_entries["시험담당자"].get().strip()
            r.overall_result = self.comp_entries["종합판정"].get().strip()
            
            r.weight_loss_rate = self.comp_entries["감량율 (45℃ 4주, %)"].get().strip()
            r.container_deformation = self.comp_entries["용기외관 변형여부"].get().strip()
            r.coating_peeling = self.comp_entries["인쇄/코팅 박리여부"].get().strip()
            r.pump_operability = self.comp_entries["펌프/토출 작동성"].get().strip()
            r.leakage_test = self.comp_entries["진공감압 누액시험"].get().strip()
            r.contents_change = self.comp_entries["내용물 변질여부"].get().strip()
            r.recommendations = self.comp_recom_textbox.get("1.0", "end-1c").strip()

            if not self.current_comp_id: session.add(r)
            session.commit()
            self.current_comp_id = r.id
            self.refresh_comp_list()
            messagebox.showinfo("저장 완료", "용기-내용물 상용성 시험보고서가 DB에 성공적으로 저장되었습니다.", parent=self)
        except Exception as e:
            session.rollback(); messagebox.showerror("오류", f"저장 실패: {e}", parent=self)
        finally:
            session.close()

    def delete_comp(self):
        if not self.current_comp_id: return
        if not messagebox.askyesno("삭제 확인", "선택한 상용성보고서를 삭제하시겠습니까?", parent=self): return
        session = db_manager.get_session()
        try:
            r = session.query(PackagingCompatibilityReport).get(self.current_comp_id)
            if r:
                session.delete(r); session.commit()
                self.clear_comp_form(); self.refresh_comp_list()
                messagebox.showinfo("완료", "삭제되었습니다.", parent=self)
        finally:
            session.close()

    def export_comp_to_excel(self):
        p_name = self.comp_entries["제품명"].get().strip() or "상용성보고서"
        wb = Workbook(); ws = wb.active; ws.title = "용기상용성보고서"
        ws.merge_cells("A1:D2")
        ws["A1"] = f"용기-내용물 상용성(적합성) 시험보고서 - {p_name}"
        ws["A1"].font = Font(name="맑은 고딕", size=16, bold=True, color="1F497D")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.append([])
        for k, v in self.comp_entries.items():
            ws.append([k, v.get(), "", ""])

        file_path = fd.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], initialfile=f"용기상용성보고서_{p_name}.xlsx", title="상용성보고서 엑셀 저장")
        if file_path:
            wb.save(file_path)
            messagebox.showinfo("완료", f"용기 상용성보고서가 저장되었습니다:\n{file_path}", parent=self)
            try: os.startfile(os.path.abspath(file_path))
            except: pass



    # =========================================================================
    # 네비게이션 및 새로고침
    # =========================================================================
    def switch_to_tab(self, tab_name):
        try:
            if hasattr(self, 'tab_view') and hasattr(self.tab_view, '_segmented_button'):
                self.tab_view._segmented_button.grid_forget()
        except Exception:
            pass

        tab_key = tab_name
        tab_name_mapping = {
            "coa": self.texts.get('coa', 'COA (시험성적서)'),
            "msds": self.texts.get('msds', 'MSDS (물질안전보건자료)'),
            "ingredient_report": self.texts.get('ingredient_report', '원료목록보고'),
            "mat_inspection": self.texts.get('mat_inspection', '원료 입고검사'),
            "prod_standard": self.texts.get('prod_standard', '제품표준서'),
            "mfg_record": self.texts.get('mfg_record', '제조관리기록서'),
            "stability_test": self.texts.get('stability_test', '안정성 시험'),
            "compatibility_test": self.texts.get('compatibility_test', '용기 상용성 시험')
        }
        
        # 키 매핑 역탐색
        for k, v in tab_name_mapping.items():
            if tab_name in (k, v):
                tab_key = k
                break

        # 해당 탭이 아직 초기화되지 않았다면 즉시 빌드
        self.ensure_tab_initialized(tab_key)

        target_tab = tab_name_mapping.get(tab_key, tab_name)
        if hasattr(self, 'tab_view') and target_tab in self.tab_view._name_list:
            self.tab_view.set(target_tab)

    def refresh_data(self):
        """품질 관리 프레임의 전체 데이터를 새로고침합니다. (초기화된 탭만 선별 갱신)"""
        try:
            if 'coa' in self._initialized_tabs:
                if hasattr(self, 'refresh_semi_coa_list'): self.refresh_semi_coa_list()
                if hasattr(self, 'refresh_finished_coa_list'): self.refresh_finished_coa_list()
            if 'msds' in self._initialized_tabs and hasattr(self, 'refresh_msds_list'):
                self.refresh_msds_list()
            if 'ingredient_report' in self._initialized_tabs and hasattr(self, 'refresh_ingredient_report_list'):
                self.refresh_ingredient_report_list()
            if 'mat_inspection' in self._initialized_tabs and hasattr(self, 'refresh_mat_insp_list'):
                self.refresh_mat_insp_list()
            if 'prod_standard' in self._initialized_tabs and hasattr(self, 'refresh_prod_std_list'):
                self.refresh_prod_std_list()
            if 'mfg_record' in self._initialized_tabs and hasattr(self, 'refresh_bmr_list'):
                self.refresh_bmr_list()
            if 'stability_test' in self._initialized_tabs and hasattr(self, 'refresh_stab_list'):
                self.refresh_stab_list()
            if 'compatibility_test' in self._initialized_tabs and hasattr(self, 'refresh_comp_list'):
                self.refresh_comp_list()
        except Exception as e:
            print(f"[오류] 품질 관리 프레임 새로고침 실패: {e}")
